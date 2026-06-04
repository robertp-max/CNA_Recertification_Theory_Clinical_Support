"""rebuild_contentv2_from_json.py

Single-source ContentV2 rebuild pipeline.

The canonical source of truth is:
    CNA-Recert-Course/ContentV2/data/courseContentV2.json

Every other ContentV2 runtime/mirror artifact is REGENERATED from that JSON by
this script. Do NOT hand-edit the generated mirror artifacts:
    - CNA-Recert-Course/ContentV2/data/courseContentV2.ts
    - standalone-course-mvp/src/data/contentV2.generated.ts
    - CNA-Recert-Course/ContentV2/narration/narration_master.csv
    - CNA-Recert-Course/ContentV2/narration/tts_narration_import.csv
    - CNA-Recert-Course/ContentV2/narration/narration_metadata.json
    - CNA-Recert-Course/ContentV2/xlsx/CNA_RECERT_CONTENTV2_MASTER.xlsx
    - CNA-Recert-Course/ContentV2/03_MODULE_LESSON_4CARD_MAP.md
    - CNA-Recert-Course/ContentV2/07_MEDIA_PROMPT_PLACEHOLDERS.md
    - CNA-Recert-Course/ContentV2/08_SME_COMPLIANCE_REVIEW_FLAGS.md
    - CNA-Recert-Course/ContentV2/10_CONTENT_COVERAGE_AND_TIME_RECONCILIATION.md
    - CNA-Recert-Course/ContentV2/data/generation_validation_summary.json

To change content: edit courseContentV2.json, then rerun this script.

Usage:
    python CNA-Recert-Course/ContentV2/tools/rebuild_contentv2_from_json.py
    python .../rebuild_contentv2_from_json.py --check   # validate only, write nothing

Exit code is non-zero if any hard-fail guardrail is violated.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
THIS = Path(__file__).resolve()
ROOT = THIS.parents[3]  # tools -> ContentV2 -> CNA-Recert-Course -> repo root
CV2 = ROOT / "CNA-Recert-Course" / "ContentV2"
DATA_DIR = CV2 / "data"
NARR_DIR = CV2 / "narration"
XLSX_DIR = CV2 / "xlsx"
APP_DATA_DIR = ROOT / "standalone-course-mvp" / "src" / "data"

CANON_JSON = DATA_DIR / "courseContentV2.json"
SCHEMA_JSON = DATA_DIR / "courseContentV2.schema.json"
CANON_TS = DATA_DIR / "courseContentV2.ts"
APP_GENERATED_TS = APP_DATA_DIR / "contentV2.generated.ts"
NARR_MASTER_CSV = NARR_DIR / "narration_master.csv"
TTS_IMPORT_CSV = NARR_DIR / "tts_narration_import.csv"
NARR_META_JSON = NARR_DIR / "narration_metadata.json"
XLSX_PATH = XLSX_DIR / "CNA_RECERT_CONTENTV2_MASTER.xlsx"
MAP_4CARD_MD = CV2 / "03_MODULE_LESSON_4CARD_MAP.md"
MEDIA_MD = CV2 / "07_MEDIA_PROMPT_PLACEHOLDERS.md"
FLAGS_MD = CV2 / "08_SME_COMPLIANCE_REVIEW_FLAGS.md"
COVERAGE_MD = CV2 / "10_CONTENT_COVERAGE_AND_TIME_RECONCILIATION.md"
VALIDATION_SUMMARY_JSON = DATA_DIR / "generation_validation_summary.json"

PROJECT_NAME = "CNA Recertification Theory + Clinical Support"

# Learner-facing strings that must never leak. This mirrors the app guardrail
# test exactly. NOTE: the protective sentence "Final exams never reveal answer
# keys" is compliant and is deliberately NOT in this set.
FORBIDDEN_LEARNER_PHRASES = [
    "Correct internal answer",
    "Rationale for instructor/internal use",
    "Correct. Continue to the next item.",
    "Review the related lesson and try again.",
    "correct_id_internal",
    "rationale_internal",
]

# Source-document scaffolding and embedded check-activity answer-key phrases that
# must never appear inside a learner-facing DELIVERY card. The source transformation
# (author_delivery_from_source.py) strips these; this guardrail prevents regression.
DELIVERY_SCAFFOLD_PHRASES = [
    "On-Screen Text:",
    "TTS/Transcript-Ready Text:",
    "Pronunciation Notes:",
    "Correct Answer:",
    "Correct Feedback:",
    "Incorrect Feedback",
    "Feedback (all correct)",
    "Feedback (any incorrect)",
]

# Internal-only fields that may carry scoring/answer data but must never be
# placed inside a learner-facing text field.
INTERNAL_ONLY_FIELDS = {"correct_id_internal", "rationale_internal"}

REQUIRED_CARD_FIELDS = [
    "card_id",
    "card_type",
    "app",
    "narration_script",
    "transcript_text",
    "estimated_narration_seconds",
    "media_prompt_placeholder",
    "source_reference",
    "sme_review_flag",
    "compliance_review_flag",
]


class HardFail(Exception):
    """Raised when a guardrail that must block the build is violated."""


# ---------------------------------------------------------------------------
# Load + validate
# ---------------------------------------------------------------------------
def load_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def schema_validate(data: dict, errors: list[str]) -> None:
    try:
        import jsonschema  # type: ignore
    except Exception:  # pragma: no cover - fallback when lib unavailable
        errors.append("[warn] jsonschema not installed; skipped formal schema validation.")
        return
    schema = load_json(SCHEMA_JSON)
    validator = jsonschema.Draft202012Validator(schema)
    for err in sorted(validator.iter_errors(data), key=lambda e: list(e.path)):
        loc = "/".join(str(p) for p in err.path) or "(root)"
        errors.append(f"[schema] {loc}: {err.message}")


def iter_cards(data: dict):
    for module in data.get("modules", []):
        for lesson in module.get("lessons", []):
            for card in lesson.get("cards", []):
                yield module, lesson, card


def validate_required_fields(data: dict, errors: list[str]) -> None:
    if not isinstance(data.get("modules"), list) or not data["modules"]:
        errors.append("[required] modules array is missing or empty.")
        return
    for module in data["modules"]:
        mid = module.get("module_id", "?")
        for key in ("module_id", "module_title", "estimated_minutes", "lessons"):
            if key not in module:
                errors.append(f"[required] module {mid} missing '{key}'.")
        for lesson in module.get("lessons", []):
            lid = lesson.get("lesson_id", "?")
            for key in ("lesson_id", "lesson_title", "cards"):
                if key not in lesson:
                    errors.append(f"[required] {mid}/{lid} missing '{key}'.")
            for card in lesson.get("cards", []):
                cid = card.get("card_id", "?")
                for key in REQUIRED_CARD_FIELDS:
                    if key not in card:
                        errors.append(f"[required] {mid}/{lid}/{cid} missing '{key}'.")
                app = card.get("app")
                if not isinstance(app, dict) or not app.get("location"):
                    errors.append(f"[required] {mid}/{lid}/{cid} missing app.location.")


# ---------------------------------------------------------------------------
# Derived metrics + guardrail checks
# ---------------------------------------------------------------------------
def card_location(card: dict) -> str:
    return (card.get("app") or {}).get("location", "")


def primary_locations(data: dict) -> list[str]:
    """All locations expected to be globally unique (excludes media placeholders,
    which intentionally reuse their parent card location)."""
    locs: list[str] = []
    for _m, _l, card in iter_cards(data):
        locs.append(card_location(card))
    ma = data.get("assessments", {}).get("module_assessments", {})
    for _mid, block in ma.items():
        if block.get("splash_app_location"):
            locs.append(block["splash_app_location"])
        for q in block.get("questions", []):
            if q.get("app_location"):
                locs.append(q["app_location"])
    fa = data.get("assessments", {}).get("final_assessment", {})
    for key in ("splash_app_location", "result_pass_app_location", "result_fail_app_location"):
        if fa.get(key):
            locs.append(fa[key])
    for q in fa.get("questions", []):
        if q.get("app_location"):
            locs.append(q["app_location"])
    cs = data.get("clinical_support", {})
    if cs.get("overview_app_location"):
        locs.append(cs["overview_app_location"])
    for unit in cs.get("units", []):
        if unit.get("app_location"):
            locs.append(unit["app_location"])
    return [l for l in locs if l]


def is_source_repair_card(card: dict) -> bool:
    text = (card.get("display_title", "") + " " + card.get("learner_facing_content", "")).lower()
    return "source repair" in text


def learner_text_blob(card: dict) -> str:
    return "\n".join(
        str(card.get(k, ""))
        for k in ("learner_facing_content", "narration_script", "transcript_text", "display_title", "learning_goal")
    )


def compute_metrics(data: dict) -> dict:
    modules = data.get("modules", [])
    lessons = [l for m in modules for l in m.get("lessons", [])]
    cards = [c for _m, _l, c in iter_cards(data)]

    narration_cards = [c for c in cards if (c.get("narration_script") or "").strip()]
    total_narr_seconds = sum(int(c.get("estimated_narration_seconds") or 0) for c in narration_cards)
    clips_over_75 = sum(1 for c in narration_cards if int(c.get("estimated_narration_seconds") or 0) > 75)

    locs = primary_locations(data)
    dup = [loc for loc, n in Counter(locs).items() if n > 1]

    source_repair = sum(1 for c in cards if is_source_repair_card(c))
    sme_flags = sum(
        1 for c in cards if c.get("sme_review_flag") and c.get("sme_review_flag") != "None identified"
    )
    compliance_flags = sum(
        1 for c in cards if c.get("compliance_review_flag") and c.get("compliance_review_flag") != "None identified"
    )
    placeholder_media = sum(
        1 for c in cards if (c.get("media_prompt_placeholder") or {}).get("media_status") == "placeholder-pending"
    )

    # Explicit time model (inherited from ContentV1 syllabus; NOT recomputed from
    # narration/card count). Source-repair status does NOT remove a module from the
    # required 720 allocation - it is a content-depth flag, not a removal.
    required_theory_minutes = sum(
        int(m.get("instructional_minutes") or 0)
        for m in modules
        if m.get("counts_toward_720_instructional_minutes") is True
    )
    instructional_minutes = required_theory_minutes  # back-compat alias
    module_assessment_minutes = sum(
        int(b.get("estimated_assessment_minutes", b.get("estimated_minutes") or 0) or 0)
        for b in data.get("assessments", {}).get("module_assessments", {}).values()
    )
    final_assessment_minutes = int(
        data.get("assessments", {}).get("final_assessment", {}).get("estimated_assessment_minutes")
        or data.get("assessments", {}).get("final_assessment", {}).get("estimated_minutes")
        or 0
    )
    # Sum of separately-tracked descriptive component minutes (theory modules only).
    narration_model_minutes = round(
        sum(float(m.get("estimated_narration_minutes") or 0) for m in modules), 1
    )
    # Content-depth gap: allocation minus what lesson cards currently allocate.
    depth_gap_minutes = 0
    for m in modules:
        if m.get("counts_toward_720_instructional_minutes") is True:
            lesson_alloc = sum(int(l.get("instructional_minutes", l.get("estimated_minutes") or 0) or 0) for l in m.get("lessons", []))
            depth_gap_minutes += max(0, int(m.get("instructional_minutes") or 0) - lesson_alloc)

    return {
        "modules_count": len(modules),
        "lessons_count": len(lessons),
        "cards_count": len(cards),
        "narration_clip_count": len(narration_cards),
        "total_narration_seconds": total_narr_seconds,
        "total_narration_minutes": round(total_narr_seconds / 60.0, 2),
        "clips_over_75_seconds": clips_over_75,
        "unique_app_location_count": len(set(locs)),
        "total_app_location_count": len(locs),
        "duplicate_app_locations": dup,
        "source_repair_flag_count": source_repair,
        "sme_review_flag_count": sme_flags,
        "compliance_flag_count": compliance_flags,
        "placeholder_media_count": placeholder_media,
        "instructional_minutes_total": instructional_minutes,
        "required_theory_instructional_minutes_total": required_theory_minutes,
        "module_assessment_minutes_excluded": module_assessment_minutes,
        "course_final_assessment_minutes_excluded": final_assessment_minutes,
        "narration_model_minutes": narration_model_minutes,
        "content_depth_gap_minutes": depth_gap_minutes,
    }


def run_guardrails(data: dict, metrics: dict) -> tuple[list[str], list[str]]:
    """Returns (hard_failures, warnings)."""
    hard: list[str] = []
    warn: list[str] = []

    # 1. Duplicate app.location.
    if metrics["duplicate_app_locations"]:
        hard.append(f"Duplicate app.location values: {metrics['duplicate_app_locations'][:10]}")

    # --- Time-model guardrails -------------------------------------------------
    modules = data.get("modules", [])
    TIME_MODEL_FIELDS = [
        "instructional_minutes",
        "estimated_narration_minutes",
        "estimated_reading_minutes",
        "estimated_interaction_minutes",
        "estimated_assessment_minutes",
        "assessment_minutes_excluded_from_instructional_total",
        "counts_toward_720_instructional_minutes",
        "counts_toward_certificate_gate",
        "counts_toward_optional_clinical_support",
        "source_time_basis",
        "time_model_status",
        "time_model_notes",
    ]
    # T-5. Every module must carry explicit time-model fields.
    for m in modules:
        missing = [f for f in TIME_MODEL_FIELDS if f not in m]
        if missing:
            hard.append(f"Module {m.get('module_id')} lacks explicit time-model field(s): {missing}")
    # T-1. Required theory total must equal exactly 720.
    if metrics["required_theory_instructional_minutes_total"] != 720:
        hard.append(
            f"Required theory instructional total is {metrics['required_theory_instructional_minutes_total']}, must equal 720."
        )
    # T-2. Optional clinical support must never contribute to 720.
    cs = data.get("clinical_support", {})
    if cs.get("counts_toward_720_instructional_minutes") is True or cs.get("counts_toward_optional_clinical_support") is False:
        hard.append("Optional Clinical Support must be flagged non-720 / optional in the time model.")
    if int(cs.get("instructional_minutes") or 0) != 0:
        hard.append("Optional Clinical Support must have 0 instructional_minutes (never counts toward 720).")
    # T-3. Assessment-only time must never be counted as instructional.
    assess_blocks = list(data.get("assessments", {}).get("module_assessments", {}).values())
    fa_block = data.get("assessments", {}).get("final_assessment", {})
    for b in assess_blocks + [fa_block]:
        if not b:
            continue
        if b.get("counts_toward_720_instructional_minutes") is True:
            hard.append(f"Assessment block '{b.get('title', b.get('module_id', '?'))}' must not count toward 720.")
        if b.get("assessment_minutes_excluded_from_instructional_total") is not True:
            hard.append(f"Assessment block '{b.get('title', b.get('module_id', '?'))}' must mark assessment minutes excluded.")
    # T-4. Narration minutes must not be treated as full instructional time.
    #      Enforced at module level (robust): module narration must be strictly less
    #      than the module's syllabus instructional allocation. Per-lesson equality is
    #      surfaced as a warning (flags under-allocated legacy lesson minutes).
    for m in modules:
        if m.get("counts_toward_720_instructional_minutes") is not True:
            continue
        m_inst = float(m.get("instructional_minutes") or 0)
        m_narr = float(m.get("estimated_narration_minutes") or 0)
        if m_inst > 0 and m_narr >= m_inst:
            hard.append(
                f"Module {m.get('module_id')} treats narration ({m_narr}m) as the full instructional allocation ({m_inst}m)."
            )
        for l in m.get("lessons", []):
            inst = float(l.get("instructional_minutes", l.get("estimated_minutes") or 0) or 0)
            narr = float(l.get("estimated_narration_minutes") or 0)
            if inst > 0 and narr >= inst:
                warn.append(
                    f"Lesson {m.get('module_id')}/{l.get('lesson_id')} narration ({narr}m) >= its allocation ({inst}m); "
                    "lesson minutes look under-allocated vs ContentV1 (review/expand)."
                )

    # 5. Final-assessment learner-facing answer keys must not be exposed.
    fa = data.get("assessments", {}).get("final_assessment", {})
    learner_final_fields = []
    for key in ("instructions", "pass_copy", "fail_copy", "answer_key_policy"):
        val = fa.get(key)
        if isinstance(val, str):
            learner_final_fields.append(val)
    rmap = fa.get("remediation_map")
    if isinstance(rmap, dict):
        learner_final_fields.extend(str(v) for v in rmap.values())
    elif isinstance(rmap, list):
        learner_final_fields.extend(str(v) for v in rmap)
    for q in fa.get("questions", []):
        for fld in ("learner_feedback_correct", "learner_feedback_incorrect"):
            v = q.get(fld)
            if isinstance(v, str):
                # learner feedback must not reveal the correct option id/letter or rationale
                if q.get("correct_id_internal") and re.search(rf"\b{re.escape(q['correct_id_internal'])}\b\s*\)?\s*is correct", v, re.I):
                    hard.append(f"Final Q {q.get('id')} learner feedback reveals correct option.")
                if q.get("rationale_internal") and q["rationale_internal"][:30] and q["rationale_internal"][:30] in v:
                    hard.append(f"Final Q {q.get('id')} learner feedback leaks internal rationale.")
    final_blob = "\n".join(learner_final_fields).lower()
    for bad in ("correct_id_internal", "rationale_internal"):
        if bad in final_blob:
            hard.append(f"Final assessment learner-facing copy leaks internal field '{bad}'.")
    policy = (fa.get("answer_key_policy") or "").lower()
    if "not revealed" not in policy and "internal only" not in policy:
        hard.append("Final assessment answer_key_policy must assert answers are not revealed.")

    # 6. Optional Clinical Support must remain optional / non-credit / non-gating.
    cs = data.get("clinical_support", {})
    if cs.get("counts_toward_theory") or cs.get("certificate_gating") or cs.get("required"):
        hard.append("Optional Clinical Support must not be required/credit/certificate-gating.")
    guardrails = " ".join(data.get("compliance_guardrails", [])).lower()
    if "optional clinical support is optional" not in guardrails:
        warn.append("compliance_guardrails no longer states clinical support optional/non-credit/non-gating.")

    # 7. Production certificate must remain disabled.
    cert = data.get("app_copy", {}).get("certificate", {})
    if cert.get("production_enabled") is True or data.get("certificate_production_enabled") is True:
        hard.append("Production certificate issuance must remain disabled.")
    if "certificate production remains disabled" not in guardrails:
        warn.append("compliance_guardrails no longer states certificate production disabled.")

    # 8. ContentV1 source gaps must remain visible (not hidden). A source-repair
    #    card is acceptable as long as it still carries a visible review flag
    #    (SME or compliance) or the card/module status declares the gap. It only
    #    fails if every flag has been cleared back to default, hiding the gap.
    for module, _l, card in iter_cards(data):
        if not is_source_repair_card(card):
            continue
        sme = (card.get("sme_review_flag") or "").strip().lower()
        comp = (card.get("compliance_review_flag") or "").strip().lower()
        status = (card.get("status") or "").strip().lower()
        mod_src = (module.get("source_status_flag") or "").strip().lower()
        visible = any(
            v and v != "none identified"
            for v in (sme, comp)
        ) or "source repair" in status or "source repair" in mod_src or "draft" in status
        if not visible:
            hard.append(
                f"Card {card_location(card)} reads as source-repair but all review flags were cleared (hidden gap)."
            )

    # Learner-facing internal-answer language (Task 3): reported as warnings so the
    # pipeline can run on pre-cleanup data; the strict app test enforces the subset.
    leak_cards = []
    for _m, _l, card in iter_cards(data):
        blob = learner_text_blob(card).lower()
        for phrase in FORBIDDEN_LEARNER_PHRASES:
            if phrase.lower() in blob:
                leak_cards.append((card_location(card), phrase))
                break
    if leak_cards:
        warn.append(
            f"{len(leak_cards)} learner-facing card(s) still contain answer-key-style phrasing (Task 3 cleanup): "
            + ", ".join(f"{loc} ~ '{p}'" for loc, p in leak_cards[:8])
            + (" ..." if len(leak_cards) > 8 else "")
        )

    # Delivery cards must not leak source scaffolding or embedded check answer keys.
    # Hard-fail for the source-rebuilt modules (M01-M06); surface remaining legacy
    # modules (M00 orientation, M07 review/affidavit) as warnings for follow-up.
    REBUILT_MODULES = {"M01", "M02", "M03", "M04", "M05", "M06"}
    scaffold_legacy = []
    for module, _l, card in iter_cards(data):
        if card.get("card_type") != "delivery":
            continue
        blob = "\n".join(
            str(card.get(f) or "") for f in ("learner_facing_content", "narration_script", "transcript_text")
        )
        hit = next((p for p in DELIVERY_SCAFFOLD_PHRASES if p in blob), None)
        if not hit:
            continue
        if module.get("module_id") in REBUILT_MODULES:
            hard.append(
                f"Delivery card {card_location(card)} leaks source scaffolding / answer-key phrase '{hit}'."
            )
        else:
            scaffold_legacy.append((card_location(card), hit))
    if scaffold_legacy:
        warn.append(
            f"{len(scaffold_legacy)} legacy delivery card(s) (M00/M07, out of M02-M06 scope) still contain source "
            f"scaffolding labels; flagged for a later source rebuild: "
            + ", ".join(f"{loc} ~ '{p}'" for loc, p in scaffold_legacy[:6])
            + (" ..." if len(scaffold_legacy) > 6 else "")
        )

    return hard, warn


# ---------------------------------------------------------------------------
# Emit: TS mirrors
# ---------------------------------------------------------------------------
def ts_body(data: dict) -> str:
    return json.dumps(data, indent=2, ensure_ascii=False)


def write_ts_mirror(path: Path, header: str, data: dict) -> None:
    body = ts_body(data)
    path.write_text(
        f"{header}\nexport const courseContentV2 = {body} as const;\nexport default courseContentV2;\n",
        encoding="utf-8",
    )


def verify_ts_matches_json(path: Path, data: dict, label: str, errors: list[str]) -> None:
    text = path.read_text(encoding="utf-8")
    m = re.search(r"export const courseContentV2 = (.*) as const;", text, re.DOTALL)
    if not m:
        errors.append(f"[verify] {label}: could not locate exported object.")
        return
    try:
        parsed = json.loads(m.group(1))
    except json.JSONDecodeError as exc:
        errors.append(f"[verify] {label}: emitted object is not valid JSON ({exc}).")
        return
    if parsed != data:
        errors.append(f"[verify] {label}: generated TS does not match canonical JSON.")


# ---------------------------------------------------------------------------
# Emit: narration artifacts
# ---------------------------------------------------------------------------
def card_title(module: dict, lesson: dict, card: dict) -> str:
    cid = card.get("card_id", "")
    parts = cid.split("_")
    code = parts[0] if parts else cid
    rest = " ".join(p.capitalize() for p in parts[1:]) if len(parts) > 1 else card.get("card_type", "").capitalize()
    return f"{module.get('module_id','')} {lesson.get('lesson_id','')} {code} {rest}".strip()


def narration_rows(data: dict) -> list[dict]:
    rows = []
    for module, lesson, card in iter_cards(data):
        if not (card.get("narration_script") or "").strip():
            continue
        rows.append(
            {
                "Project": PROJECT_NAME,
                "Module ID": module.get("module_id", ""),
                "Module Title": module.get("module_title", ""),
                "Lesson ID": lesson.get("lesson_id", ""),
                "Lesson Title": lesson.get("lesson_title", ""),
                "Card ID": card.get("card_id", ""),
                "Card Type": card.get("card_type", ""),
                "Title": card_title(module, lesson, card),
                "app.location": card_location(card),
                "Narration": card.get("narration_script", ""),
                "Transcript": card.get("transcript_text", ""),
                "Estimated Seconds": int(card.get("estimated_narration_seconds") or 0),
                "Estimated Word Count": int(card.get("estimated_word_count") or 0),
                "Target Duration Status": card.get("target_duration_status", ""),
                "Voice Notes": "Neutral instructional tone; no PHI; planning transcript only (no audio authorized).",
                "Source Reference": card.get("source_reference", ""),
                "SME Review Flag": card.get("sme_review_flag", ""),
                "Compliance Review Flag": card.get("compliance_review_flag", ""),
                "Status": card.get("status", ""),
            }
        )
    return rows


def write_narration_csvs(rows: list[dict]) -> None:
    master_cols = [
        "Project", "Module ID", "Module Title", "Lesson ID", "Lesson Title", "Card ID",
        "Card Type", "Title", "app.location", "Narration", "Transcript", "Estimated Seconds",
        "Estimated Word Count", "Target Duration Status", "Voice Notes", "Source Reference",
        "SME Review Flag", "Compliance Review Flag", "Status",
    ]
    with NARR_MASTER_CSV.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=master_cols)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

    tts_cols = ["Project", "Title", "app.location", "Narration"]
    with TTS_IMPORT_CSV.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=tts_cols, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row[k] for k in tts_cols})


def write_narration_metadata(data: dict, metrics: dict, rows: list[dict]) -> None:
    modules_covered = sorted({r["Module ID"] for r in rows})
    meta = {
        "generated_at": now_iso(),
        "source_folder": data.get("source_folder", ""),
        "generator": "CNA-Recert-Course/ContentV2/tools/rebuild_contentv2_from_json.py",
        "total_clips": metrics["narration_clip_count"],
        "total_estimated_seconds": metrics["total_narration_seconds"],
        "total_estimated_minutes": metrics["total_narration_minutes"],
        "clips_over_target": metrics["clips_over_75_seconds"],
        "unique_app_locations": metrics["unique_app_location_count"],
        "modules_covered": modules_covered,
        "audio_production": "Not authorized. Transcripts/manifest are planning artifacts only; no audio generated.",
        "validation_status": "Regenerated from canonical courseContentV2.json by single-source pipeline.",
    }
    NARR_META_JSON.write_text(json.dumps(meta, indent=2, ensure_ascii=False), encoding="utf-8")


# ---------------------------------------------------------------------------
# Emit: XLSX
# ---------------------------------------------------------------------------
def write_xlsx(data: dict, metrics: dict, rows: list[dict]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    ws_mod = wb.active
    ws_mod.title = "Modules"
    ws_mod.append([
        "Module ID", "Title", "Counts toward 720", "Instructional min",
        "Narration min", "Reading min", "Interaction min", "Assessment min (excl.)",
        "Lessons", "Cards", "Time-model status", "SME Flag", "Compliance Flag",
    ])
    for m in data.get("modules", []):
        lessons = m.get("lessons", [])
        ncards = sum(len(l.get("cards", [])) for l in lessons)
        ws_mod.append([
            m.get("module_id", ""), m.get("module_title", ""),
            m.get("counts_toward_720_instructional_minutes") is True,
            m.get("instructional_minutes", m.get("estimated_minutes", "")),
            m.get("estimated_narration_minutes", ""), m.get("estimated_reading_minutes", ""),
            m.get("estimated_interaction_minutes", ""), m.get("estimated_assessment_minutes", ""),
            len(lessons), ncards, m.get("time_model_status", ""),
            m.get("sme_review_flag", ""), m.get("compliance_review_flag", ""),
        ])

    ws_cards = wb.create_sheet("Cards")
    ws_cards.append([
        "Module ID", "Lesson ID", "Card ID", "Card Type", "Title", "app.location",
        "Est. narration s", "Word count", "Source Reference", "SME Flag", "Compliance Flag", "Status",
    ])
    for module, lesson, card in iter_cards(data):
        ws_cards.append([
            module.get("module_id", ""), lesson.get("lesson_id", ""), card.get("card_id", ""),
            card.get("card_type", ""), card.get("display_title", ""), card_location(card),
            card.get("estimated_narration_seconds", 0), card.get("estimated_word_count", 0),
            card.get("source_reference", ""), card.get("sme_review_flag", ""),
            card.get("compliance_review_flag", ""), card.get("status", ""),
        ])

    ws_narr = wb.create_sheet("Narration")
    narr_cols = ["Module ID", "Lesson ID", "Card ID", "Title", "app.location", "Estimated Seconds", "Narration", "Transcript"]
    ws_narr.append(narr_cols)
    for r in rows:
        ws_narr.append([r["Module ID"], r["Lesson ID"], r["Card ID"], r["Title"], r["app.location"], r["Estimated Seconds"], r["Narration"], r["Transcript"]])

    ws_media = wb.create_sheet("Media Placeholders")
    ws_media.append(["app.location", "Scene Title", "Media Status", "Required for MVP", "Asset Path", "Alt Text", "PHI Safety Note"])
    for _m, _l, card in iter_cards(data):
        mp = card.get("media_prompt_placeholder") or {}
        ws_media.append([
            mp.get("app_location", card_location(card)), mp.get("scene_title", ""), mp.get("media_status", ""),
            mp.get("required_for_mvp", False), mp.get("asset_path", ""), mp.get("alt_text", ""), mp.get("phi_safety_note", ""),
        ])

    ws_sum = wb.create_sheet("Runtime Summary")
    ws_sum.append(["Metric", "Value"])
    for key in [
        "modules_count", "lessons_count", "cards_count", "narration_clip_count",
        "total_narration_seconds", "total_narration_minutes", "clips_over_75_seconds",
        "unique_app_location_count", "source_repair_flag_count", "sme_review_flag_count",
        "compliance_flag_count", "placeholder_media_count",
        "required_theory_instructional_minutes_total", "module_assessment_minutes_excluded",
        "course_final_assessment_minutes_excluded", "narration_model_minutes",
        "content_depth_gap_minutes",
    ]:
        ws_sum.append([key, metrics[key]])
    ws_sum.append(["generated_at", now_iso()])
    ws_sum.append(["audio_production", "Not authorized - transcripts/manifest only"])

    header_fill = PatternFill("solid", fgColor="DCE6F1")
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        if ws.max_row >= 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    XLSX_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_PATH)


# ---------------------------------------------------------------------------
# Emit: markdown maps
# ---------------------------------------------------------------------------
def now_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S+00:00")


def md_header(title: str) -> str:
    return (
        f"# {title}\n\n"
        f"> Generated from `CNA-Recert-Course/ContentV2/data/courseContentV2.json` by "
        f"`tools/rebuild_contentv2_from_json.py` on {now_iso()}. Do not hand-edit.\n\n"
    )


def write_map_4card(data: dict) -> None:
    lines = [md_header("03_MODULE_LESSON_4CARD_MAP — Module / Lesson / Card Map")]
    for m in data.get("modules", []):
        lines.append(
            f"## {m.get('module_id')} — {m.get('module_title')} "
            f"({m.get('instructional_minutes', m.get('estimated_minutes'))} min, "
            f"counts_toward_720={bool(m.get('counts_toward_720_instructional_minutes', m.get('counts_toward_theory')))})\n"
        )
        for l in m.get("lessons", []):
            cards = l.get("cards", [])
            lines.append(f"### {l.get('lesson_id')} — {l.get('lesson_title')} ({l.get('estimated_minutes')} min, {len(cards)} cards)\n")
            lines.append("| Card ID | Type | Title | Est. s | Source Ref | SME Flag |")
            lines.append("|---|---|---|---:|---|---|")
            for c in cards:
                lines.append(
                    f"| {c.get('card_id')} | {c.get('card_type')} | {sanitize_cell(c.get('display_title',''))} | "
                    f"{c.get('estimated_narration_seconds',0)} | {sanitize_cell(c.get('source_reference',''))} | "
                    f"{sanitize_cell(c.get('sme_review_flag',''))} |"
                )
            lines.append("")
    MAP_4CARD_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_media_md(data: dict) -> None:
    lines = [md_header("07_MEDIA_PROMPT_PLACEHOLDERS — Per-Card Media Placeholders")]
    lines.append(
        "No media assets are generated for the MVP. Every card carries a safe 16:9 placeholder with alt text "
        "and a PHI-safety note. No real charts, facility names, patient identifiers, gore, or unsafe actions.\n"
    )
    lines.append("| app.location | Scene Title | Media Status | Required for MVP | Asset Path |")
    lines.append("|---|---|---|---|---|")
    for _m, _l, card in iter_cards(data):
        mp = card.get("media_prompt_placeholder") or {}
        lines.append(
            f"| {mp.get('app_location', card_location(card))} | {sanitize_cell(mp.get('scene_title',''))} | "
            f"{mp.get('media_status','')} | {mp.get('required_for_mvp', False)} | {mp.get('asset_path','')} |"
        )
    MEDIA_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_flags_md(data: dict) -> None:
    lines = [md_header("08_SME_COMPLIANCE_REVIEW_FLAGS — Open Review Flags")]
    lines.append("## Module-level flags\n")
    lines.append("| Module | SME Review Flag | Compliance Review Flag | Source Status Flag |")
    lines.append("|---|---|---|---|")
    for m in data.get("modules", []):
        lines.append(
            f"| {m.get('module_id')} | {sanitize_cell(m.get('sme_review_flag',''))} | "
            f"{sanitize_cell(m.get('compliance_review_flag',''))} | {sanitize_cell(m.get('source_status_flag',''))} |"
        )
    lines.append("\n## Card-level flags (non-default only)\n")
    lines.append("| app.location | SME Review Flag | Compliance Review Flag | Source-Repair? |")
    lines.append("|---|---|---|---|")
    for _m, _l, card in iter_cards(data):
        sme = card.get("sme_review_flag", "")
        comp = card.get("compliance_review_flag", "")
        repair = is_source_repair_card(card)
        if (sme and sme != "None identified") or repair:
            lines.append(
                f"| {card_location(card)} | {sanitize_cell(sme)} | {sanitize_cell(comp)} | {'YES' if repair else ''} |"
            )
    lines.append("\n## Compliance guardrails (control facts)\n")
    for g in data.get("compliance_guardrails", []):
        lines.append(f"- {g}")
    FLAGS_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_coverage_md(data: dict, metrics: dict) -> None:
    tm = data.get("time_model", {})
    lines = [md_header("10_CONTENT_COVERAGE_AND_TIME_RECONCILIATION — Coverage & Time")]
    lines.append(
        "Instructional minutes are inherited from the ContentV1 syllabus "
        "(`02_THEORY_SYLLABUS_TABLE.md`). The 720-minute total is **not** recomputed from narration length or card "
        "count. Narration, reading/review, interaction, and assessment minutes are tracked **separately** and never "
        "replace the authoritative allocation. Optional Clinical Support and all assessment time are excluded.\n"
    )
    lines.append("## Required theory instructional allocation (counts toward 720)\n")
    lines.append("| Module | Counts→720 | Instructional min | Lesson-allocated min | Depth gap | Narration min | Reading min | Interaction min | Assessment min (excl.) | Status |")
    lines.append("|---|:--:|---:|---:|---:|---:|---:|---:|---:|---|")
    inst_total = 0
    gap_total = 0
    for m in data.get("modules", []):
        lessons = m.get("lessons", [])
        counts = m.get("counts_toward_720_instructional_minutes") is True
        inst = int(m.get("instructional_minutes") or 0)
        lesson_alloc = sum(int(l.get("instructional_minutes", l.get("estimated_minutes") or 0) or 0) for l in lessons)
        gap = max(0, inst - lesson_alloc)
        if counts:
            inst_total += inst
            gap_total += gap
        lines.append(
            f"| {m.get('module_id')} | {'yes' if counts else 'no'} | {inst} | {lesson_alloc} | {gap} | "
            f"{m.get('estimated_narration_minutes', 0)} | {m.get('estimated_reading_minutes', 0)} | "
            f"{m.get('estimated_interaction_minutes', 0)} | {m.get('estimated_assessment_minutes', 0)} | "
            f"{sanitize_cell(m.get('time_model_status', ''))} |"
        )
    lines.append(f"\n**Required theory instructional minutes counted toward 720:** {inst_total}  (target 720)")
    lines.append(f"\n**Module assessment minutes (tracked separately, excluded):** {metrics['module_assessment_minutes_excluded']}")
    lines.append(f"\n**Course final assessment minutes (tracked separately, excluded):** {metrics['course_final_assessment_minutes_excluded']}")
    lines.append(f"\n**Optional Clinical Support counts toward 720:** {tm.get('optional_clinical_support_counts_toward_720', False)}")
    lines.append(f"\n**Total authored narrated lesson minutes (descriptive only):** {metrics['total_narration_minutes']}")
    lines.append(f"\n**Open content-depth gap (allocation minus authored lesson minutes):** {gap_total} min")
    lines.append(
        "\n> Narration minutes are descriptive and are far below the instructional allocation by design; they are NOT "
        "the full lesson time and must not be used to reduce the 720 model. Content-depth gaps above are reported, not "
        "padded; expand only from ContentV1 source. SME/source-repair flags are preserved.\n"
    )
    lines.append("## Time-model rules (enforced by the rebuild pipeline)\n")
    for rule in tm.get("rules", []):
        lines.append(f"- {rule}")
    lines.append("\n## Pipeline counts\n")
    for k in [
        "modules_count", "lessons_count", "cards_count", "narration_clip_count",
        "required_theory_instructional_minutes_total", "module_assessment_minutes_excluded",
        "course_final_assessment_minutes_excluded", "content_depth_gap_minutes",
        "total_narration_minutes", "clips_over_75_seconds", "unique_app_location_count",
        "source_repair_flag_count", "sme_review_flag_count", "compliance_flag_count",
    ]:
        lines.append(f"- **{k}**: {metrics.get(k)}")
    COVERAGE_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def sanitize_cell(value: str) -> str:
    return str(value).replace("\n", " ").replace("|", "\\|").strip()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> int:
    parser = argparse.ArgumentParser(description="Rebuild ContentV2 mirror artifacts from canonical JSON.")
    parser.add_argument("--check", action="store_true", help="Validate only; write no artifacts.")
    args = parser.parse_args()

    if not CANON_JSON.exists():
        print(f"ERROR: canonical JSON not found at {CANON_JSON}", file=sys.stderr)
        return 2

    data = load_json(CANON_JSON)

    errors: list[str] = []
    schema_validate(data, errors)
    validate_required_fields(data, errors)

    metrics = compute_metrics(data)
    hard, warn = run_guardrails(data, metrics)

    schema_hard = [e for e in errors if e.startswith("[schema]") or e.startswith("[required]")]
    hard.extend(schema_hard)

    if args.check:
        print(json.dumps({"mode": "check", "metrics": metrics, "hard_failures": hard, "warnings": warn + [e for e in errors if e.startswith('[warn]')]}, indent=2))
        return 1 if hard else 0

    if hard:
        print(json.dumps({"status": "BLOCKED", "hard_failures": hard, "warnings": warn}, indent=2), file=sys.stderr)
        return 1

    # Emit mirrors.
    write_ts_mirror(CANON_TS, "/* Generated from ContentV2 canonical JSON. Do not hand-edit. */", data)
    write_ts_mirror(
        APP_GENERATED_TS,
        "/* Generated from CNA-Recert-Course/ContentV2/data/courseContentV2.json. Do not hand-edit. */",
        data,
    )

    verify_errors: list[str] = []
    verify_ts_matches_json(CANON_TS, data, "courseContentV2.ts", verify_errors)
    verify_ts_matches_json(APP_GENERATED_TS, data, "contentV2.generated.ts", verify_errors)
    if verify_errors:
        print(json.dumps({"status": "VERIFY_FAILED", "errors": verify_errors}, indent=2), file=sys.stderr)
        return 1

    rows = narration_rows(data)
    if len(rows) != metrics["narration_clip_count"]:
        print(
            f"ERROR: narration row count {len(rows)} != expected {metrics['narration_clip_count']}",
            file=sys.stderr,
        )
        return 1
    write_narration_csvs(rows)
    write_narration_metadata(data, metrics, rows)
    write_xlsx(data, metrics, rows)
    write_map_4card(data)
    write_media_md(data)
    write_flags_md(data)
    write_coverage_md(data, metrics)

    summary = {
        "status": "OK",
        "generated_at": now_iso(),
        "canonical": str(CANON_JSON.relative_to(ROOT)).replace("\\", "/"),
        "metrics": metrics,
        "warnings": warn + [e for e in errors if e.startswith("[warn]")],
        "artifacts_written": [
            str(p.relative_to(ROOT)).replace("\\", "/")
            for p in [
                CANON_TS, APP_GENERATED_TS, NARR_MASTER_CSV, TTS_IMPORT_CSV, NARR_META_JSON,
                XLSX_PATH, MAP_4CARD_MD, MEDIA_MD, FLAGS_MD, COVERAGE_MD, VALIDATION_SUMMARY_JSON,
            ]
        ],
    }
    VALIDATION_SUMMARY_JSON.write_text(json.dumps({"generated_at": now_iso(), **metrics, "warnings_count": len(summary["warnings"])}, indent=2), encoding="utf-8")
    print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
