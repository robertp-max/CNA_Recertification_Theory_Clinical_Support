"""Focused, additive ContentV2 hardening pass.

This does NOT rerun the full ContentV2 generation. It reads the existing
canonical courseContentV2.json, transforms debrief cards into polished
course-extension remediation (removing all internal answer-key language), adds
remediation + media readiness fields, and deterministically regenerates the
derived artifacts (app generated.ts, courseContentV2.ts, narration CSVs,
narration metadata, media doc, and the XLSX workbook) from that JSON.

Run from repo root:  python CNA-Recert-Course/ContentV2/tools/apply_remediation.py
"""
from pathlib import Path
import csv
import json
import math
import re

ROOT = Path(__file__).resolve().parents[3]
V2 = ROOT / "CNA-Recert-Course" / "ContentV2"
APP = ROOT / "standalone-course-mvp"
JSON_PATH = V2 / "data" / "courseContentV2.json"
PROJECT = "CNA Recertification Theory + Clinical Support"

DEFAULT_SAFETY_PRINCIPLE = (
    "When you are unsure, choose the action that protects the resident first, stays inside CNA scope, "
    "and keeps required observe-report-document and care-plan steps in the right order."
)
DEFAULT_SCOPE_NOTE = (
    "As a CNA you observe, report, document, and follow the care plan. You do not diagnose, change orders, "
    "or perform tasks outside your training and assignment."
)
DEFAULT_RESIDENT_SAFETY_NOTE = (
    "Resident safety and dignity come first. Prevent avoidable harm, protect privacy, and report any change "
    "of condition promptly to the licensed nurse."
)

# Authored override mirrors standalone-course-mvp/src/data/remediationOverrides.ts (keyed by challenge id).
OVERRIDES = {
    "Q01": {
        "safety_principle": (
            "Infection prevention starts with the action that breaks the chain of transmission for every resident, "
            "not just the ones who look sick. Hand hygiene, done correctly and at the right moments, is that action."
        ),
        "why_safest": (
            "Hand hygiene performed correctly and at the right times is the single most effective way to prevent the "
            "spread of infection in long-term care. Hands are the main way germs travel between residents, surfaces, "
            "and you, so clean hands before and after every contact stop transmission at its most common point."
        ),
        "cna_scope_note": (
            "Performing hand hygiene at the right moments is squarely within CNA scope. You are expected to clean your "
            "hands before and after resident contact, after removing gloves, and after contact with body fluids or "
            "contaminated surfaces."
        ),
        "resident_safety_note": (
            "Every resident is protected when you clean your hands at the right times. This is also how you protect "
            "yourself, your coworkers, and residents who cannot fight infection well."
        ),
        "remember_this": (
            "Gloves, masks, and isolation all have a place, but none of them replace hand hygiene. Clean hands at the "
            "right moments are your strongest, most reliable infection-control tool on every shift."
        ),
        "options": {
            "A": {
                "why": "Gloves are useful, but wearing them for all contact does not replace hand hygiene. Germs reach the skin when gloves are removed or reused, and gloves give a false sense of safety if hands are not cleaned before and after.",
                "remember": "Always perform hand hygiene before applying and after removing gloves - gloves are an addition to clean hands, not a substitute.",
            },
            "B": {
                "why": "Hand hygiene performed correctly and at the right times is the single most effective measure for preventing the spread of infection, because hands are the most common way germs move between residents and surfaces.",
                "remember": "Make correct, well-timed hand hygiene your automatic first infection-control habit on every shift.",
            },
            "C": {
                "why": "A surgical mask protects against certain droplet exposures but does nothing about germs carried on the hands, which cause most transmission in long-term care. Routine masking for all care is not the most effective general measure.",
                "remember": "Use masks when the situation or precautions call for them - they support, but never replace, hand hygiene.",
            },
            "D": {
                "why": "Isolating residents with known infections is only part of the picture. Many residents carry germs without obvious signs, so relying on isolation alone misses the everyday transmission that hand hygiene prevents.",
                "remember": "Standard precautions and hand hygiene apply to every resident, not only those already known to be infected.",
            },
        },
    }
}


def words(text):
    return re.findall(r"[A-Za-z0-9']+", text or "")


def estimate_seconds(text):
    wc = len(words(text))
    return max(8, int(math.ceil(wc / 2.35))), wc


def target_status(seconds, status_flags=""):
    if seconds <= 75:
        if seconds < 35:
            return "Short - acceptable if intentionally brief"
        if seconds <= 65:
            return "Within target"
        return "Acceptable near upper target"
    if "Source Repair" in status_flags:
        return "Source repair placeholder - timing review required"
    return "Over 75 seconds - split/review required"


def sentence(text, fallback):
    t = (text or "").strip()
    if not t:
        return fallback
    return t if t[-1] in ".!?" else t + "."


def lower_label(label):
    return label.rstrip(".").strip().lower()


def title_fmt(module_id, lesson_id, card_id, card_type):
    return f"{module_id} {lesson_id} {card_id.split('_')[0]} {card_type.title()}"


def build_remediation(challenge):
    cid = challenge.get("id")
    ov = OVERRIDES.get(cid, {})
    choices = challenge.get("choices", [])
    safest_id = challenge.get("correct_id_internal") or (choices[0]["id"] if choices else "A")
    safest = next((c for c in choices if c["id"] == safest_id), choices[0] if choices else {"id": "A", "label": ""})
    safest_label = safest.get("label", "")

    why_safest = sentence(
        ov.get("why_safest") or challenge.get("rationale_internal"),
        f"Choosing to {lower_label(safest_label)} keeps the resident safe and stays within CNA scope, which is why it is the safest response in this situation.",
    )
    safety_principle = ov.get("safety_principle", DEFAULT_SAFETY_PRINCIPLE)
    scope_note = ov.get("cna_scope_note", DEFAULT_SCOPE_NOTE)
    resident_note = ov.get("resident_safety_note", DEFAULT_RESIDENT_SAFETY_NOTE)
    remember_this = sentence(
        ov.get("remember_this"),
        f"Lead with the action that keeps the resident safest, then complete the remaining observe-report-document steps in order. {why_safest}",
    )

    option_rationales = []
    why_incorrect = {}
    for c in choices:
        is_safest = c["id"] == safest_id
        a = (ov.get("options", {}) or {}).get(c["id"], {})
        if is_safest:
            why = sentence(a.get("why") or ov.get("why_safest") or challenge.get("rationale_internal"), why_safest)
            remember = sentence(a.get("remember"), "This is the response to carry into real practice. Make it your default first move whenever this situation comes up.")
        else:
            why = sentence(a.get("why"), "This option is not the safest first response here. On its own it can skip a required step, act out of sequence, or reach past what a CNA should do alone. Read it next to the safest response and notice what it leaves out.")
            remember = sentence(a.get("remember"), "Before choosing an action, check it against CNA scope and resident safety, and confirm nothing required is being skipped.")
            why_incorrect[c["id"]] = why
        option_rationales.append({"id": c["id"], "label": c["label"], "status": "safest" if is_safest else "needs-review", "why": why, "remember": remember})

    narration = ov.get("narration") or (
        f"Challenge debrief. Your response has been submitted. {safety_principle} "
        f"The safest response is to {lower_label(safest_label)}. {why_safest} "
        "As a CNA, stay within scope: observe, report, document, and follow the care plan. "
        "Resident safety and dignity come first, so report any change of condition promptly to the nurse. "
        "Before you continue, review the safest response and your own choice so the reason stays with you on the floor."
    )

    return {
        "safest_id": safest_id,
        "safest_label": safest_label,
        "safety_principle": safety_principle,
        "why_safest": why_safest,
        "cna_scope_note": scope_note,
        "resident_safety_note": resident_note,
        "remember_this": remember_this,
        "option_rationales": option_rationales,
        "why_incorrect": why_incorrect,
        "narration": narration,
    }


def media_for(card, scene_title, debrief=False):
    base = card.get("media_prompt_placeholder", {})
    app_loc = card["app"]["location"]
    scene = scene_title
    image_prompt = (
        f"Realistic, de-identified long-term care training scene illustrating: {scene}; "
        "calm instructional tone; no readable charts, no facility logos, no PHI."
    )
    return {
        "app_location": app_loc,
        "scene_title": scene,
        "image_16_9_prompt": image_prompt if debrief else base.get("image_16_9_prompt", image_prompt),
        "optional_video_prompt": base.get("optional_video_prompt", f"Optional short instructional motion for {scene}; fictional/de-identified CNA training context only."),
        "negative_prompt": base.get("negative_prompt", "No PHI, no real patient records, no readable charts, no facility names, no gore, no unsafe clinical actions, no plastic-looking people."),
        "alt_text": f"Training visual: {scene}. De-identified illustration; no PHI.",
        "media_status": "placeholder-pending",
        "asset_path": f"/assets/media/{app_loc.lower()}.png",
        "required_for_mvp": False,
        "phi_safety_note": base.get("phi_safety_note", "No real patient/resident identifiers or facility data."),
        "status": "Placeholder only - no media generated.",
    }


NEUTRAL_FEEDBACK = "Your response has been submitted. Continue to the Challenge Debrief to review the safest response and why."


def normalize_feedback(challenge):
    """Replace inline answer-key/internal feedback phrases. These are not rendered
    to learners (the Challenge Debrief course-extension model is used instead), but
    we neutralize the data so no answer-key/internal language ships."""
    if not challenge:
        return
    for key in ("learner_feedback_correct", "learner_feedback_incorrect"):
        val = challenge.get(key, "")
        if val in (
            "Correct. Continue to the next item.",
            "Review the related lesson and try again.",
        ):
            challenge[key] = NEUTRAL_FEEDBACK


def main():
    content = json.loads(JSON_PATH.read_text(encoding="utf-8"))
    modules = content["modules"]

    debrief_count = 0
    for module in modules:
        for lesson in module["lessons"]:
            cards = lesson["cards"]
            challenge_card = next((c for c in cards if c.get("internal_challenge")), None)
            challenge = challenge_card["internal_challenge"] if challenge_card else None
            normalize_feedback(challenge)
            for card in cards:
                if card.get("internal_challenge"):
                    normalize_feedback(card["internal_challenge"])
                # Enrich every card's media placeholder with the readiness fields.
                card["media_prompt_placeholder"] = media_for(card, card.get("media_prompt_placeholder", {}).get("scene_title", card.get("display_title", "")))
                if card.get("card_type") != "debrief":
                    continue
                debrief_count += 1
                lesson_title = lesson["lesson_title"]
                if not challenge:
                    continue
                r = build_remediation(challenge)
                card["display_title"] = f"{lesson_title} - Challenge Debrief"
                card["learner_facing_content"] = (
                    "Challenge Debrief - lesson re-teaching extension. Your response has been submitted.\n\n"
                    f"Safety Principle: {r['safety_principle']}\n\n"
                    f"Safest Response: {r['safest_label']}\n"
                    f"Why This Is Safest: {r['why_safest']}\n\n"
                    f"CNA Scope Note: {r['cna_scope_note']}\n"
                    f"Resident Safety Note: {r['resident_safety_note']}\n\n"
                    f"What to Remember: {r['remember_this']}"
                )
                card["narration_script"] = r["narration"]
                card["transcript_text"] = r["narration"]
                sec, wc = estimate_seconds(r["narration"])
                card["estimated_narration_seconds"] = sec
                card["estimated_word_count"] = wc
                card["target_duration_status"] = target_status(sec, " ".join([card.get("status", ""), card.get("sme_review_flag", ""), card.get("compliance_review_flag", "")]))
                card["completion_condition"] = (
                    "Submit a response, then review the safest response and your own choice before completing the lesson."
                )
                # Internal rationale leak removed from rendered/learner fields.
                card["debrief_rationale"] = None
                card["media_prompt_placeholder"] = media_for(card, f"{lesson_title} challenge debrief - safest CNA response", debrief=True)
                # New course-extension remediation model fields.
                card["remediation"] = {
                    "remediation_model": "course_extension",
                    "safety_principle": r["safety_principle"],
                    "safest_response": {"option_id": r["safest_id"], "text": r["safest_label"]},
                    "selected_response": None,
                    "why_correct": r["why_safest"],
                    "why_incorrect": r["why_incorrect"],
                    "option_rationales": r["option_rationales"],
                    "resident_safety_note": r["resident_safety_note"],
                    "cna_scope_note": r["cna_scope_note"],
                    "remember_this": r["remember_this"],
                    "remediation_narration": r["narration"],
                    "remediation_transcript": r["narration"],
                    "remediation_media_prompt": card["media_prompt_placeholder"],
                    "remediation_completion_condition": card["completion_condition"],
                }

    # Normalize internal feedback phrases on module-assessment questions too.
    for a in content["assessments"]["module_assessments"].values():
        for q in a.get("questions", []):
            normalize_feedback(q)

    # ---- Recompute narration rows from JSON (mirror generator order) ----
    all_cards = [c for m in modules for l in m["lessons"] for c in l["cards"]]
    master_cols = [
        "Project", "Module ID", "Module Title", "Lesson ID", "Lesson Title", "Card ID", "Card Type", "Title",
        "app.location", "Narration", "Transcript", "Estimated Seconds", "Estimated Word Count",
        "Target Duration Status", "Voice Notes", "Source Reference", "SME Review Flag", "Compliance Review Flag", "Status",
    ]
    narration_rows = []
    for c in all_cards:
        narration_rows.append({
            "Project": PROJECT, "Module ID": c["module_id"], "Module Title": c["module_title"],
            "Lesson ID": c["lesson_id"], "Lesson Title": c["lesson_title"], "Card ID": c["card_id"],
            "Card Type": c["card_type"], "Title": title_fmt(c["module_id"], c["lesson_id"], c["card_id"], c["card_type"]),
            "app.location": c["app"]["location"], "Narration": c["narration_script"], "Transcript": c["transcript_text"],
            "Estimated Seconds": c["estimated_narration_seconds"], "Estimated Word Count": c["estimated_word_count"],
            "Target Duration Status": c["target_duration_status"], "Voice Notes": "Plain professional narration; no audio production authorized.",
            "Source Reference": c["source_reference"], "SME Review Flag": c["sme_review_flag"],
            "Compliance Review Flag": c["compliance_review_flag"], "Status": c["status"],
        })

    app_copy = content["app_copy"]
    module_assessments = content["assessments"]["module_assessments"]
    final_assessment = content["assessments"]["final_assessment"]
    final_questions = final_assessment["questions"]

    surface_rows = [
        ("Dashboard Hero", "dashboard.hero", app_copy["dashboard"]["summary"]),
        ("Certificate Gate Status", "certificate.gate.status", app_copy["certificate"]["intro"]),
        ("Mock Certificate Preview", "certificate.preview.mock", "Mock certificate preview only. Production print, download, export, and issuance remain disabled pending approval metadata, NAC#, approved wording, affidavit method, and active-time validation."),
        ("Clinical Hub Overview", "clinical.hub.overview", app_copy["clinicalHub"]["warning"]),
        ("Course Final Assessment Splash", "course.final.splash", app_copy["final"]["summary"]),
        ("Course Final Result Pass", "course.final.result.pass", app_copy["final"]["pass_body"]),
        ("Course Final Result Fail", "course.final.result.fail", app_copy["final"]["fail_body"]),
    ]
    for a in module_assessments.values():
        surface_rows.append((f"{a['title']} Splash", a["splash_app_location"], f"{a['title']}. {a['remediation_rule']} Correct answer keys are internal only."))
    for title, loc, text in surface_rows:
        sec, wc = estimate_seconds(text)
        narration_rows.append({
            "Project": PROJECT, "Module ID": "APP", "Module Title": "App Surface", "Lesson ID": "SURFACE",
            "Lesson Title": title, "Card ID": "SURFACE", "Card Type": "surface", "Title": title, "app.location": loc,
            "Narration": text, "Transcript": text, "Estimated Seconds": sec, "Estimated Word Count": wc,
            "Target Duration Status": target_status(sec), "Voice Notes": "Planning only; no audio production authorized.",
            "Source Reference": "Generated from ContentV2 app_copy", "SME Review Flag": "None identified",
            "Compliance Review Flag": "Compliance review required before production.", "Status": "Draft",
        })
    for q in final_questions:
        text = f"{q['prompt']} Response choices are presented to the learner and scored internally. Final exam correct answers and rationales are not revealed in learner-facing results."
        sec, wc = estimate_seconds(text)
        narration_rows.append({
            "Project": PROJECT, "Module ID": "FINAL", "Module Title": "Course Final Assessment", "Lesson ID": "FINAL",
            "Lesson Title": "Final Assessment Question Bank", "Card ID": q["id"], "Card Type": "final_question",
            "Title": f"Course Final {q['id']}", "app.location": q["app_location"], "Narration": text, "Transcript": text,
            "Estimated Seconds": sec, "Estimated Word Count": wc, "Target Duration Status": target_status(sec),
            "Voice Notes": "Planning only; no audio production authorized.", "Source Reference": q["source_reference"],
            "SME Review Flag": q["sme_review_flag"], "Compliance Review Flag": q["compliance_review_flag"],
            "Status": "Internal scoring only - answer key protected",
        })

    total_seconds = sum(int(r["Estimated Seconds"]) for r in narration_rows)
    narration_summary = {
        "total_clips": len(narration_rows),
        "total_estimated_seconds": total_seconds,
        "total_estimated_minutes": round(total_seconds / 60, 2),
        "clips_over_target": sum(1 for r in narration_rows if int(r["Estimated Seconds"]) > 75),
        "unique_app_locations": len({r["app.location"] for r in narration_rows}),
        "modules_covered": sorted({r["Module ID"] for r in narration_rows}),
        "validation_status": "Remediation hardening pass - additive update of debrief/remediation cards",
    }
    content["narration_summary"] = narration_summary

    # ---- Write JSON + TS mirrors ----
    JSON_PATH.write_text(json.dumps(content, indent=2, ensure_ascii=False), encoding="utf-8")
    ts_body = json.dumps(content, indent=2, ensure_ascii=False)
    (V2 / "data" / "courseContentV2.ts").write_text(
        "/* Generated from ContentV2 canonical JSON. Do not hand-edit. */\nexport const courseContentV2 = " + ts_body + " as const;\nexport default courseContentV2;\n",
        encoding="utf-8",
    )
    (APP / "src" / "data" / "contentV2.generated.ts").write_text(
        "/* Generated from CNA-Recert-Course/ContentV2/data/courseContentV2.json. Do not hand-edit. */\nexport const courseContentV2 = " + ts_body + " as const;\nexport default courseContentV2;\n",
        encoding="utf-8",
    )

    # ---- Narration CSVs + metadata ----
    with (V2 / "narration" / "narration_master.csv").open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=master_cols)
        w.writeheader()
        w.writerows(narration_rows)
    with (V2 / "narration" / "tts_narration_import.csv").open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["Project", "Title", "app.location", "Narration"])
        w.writeheader()
        for r in narration_rows:
            w.writerow({"Project": r["Project"], "Title": r["Title"], "app.location": r["app.location"], "Narration": r["Narration"]})
    meta = json.loads((V2 / "narration" / "narration_metadata.json").read_text(encoding="utf-8"))
    meta.update(narration_summary)
    (V2 / "narration" / "narration_metadata.json").write_text(json.dumps(meta, indent=2, ensure_ascii=False), encoding="utf-8")

    # ---- Media doc + rows ----
    media_rows = []
    for c in all_cards:
        m = c["media_prompt_placeholder"]
        media_rows.append(m)

    def md_table(headers, rows):
        out = ["| " + " | ".join(headers) + " |", "| " + " | ".join(["---"] * len(headers)) + " |"]
        for row in rows:
            out.append("| " + " | ".join(str(x).replace("\n", "<br>") for x in row) + " |")
        return "\n".join(out)

    (V2 / "07_MEDIA_PROMPT_PLACEHOLDERS.md").write_text(
        "# Media Prompt Placeholders\n\nNo images, videos, audio, PDFs, or certificate files were generated. These are placeholders only. "
        "Every card has a safe 16:9 placeholder with alt text; the app never shows empty black boxes or broken-image icons.\n\n"
        "Required media fields per card: app.location, scene_title, image_16_9_prompt, optional_video_prompt, negative_prompt, alt_text, media_status, asset_path, required_for_mvp, PHI safety note.\n\n"
        + md_table(
            ["app.location", "Scene", "Image Prompt", "Alt Text", "Media Status", "Asset Path", "Required for MVP", "PHI Safety Note", "Status"],
            [(m["app_location"], m["scene_title"], m["image_16_9_prompt"], m["alt_text"], m["media_status"], m["asset_path"], m["required_for_mvp"], m["phi_safety_note"], m["status"]) for m in media_rows],
        ),
        encoding="utf-8",
    )

    # ---- XLSX rebuild from JSON ----
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except Exception:
        openpyxl = None

    source_files = content.get("source_manifest", [])
    coverage_rows = [
        ("Source files represented", len(source_files)),
        ("Required theory modules", len(modules)),
        ("Required theory minutes", 720),
        ("Clinical support units", len(content["clinical_support"]["units"])),
        ("Confidence checks", len(content["clinical_support"]["confidence_checks"])),
        ("Final exam questions", len(final_questions)),
        ("Narration clips", len(narration_rows)),
        ("Debrief remediation cards", debrief_count),
    ]
    sme_rows, comp_rows = [], []
    for c in all_cards:
        if c["sme_review_flag"] != "None identified":
            sme_rows.append((c["module_id"], c["lesson_id"], c["card_id"], c["app"]["location"], c["sme_review_flag"], c["source_reference"]))
        if c["compliance_review_flag"] != "None identified":
            comp_rows.append((c["module_id"], c["lesson_id"], c["card_id"], c["app"]["location"], c["compliance_review_flag"], c["source_reference"]))
    for q in final_questions:
        if q["sme_review_flag"] != "None identified":
            sme_rows.append(("FINAL", "FINAL", q["id"], f"course.final.{q['id'].lower()}", q["sme_review_flag"], q["source_reference"]))
        comp_rows.append(("FINAL", "FINAL", q["id"], f"course.final.{q['id'].lower()}", q["compliance_review_flag"], q["source_reference"]))

    if openpyxl:
        def write_sheet(wb, name, headers, rows):
            ws = wb[name]
            ws.append(headers)
            for row in rows:
                ws.append(list(row))
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="5C1111")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.freeze_panes = "A2"
            for col in range(1, min(len(headers), 12) + 1):
                ws.column_dimensions[get_column_letter(col)].width = min(55, max(14, len(str(headers[col - 1])) + 4))
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "README"
        for name in ["ContentV2 Master", "Narration Only", "TTS Import Preview", "Image + Video Prompt Placeholders",
                     "Assessment Map", "Final Assessment Map", "App Location Map", "Source Coverage Map", "Runtime Summary",
                     "SME Review Flags", "Compliance Review Flags", "Deferred Items", "Agent Contribution Map"]:
            wb.create_sheet(name)
        ws.append(["CNA Recertification ContentV2 Master Workbook"])
        ws.append(["Generated At", content.get("generated_at", "")])
        ws.append(["Last Hardening Pass", "Remediation, TTS, media, and MVP wiring hardening (additive)"])
        ws.append(["Compliance", "Planning/staging artifact only. Not production-ready. No media/audio/certificate generated."])
        ws.append(["Remediation", "Debrief cards are course-extension remediation. No learner-facing answer-key/internal language. Final-exam answer key remains protected."])

        write_sheet(wb, "ContentV2 Master", ["Module", "Lesson", "Card", "Type", "app.location", "Title", "Content", "Seconds", "Words", "Source", "Status", "SME Flag", "Compliance Flag"],
                    [(c["module_id"], c["lesson_id"], c["card_id"], c["card_type"], c["app"]["location"], c["display_title"], c["learner_facing_content"], c["estimated_narration_seconds"], c["estimated_word_count"], c["source_reference"], c["status"], c["sme_review_flag"], c["compliance_review_flag"]) for c in all_cards])
        write_sheet(wb, "Narration Only", master_cols, [[r[h] for h in master_cols] for r in narration_rows])
        write_sheet(wb, "TTS Import Preview", ["Project", "Title", "app.location", "Narration"], [[r["Project"], r["Title"], r["app.location"], r["Narration"]] for r in narration_rows])
        write_sheet(wb, "Image + Video Prompt Placeholders",
                    ["app.location", "scene_title", "image_16_9_prompt", "optional_video_prompt", "negative_prompt", "alt_text", "media_status", "asset_path", "required_for_mvp", "phi_safety_note", "status"],
                    [(m["app_location"], m["scene_title"], m["image_16_9_prompt"], m["optional_video_prompt"], m["negative_prompt"], m["alt_text"], m["media_status"], m["asset_path"], m["required_for_mvp"], m["phi_safety_note"], m["status"]) for m in media_rows])
        write_sheet(wb, "Assessment Map", ["Module", "Assessment", "Question", "Prompt", "Choices", "Correct Answer (Internal)", "Rationale (Internal)", "Source", "SME Flag", "Compliance Flag"],
                    [(mid, a["title"], q["id"], q["prompt"], " | ".join(f"{c['id']}:{c['label']}" for c in q["choices"]), q["correct_id_internal"], q["rationale_internal"], q["source_reference"], q["sme_review_flag"], q["compliance_review_flag"]) for mid, a in module_assessments.items() for q in a["questions"]])
        write_sheet(wb, "Final Assessment Map", ["Question", "Prompt", "Choices", "Correct Answer (Internal)", "Rationale (Internal)", "Difficulty", "Source", "SME Flag", "Compliance Flag"],
                    [(q["id"], q["prompt"], " | ".join(f"{c['id']}:{c['label']}" for c in q["choices"]), q["correct_id_internal"], q["rationale_internal"], q.get("difficulty", ""), q["source_reference"], q["sme_review_flag"], q["compliance_review_flag"]) for q in final_questions])
        write_sheet(wb, "App Location Map", ["app.location", "Module", "Lesson", "Card", "Type", "Title"],
                    [(c["app"]["location"], c["module_id"], c["lesson_id"], c["card_id"], c["card_type"], c["display_title"]) for c in all_cards] + [(r["app.location"], r["Module ID"], r["Lesson ID"], r["Card ID"], r["Card Type"], r["Title"]) for r in narration_rows if r["Module ID"] == "APP"])
        write_sheet(wb, "Source Coverage Map", ["Path", "Type", "Role", "Canonical Status", "Flags", "Representation", "SHA256"],
                    [(s["path"], s["file_type"], s["role"], s["canonical_status"], s["source_flags"], s["contentv2_representation"], s["sha256"]) for s in source_files])
        write_sheet(wb, "Runtime Summary", ["Metric", "Value"],
                    coverage_rows + [("Total cards/subcards", len(all_cards)), ("SME flags", len(sme_rows)), ("Compliance flags", len(comp_rows)), ("Narration minutes", narration_summary["total_estimated_minutes"]), ("Clips over 75s", narration_summary["clips_over_target"]), ("Source repair required items", sum(1 for c in all_cards if "Source Repair Required" in c["status"]))])
        write_sheet(wb, "SME Review Flags", ["Module", "Lesson", "Item", "app.location", "Flag", "Source"], sme_rows)
        write_sheet(wb, "Compliance Review Flags", ["Module", "Lesson", "Item", "app.location", "Flag", "Source"], comp_rows)
        write_sheet(wb, "Deferred Items", ["Item", "Resolution Needed"], [
            ("Module 1 infection control", "SME/source review required before production."),
            ("Module 5 skin integrity", "SME/source review required before production."),
            ("Module 3 source repair", "Canonical Module 3 file stops mid Screen 3.2.3; repair source before production."),
            ("CDPH/provider placeholders", "Do not fabricate CDPH contact info, NAC#, provider number, approval metadata, certificate wording."),
            ("TTS authorization", "No production audio until authorized voice approval and transcript pairing."),
            ("Certificate production", "Disabled until all legal/compliance/CDPH gates are approved."),
        ])
        write_sheet(wb, "Agent Contribution Map", ["Agent", "Assignment", "Exit", "Contribution"], [("hardening_pass", "Remediation/TTS/media/MVP wiring", "0", "Additive course-extension remediation + readiness fields")])
        wb.save(V2 / "xlsx" / "CNA_RECERT_CONTENTV2_MASTER.xlsx")

    print(json.dumps({
        "debrief_remediation_cards": debrief_count,
        "narration_clips": narration_summary["total_clips"],
        "narration_minutes": narration_summary["total_estimated_minutes"],
        "clips_over_target": narration_summary["clips_over_target"],
        "unique_app_locations": narration_summary["unique_app_locations"],
        "cards": len(all_cards),
        "xlsx": bool(openpyxl),
    }, indent=2))


if __name__ == "__main__":
    main()
