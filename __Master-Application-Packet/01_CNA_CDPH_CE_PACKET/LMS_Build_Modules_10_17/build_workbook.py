"""
Build the supporting XLSX workbook for the CNA Recertification Modules 10-17
24-hour LMS build package.

Support/defensibility artifact only. The markdown course package is the primary
deliverable. All content is Draft / Pending Approval. No real approval/provider/
course numbers. Fictional, de-identified scenarios only; no PHI.

Source of truth: CCCCO Nurse Assistant Model Curriculum (Revised Dec 2018),
Modules 10-17 (CNA-Recert-Course/ContentV2/survey-evidence/_source_text/).

Run:  python build_workbook.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT = "CNA_Recert_Modules_10_17_24HR_LMS_BUILD_WORKBOOK.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=13, color="1F4E78")
NOTE_FONT = Font(italic=True, size=9, color="7F7F7F")
WRAP = Alignment(vertical="top", wrap_text=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

BANNER = ("DRAFT / PENDING APPROVAL - Do not advertise, issue certificates, or "
          "represent this course as approved until approval status, "
          "provider/course identifiers, and certificate wording are confirmed. "
          "Online CE only; not clinical hours. Fictional data only; no PHI.")

MODULES = [
    (10, "Vital Signs", 1),
    (11, "Nutrition", 1),
    (12, "Emergency Procedures", 1),
    (13, "Long Term Care Patient/Resident", 1),
    (14, "Rehabilitative Nursing", 2),
    (15, "Observation and Charting", 2),
    (16, "Death and Dying", 2),
    (17, "Patient/Resident Abuse", 2),
]


def style_sheet(ws, headers, widths, title):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = BANNER
    ws["A2"].font = NOTE_FONT
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.row_dimensions[2].height = 42
    header_row = 4
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    return header_row


def write_rows(ws, header_row, rows):
    for r, row in enumerate(rows, start=header_row + 1):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = WRAP
            cell.border = BORDER


wb = Workbook()

# ---------------------------------------------------------------- Course_Map
ws = wb.active
ws.title = "Course_Map"
headers = ["Year", "Source Module #", "Source Module Title", "Online CE Hours",
           "Estimated Minutes", "Source Reference", "Completion Requirement",
           "Evidence Artifact", "SME/Compliance Flag"]
hr = style_sheet(ws, headers,
                 [6, 14, 30, 14, 16, 34, 40, 40, 34],
                 "Course_Map - 24-Hour CNA Recertification Online CE (Modules 10-17)")
src = "ContentV2/survey-evidence/_source_text/module-{n}.txt"
comp = ("All lessons viewed + knowledge check attempted + scenario completed + "
        "documentation exercise submitted")
evid = ("Activity-completion timestamps, KC attempt record, scenario branch log, "
        "submitted simulated artifact, time-in-module log")
flags = {
    10: "Confirm normal-range values; no hands-on competency implied",
    11: "Confirm therapeutic-diet list vs facility diet manual; assistance level only",
    12: "HIGH: state no hands-on CPR/Heimlich competency; confirm code list/AHA-ARC",
    13: "Confirm person-centered dementia language; CNA scope (no diagnosis)",
    14: "ROM/restorative at assist/observe level only; confirm device list",
    15: "HIGH PHI emphasis (charting); document observations not diagnoses",
    16: "Add sensitivity note; confirm postmortem steps; no pronouncement of death",
    17: "HIGH: confirm CA mandated-reporter channels/timelines/ombudsman framing",
}
rows = []
for num, title, yr in MODULES:
    rows.append([yr, num, title, 3, 180, src.format(n=num), comp, evid, flags[num]])
rows.append(["", "", "TOTAL", 24, 1440,
             "8 modules x 3 online CE hours = 24 (Year1=12, Year2=12)",
             "Both Year assessments passed at 80% + attestation + seat-time",
             "Course completion report + certificate record (post-approval)",
             "Master-packet structure conflict (12-Unit vs Modules 10-17): confirm"])
write_rows(ws, hr, rows)

# --------------------------------------------------------- Module_Source_Map
ws = wb.create_sheet("Module_Source_Map")
headers = ["Source Module #", "Source Module Title", "Source Topic",
           "Required Content to Preserve", "LMS Destination",
           "Assessment Linkage", "Notes / Risks"]
hr = style_sheet(ws, headers, [14, 28, 40, 50, 18, 26, 40],
                 "Module_Source_Map - CCCCO NA Model Curriculum (Rev. Dec 2018)")
msm = [
    (10, "Vital Signs",
     "How/when/why vitals taken; T/P/R/BP; recognize & report normal/abnormal; pain; record vitals",
     "TPR + BP concepts, adult normal ranges, factors raising/lowering each vital, pain as observation, scope-safe reporting & charting",
     "Y1 > M10", "KC-M10 + S-M10 + Year 1 Assessment",
     "Keep at observe/report level; no hands-on measurement competency claimed"),
    (11, "Nutrition",
     "Body need for food/fluids; nutrients; therapeutic diets; feeding; hydration; cultural/religious",
     "Nutrients/food groups, common therapeutic diets, dysphagia/aspiration awareness, hydration/I&O, feeding assistance, cultural respect, alternative routes (observe/report)",
     "Y1 > M11", "KC-M11 + S-M11 + Year 1 Assessment",
     "Feeding stays at assistance level; alternative nutrition routes are observe/report only"),
    (12, "Emergency Procedures",
     "Signs/symptoms of distress; CNA immediate/temporary response; choking; emergency codes",
     "Recognize distress, activate help/EMS chain, choking response within scope, facility codes, scope boundaries",
     "Y1 > M12", "KC-M12 + S-M12 + Year 1 Assessment",
     "HIGH RISK: explicitly NOT a hands-on CPR/first-aid certification"),
    (13, "Long Term Care Patient/Resident",
     "Aging effects; common conditions; dementia care; community resources; basic A&P",
     "Human needs of elderly, dementia-sensitive approaches, special-needs populations, body systems overview, immobility complications",
     "Y1 > M13", "KC-M13 + S-M13 + Year 1 Assessment",
     "Source hours vary (5 SNF/ICF or 3 non-SNF/ICF + 2) - confirm variant"),
    (14, "Rehabilitative Nursing",
     "Restorative care; independence; rehab team; ADLs; ROM; adaptive devices; immobility prevention",
     "Restorative philosophy, CNA role on rehab team, ADLs, ROM types, adaptive/comfort devices, immobility prevention, documentation/care-plan role",
     "Y2 > M14", "KC-M14 + S-M14 + Year 2 Assessment",
     "ROM at assist/observe level; no independent therapy claim"),
    (15, "Observation and Charting",
     "Objective/subjective observation; medical terminology/abbreviations; charting docs; MDS/ADL",
     "Senses used in observation, objective vs subjective, charting document types, accurate/timely/factual recording, incident reporting, scope-safe documentation",
     "Y2 > M15", "KC-M15 + S-M15 + Year 2 Assessment",
     "HIGH PHI emphasis; CNAs document observations not diagnoses"),
    (16, "Death and Dying",
     "Grief stages (Kubler-Ross); emotional/spiritual needs; dying rights; signs of death; hospice; postmortem",
     "Five stages of grief, comfort/dignity measures, dying rights, approaching vs biological death, hospice role, postmortem responsibilities",
     "Y2 > M16", "KC-M16 + S-M16 + Year 2 Assessment",
     "Add learner sensitivity/content note; no pronouncement of death"),
    (17, "Patient/Resident Abuse",
     "Types of abuse; issues; CNA prevention role; CNA reporting role",
     "Abuse types/definitions, mandated-reporter duty, prevention, recognition, reporting pathways (ombudsman/supervisor/state), confidentiality/HIPAA basics",
     "Y2 > M17", "KC-M17 + S-M17 + Year 2 Assessment",
     "HIGH: verify current CA mandated-reporter channels and timelines"),
]
write_rows(ws, hr, msm)

# ----------------------------------------------------------- LMS_Activity_Map
ws = wb.create_sheet("LMS_Activity_Map")
headers = ["Activity ID", "Year", "Source Module #", "Activity Name",
           "Moodle Activity Type", "Learner Action Required",
           "Completion Tracking Setting", "Restrict Access Rule",
           "Gradebook Category", "Evidence Generated"]
hr = style_sheet(ws, headers, [14, 6, 14, 30, 18, 30, 26, 32, 22, 30],
                 "LMS_Activity_Map - Moodle build (Draft)")
act_rows = []
for num, title, yr in MODULES:
    cat = f"Year {yr} - Formative"
    act_rows += [
        [f"M{num}-PAGE", yr, num, f"M{num} Objectives & Banner", "Page",
         "View objectives and compliance banner", "Mark as viewed",
         "Available after Section 0 complete", cat,
         "View timestamp"],
        [f"M{num}-LSN", yr, num, f"M{num} Lesson", "Lesson",
         "Work through branching lesson content", "Require view (+min pages/time)",
         "After M{}-PAGE viewed".format(num), cat,
         "Lesson completion + time"],
        [f"M{num}-H5P", yr, num, f"M{num} Interactive Practice", "H5P",
         "Complete interactive practice", "Require attempt",
         "After M{}-LSN viewed".format(num), cat,
         "Interaction/xAPI record"],
        [f"S-M{num}", yr, num, f"M{num} Scenario (fictional)", "Lesson (branching)",
         "Make scope-safe decisions", "Require completion",
         "After M{}-LSN viewed".format(num), cat,
         "Scenario branch log"],
        [f"KC-M{num}", yr, num, f"M{num} Knowledge Check", "Quiz",
         "Attempt low-stakes quiz (unlimited)", "Require attempt",
         "After S-M{} complete".format(num), cat,
         "Attempt record + score"],
        [f"DOC-M{num}", yr, num, f"M{num} Documentation Exercise", "Assignment",
         "Submit simulated documentation (fictional)", "Require submission",
         "Available with M{}-LSN".format(num), cat,
         "Submitted artifact"],
    ]
# Year-level + course-level activities
act_rows += [
    ["SEC0-ID", 0, 0, "Identity Confirmation", "Feedback/Choice",
     "Confirm legal name (+CNA # if required)", "Require submission",
     "Available at enrollment", "Orientation",
     "Identity record"],
    ["SEC0-NOPHI", 0, 0, "No-PHI Agreement", "Feedback",
     "Accept no-PHI agreement", "Require submission",
     "Available at enrollment", "Orientation", "Agreement record"],
    ["Y1-ATTEST", 1, 0, "Year 1 Integrity Attestation", "Feedback",
     "Re-attest before Year 1 assessment", "Require submission",
     "After M10-M13 complete", "Year 1 - Summative", "Attestation record"],
    ["Y1-ASSESS", 1, 0, "Year 1 Assessment", "Quiz",
     "Pass summative at 80%", "Require pass grade (>=80%)",
     "After M10-M13 complete + Y1-ATTEST", "Year 1 - Summative",
     "Attempts, score, duration"],
    ["Y2-ATTEST", 2, 0, "Year 2 Integrity Attestation", "Feedback",
     "Re-attest before Year 2 assessment", "Require submission",
     "After M14-M17 complete", "Year 2 - Summative", "Attestation record"],
    ["Y2-ASSESS", 2, 0, "Year 2 Assessment", "Quiz",
     "Pass summative at 80%", "Require pass grade (>=80%)",
     "After M14-M17 complete + Y2-ATTEST", "Year 2 - Summative",
     "Attempts, score, duration"],
    ["FINAL-AFF", 2, 0, "Final Affidavit", "Assignment",
     "Submit completion affidavit", "Require submission",
     "After both Year assessments passed", "Course",
     "Affidavit submission"],
    ["EVAL", 2, 0, "Course Evaluation", "Feedback",
     "Complete evaluation (no PHI)", "Require submission",
     "After final affidavit", "Course", "Evaluation responses"],
    ["CERT", 2, 0, "Certificate (gated, post-approval)", "Custom certificate",
     "Download once gate met + approval confirmed", "Restricted by course completion",
     "After all gate items incl. APPROVAL CONFIRMED (currently OFF)", "Course",
     "Certificate record + unique ID"],
]
write_rows(ws, hr, act_rows)

# -------------------------------------------------------- Assessment_Blueprint
ws = wb.create_sheet("Assessment_Blueprint")
headers = ["Assessment ID", "Source Module #", "Question Type", "Question Stem",
           "Answer Choices", "Correct Answer", "Rationale",
           "Remediation Guidance", "Learner Visibility",
           "Admin/Faculty Visibility"]
hr = style_sheet(ws, headers, [14, 14, 16, 46, 40, 14, 40, 34, 22, 22],
                 "Assessment_Blueprint - seed items (Internal answer key - DO NOT publish)")
# Two seed items per module (low-stakes), objective-aligned, fictional, scope-safe.
ab = [
    ("KC-M10-01", 10, "MCQ",
     "Which adult finding should the CNA report to the licensed nurse as abnormal?",
     "A) Pulse 72  B) Resp 16  C) Pulse 124  D) Oral temp 98.6F",
     "C", "Adult pulse >100 (tachycardia) is abnormal and reported.",
     "Review M10 normal ranges and reporting (L10.2-L10.3).",
     "Score + remediation only", "Full key"),
    ("KC-M10-02", 10, "True/False",
     "A CNA may adjust a resident's blood pressure medication if the reading is high.",
     "True / False", "False",
     "Out of CNA scope; CNAs observe and report, never medicate/adjust orders.",
     "Review M10 scope-of-practice and reporting (L10.5-L10.6).",
     "Score + remediation only", "Full key"),
    ("KC-M11-01", 11, "Matching",
     "Match the therapeutic diet to its description.",
     "Pureed / Clear liquid / Low-sodium / Mechanical soft (descriptions provided)",
     "Correct pairings per source", "Diets are ordered; CNA assists and reports intake.",
     "Review M11 therapeutic diets (L11.4).",
     "Score + remediation only", "Full key"),
    ("KC-M11-02", 11, "MCQ",
     "A resident on a thickened-liquid diet is served thin juice. The CNA should:",
     "A) Serve it  B) Withhold and verify the diet order with the nurse  C) Thin it more  D) Ignore",
     "B", "Aspiration risk; verify order and report before serving.",
     "Review M11 aspiration precautions and scope (L11.5).",
     "Score + remediation only", "Full key"),
    ("KC-M12-01", 12, "MCQ",
     "A resident returns from a walk reporting shortness of breath and chest tightness. FIRST CNA action?",
     "A) Reassure they'll be fine  B) Take vitals  C) Stay and call the nurse immediately  D) Call family",
     "C", "Stay with resident and summon the nurse/EMS chain immediately.",
     "Review M12 immediate interventions (L12.2).",
     "Score + remediation only", "Full key"),
    ("KC-M12-02", 12, "True/False",
     "Completing this online module certifies the CNA in hands-on CPR.",
     "True / False", "False",
     "This is online CE only; hands-on CPR certification is separate.",
     "Review M12 banner and scope note (L12.1).",
     "Score + remediation only", "Full key"),
    ("KC-M13-01", 13, "MCQ",
     "A resident with dementia is agitated during evening care. Best scope-safe approach?",
     "A) Argue with the facts  B) Use a calm, simple, reassuring approach and reduce stimulation  C) Restrain  D) Leave alone",
     "B", "Person-centered, dignity-preserving de-escalation within scope.",
     "Review M13 dementia-sensitive care (L13.3).",
     "Score + remediation only", "Full key"),
    ("KC-M13-02", 13, "MCQ",
     "Which is a complication of immobility the CNA observes and reports?",
     "A) Pressure injury / skin breakdown  B) Improved appetite  C) Lower pulse  D) Faster healing",
     "A", "Immobility raises risk of pressure injuries; observe and report.",
     "Review M13 complications of immobility (L13.6).",
     "Score + remediation only", "Full key"),
    ("KC-M14-01", 14, "MCQ",
     "Restorative care for a resident who can do part of their ADLs means the CNA should:",
     "A) Do everything for them  B) Assist only as needed and encourage independence  C) Skip the ADL  D) Rush them",
     "B", "Restorative philosophy promotes maximum independence.",
     "Review M14 restorative philosophy & self-care (L14.1-L14.3).",
     "Score + remediation only", "Full key"),
    ("KC-M14-02", 14, "Matching",
     "Match adaptive/comfort device to its purpose.",
     "Trochanter roll / Foot board / Bed cradle / Heel protector (purposes provided)",
     "Correct pairings per source", "Devices prevent complications and aid mobility.",
     "Review M14 devices (L14.4).",
     "Score + remediation only", "Full key"),
    ("KC-M15-01", 15, "MCQ",
     "Which is an OBJECTIVE observation?",
     "A) 'Resident seems sad'  B) 'Resident ate 50% of lunch'  C) 'Resident is depressed'  D) 'Resident is fine'",
     "B", "Objective = measurable/observed fact; others are subjective/diagnostic.",
     "Review M15 objective vs subjective (L15.3).",
     "Score + remediation only", "Full key"),
    ("KC-M15-02", 15, "True/False",
     "A CNA may chart 'resident has pneumonia' based on a cough.",
     "True / False", "False",
     "CNAs document observations, not diagnoses.",
     "Review M15 scope-safe recording (L15.5).",
     "Score + remediation only", "Full key"),
    ("KC-M16-01", 16, "Matching",
     "Match the Kubler-Ross grief stage to its description.",
     "Denial / Anger / Bargaining / Depression / Acceptance (descriptions provided)",
     "Correct pairings per source", "Five stages of grief per source.",
     "Review M16 grief stages (L16.2).",
     "Score + remediation only", "Full key"),
    ("KC-M16-02", 16, "MCQ",
     "A dying resident's family member is angry at staff. Best CNA response?",
     "A) Argue back  B) Respond with compassion, listen, and report concerns to the nurse  C) Ignore them  D) Leave the unit",
     "B", "Compassionate, scope-safe support; escalate appropriately.",
     "Review M16 emotional/spiritual needs (L16.3).",
     "Score + remediation only", "Full key"),
    ("KC-M17-01", 17, "MCQ",
     "A CNA witnesses a coworker handling a resident roughly and threatening them. The CNA must:",
     "A) Say nothing  B) Ensure resident safety and report immediately per policy/mandated-reporter duty  C) Confront only the coworker  D) Wait a week",
     "B", "CNAs are mandated reporters; ensure safety and report promptly.",
     "Review M17 recognizing & reporting (L17.5).",
     "Score + remediation only", "Full key"),
    ("KC-M17-02", 17, "MCQ",
     "Which is an example of involuntary seclusion / a resident-rights violation?",
     "A) Offering activities  B) Confining a resident to a room against their will as punishment  C) Assisting to the dining room  D) Honoring a preference",
     "B", "Involuntary seclusion is a form of abuse/rights violation.",
     "Review M17 abuse types & resident rights (L17.2, L17.4).",
     "Score + remediation only", "Full key"),
]
write_rows(ws, hr, ab)

# ------------------------------------------------------------- Certificate_Gate
ws = wb.create_sheet("Certificate_Gate")
headers = ["Gate Item", "Required Condition", "LMS Evidence Source",
           "Manual Review Needed", "Approval Placeholder", "Risk if Omitted"]
hr = style_sheet(ws, headers, [26, 34, 30, 18, 26, 36],
                 "Certificate_Gate - certificate blocked until ALL items pass (Draft)")
cg = [
    ["Identity confirmed", "Section 0 identity checkpoint complete",
     "Section 0 activity completion + attestation", "Optional",
     "N/A", "Cannot tie completion to a verified learner"],
    ["No-PHI agreement", "Accepted in Section 0", "Activity completion record",
     "No", "N/A", "Privacy/PHI exposure risk"],
    ["Year 1 modules complete", "M10-M13 all activities complete",
     "Activity completion roll-up", "No", "N/A", "Incomplete content coverage"],
    ["Year 1 assessment passed", ">= 80%", "Gradebook (grade-to-pass)", "No",
     "N/A", "Unverified competency at the knowledge level"],
    ["Year 2 modules complete", "M14-M17 all activities complete",
     "Activity completion roll-up", "No", "N/A", "Incomplete content coverage"],
    ["Year 2 assessment passed", ">= 80%", "Gradebook (grade-to-pass)", "No",
     "N/A", "Unverified competency at the knowledge level"],
    ["Seat-time evidence", "Minimum documented engaged time met",
     "Logs / time reports / min-time conditions", "Recommended spot-check",
     "N/A", "Hour claims not defensible"],
    ["Learner attestation/affidavit", "Final affidavit submitted",
     "Assignment/Feedback submission", "Recommended", "N/A",
     "Integrity of completion not attested"],
    ["Approval status confirmed",
     "Provider/course approval + certificate wording confirmed",
     "Master packet / CDPH confirmation", "REQUIRED",
     "APPROVAL PLACEHOLDER - PENDING (currently NOT met)",
     "Issuing an unapproved certificate = compliance violation"],
]
write_rows(ws, hr, cg)

# ------------------------------------------------------ Application_Packet_Checklist
ws = wb.create_sheet("Application_Packet_Checklist")
headers = ["Packet Item", "Draft Status", "Owner", "Source Reference",
           "Missing Information", "Approval Status", "Notes"]
hr = style_sheet(ws, headers, [34, 14, 16, 36, 34, 16, 34],
                 "Application_Packet_Checklist (Draft)")
mp = "__Master-Application-Packet/01_CNA_CDPH_CE_PACKET/"
apc = [
    ["Course outline (Modules 10-17)", "Draft", "Owner/SME",
     "This package Section M; _source_text/module-10..17",
     "CDPH-facing structure decision (Unit vs Module)", "Pending",
     "Top open item #1"],
    ["24-hour allocation table", "Draft", "Owner/SME", "Section D",
     "Whether CDPH wants source-proportional or normalized hours", "Pending",
     "Even 3h x 8 used as draft"],
    ["Year 1 / Year 2 split", "Draft", "Owner", "Section E", "None", "Pending",
     "12 / 12 split"],
    ["Module mapping table", "Draft", "SME", "Sections C & F", "None", "Pending",
     "Source titles preserved"],
    ["Flyer / brochure copy", "Draft", "Owner/Marketing", "Section N",
     "Final approval + provider/course IDs; refund policy", "Pending",
     "Marked Draft-Pending Approval"],
    ["Certificate text", "Draft", "Owner", "Section O",
     "Approval/provider/course #, signer, CNA # requirement", "Pending",
     "Placeholders only"],
    ["Course evaluation form", "Draft", "SME", "Section P", "None", "Pending", ""],
    ["Attendance/proof-of-participation policy", "Draft", "Owner/LMS admin",
     "Section M.10 / Q", "Exact seat-time definition", "Pending", ""],
    ["Learner attestation/affidavit", "Draft", "Owner/Legal", "Section L / M.11",
     "CNA # requirement; legal review", "Pending", ""],
    ["Recordkeeping policy", "Draft", "Owner/LMS admin", "Section Q",
     "Retention period (CDPH vs BPPE)", "Pending", ""],
    ["Refund / cancellation policy", "Placeholder", "Owner/Finance",
     "BPPE refund/pro-rata source files (01_SOURCE_INVENTORY.md)",
     "Pull institution's approved policy", "Pending", "Placeholder only"],
    ["Advertising compliance checklist", "Draft", "Compliance", "Section T",
     "None", "Pending", ""],
    ["LMS evidence checklist", "Draft", "LMS admin", "Section Q", "None", "Pending", ""],
    ["SME/compliance review checklist", "Draft", "SME/Compliance",
     "Sections T & U", "SME of record sign-off", "Pending",
     mp + " references Vanessa Valerio for CNA lane - confirm"],
    ["Owner approval checklist", "Draft", "Owner", "Section W", "None", "Pending", ""],
    ["Provider/course approval evidence", "MISSING", "Owner",
     mp, "No approved provider/course # for Modules 10-17 lane found", "Pending",
     "Course stays Draft until obtained"],
]
write_rows(ws, hr, apc)

# ------------------------------------------------------------ SME_Compliance_Flags
ws = wb.create_sheet("SME_Compliance_Flags")
headers = ["Item ID", "Source Module #", "Flag Type", "Issue", "Risk Level",
           "Reviewer Needed", "Required Action", "Status"]
hr = style_sheet(ws, headers, [12, 14, 18, 44, 12, 18, 40, 12],
                 "SME_Compliance_Flags (Draft)")
sf = [
    ["F-STRUCT", "ALL", "Compliance/Structure",
     "Master packet official allocation is 12 Units x 2h, not Modules 10-17",
     "HIGH", "Owner + SME + CDPH",
     "Resolve which structure is submitted; reconcile before any CDPH use", "Open"],
    ["F-HOURS", "ALL", "Compliance",
     "Even 3h/module is normalized; source CCCCO theory hours are uneven",
     "MEDIUM", "SME/CDPH", "Confirm expected hour basis", "Open"],
    ["F-APPROVAL", "ALL", "Compliance",
     "No approved provider/course number; certificate/advertising are Draft",
     "HIGH", "Owner", "Obtain approval before issuing/advertising", "Open"],
    ["F-M10", 10, "Clinical accuracy",
     "Vital-signs normal ranges must match current standards", "MEDIUM",
     "Clinical SME", "Verify ranges/factors; ensure observe/report framing", "Open"],
    ["F-M11", 11, "Clinical accuracy",
     "Therapeutic-diet list and aspiration precautions", "MEDIUM",
     "Clinical SME", "Confirm vs facility diet manual; assistance level only", "Open"],
    ["F-M12", 12, "Scope/Clinical",
     "Must not imply hands-on CPR/Heimlich certification", "HIGH",
     "Clinical SME", "Add explicit disclaimer; verify codes + AHA/ARC language", "Open"],
    ["F-M13", 13, "Scope/Content",
     "Dementia-sensitive language; SNF/ICF hour variant", "MEDIUM",
     "Clinical SME", "Confirm person-centered language and hour variant", "Open"],
    ["F-M14", 14, "Scope",
     "ROM/restorative must stay assist/observe level", "MEDIUM",
     "Clinical SME", "Ensure no independent-therapy claim; confirm devices", "Open"],
    ["F-M15", 15, "PHI/Scope",
     "Charting module; observations not diagnoses; PHI risk", "HIGH",
     "SME/Compliance", "Reinforce no-PHI + objective documentation rules", "Open"],
    ["F-M16", 16, "Sensitivity/Scope",
     "End-of-life content; no pronouncement of death", "MEDIUM",
     "Clinical SME", "Add sensitivity note; confirm postmortem steps", "Open"],
    ["F-M17", 17, "Legal/Compliance",
     "CA mandated-reporter channels, timelines, ombudsman framing", "HIGH",
     "Compliance/Legal", "Verify current CA reporting requirements", "Open"],
    ["F-PHI", "ALL", "Privacy",
     "No real resident/facility data anywhere; fictional only", "HIGH",
     "Compliance", "Verify banners/agreements; periodic free-text review", "Open"],
]
write_rows(ws, hr, sf)

# -------------------------------------------------------------- Audit_Evidence_Log
ws = wb.create_sheet("Audit_Evidence_Log")
headers = ["Evidence Item", "Generated By", "Stored Where", "Retention Period",
           "Review Responsibility", "Export Method", "Notes"]
hr = style_sheet(ws, headers, [30, 26, 28, 18, 22, 28, 30],
                 "Audit_Evidence_Log (Draft)")
ael = [
    ["Enrollment & identity record", "Section 0 / LMS enrollment",
     "Moodle user + activity records", "[confirm]", "LMS admin",
     "CSV/PDF export", "CNA # only if required"],
    ["Activity completion data", "Module activities", "Moodle completion report",
     "[confirm]", "LMS admin/faculty", "Completion report (CSV)", ""],
    ["Time / participation logs", "Moodle logs", "Moodle logs/reports",
     "[confirm]", "LMS admin", "Logs export (CSV)", "Supplemental to attestation"],
    ["Knowledge-check attempts", "Formative quizzes", "Gradebook/quiz reports",
     "[confirm]", "Faculty", "Quiz report export", "Unlimited attempts"],
    ["Year assessment attempts/scores", "Summative quizzes", "Gradebook",
     "[confirm]", "Faculty", "Grade export", "80% pass"],
    ["Scenario decision logs", "Lesson/H5P", "Activity reports", "[confirm]",
     "Faculty", "Report export", "Branch choices"],
    ["Learner attestation / affidavit", "Assignment/Feedback", "Submission records",
     "[confirm]", "LMS admin", "PDF/CSV export", ""],
    ["Course evaluation responses", "Feedback module", "Feedback module",
     "[confirm]", "Admin", "CSV export", "No PHI"],
    ["Certificate record (post-approval)", "Certificate activity",
     "Certificate logs", "[confirm]", "LMS admin", "PDF + verification ID",
     "Only after approval confirmed"],
]
write_rows(ws, hr, ael)

wb.save(OUT)
print("Wrote", OUT, "with sheets:", [s.title for s in wb.worksheets])
