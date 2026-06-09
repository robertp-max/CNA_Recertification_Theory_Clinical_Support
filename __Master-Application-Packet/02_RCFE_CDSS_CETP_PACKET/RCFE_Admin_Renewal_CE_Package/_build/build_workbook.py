# -*- coding: utf-8 -*-
"""
Build RCFE_Admin_CE_LMS_BUILD_WORKBOOK.xlsx from rcfe_ce_data.py (single source).

Support/defensibility artifact only. The markdown course package is the primary deliverable.
Run: python build_workbook.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import rcfe_ce_data as D

OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                   "RCFE_Admin_CE_LMS_BUILD_WORKBOOK.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=13, color="1F3864")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def add_sheet(wb, name, headers, rows, widths=None, title=None, note=None):
    ws = wb.create_sheet(name)
    r = 1
    if title:
        ws.cell(row=r, column=1, value=title).font = TITLE_FONT
        r += 1
    if note:
        c = ws.cell(row=r, column=1, value=note)
        c.font = Font(italic=True, size=9, color="C00000")
        c.alignment = WRAP
        r += 1
    header_row = r
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=j, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = BORDER
    for i, row in enumerate(rows, start=header_row + 1):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.alignment = WRAP
            c.border = BORDER
    widths = widths or [22] * len(headers)
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.row_dimensions[header_row].height = 30
    return ws


def sheet_cover(wb):
    ws = wb.active
    ws.title = "Cover"
    ws["A1"] = D.DRAFT_BANNER
    ws["A1"].font = Font(bold=True, size=14, color="C00000")
    info = [
        ("Workbook", "RCFE_Admin_CE_LMS_BUILD_WORKBOOK.xlsx"),
        ("Purpose", "Build-support & defensibility artifact only. Primary deliverable is the markdown course package."),
        ("Course title", D.COURSE["title"]),
        ("Audience", D.COURSE["audience"]),
        ("Status", D.COURSE["status"]),
        ("Total CE hours (draft math)", D.COURSE["total_hours"]),
        ("Self-paced cap", f"<= {D.COURSE['self_paced_cap_hours']} hours"),
        ("Self-paced planned", f"{D.COURSE['self_paced_planned_hours']} hours (buffer {D.COURSE['self_paced_buffer_hours']}h)"),
        ("Instructor-interactive planned", f"{D.COURSE['instructor_interactive_hours']} hours"),
        ("Dementia hours", f"{D.COURSE['dementia_planned_hours']} (required >= {D.COURSE['dementia_required_hours']})"),
        ("Pass threshold", f"{D.COURSE['pass_threshold_pct']}%"),
        ("Provider/Vendor (placeholder)", D.PACKET_FACTS["provider_vendor_name"]),
        ("Vendor number", D.PACKET_FACTS["vendor_number"]),
        ("Course number", D.PACKET_FACTS["course_number"]),
        ("Instructor of record", D.PACKET_FACTS["instructor_of_record"]),
        ("Source of truth", D.PACKET_FACTS["source_path"]),
        ("Required placeholder language", D.PENDING_LANGUAGE),
        ("Closing language", D.CLOSING_LANGUAGE),
    ]
    r = 3
    for k, v in info:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        c = ws.cell(row=r, column=2, value=v)
        c.alignment = WRAP
        r += 1
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 90


def s_course_map():
    headers = ["Module ID", "Module title", "CE hours", "Estimated minutes", "Delivery mode",
               "RCFE core knowledge area", "Dementia-hour flag", "Self-paced flag",
               "Completion requirement", "Evidence artifact"]
    rows = []
    for m in D.MODULES:
        rows.append([m["id"], m["title"], m["ce_hours"], m["minutes"], m["delivery"],
                     "; ".join(m["core_knowledge"]),
                     f"Yes ({m['dementia_hours']}h)" if m["dementia_hours"] else "No",
                     "Yes" if m["delivery"] == "Self-paced online" else "No",
                     m["completion_requirement"], m["evidence_artifact"]])
    rows.append(["TOTAL", "12 modules", D.COURSE["total_hours"], D.COURSE["total_minutes"],
                 f"{D.COURSE['self_paced_planned_hours']}h self-paced / {D.COURSE['instructor_interactive_hours']}h live",
                 "All 12 core areas", f"{D.COURSE['dementia_planned_hours']}h", f"{D.COURSE['self_paced_planned_hours']}h",
                 "All gates met + signed statement", "Full audit evidence set"])
    widths = [10, 34, 9, 12, 22, 30, 14, 11, 40, 40]
    return headers, rows, widths


def s_module_blueprints():
    headers = ["Module ID", "Lesson ID", "Lesson title", "Learning objective", "Activity type",
               "Estimated minutes", "Scenario or exercise summary", "Moodle activity recommendation",
               "Completion condition"]
    rows = []
    for m in D.MODULES:
        for l in m["lessons"]:
            rows.append([m["id"], l["id"], l["title"], l["objective"], l["activity_type"],
                         l["minutes"], l["summary"], l["moodle"], l["completion"]])
    widths = [10, 10, 28, 34, 22, 11, 38, 26, 30]
    return headers, rows, widths


def s_core_knowledge_map():
    headers = ["Core knowledge area", "Module ID", "Lesson ID", "Evidence of coverage",
               "Assessment linkage", "Compliance note"]
    rows = []
    for area in D.CORE_KNOWLEDGE_AREAS:
        covering = [m for m in D.MODULES if area in m["core_knowledge"]]
        if not covering:
            rows.append([area, "Needs confirmation", "-", "Not yet mapped to a module",
                         "-", "Needs confirmation - confirm area wording vs CDSS source"])
            continue
        for m in covering:
            lesson_ids = ", ".join(l["id"] for l in m["lessons"])
            rows.append([area, m["id"], lesson_ids,
                         f"Lessons + {m['doc_exercise'][:60]}...",
                         f"{m['id']} knowledge check + final assessment",
                         "; ".join(m["compliance_flags"])[:90]])
    widths = [38, 10, 24, 40, 28, 40]
    return headers, rows, widths


def s_dementia_hour_map():
    headers = ["Module ID", "Lesson ID", "Dementia topic", "Minutes", "CE hours",
               "Direct care linkage", "Physical environment linkage",
               "Admissions/assessment linkage", "Evidence artifact"]
    rows = []
    total_min = 0
    for m in D.MODULES:
        if not m["dementia_hours"]:
            continue
        for l in m["lessons"]:
            total_min += l["minutes"]
            direct = "Yes" if m["id"] == "M08" else ("Partial" if "Behavior" in l["title"] else "Linked via M08")
            env = "Yes" if (m["id"] == "M09" and "Environment" in l["title"]) else ("Yes" if m["id"] == "M09" else "No")
            adm = "Yes" if (m["id"] == "M09" and ("Admissions" in l["title"] or "Assessment" in l["title"])) else "No"
            rows.append([m["id"], l["id"], l["title"], l["minutes"], round(l["minutes"] / 60, 2),
                         direct, env, adm, m["evidence_artifact"][:50]])
    rows.append(["TOTAL", "-", "Dementia instruction total", total_min, round(total_min / 60, 2),
                 "M08 (4h)", "M09 environment", "M09 admissions/assessment",
                 f"Required >= {D.COURSE['dementia_required_hours']}h; planned {D.COURSE['dementia_planned_hours']}h"])
    widths = [10, 10, 34, 9, 9, 18, 22, 24, 34]
    return headers, rows, widths


def s_delivery_mode_map():
    headers = ["Module ID", "Delivery mode", "Self-paced hours", "Instructor-interactive hours",
               "Interaction method", "Identity confirmation method", "Signed statement required",
               "Compliance note"]
    rows = []
    sp_total = it_total = 0
    for m in D.MODULES:
        sp = m["ce_hours"] if m["delivery"] == "Self-paced online" else 0
        it = m["ce_hours"] if m["delivery"] != "Self-paced online" else 0
        sp_total += sp
        it_total += it
        rows.append([m["id"], m["delivery"], sp, it, m["interactive_feedback"],
                     m["identity_confirmation"],
                     "Yes (course-level, before certificate)",
                     "; ".join(m["compliance_flags"])[:90]])
    rows.append(["TOTAL", f"<= {D.COURSE['self_paced_cap_hours']}h self-paced cap", sp_total, it_total,
                 "Mixed", "PIN (self-paced) + live roster (live)", "Yes",
                 f"Self-paced {sp_total}h <= cap {D.COURSE['self_paced_cap_hours']}h"])
    widths = [10, 24, 13, 16, 34, 30, 22, 34]
    return headers, rows, widths


def s_lms_activity_map():
    headers = ["Activity ID", "Module ID", "Activity name", "Moodle activity type",
               "Learner action required", "Completion tracking setting", "Restrict access rule",
               "Gradebook category", "Evidence generated"]
    rows = []
    for m in D.MODULES:
        live = m["delivery"] != "Self-paced online"
        cat = m["id"]
        n = 1
        for l in m["lessons"]:
            aid = f"{m['id']}-A{n}"; n += 1
            restrict = ("Identity confirmation completed" if not live else "Join live session window")
            rows.append([aid, m["id"], l["title"], l["moodle"], l["activity_type"],
                         "View/complete + activity", restrict, cat, l["completion"]])
        # knowledge check activity
        aid = f"{m['id']}-Q"; 
        rows.append([aid, m["id"], f"{m['id']} Knowledge Check", "Quiz",
                     "Pass at >=80%", "Passing grade required",
                     "All module lessons complete", cat, "Quiz score + attempt log"])
    # course-level activities
    rows.append(["CERT-IDP", "Course", "Identity confirmation (self-paced)", "Choice/Custom + attestation",
                 "Enter PIN / attest identity", "Activity completed", "Module start",
                 "Compliance", "Identity-confirmation log"])
    rows.append(["FINAL-Q", "M12", "Cumulative Final Assessment", "Quiz (secure)",
                 "Pass at >=80%", "Passing grade required",
                 "All modules complete + identity confirmed", "Final", "Final score + attempt log"])
    rows.append(["SIGN-STMT", "Course", "Signed Completion Statement", "Assignment/e-sign",
                 "Sign & submit completion statement", "Submission required + manual accept",
                 "Final passed", "Compliance", "Signed statement record"])
    rows.append(["CERT", "Course", "Certificate (gated)", "Certificate (custom cert)",
                 "Download after all gates", "Manual / restricted",
                 "All gates + approval issued", "n/a", "Issued-certificate record"])
    widths = [11, 9, 30, 24, 26, 26, 28, 14, 28]
    return headers, rows, widths


def s_assessment_blueprint():
    headers = ["Assessment ID", "Module ID", "Question type", "Question stem", "Answer choices",
               "Correct answer", "Rationale", "Remediation guidance", "Learner visibility",
               "Admin/faculty visibility"]
    rows = []
    for m in D.MODULES:
        kc = m["knowledge_check"]
        # blueprint summary row
        rows.append([f"{m['id']}-KC", m["id"], "Knowledge check (blueprint)", kc["blueprint"], "-",
                     "-", "-", "Failed attempt -> targeted lesson re-review",
                     "Score + per-item feedback (no full key)", "Full key + item analysis"])
        for k, it in enumerate(kc["items"], start=1):
            rows.append([f"{m['id']}-KC{k}", m["id"], it["type"], it["stem"],
                         " | ".join(it["choices"]), it["answer"], it["rationale"], it["remediation"],
                         "Feedback after submission; no answer key", "Full key + rationale"])
    for it in D.FINAL_ASSESSMENT_ITEMS:
        rows.append([it["id"], it["module"], it["type"], it["stem"], it["choices"], it["answer"],
                     it["rationale"], it["remediation"], it["learner_vis"], it["admin_vis"]])
    rows.append(["FINAL-BP", "M12", "Final (blueprint)",
                 "30-item cumulative; >=80% to pass; up to 3 attempts; remediation after fail.",
                 "-", "-", "-", "Re-review modules then re-attempt",
                 "Score only; NO answer key", "Full internal answer key (separate)"])
    widths = [12, 9, 18, 40, 36, 9, 34, 30, 26, 24]
    return headers, rows, widths


def s_capstone_map():
    headers = ["Capstone task ID", "Scenario summary", "Learner deliverable", "Rubric criterion",
               "Passing standard", "Documentation expectation", "Resident privacy note", "Evidence artifact"]
    rows = [[c["id"], c["scenario"], c["deliverable"], c["rubric"], c["passing"],
             c["doc_expectation"], c["privacy"], c["evidence"]] for c in D.CAPSTONE_TASKS]
    widths = [13, 34, 30, 30, 26, 30, 26, 26]
    return headers, rows, widths


def s_application_packet():
    headers = ["Packet item", "Draft status", "Owner", "Source reference", "Missing information",
               "Approval status", "Notes"]
    rows = [[c["item"], c["status"], c["owner"], c["source"], c["missing"], c["approval"], c["notes"]]
            for c in D.APPLICATION_PACKET_CHECKLIST]
    widths = [34, 16, 30, 26, 30, 14, 30]
    return headers, rows, widths


def s_certificate_gate():
    headers = ["Gate item", "Required condition", "LMS evidence source", "Manual review needed",
               "Approval placeholder", "Risk if omitted"]
    rows = [[c["gate"], c["condition"], c["source"], c["manual"], c["approval"], c["risk"]]
            for c in D.CERTIFICATE_GATE]
    widths = [28, 34, 28, 14, 26, 34]
    return headers, rows, widths


def s_sme_flags():
    headers = ["Item ID", "Module ID", "Flag type", "Issue", "Risk level", "Reviewer needed",
               "Required action", "Status"]
    rows = []
    n = 1
    for m in D.MODULES:
        for f in m["sme_flags"]:
            rows.append([f"SME-{n:02d}", m["id"], "SME (content accuracy)", f, "Medium",
                         "RN / RCFE SME", "Verify against CDSS/Title 22 source", "Open"]); n += 1
        for f in m["compliance_flags"]:
            risk = "High" if ("identity" in f.lower() or "certificate" in f.lower() or "dementia" in f.lower()) else "Medium"
            rows.append([f"CMP-{n:02d}", m["id"], "Compliance", f, risk,
                         "Compliance reviewer", "Confirm requirement & implement control", "Open"]); n += 1
    widths = [10, 9, 22, 44, 11, 18, 30, 10]
    return headers, rows, widths


def s_audit_evidence():
    headers = ["Evidence item", "Generated by", "Stored where", "Retention period",
               "Review responsibility", "Export method", "Notes"]
    rows = [[c["item"], c["by"], c["where"], c["retention"], c["review"], c["export"], c["notes"]]
            for c in D.AUDIT_EVIDENCE_LOG]
    widths = [30, 26, 28, 24, 22, 26, 30]
    return headers, rows, widths


def main():
    wb = Workbook()
    sheet_cover(wb)
    builders = [
        ("Course_Map", s_course_map, "40-Hour Course Map", None),
        ("Module_Blueprints", s_module_blueprints, "Module/Lesson Blueprints", None),
        ("Core_Knowledge_Map", s_core_knowledge_map, "RCFE Uniform Core of Knowledge Mapping",
         "Core-area wording: Needs confirmation against current CDSS/ACS source materials."),
        ("Dementia_Hour_Map", s_dementia_hour_map, "Dementia-Hour Mapping (>=8 required)", None),
        ("Delivery_Mode_Map", s_delivery_mode_map, "Self-Paced vs Instructor-Interactive Delivery", None),
        ("LMS_Activity_Map", s_lms_activity_map, "Moodle Activity Map", None),
        ("Assessment_Blueprint", s_assessment_blueprint, "Assessment Blueprint (sample items; full bank pending SME)",
         "Sample items only. Full item bank to be authored and SME-reviewed before submission."),
        ("Capstone_Map", s_capstone_map, "Capstone Activity Map", None),
        ("Application_Packet_Checklist", s_application_packet, "Application Packet Checklist", None),
        ("Certificate_Gate", s_certificate_gate, "Certificate Gate & Completion Logic", None),
        ("SME_Compliance_Flags", s_sme_flags, "SME / Compliance Review Flags", None),
        ("Audit_Evidence_Log", s_audit_evidence, "Audit Evidence Log", None),
    ]
    for name, fn, title, note in builders:
        headers, rows, widths = fn()
        add_sheet(wb, name, headers, rows, widths, title=title, note=note)
    wb.save(OUT)
    print("Wrote", OUT)
    print("Sheets:", wb.sheetnames)


if __name__ == "__main__":
    main()
