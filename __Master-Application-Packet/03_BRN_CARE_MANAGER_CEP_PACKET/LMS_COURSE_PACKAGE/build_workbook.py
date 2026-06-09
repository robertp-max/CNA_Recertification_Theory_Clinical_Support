"""
Build the BRN Case Management CE LMS build-support workbook.

Support/defensibility artifact only. The markdown course package
(BRN_CM_CE_LMS_COURSE_PACKAGE.md) is the primary deliverable.

Status: Draft / Pending BRN CEP Approval. No CEP number has been issued.
Do not advertise, issue certificates, or represent this course as BRN-approved
until the official BRN CEP approval and provider number are issued.

Usage: python build_workbook.py
Requires: openpyxl  (pip install openpyxl)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUTFILE = "BRN_CM_CE_LMS_BUILD_WORKBOOK.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, color="1F3864", size=13)
NOTE_FONT = Font(italic=True, color="7F0000", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

BANNER = ("BRN Case Management CE Package - Draft / Pending BRN CEP Approval. "
          "Support artifact only; markdown package governs. No CEP number issued.")


def add_sheet(wb, title, headers, rows, widths=None):
    ws = wb.create_sheet(title=title)
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = BANNER
    ws["A2"].font = NOTE_FONT
    header_row = 4
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
        cell.border = BORDER
    for r, row in enumerate(rows, start=header_row + 1):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = WRAP
            cell.border = BORDER
    for c in range(1, len(headers) + 1):
        letter = get_column_letter(c)
        if widths and c <= len(widths):
            ws.column_dimensions[letter].width = widths[c - 1]
        else:
            ws.column_dimensions[letter].width = 24
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    return ws


wb = Workbook()
wb.remove(wb.active)

# ---------------------------------------------------------------- Course_Map
course_map_headers = ["Module ID", "Module Title", "Contact Hours",
                      "Estimated Minutes", "BRN Content Rationale",
                      "Direct/Indirect Patient Care Linkage",
                      "Completion Requirement", "Evidence Artifact"]
course_map_rows = [
    ["CM-M01", "Foundations of RN Case Management and Nursing Scope", 3, 165,
     "Defines RN scope/role boundaries enabling safe nurse-led case management.",
     "Indirect", "All activities + KC1 >=80%", "Activity-completion log, KC1 score, reflection"],
    ["CM-M02", "Assessment, Intake, and Patient-Centered Care Planning", 3, 170,
     "Nursing assessment and patient-centered goal setting are core RN clinical practice.",
     "Direct", "All activities + intake exercise + KC2 >=80%", "Care-plan worksheet, KC2 score"],
    ["CM-M03", "Care Coordination Across Disciplines and Settings", 3, 165,
     "Nurse-led interdisciplinary coordination protects continuity of patient care.",
     "Direct + Indirect", "All activities + coordination map + KC3 >=80%", "Coordination map, SBAR, KC3 score"],
    ["CM-M04", "Documentation, Legal Charting, and Audit-Ready Records", 3, 170,
     "Accurate nursing documentation is a legal/clinical patient-safety function.",
     "Indirect", "All activities + charting exercise + KC4 >=80%", "De-identified chart-note sample, KC4 score"],
    ["CM-M05", "Patient Rights, Consent, Advocacy, and Ethical Boundaries", 3, 165,
     "Advocacy, consent, and ethical boundaries are direct RN duties.",
     "Direct", "All activities + ethics reflection + KC5 >=80%", "Reflection submission, KC5 score"],
    ["CM-M06", "Transitions of Care, Discharge Planning, and Community Resources", 3, 170,
     "Discharge planning prevents readmission harm - direct patient care.",
     "Direct", "All activities + discharge plan + KC6 >=80%", "Discharge-plan artifact, KC6 score"],
    ["CM-M07", "Utilization Review, Medical Necessity, and Payer-Safe Communication", 3, 165,
     "UR tied to nursing judgment ensures appropriate, safe patient care.",
     "Indirect", "All activities + necessity rationale + KC7 >=80%", "Medical-necessity rationale note, KC7 score"],
    ["CM-M08", "Quality Assurance, Risk Management, Incident Reporting, and Escalation", 3, 165,
     "QA/risk/escalation are nursing patient-safety and legal-aspects content.",
     "Indirect", "All activities + incident report + KC8 >=80%", "De-identified incident report, KC8 score"],
    ["CM-M09", "Chronic Disease, Geriatric, and High-Risk Patient Management", 3, 170,
     "Applies scientific knowledge to high-risk patient populations.",
     "Direct", "All activities + risk-stratification exercise + KC9 >=80%", "Risk-stratification worksheet, KC9 score"],
    ["CM-M10", "Integrated Case Management Capstone and Final Assessment", 3, 175,
     "Synthesizes nurse-led care management across the patient journey.",
     "Direct + Indirect", "Capstone submitted + final exam >=80%", "Capstone package + final-exam record"],
    ["TOTAL", "10 modules", 30, 1680, "All modules nursing-relevant (see rationale per module).",
     "Direct + Indirect", "All modules complete + final >=80% + evaluation", "Full course audit export"],
]
add_sheet(wb, "Course_Map", course_map_headers, course_map_rows,
          widths=[10, 42, 12, 12, 40, 18, 32, 34])

# --------------------------------------------------------- Module_Blueprints
mb_headers = ["Module ID", "Lesson ID", "Lesson Title", "Learning Objective",
              "Activity Type", "Estimated Minutes", "Scenario/Exercise Summary",
              "Moodle Activity Recommendation", "Completion Condition"]
mb_rows = [
    # M01
    ["CM-M01", "CM-M01-L1", "What RN case management is (and is not)",
     "Define RN case management vs. social work and unlicensed coordination.",
     "Lesson content", 35, "Overview of nurse-led case management role.", "Book/Page", "Require view"],
    ["CM-M01", "CM-M01-L2", "Nursing scope of practice and legal boundaries",
     "Identify three RN scope boundaries constraining CM decisions.",
     "Lesson content", 40, "Scope boundary mini-cases (fictional).", "Book", "Require view"],
    ["CM-M01", "CM-M01-L3", "Roles on the care team",
     "Distinguish RN vs SW vs unlicensed roles.",
     "Lesson content", 35, "Team-role matching.", "Book/Glossary", "Require view"],
    ["CM-M01", "CM-M01-L4", "The case-management process model",
     "Describe screen-assess-plan-coordinate-monitor-evaluate.",
     "Lesson content", 35, "Walkthrough of the CM process.", "Book", "Require view"],
    ["CM-M01", "CM-M01-KC", "Knowledge check + scope/escalation note",
     "Recognize when to escalate beyond RN scope.",
     "Quiz + Assignment", 20, "Fictional 72-yr-old hip-recovery referral asking for med change.",
     "Quiz + Assignment", "KC >=80% + note submitted"],
    # M02
    ["CM-M02", "CM-M02-L1", "Intake workflow & consent-to-assess",
     "Conduct structured intake using a nursing framework.",
     "Lesson content", 35, "Intake sequence.", "Book", "Require view"],
    ["CM-M02", "CM-M02-L2", "Nursing assessment frameworks / SDOH",
     "Differentiate subjective vs objective data; screen SDOH.",
     "Lesson content", 40, "Fictional diabetic client with transport barrier.", "Book/Glossary", "Require view"],
    ["CM-M02", "CM-M02-L3", "From data to nursing problems",
     "Translate assessment data into nursing problems.",
     "Lesson content", 35, "Problem identification practice.", "Book", "Require view"],
    ["CM-M02", "CM-M02-L4", "Writing measurable, patient-centered goals",
     "Write >=2 measurable patient-centered goals.",
     "Lesson content", 40, "Goal-writing with bias check.", "Book", "Require view"],
    ["CM-M02", "CM-M02-KC", "Knowledge check + care-plan worksheet",
     "Build a care plan from assessment findings.",
     "Quiz + Assignment", 20, "Care-Plan Worksheet: 3 problems, 2 goals, 1 SDOH intervention, bias note.",
     "Quiz + Assignment", "KC >=80% + worksheet submitted"],
    # M03
    ["CM-M03", "CM-M03-L1", "Interdisciplinary team & RN coordinating role",
     "Map team members and contributions.",
     "Lesson content", 35, "Team map.", "Book/Wiki", "Require view"],
    ["CM-M03", "CM-M03-L2", "Structured communication / SBAR",
     "Apply SBAR to a coordination event.",
     "Lesson content", 40, "SBAR practice.", "Book", "Require view"],
    ["CM-M03", "CM-M03-L3", "Cross-setting coordination",
     "Identify transition failure points + mitigations.",
     "Branching scenario", 35, "Fictional hospital-to-home-health med-rec gap.", "Lesson (branching)", "Require view/complete"],
    ["CM-M03", "CM-M03-L4", "Closing the loop & follow-up tracking",
     "Apply minimum-necessary sharing; track follow-up.",
     "Lesson content", 35, "Loop-closure tracking.", "Book", "Require view"],
    ["CM-M03", "CM-M03-KC", "Knowledge check + coordination map",
     "Produce an interdisciplinary coordination map + SBAR.",
     "Quiz + Assignment", 20, "Coordination Map (who/what/when/loop) + 3-line SBAR.",
     "Quiz + Assignment", "KC >=80% + map submitted"],
    # M04
    ["CM-M04", "CM-M04-L1", "Purpose & legal weight of the record",
     "Apply documentation standards.",
     "Lesson content", 35, "Why the record matters.", "Book", "Require view"],
    ["CM-M04", "CM-M04-L2", "Standards & charting do/don't",
     "Identify audit-proof vs weak documentation.",
     "Lesson content", 40, "Do/don't examples.", "Book", "Require view"],
    ["CM-M04", "CM-M04-L3", "Error correction, late entries, addenda",
     "Correct a non-compliant entry properly.",
     "Lesson content", 35, "Correction conventions.", "Book", "Require view"],
    ["CM-M04", "CM-M04-L4", "Audit-readiness & retention",
     "Explain 4-year retention posture.",
     "Lesson content", 35, "Retention overview.", "Book", "Require view"],
    ["CM-M04", "CM-M04-KC", "Knowledge check + chart-note rewrite",
     "Rewrite a weak note to audit-ready standard.",
     "Quiz + Assignment", 25, "Rewrite 'pt seems fine, will follow up' into compliant de-identified note + addendum.",
     "Quiz + Assignment", "KC >=80% + note submitted"],
    # M05
    ["CM-M05", "CM-M05-L1", "Patient rights & confidentiality",
     "Describe core patient rights in CM.",
     "Lesson content", 35, "Rights overview.", "Book", "Require view"],
    ["CM-M05", "CM-M05-L2", "Informed consent / refusal",
     "Distinguish consent, assent, refusal; RN role.",
     "Lesson content", 35, "Consent scenarios.", "Book", "Require view"],
    ["CM-M05", "CM-M05-L3", "Advocacy in practice",
     "Apply advocacy to a boundary conflict.",
     "Lesson content", 35, "Advocacy vs autonomy.", "Book/Choice", "Require view"],
    ["CM-M05", "CM-M05-L4", "Ethical boundaries & conflicts of interest",
     "Apply an ethical decision framework.",
     "Lesson content", 40, "Boundary conflict.", "Book", "Require view"],
    ["CM-M05", "CM-M05-KC", "Knowledge check + ethics reflection",
     "Apply a framework + document consent/refusal.",
     "Quiz + Assignment", 20, "Fictional client refuses service family wants; 250-350 word reflection + consent note.",
     "Quiz + Assignment", "KC >=80% + reflection submitted"],
    # M06
    ["CM-M06", "CM-M06-L1", "Transition risk & readmission drivers",
     "Identify readmission drivers.",
     "Lesson content", 35, "Risk drivers.", "Book", "Require view"],
    ["CM-M06", "CM-M06-L2", "Discharge-plan components",
     "Build a discharge/transition plan.",
     "Lesson content", 40, "Plan components.", "Book", "Require view"],
    ["CM-M06", "CM-M06-L3", "Teach-back & caregiver readiness",
     "Apply teach-back to confirm understanding.",
     "Lesson content", 35, "Teach-back practice.", "Book", "Require view"],
    ["CM-M06", "CM-M06-L4", "Community resources & referrals",
     "Match >=3 resources to needs.",
     "Lesson content", 35, "Fictional resource directory.", "Book/Database", "Require view"],
    ["CM-M06", "CM-M06-KC", "Knowledge check + discharge plan",
     "Complete a discharge/transition plan.",
     "Quiz + Assignment", 25, "Fictional client home with new equipment + working caregiver; identify gaps/resources.",
     "Quiz + Assignment", "KC >=80% + plan submitted"],
    # M07
    ["CM-M07", "CM-M07-L1", "What UR is and the RN's lane",
     "Define medical necessity and RN's supporting role.",
     "Lesson content", 40, "UR fundamentals.", "Book", "Require view"],
    ["CM-M07", "CM-M07-L2", "Medical-necessity criteria & evidence",
     "Apply criteria-based review.",
     "Lesson content", 40, "Fictional extended-visit request.", "Book", "Require view"],
    ["CM-M07", "CM-M07-L3", "Documentation that supports necessity",
     "Write a defensible necessity rationale.",
     "Lesson content", 35, "Rationale structure.", "Book", "Require view"],
    ["CM-M07", "CM-M07-L4", "Payer-safe, accurate communication",
     "Communicate without misrepresentation/overreach.",
     "Lesson content", 30, "Payer-safe phrasing.", "Book", "Require view"],
    ["CM-M07", "CM-M07-KC", "Knowledge check + necessity rationale",
     "Draft medical-necessity rationale.",
     "Quiz + Assignment", 20, "Clinical facts -> criteria -> conclusion + payer-safe snippet (fictional).",
     "Quiz + Assignment", "KC >=80% + rationale submitted"],
    # M08
    ["CM-M08", "CM-M08-L1", "QA fundamentals & indicators",
     "Identify CM quality indicators.",
     "Lesson content", 35, "QA basics.", "Book", "Require view"],
    ["CM-M08", "CM-M08-L2", "Risk identification",
     "Recognize risks in CM.",
     "Lesson content", 35, "Risk spotting.", "Book", "Require view"],
    ["CM-M08", "CM-M08-L3", "Incident reporting standards",
     "Complete a compliant incident report.",
     "Lesson content", 40, "Reporting standards.", "Book", "Require view"],
    ["CM-M08", "CM-M08-L4", "Escalation & just culture",
     "Apply an escalation pathway; just-culture response.",
     "Lesson content", 35, "Escalation pathways.", "Book", "Require view"],
    ["CM-M08", "CM-M08-KC", "Knowledge check + incident report",
     "Report and escalate a near-miss.",
     "Quiz + Assignment", 20, "Fictional near-miss med discrepancy on home visit; de-identified incident report + escalation note.",
     "Quiz + Assignment", "KC >=80% + report submitted"],
    # M09
    ["CM-M09", "CM-M09-L1", "Risk stratification",
     "Apply risk stratification to a panel.",
     "Lesson content", 40, "Stratification method.", "Book", "Require view"],
    ["CM-M09", "CM-M09-L2", "Chronic-disease management priorities",
     "Describe nursing priorities for chronic conditions.",
     "Lesson content", 40, "Chronic care priorities.", "Book", "Require view"],
    ["CM-M09", "CM-M09-L3", "Geriatric syndromes & polypharmacy",
     "Identify geriatric risks + mitigations.",
     "Lesson content", 35, "Falls, polypharmacy, cognition.", "Book", "Require view"],
    ["CM-M09", "CM-M09-L4", "Monitoring plans & triggers",
     "Build a monitoring plan with escalation triggers.",
     "Lesson content", 35, "Trigger setting.", "Book", "Require view"],
    ["CM-M09", "CM-M09-KC", "Knowledge check + risk-stratification worksheet",
     "Stratify risk and set monitoring triggers.",
     "Quiz + Assignment", 20, "Fictional 80-yr-old, multi-condition, 9 meds; stratify + >=3 triggers.",
     "Quiz + Assignment", "KC >=80% + worksheet submitted"],
    # M10
    ["CM-M10", "CM-M10-L1", "Capstone briefing & rubric walkthrough",
     "Integrate all domains into one case.",
     "Briefing", 30, "Capstone briefing.", "Page", "Require view"],
    ["CM-M10", "CM-M10-L2", "Building the integrated case package",
     "Produce an audit-ready de-identified package.",
     "Capstone assignment", 60, "End-to-end fictional home-health case.", "Assignment (rubric)", "Require submission + grade"],
    ["CM-M10", "CM-M10-L3", "Self-review against rubric",
     "Self-assess against rubric.",
     "Self-check", 25, "Rubric self-review.", "Page/Checklist", "Require view"],
    ["CM-M10", "CM-M10-FE", "Final cumulative assessment",
     "Demonstrate >=80% cumulative mastery.",
     "Final exam", 60, "40-50 item cumulative exam, bank draw.", "Quiz (gated)", "Pass >=80%"],
]
add_sheet(wb, "Module_Blueprints", mb_headers, mb_rows,
          widths=[10, 12, 34, 38, 16, 12, 44, 20, 22])

# ----------------------------------------------------------- LMS_Activity_Map
am_headers = ["Activity ID", "Module ID", "Activity Name", "Moodle Activity Type",
              "Learner Action Required", "Completion Tracking Setting",
              "Restrict Access Rule", "Gradebook Category", "Evidence Generated"]
am_rows = []
module_titles_short = {
    "CM-M01": "Foundations/Scope", "CM-M02": "Assessment/Care Plan",
    "CM-M03": "Care Coordination", "CM-M04": "Documentation",
    "CM-M05": "Rights/Ethics", "CM-M06": "Transitions",
    "CM-M07": "Utilization Review", "CM-M08": "QA/Risk",
    "CM-M09": "Chronic/Geriatric", "CM-M10": "Capstone/Final",
}
for i in range(1, 10):
    mid = f"CM-M{i:02d}"
    prior = f"CM-M{i-1:02d}" if i > 1 else "Start Here"
    am_rows.append([f"ACT-{mid}-LSN", mid, f"{module_titles_short[mid]} lessons", "Book/Page/Lesson",
                    "View all lesson pages", "Require view",
                    f"Completion of {prior}", "Not graded (view)", "Activity completion timestamp"])
    am_rows.append([f"ACT-{mid}-KC", mid, f"{module_titles_short[mid]} knowledge check", "Quiz",
                    "Pass quiz >=80% (unlimited attempts)", "Require passing grade",
                    "Completion of module lessons", "Knowledge Checks", "Quiz score + attempt log"])
    am_rows.append([f"ACT-{mid}-EX", mid, f"{module_titles_short[mid]} applied exercise", "Assignment",
                    "Submit documentation exercise", "Require submission (+grade)",
                    "Completion of knowledge check", "Applied Exercises", "Submission + faculty score"])
# Module 10 / capstone / final / certificate
am_rows.append(["ACT-CM-M10-BRIEF", "CM-M10", "Capstone briefing", "Page",
                "View briefing & rubric", "Require view", "Completion of CM-M09",
                "Not graded (view)", "Activity completion timestamp"])
am_rows.append(["ACT-CM-CAP-01", "CM-M10", "Capstone documentation package", "Assignment (rubric)",
                "Submit integrated case package + no-PHI attestation", "Require submission + grade",
                "Completion of capstone briefing", "Capstone", "Capstone submission + rubric scoring + attestation"])
am_rows.append(["ACT-CM-FINAL", "CM-M10", "Final cumulative exam", "Quiz (bank draw)",
                "Pass >=80% (max 3 attempts)", "Require passing grade",
                "Modules 1-9 complete + capstone submitted", "Final Exam", "Quiz attempts + item analysis"])
am_rows.append(["ACT-EVAL", "Course", "Course evaluation", "Feedback",
                "Complete evaluation + attestation", "Require submission",
                "Final exam passed", "Not graded", "Evaluation responses"])
am_rows.append(["ACT-CERT", "Course", "Certificate of completion (gated)", "Certificate/Custom cert",
                "Download after gate met", "Require all course criteria",
                "Course completion + evaluation submitted", "Not graded", "Issued-certificate record"])
add_sheet(wb, "LMS_Activity_Map", am_headers, am_rows,
          widths=[18, 10, 30, 22, 34, 24, 30, 18, 34])

# ------------------------------------------------------- Assessment_Blueprint
ab_headers = ["Assessment ID", "Module ID", "Question Type", "Question Stem",
              "Answer Choices", "Correct Answer (ADMIN ONLY)", "Rationale (ADMIN ONLY)",
              "Remediation Guidance", "Learner Visibility", "Admin/Faculty Visibility"]
ab_rows = [
    ["KC1-Q1", "CM-M01", "MCQ",
     "A home-health referral asks the RN case manager to 'adjust the pain medication.' What is the correct first action?",
     "A) Adjust the dose; B) Decline and route to the prescriber/RN supervisor; C) Ask the family to decide; D) Ignore the referral",
     "B", "Medication changes are outside RN case-management scope; route to prescriber.",
     "Review CM-M01-L2 (scope boundaries).", "Feedback only (no key)", "Full key + rationale"],
    ["KC2-Q1", "CM-M02", "Multi-select",
     "Which are examples of objective assessment data? (select all)",
     "A) Reported pain 7/10; B) BP reading; C) Observed gait; D) 'Feels tired'",
     "B, C", "Objective data are measurable/observed; A and D are subjective.",
     "Review CM-M02-L2 (subjective vs objective).", "Feedback only", "Full key + rationale"],
    ["KC3-Q1", "CM-M03", "MCQ",
     "Which tool best structures a nurse-to-nurse handoff?",
     "A) SOAP; B) SBAR; C) PDSA; D) SWOT",
     "B", "SBAR is a structured handoff communication tool.",
     "Review CM-M03-L2 (SBAR).", "Feedback only", "Full key + rationale"],
    ["KC4-Q1", "CM-M04", "MCQ",
     "Which note best meets audit-ready documentation standards?",
     "A) 'Pt seems fine, will follow up'; B) Objective findings, action, plan, time, signature; C) 'No issues'; D) 'See chart'",
     "B", "Audit-ready notes are objective, complete, timed, attributable.",
     "Review CM-M04-L2 (standards).", "Feedback only", "Full key + rationale"],
    ["KC5-Q1", "CM-M05", "MCQ",
     "A competent client refuses a service the family wants. The RN's role is to:",
     "A) Override the client; B) Support informed autonomy and document; C) Withdraw care; D) Defer to family",
     "B", "RN advocates for informed autonomy and documents the refusal.",
     "Review CM-M05-L2/L3 (consent/advocacy).", "Feedback only", "Full key + rationale"],
    ["KC6-Q1", "CM-M06", "MCQ",
     "Which best confirms a caregiver understands new equipment instructions?",
     "A) Ask 'Do you understand?'; B) Teach-back demonstration; C) Hand them a pamphlet; D) Assume understanding",
     "B", "Teach-back verifies understanding; yes/no questions do not.",
     "Review CM-M06-L3 (teach-back).", "Feedback only", "Full key + rationale"],
    ["KC7-Q1", "CM-M07", "MCQ",
     "When supporting medical necessity, the RN case manager should:",
     "A) Decide coverage; B) Document clinical facts against criteria; C) Tell the payer what they want; D) Omit unfavorable facts",
     "B", "RN supports necessity with accurate clinical documentation; does not determine coverage.",
     "Review CM-M07-L2/L3.", "Feedback only", "Full key + rationale"],
    ["KC8-Q1", "CM-M08", "MCQ",
     "On finding a near-miss medication discrepancy, the RN should first:",
     "A) Ignore it (no harm); B) Address safety then file an incident report and escalate; C) Blame the prior nurse; D) Delete the note",
     "B", "Patient safety first, then report and escalate within a just-culture framework.",
     "Review CM-M08-L3/L4.", "Feedback only", "Full key + rationale"],
    ["KC9-Q1", "CM-M09", "Multi-select",
     "Which raise a geriatric client's risk profile? (select all)",
     "A) Polypharmacy (9 meds); B) Recent falls; C) Cognitive changes; D) Stable single condition",
     "A, B, C", "A-C are high-risk geriatric markers; D is lower risk.",
     "Review CM-M09-L1/L3.", "Feedback only", "Full key + rationale"],
    ["FINAL-SCN1", "CM-M10", "Scenario",
     "Given a fictional end-to-end case, select the sequence that best protects a safe transition.",
     "Four ordered-sequence options (A-D)",
     "Per validated key (TBD by SME)", "Integration of coordination, documentation, teach-back, escalation.",
     "Route to weakest objective area; require re-do before retry.", "Pass/fail + objective feedback", "Full key + item analysis"],
    ["FINAL-BANK", "CM-M10", "Bank (40-50 items)",
     "Cumulative bank: ~4-5 items per module, randomized draw.",
     "Mixed MCQ / multi-select / T-F / scenario",
     "Maintained admin-only; SME-validated", "Per item; SME-authored rationales.",
     "Remediation routing between attempts; lock after 3.", "Pass/fail only", "Full key + analytics"],
]
ws_ab = add_sheet(wb, "Assessment_Blueprint", ab_headers, ab_rows,
                  widths=[14, 10, 16, 44, 40, 20, 36, 30, 20, 22])
# Admin-only warning row
warn_row = ws_ab.max_row + 2
ws_ab.cell(row=warn_row, column=1,
           value=("INTERNAL ANSWER KEY - ADMIN/FACULTY ONLY. Do NOT expose the "
                  "'Correct Answer' / 'Rationale' columns to learners. Items require SME validation."))
ws_ab.cell(row=warn_row, column=1).font = NOTE_FONT

# --------------------------------------------------------------- Capstone_Map
cap_headers = ["Capstone Task ID", "Scenario Summary", "Learner Deliverable",
               "Rubric Criterion", "Passing Standard", "Documentation Expectation",
               "PHI Safety Note", "Evidence Artifact"]
SCN = "Fictional, de-identified end-to-end home-health case: high-risk client from intake to safe transition."
PASS = "Meets minimum on every criterion; overall rubric >=80%."
PHI = "Fictional data only; no PHI; no-PHI attestation + faculty manual review."
cap_rows = [
    ["CM-CAP-01-A", SCN, "Intake + nursing assessment summary", "Clinical accuracy", PASS,
     "Objective, complete, internally consistent", PHI, "Capstone package + rubric record"],
    ["CM-CAP-01-B", SCN, "Patient-centered care plan w/ measurable goals", "Nursing-scope appropriateness", PASS,
     "Goals measurable & patient-centered", PHI, "Capstone package + rubric record"],
    ["CM-CAP-01-C", SCN, "Interdisciplinary coordination map + SBAR", "Coordination/communication", PASS,
     "Roles, loop closure, structured handoff", PHI, "Capstone package + rubric record"],
    ["CM-CAP-01-D", SCN, "Audit-ready chart note", "Documentation quality/audit-readiness", PASS,
     "Standards-compliant, attributable", PHI, "Capstone package + rubric record"],
    ["CM-CAP-01-E", SCN, "Consent/advocacy + ethics reflection", "Advocacy/ethics & bias-awareness", PASS,
     "Framework applied; bias addressed", PHI, "Capstone package + rubric record"],
    ["CM-CAP-01-F", SCN, "Discharge/transition plan w/ teach-back", "Transition safety", PASS,
     "Meds, follow-up, equipment, caregiver, teach-back", PHI, "Capstone package + rubric record"],
    ["CM-CAP-01-G", SCN, "Medical-necessity rationale", "Nursing-scope appropriateness", PASS,
     "Facts -> criteria -> conclusion", PHI, "Capstone package + rubric record"],
    ["CM-CAP-01-H", SCN, "Incident/escalation note", "Clinical accuracy / safety", PASS,
     "Compliant report + escalation path", PHI, "Capstone package + rubric record"],
    ["CM-CAP-01-I", SCN, "Risk-stratification + monitoring plan", "Completeness/integration", PASS,
     ">=3 escalation triggers; coherent plan", PHI, "Capstone package + rubric record"],
]
add_sheet(wb, "Capstone_Map", cap_headers, cap_rows,
          widths=[16, 40, 36, 28, 30, 32, 34, 26])

# --------------------------------------------- Application_Packet_Checklist
apc_headers = ["Packet Item", "Draft Status", "Owner", "Source Reference",
               "Missing Information", "Approval Status", "Notes"]
apc_rows = [
    ["Course Information Form support", "Drafted", "SME/Compliance", "Section K; BRN_TRANSFER_WORKSHEET.md",
     "Offer dates; reconcile 30hr vs 4hr packet course", "Pending BRN CEP", "Offering type/contact hours differ from packet"],
    ["Instructor Information checklist", "Drafted", "Records custodian", "Section L; BRN_INSTRUCTOR_EVIDENCE_INDEX.md",
     "License expiry, education, experience, teaching, adult-ed", "Pending verification", "Vanessa Valerio RN 788389"],
    ["Flyer/Brochure copy", "Drafted", "Compliance/Legal", "Section M; FLYER_BROCHURE_DRAFT.md",
     "Final provider number (after approval)", "Pending BRN CEP", "Placeholder approval language only"],
    ["Sample certificate text", "Drafted", "Compliance/Legal", "Section N; SAMPLE_CERTIFICATE_DRAFT.md",
     "Real CEP number", "Pending BRN CEP", "4-yr retention; no blank certificates"],
    ["Course outline", "Drafted", "SME", "Sections D & E", "SME sign-off", "Pending SME", "30hr / 10 modules"],
    ["Course objectives", "Drafted", "SME", "Section E", "SME sign-off", "Pending SME", "Measurable per module"],
    ["Course evaluation form", "Drafted", "Compliance", "Section O", "None", "Pending", "Moodle Feedback"],
    ["Attendance/completion/proof policy", "Drafted", "Compliance", "Section J & P; POLICIES_PACKET.md",
     "None", "Pending legal sign-off", "Time-on-task via LMS logs"],
    ["Recordkeeping policy", "Drafted (packet)", "Records custodian", "Section P; POLICIES_PACKET.md",
     "None", "Pending legal sign-off", ">=4 yr; custodian Yadvir Saandal"],
    ["Refund/cancellation policy", "Drafted (packet)", "Legal", "Section M & P; POLICIES_PACKET.md",
     "None", "Pending legal sign-off", "Pro-rata; 45-day refund"],
    ["Advertising compliance checklist", "Drafted", "Compliance", "Section R", "None", "Pending", "No CEU/approval claims"],
    ["LMS evidence checklist", "Drafted", "LMS admin", "Sections F & P", "None", "Pending", "Audit export plan"],
    ["SME/compliance review checklist", "Drafted", "SME/Compliance", "Sections R & S", "SME sign-off", "Pending", "High-scrutiny: M07, M08"],
    ["Owner approval before submission", "Drafted", "Executive/Authorized signer", "Section U; 08_DO_NOT_SUBMIT_UNTIL.md",
     "Exec go-ahead; payment; signature", "Pending", "$750 fee; do-not-submit gates"],
]
add_sheet(wb, "Application_Packet_Checklist", apc_headers, apc_rows,
          widths=[30, 18, 22, 36, 32, 18, 34])

# ------------------------------------------------------------- Certificate_Gate
cg_headers = ["Gate Item", "Required Condition", "LMS Evidence Source",
              "Manual Review Needed", "Approval Placeholder", "Risk If Omitted"]
cg_rows = [
    ["All modules complete", "CM-M01..CM-M10 activity completion = done", "Activity completion report",
     "No", "N/A", "Incomplete instruction; contact-hour claim invalid"],
    ["Knowledge checks passed", "Each KC >=80%", "Gradebook", "No", "N/A", "Mastery not demonstrated"],
    ["Capstone meets rubric", "Minimum met on all criteria", "Assignment grade", "Yes (faculty)",
     "N/A", "Competency not verified"],
    ["Final exam passed", "Final >=80%", "Quiz grade", "No", "N/A", "Summative competency not met"],
    ["Evaluation submitted", "Course evaluation complete", "Feedback activity", "No", "N/A", "Missing evaluation evidence"],
    ["Time/participation", "Completion timestamps + attempt durations present", "Logs/completion", "Spot-check",
     "N/A", "Weak time-on-task evidence"],
    ["No-PHI attestation", "Learner attestation + faculty PHI review", "Assignment attestation", "Yes",
     "N/A", "Privacy exposure"],
    ["Approval placeholder", "CEP number field shows placeholder until issued", "Certificate template", "Yes",
     "[CEP # PENDING - placeholder until issued]", "Misrepresentation of approval"],
    ["Retention statement", "Certificate shows 4-year retention notice", "Certificate template", "Yes",
     "N/A", "Recordkeeping non-compliance"],
]
add_sheet(wb, "Certificate_Gate", cg_headers, cg_rows,
          widths=[24, 38, 26, 18, 34, 36])

# --------------------------------------------------------- SME_Compliance_Flags
scf_headers = ["Item ID", "Module ID", "Flag Type", "Issue", "Risk Level",
               "Reviewer Needed", "Required Action", "Status"]
scf_rows = [
    ["FLAG-001", "All", "Scope reconciliation",
     "Packet documents 4hr live webinar; this design is 30hr online self-paced.",
     "High", "SME + Compliance + Exec", "Reconcile & re-approve before BRN submission", "Open"],
    ["FLAG-002", "CM-M07", "BRN content fit",
     "UR/medical-necessity could read as administrative, not patient care.",
     "High", "SME + Compliance", "Keep nursing-judgment/patient-safety thread explicit", "Open"],
    ["FLAG-003", "CM-M08", "BRN content fit",
     "QA/risk could read as administrative, not patient care.",
     "High", "SME + Compliance", "Anchor to patient-safety/nursing practice", "Open"],
    ["FLAG-004", "CM-M04", "Legal aspects",
     "Documentation/charting legal accuracy.",
     "Medium", "SME + Legal", "Verify charting/correction standards", "Open"],
    ["FLAG-005", "CM-M05", "Legal/ethics",
     "Consent/advocacy accuracy; avoid legal-advice overreach.",
     "Medium", "SME + Legal", "Verify consent/advocacy content", "Open"],
    ["FLAG-006", "CM-M09", "Clinical currency",
     "Chronic/geriatric clinical content must be current; education-only.",
     "Medium", "SME", "Verify clinical currency + add education-only note", "Open"],
    ["FLAG-007", "CM-M10", "Assessment validity",
     "Final exam + capstone rubric require validated items/key.",
     "High", "SME", "Author + validate items and internal key", "Open"],
    ["FLAG-008", "All", "Instructor evidence",
     "Vanessa Valerio evidence (expiry, education, experience, teaching, adult-ed) incomplete.",
     "Medium", "Records custodian", "Attach instructor evidence packet", "Open"],
    ["FLAG-009", "All", "Privacy/PHI",
     "Learner free-text/capstone could introduce PHI.",
     "Medium", "Compliance/Privacy", "No-PHI banners, attestation, manual review", "Mitigated-design"],
    ["FLAG-010", "All", "Advertising",
     "Risk of premature approval/CEU language.",
     "High", "Compliance", "Placeholder-only language until CEP issued", "Mitigated-design"],
    ["FLAG-011", "All", "Accessibility",
     "WCAG 2.1 AA must be verified on built course.",
     "Medium", "Accessibility", "Run accessibility + mobile QA", "Open"],
    ["FLAG-012", "All", "SME of record",
     "Confirm SME for full 30hr scope (Vanessa or additional).",
     "Medium", "SME/Exec", "Confirm/assign SME of record", "Open"],
]
add_sheet(wb, "SME_Compliance_Flags", scf_headers, scf_rows,
          widths=[10, 10, 20, 44, 12, 22, 36, 18])

# ----------------------------------------------------------- Audit_Evidence_Log
ael_headers = ["Evidence Item", "Generated By", "Stored Where", "Retention Period",
               "Review Responsibility", "Export Method", "Notes"]
ael_rows = [
    ["Activity completion report", "Moodle completion", "LMS + secured backup", ">=4 years",
     "Records custodian (Yadvir Saandal)", "CSV/PDF export", "No PHI"],
    ["Knowledge-check scores", "Moodle gradebook", "LMS", ">=4 years", "Faculty/admin", "Gradebook export", "No PHI"],
    ["Assignment submissions", "Moodle Assignment", "LMS", ">=4 years", "Faculty", "Bulk download", "Fictional data only"],
    ["Capstone + rubric scoring", "Moodle Assignment", "LMS", ">=4 years", "Faculty/SME", "PDF export", "Manual PHI review"],
    ["Final-exam attempt records", "Moodle Quiz", "LMS", ">=4 years", "Faculty/admin", "Quiz report export", "No key exposure"],
    ["Time-on-task logs", "Moodle logs", "LMS", ">=4 years", "Admin", "Logs export", "Time evidence"],
    ["Evaluation responses", "Moodle Feedback", "LMS", ">=4 years", "Admin", "Feedback export", "Includes attestation"],
    ["Issued certificate record", "Certificate activity", "LMS + records file", ">=4 years", "Records custodian",
     "PDF + register", "No blank certificates; 4-yr retention notice"],
    ["No-PHI attestations", "Assignment attestation", "LMS", ">=4 years", "Faculty", "Export", "Required for capstone"],
    ["Provider/CE records (overall)", "Provider", "419 E. Hamilton Ave., Campbell, CA 95008",
     ">=4 years (institutional 5 yr)", "Records custodian", "Records file", "Per POLICIES_PACKET.md"],
]
add_sheet(wb, "Audit_Evidence_Log", ael_headers, ael_rows,
          widths=[28, 22, 34, 22, 26, 20, 30])

wb.save(OUTFILE)
print(f"Wrote {OUTFILE} with sheets: {wb.sheetnames}")
