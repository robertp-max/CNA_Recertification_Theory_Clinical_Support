from __future__ import annotations

import datetime as dt
import re
import shutil
import textwrap
import urllib.request
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.colors import HexColor
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.graphics import renderPDF
from svglib.svglib import svg2rlg


ROOT = Path(r"C:\AI\Git\CNA_Recertification_Theory_Clinical_Support")
COMBINED = ROOT / "RCFE CETP Vendor Packet for CDSS"
ARCHIVE = ROOT / f"RCFE CETP Vendor Packet for CDSS_ARCHIVE_COMBINED_{dt.date.today():%Y%m%d}"
VENDOR = ROOT / "RCFE CETP Vendor Application - LIC 9141"
COURSE = ROOT / "RCFE CETP Course Approval Applications - LIC 9140"
PENDING = "[PENDING CONFIRMATION]"
PENDING_CDSS = "[PENDING CDSS CONFIRMATION]"
TODAY = f"{dt.date.today():%B} {dt.date.today().day}, {dt.date.today().year}"

INST = {
    "name": "CI Institute of Nursing",
    "legal": "CI INSTITUTE OF NURSING",
    "entity_type": PENDING,
    "website": "ciinstituteofnursing.com",
    "physical": "419 E Hamilton Ave, Campbell, CA 95008",
    "mailing": PENDING,
    "records": "419 E Hamilton Ave, Campbell, CA 95008",
    "phone": "(650) 449-6706",
    "email": PENDING,
    "prepared_by": "TJ Padilla",
    "oversight": "Dee Bustos",
}

COURSES = [
    ("RCFE-CETP-001", "RCFE Laws, Regulations, Policies, and Procedural Standards", 4, "Laws, Regulations, Policies, and Procedural Standards Impacting RCFE"),
    ("RCFE-CETP-002", "Alzheimer's Disease and Related Dementias: Person-Centered Care Foundations", 4, "Managing Alzheimer's Disease and Related Dementias"),
    ("RCFE-CETP-003", "Alzheimer's Disease and Related Dementias: Safety, Environment, and Documentation", 4, "Managing Alzheimer's Disease and Related Dementias"),
    ("RCFE-CETP-004", "Resident Rights, Dignity, Councils, and Abuse Prevention", 2, "Residents' Rights"),
    ("RCFE-CETP-005", "Medication Management in RCFE Operations", 4, "Medication Management"),
    ("RCFE-CETP-006", "Admission, Retention, Reappraisal, and Needs and Services Plans", 3, "Resident Admission, Retention, and Assessment Procedures"),
    ("RCFE-CETP-007", "Emergency Procedures and Physical Environment Controls", 2, "Managing the Physical Environment"),
    ("RCFE-CETP-008", "Staff Supervision, Training Records, and Administrator Accountability", 2, "Management/Supervision of Staff"),
    ("RCFE-CETP-009", "RCFE Business Operations, Records, and Claim-Safe Communications", 2, "Business Operations"),
]
TOTAL_HOURS = sum(c[2] for c in COURSES)
TOTAL_COURSE_FEES = TOTAL_HOURS * 10

SOURCE_URLS = [
    ("CDSS Vendor Information page", "https://www.cdss.ca.gov/inforesources/community-care-licensing/administrator-certification/vendor-information"),
    ("CDSS Certification Forms page", "https://www.cdss.ca.gov/inforesources/community-care/administrator-certification/certification-forms"),
    ("CDSS Vendor FAQ", "http://cdss.ca.gov/inforesources/community-care/administrator-certification/administrator-information/faq-vendors"),
    ("LIC 9141 Vendor Application/Renewal", "https://www.cdss.ca.gov/Portals/9/CCL/ACP/2021/LIC9141%28fillable%29.PDF"),
    ("LIC 9140 Request for Course Approval", "https://cdss.ca.gov/Portals/9/CCL/ACP/2023/LIC9140%28fillable%29.pdf?ver=2024-04-16-083715-430"),
    ("LIC 9140A Request to Add or Replace Instructor", "https://www.cdss.ca.gov/Portals/9/180427%20LIC%209140A%20%28Fillable%29.pdf?ver=2018-05-01-102125-973"),
    ("LIC 9139 Renewal of Continuing Education Course Approval", "https://www.cdss.ca.gov/Portals/9/CCL/ACP/2021/LIC9139%28fillable%29.PDF"),
    ("LIC 9142A Roster of Participants", "https://www.cdss.ca.gov/Portals/9/180427%20LIC%209142A%20%28Fillable%29.pdf?ver=2018-05-01-102950-963"),
    ("Vendor Course Schedule Notification", "https://www.cdss.ca.gov/Portals/9/CCL/ACP/2023/CourseScheduleNotification.pdf?ver=2024-07-09-130539-477"),
    ("Notice of Payment Information: Vendor Fees", "https://www.cdss.ca.gov/Portals/9/CCL/ACP/2023/NoticeofPaymentInformationVendorFees.pdf?ver=2024-07-09-130540-180"),
    ("Notice of Payment Information: Course Fees", "https://www.cdss.ca.gov/Portals/9/CCL/ACP/2023/NoticeofPaymentInformationCourseFees.pdf?ver=2024-07-09-130539-853"),
    ("RCFE Core of Knowledge", "https://www.cdss.ca.gov/PORTALS/9/CCL/ACP/2019/CoreofKnowledgeRCFE.pdf"),
    ("ACB Guide for Course Review", "https://www.cdss.ca.gov/Portals/9/CCL/ACP/2023/ACBGuideforCourseReview.pdf?ver=2024-07-07-120508-043"),
    ("Sample Course Outline", "https://www.cdss.ca.gov/Portals/9/CCL/ACP/2019/200505%20Sample%20Course%20Outline_PDF.pdf"),
    ("Sample Certificate of Completion", "https://www.cdss.ca.gov/Portals/9/ACS%20Sample%20Certificate%20of%20Completion.pdf?ver=2018-10-05-144756-173"),
    ("Vendor Automation Platform", "https://ca-dss.prod.simpligov.com"),
]


def md_table(headers, rows):
    out = ["| " + " | ".join(headers) + " |", "| " + " | ".join(["---"] * len(headers)) + " |"]
    for row in rows:
        out.append("| " + " | ".join(str(v).replace("|", "/").replace("\n", " ") for v in row) + " |")
    return "\n".join(out)


def write(path: Path, text: str):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(textwrap.dedent(text).strip() + "\n", encoding="utf-8")


def copy_if_exists(src: Path, dst: Path):
    if src.exists():
        try:
            if src.resolve() == dst.resolve():
                return
        except FileNotFoundError:
            pass
        dst.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(src, dst)


def slug(text: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "_", text).strip("_")


def section_from_packet(packet: str, heading: str) -> str:
    pattern = rf"### {re.escape(heading)}\n\n(.*?)(?=\n### |\Z)"
    match = re.search(pattern, packet, flags=re.S)
    return match.group(1).strip() if match else PENDING


def official_research_index(package_name: str) -> str:
    return f"""
    # CDSS Source Research Index - {package_name}

    Updated source baseline: {TODAY}. Official CDSS sources control over internal drafts.

    {md_table(["Source", "URL"], SOURCE_URLS)}

    ## Current CDSS Requirements Confirmed

    - LIC 9141 is the Vendor Application/Renewal form used to request approval to become a vendor.
    - LIC 9140 is the Request for Course Approval used for ACB review and approval of each new or updated course.
    - LIC 9140A is used to add or replace an instructor of an approved course.
    - LIC 9139 is used to renew previously approved, unmodified continuing education courses.
    - LIC 9142A or another format containing at least the same information is used for participant reporting.
    - Vendor Course Schedule Notification may be used for quarterly course schedules; alternate documents may be used if they contain the same required information.
    - A separate LIC 9141 application is required for each proposed vendor type and program type.
    - CDSS recommends vendor applications and/or course requests be submitted at least 60 days before the proposed course offering.
    - If an application or course request is incomplete, ACB notifies the vendor and gives 30 days from receipt of notice to submit missing information.
    - CDSS currently lists RCFE CETP initial/renewal vendor fee as $140.
    - CDSS currently lists CETP course fees as $10 per unit/hour for new or updated CETP courses.
    - Credit card payments are available only for applications submitted online through the Vendor Automation Platform; mailed applications use check or money order with the matching payment notice.
    - Live-stream means a real-time platform where participants can interact with the instructor and other learners; self-paced means individual asynchronous coursework without real-time interaction.
    - Course credits may only be issued in whole-hour increments.
    - Courses may not be offered for administrator ICTP or CETP credit until approved by ACB.
    - Vendor and course numbers are assigned only after approval and must not be invented or used early.

    ## Items Marked {PENDING_CDSS}

    - Final Vendor Automation Platform workflow requires current account access and ACB/assigned analyst confirmation.
    - Any CDSS interpretation not explicit in the cited public sources remains {PENDING_CDSS}.
    """


def make_doc_library(folder: Path):
    rows = []
    for path in sorted(folder.rglob("*")):
        if path.is_file() and path.name != "LINKED_DOCUMENT_LIBRARY.xlsx":
            rows.append([path.parent.relative_to(folder).as_posix() or ".", path.name, str(path), path.suffix.replace(".", "").upper() or "FILE", "Prepared"])
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Linked Document Library"
    ws.append(["Section", "Document Name", "Path", "Type", "Status"])
    for row in rows:
        ws.append(row)
    for col, width in enumerate([34, 56, 110, 14, 20], start=1):
        ws.cell(1, col).font = Font(bold=True, color="FFFFFF")
        ws.cell(1, col).fill = PatternFill("solid", fgColor="8B1515")
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(folder / "LINKED_DOCUMENT_LIBRARY.xlsx")


class SimplePDF:
    def __init__(self, path: Path, title: str, subtitle: str, source_md: str, asset_png: Path | None):
        self.path = path
        self.title = title
        self.subtitle = subtitle
        self.source_md = source_md
        self.asset_png = asset_png
        self.c = canvas.Canvas(str(path), pagesize=letter)
        self.w, self.h = letter
        self.y = self.h - 64
        self.page = 0

    def cover(self):
        maroon = HexColor("#8B1515")
        gold = HexColor("#C9A227")
        self.c.setFillColor(maroon)
        self.c.rect(0, 0, self.w, self.h, fill=1, stroke=0)
        self.c.setFillColor(gold)
        self.c.rect(0, self.h - 10, self.w, 10, fill=1, stroke=0)
        self.c.rect(0, 76, self.w, 3, fill=1, stroke=0)
        if self.asset_png and self.asset_png.exists():
            if self.asset_png.suffix.lower() == ".svg":
                drawing = svg2rlg(str(self.asset_png))
                if drawing:
                    scale = 84 / max(drawing.width, drawing.height)
                    drawing.width *= scale
                    drawing.height *= scale
                    drawing.scale(scale, scale)
                    renderPDF.draw(drawing, self.c, self.w / 2 - 42, self.h - 158)
            else:
                self.c.drawImage(str(self.asset_png), self.w / 2 - 42, self.h - 158, width=84, height=84, preserveAspectRatio=True, mask="auto")
        self.c.setFillColor(colors.white)
        self.c.setFont("Helvetica-Bold", 28)
        self.c.drawCentredString(self.w / 2, self.h - 230, "CI Institute of Nursing")
        self.c.setFont("Helvetica-Bold", 12)
        for i, line in enumerate(self.subtitle.split("\n")):
            self.c.drawCentredString(self.w / 2, self.h - 270 - i * 18, line)
        self.c.setFont("Helvetica", 10)
        footer = [
            "Prepared for Team Execution and Compliance Review",
            f"Prepared by: {INST['prepared_by']}",
            f"Project Oversight: {INST['oversight']}",
            INST["website"],
            INST["physical"],
            f"Prepared: {TODAY}",
        ]
        for i, line in enumerate(footer):
            self.c.drawCentredString(self.w / 2, 205 - i * 17, line)
        self.c.showPage()

    def new_page(self):
        if self.page:
            self.footer()
            self.c.showPage()
        self.page += 1
        self.y = self.h - 58
        self.c.setStrokeColor(HexColor("#8B1515"))
        self.c.line(54, self.h - 36, self.w - 54, self.h - 36)
        self.c.setFont("Helvetica", 7)
        self.c.setFillColor(colors.grey)
        self.c.drawString(54, self.h - 28, f"{INST['name']} | {INST['website']} | {INST['physical']}")

    def footer(self):
        self.c.setStrokeColor(HexColor("#C9A227"))
        self.c.line(54, 34, self.w - 54, 34)
        self.c.setFont("Helvetica", 7)
        self.c.setFillColor(colors.grey)
        self.c.drawString(54, 24, f"{INST['name']} / {INST['prepared_by']}")
        self.c.drawRightString(self.w - 54, 24, f"Page {self.page}")

    def text(self, line: str, font="Helvetica", size=8.6):
        if self.y < 58:
            self.new_page()
        self.c.setFont(font, size)
        self.c.setFillColor(HexColor("#212529"))
        for wrapped in textwrap.wrap(line, width=96) or [""]:
            if self.y < 58:
                self.new_page()
            self.c.drawString(60, self.y, wrapped)
            self.y -= size * 1.35

    def build(self):
        self.cover()
        self.new_page()
        for raw in self.source_md.splitlines():
            line = raw.strip()
            if not line:
                self.y -= 6
            elif line.startswith("# "):
                self.y -= 8
                self.text(line[2:], "Helvetica-Bold", 15)
                self.y -= 4
            elif line.startswith("## "):
                self.text(line[3:], "Helvetica-Bold", 11)
            elif line.startswith("### "):
                self.text(line[4:], "Helvetica-Bold", 9.5)
            elif line.startswith("|"):
                self.text(re.sub(r"\s*\|\s*", " | ", line).strip(" |"), "Helvetica", 6.8)
            else:
                self.text(line)
        self.footer()
        self.c.save()


def convert_logo(svg: Path, png: Path):
    try:
        import cairosvg

        cairosvg.svg2png(url=str(svg), write_to=str(png), output_width=256, output_height=256)
    except Exception:
        # Keep the SVG as the canonical asset even if local conversion support is unavailable.
        pass


def setup_common(folder: Path, package_name: str):
    for sub in ["assets", "appendices/official_forms", "appendices/filled_drafts", "source_research", "working_notes"]:
        (folder / sub).mkdir(parents=True, exist_ok=True)
    svg = folder / "assets" / "ci-ion-logomark-white.svg"
    try:
        urllib.request.urlretrieve("https://ciinstituteofnursing.com/assets/logos/ci-ion-logomark-white.svg", svg)
    except Exception:
        copy_if_exists(ROOT / "CNA-Recert-Course/Branding_Kit/logos/ci-ion-logomark-white.svg", svg)
    convert_logo(svg, folder / "assets" / "ci-ion-logomark-white.png")
    write(folder / "source_research" / "CDSS_SOURCE_RESEARCH_INDEX.md", official_research_index(package_name))
    copy_if_exists(COMBINED / "source_research/Vendor_Automation_Platform_Manual.pdf", folder / "source_research/Vendor_Automation_Platform_Manual.pdf")
    copy_if_exists(Path(__file__), folder / "working_notes" / "split_rcfe_cetp_packages.py")
    write(
        folder / "working_notes" / "GENERATION_SUMMARY.md",
        f"""
        # Generation Summary

        Generated on {TODAY}.

        - Source combined folder preserved at: {COMBINED}
        - Archive copy: {ARCHIVE}
        - Package folder: {folder}
        - No HTML generated.
        - Missing source-backed values remain marked {PENDING}.
        - Unclear CDSS interpretations remain marked {PENDING_CDSS}.
        """,
    )


def split_existing_packet(course_id: str, title: str, packet_path: Path, target: Path):
    packet = packet_path.read_text(encoding="utf-8")
    desc = section_from_packet(packet, "Course Description")
    objectives = section_from_packet(packet, "Measurable Learning Objectives")
    outline = section_from_packet(packet, "Hour-by-Hour Outline")
    methods = section_from_packet(packet, "Teaching Methods")
    eval_method = section_from_packet(packet, "Participant Evaluation Method")
    post_test = section_from_packet(packet, "Post-Test Questions and Answer Key")
    participation = section_from_packet(packet, "Active Participation Verification Method")
    sources = section_from_packet(packet, "Works Cited and Source List")
    qa = section_from_packet(packet, "Course Approval Checklist")
    files = {
        "README.md": f"# {course_id} - {title}\n\nThis folder contains the course-specific LIC 9140 approval packet. One LIC 9140 is required for this course; this folder does not imply one LIC 9140 covers multiple courses.\n",
        f"COURSE_APPROVAL_PACKET_{course_id}.md": packet,
        f"COURSE_OUTLINE_{course_id}.md": f"# Course Outline - {course_id}\n\n{desc}\n\n## Hour-by-Hour Outline\n\n{outline}\n",
        f"CORE_OF_KNOWLEDGE_MAPPING_{course_id}.md": f"# Core of Knowledge Mapping - {course_id}\n\n| Field | Value |\n| --- | --- |\n| Course ID | {course_id} |\n| Course Title | {title} |\n| Core Category | {dict((c[0], c[3]) for c in COURSES)[course_id]} |\n| Mapping Rule | This course is mapped to one RCFE Core of Knowledge category. |\n",
        f"LEARNING_OBJECTIVES_{course_id}.md": f"# Learning Objectives - {course_id}\n\n{objectives}\n",
        f"HOUR_BY_HOUR_OUTLINE_{course_id}.md": f"# Hour-by-Hour Outline - {course_id}\n\n{outline}\n",
        f"TEACHING_METHODS_{course_id}.md": f"# Teaching Methods - {course_id}\n\n{methods}\n",
        f"PARTICIPANT_EVALUATION_METHOD_{course_id}.md": f"# Participant Evaluation Method - {course_id}\n\n{eval_method}\n",
        f"POST_TEST_AND_ANSWER_KEY_{course_id}.md": f"# Post-Test and Answer Key - {course_id}\n\n{post_test}\n",
        f"ACTIVE_PARTICIPATION_VERIFICATION_{course_id}.md": f"# Active Participation Verification - {course_id}\n\n{participation}\n",
        f"INSTRUCTOR_ASSIGNMENT_{course_id}.md": f"# Instructor Assignment - {course_id}\n\nInstructor of record: {PENDING}\n\nSegment instructor assignments: {PENDING}\n\nDo not insert instructor names, licenses, resumes, or disclosures unless source-backed.\n",
        f"RECORDKEEPING_STATEMENT_{course_id}.md": f"# Recordkeeping Statement - {course_id}\n\nRecords maintained: course outline, schedule, roster, completion documentation, evaluations, post-test results, instructor qualification documentation, and reviewer materials.\n\nRecords Retention Address: {INST['records']}\n\nRecordkeeping owner: {PENDING}\n",
        f"REVIEWER_ACCESS_INSTRUCTIONS_{course_id}.md": f"# Reviewer Access Instructions - {course_id}\n\nLMS URL: {PENDING}\n\nReviewer login: {PENDING}\n\nReviewer password handling instruction: Provide credentials through a controlled channel; do not publish passwords in binder files.\n\nLive-stream platform: {PENDING}\n",
        f"CERTIFICATE_COMPLETION_RULE_{course_id}.md": f"# Certificate Completion Rule - {course_id}\n\nNo certificate may be issued until CDSS approval is issued, the participant attends the full approved duration, active participation is verified, the post-test is passed, and issued vendor/course numbers are available.\n",
        f"SOURCES_AND_WORKS_CITED_{course_id}.md": f"# Sources and Works Cited - {course_id}\n\n{sources}\n",
        f"COURSE_APPROVAL_QA_CHECKLIST_{course_id}.md": f"# Course Approval QA Checklist - {course_id}\n\n{qa}\n",
    }
    for name, content in files.items():
        write(target / name, content)


def build_vendor_package():
    setup_common(VENDOR, "LIC 9141 Vendor Application")
    for sub in ["appendices/corporate_documentation", "appendices/signer_authority", "appendices/fee_packet", "appendices/vendor_automation_platform"]:
        (VENDOR / sub).mkdir(parents=True, exist_ok=True)
    old_official = COMBINED / "appendices/official_forms"
    copy_if_exists(old_official / "LIC_9141_6-23_Vendor_Application_Renewal_fillable.pdf", VENDOR / "appendices/official_forms/LIC_9141_6-23_Vendor_Application_Renewal_fillable.pdf")
    copy_if_exists(old_official / "Notice_of_Payment_Information_Vendor_Fees.pdf", VENDOR / "appendices/official_forms/Notice_of_Payment_Information_Vendor_Fees.pdf")
    copy_if_exists(old_official / "LIC_9141_FILLED_DRAFT_CIIN_RCFE_CETP.pdf", VENDOR / "appendices/filled_drafts/LIC_9141_FILLED_DRAFT_CIIN_RCFE_CETP.pdf")
    copy_if_exists(old_official / "Notice_of_Payment_Information_Vendor_Fees_FILLED_DRAFT.pdf", VENDOR / "appendices/filled_drafts/Notice_of_Payment_Information_Vendor_Fees_FILLED_DRAFT.pdf")
    copy_if_exists(old_official / "Credit_Card_Payments_VAP.pdf", VENDOR / "appendices/vendor_automation_platform/Credit_Card_Payments_VAP.pdf")

    field_rows = [
        ("Legal entity name", INST["legal"], "Inherited from existing RCFE packet; confirm against corporate documents before signature."),
        ("Entity type", INST["entity_type"], "Not found in repo source materials."),
        ("Secretary of State entity number", PENDING, "Not found in repo source materials."),
        ("Physical Address", INST["physical"], "Supported by CI Institute of Nursing syllabus."),
        ("Mailing Address", INST["mailing"], "Menlo Park address found only in inherited RCFE draft, not source-backed."),
        ("Records Retention Address", INST["records"], "Supported by CI Institute of Nursing syllabus; recordkeeping owner pending."),
        ("Main phone", INST["phone"], "Supported by CI Institute of Nursing syllabus."),
        ("Main email", INST["email"], "Not found in repo source materials."),
        ("Website", INST["website"], "Supported by CI Institute of Nursing syllabus."),
        ("Authorized representative", PENDING, "Not found in repo source materials."),
        ("Signer authority", PENDING, "Not found in repo source materials."),
        ("Executive signer", PENDING, "Not found in repo source materials."),
        ("Finance/payment contact", PENDING, "Not found in repo source materials."),
    ]
    common_status = md_table(["Field", "Value / Status", "Source / Note"], field_rows)
    steps = "\n".join(f"{i}. {s}" for i, s in enumerate([
        "Confirm legal entity and authority to conduct business.",
        "Confirm authorized representative and signer.",
        "Gather corporate documentation.",
        "Complete LIC 9141.",
        "Complete vendor fee notice or Vendor Automation Platform payment process.",
        "Assemble vendor application appendices.",
        "Run QA check.",
        "Obtain Compliance clearance.",
        "Obtain Dee approval.",
        "Submit through Vendor Automation Platform or mail.",
        "Track CDSS response.",
        "Cure incomplete notice within deadline if needed.",
        "Record vendor number only after issued.",
    ], 1))
    section_texts = {
        "00_VENDOR_APPLICATION_EXECUTIVE_SUMMARY.md": f"""
        # Vendor Application Executive Summary

        CI Institute of Nursing is preparing a separate RCFE Administrator Continuing Education Training Program (CETP) vendor application package for CDSS review.

        ## What Is Being Turned Over

        - LIC 9141 vendor application binder and section files.
        - Official blank LIC 9141 and vendor fee notice.
        - Filled draft LIC 9141 and filled draft vendor fee notice with unknown fields marked {PENDING}.
        - Corporate documentation, signer authority, fee/payment, Vendor Automation Platform, QA, and final gate packets.
        - CDSS source research notes.

        ## Track Status

        {md_table(["Area", "Status"], [
            ("Vendor application track", "Drafted; pending legal/entity/signer/payment fields remain."),
            ("Course approval track", "Separated into the LIC 9140 package; not embedded as primary body content here."),
            ("Fee baseline", "$140 RCFE CETP vendor initial application fee confirmed in current CDSS Vendor Information and Vendor FAQ."),
            ("Final gate", "Required before submission, signature, fee payment, approval claims, or vendor-number use."),
        ])}

        ## Gate Status

        Execution may continue internally. No CDSS submission, signatures, approval claims, payment, or issued-number use until final clearance.
        """,
        "01_VENDOR_APPLICATION_PROJECT_CHARTER.md": f"""
        # Vendor Application Project Charter

        {md_table(["Field", "Value"], [
            ("Institution", INST["name"]),
            ("Primary form", "LIC 9141 Vendor Application/Renewal"),
            ("Purpose", "Become a CDSS-approved RCFE CETP vendor."),
            ("Physical Address", INST["physical"]),
            ("Mailing Address", INST["mailing"]),
            ("Records Retention Address", INST["records"]),
            ("Phone", INST["phone"]),
            ("Website", INST["website"]),
        ])}

        ## Scope

        This binder is limited to the vendor application. It may cross-reference the LIC 9140 course approval package, but it does not duplicate course approval packets as vendor-binder body content.
        """,
        "02_VENDOR_APPLICATION_TEAM_EXECUTION_SUMMARY.md": f"""
        # Vendor Application Team Execution Summary

        {md_table(["Team / Owner", "Action", "Do Not"], [
            ("Compliance", "Validate LIC 9141, source research, pending fields, and claim controls.", "Do not submit while pending fields remain unresolved."),
            ("Corporate / Records", "Collect entity records, authority documentation, and signer authorization.", "Do not invent entity type, SOS number, EIN, DBA, license, or signer authority."),
            ("Finance", "Confirm $140 vendor fee payment path and payment authority.", "Do not pay until Compliance and Dee approval are complete."),
            ("Leadership", "Approve final submission package after QA.", "Do not authorize signatures or approval claims before final clearance."),
        ])}
        """,
        "03_VENDOR_APPLICATION_CRITICAL_PATH.md": f"""
        # Vendor Application Critical Path

        {md_table(["Step", "Owner", "Dependency", "Status"], [
            ("Confirm legal entity and authority", "Corporate / Compliance", "Corporate records", PENDING),
            ("Confirm authorized representative/signature authority", "Leadership / Compliance", "Board or entity authorization", PENDING),
            ("Complete LIC 9141", "Compliance", "Confirmed entity/contact data", "Drafted with pending fields"),
            ("Prepare vendor fee packet", "Finance", "Payment authority", "Drafted; payment details pending"),
            ("Final QA and approval", "Compliance / Dee", "Complete appendices", PENDING),
        ])}
        """,
        "04_LIC_9141_APPLICATION_TASKS.md": f"""
        # LIC 9141 Application Tasks

        ## Step-by-Step Vendor Application Process

        {steps}

        ## LIC 9141 Field Map

        {common_status}
        """,
        "05_CORPORATE_DOCUMENTATION_CHECKLIST.md": f"""
        # Corporate Documentation Checklist

        {md_table(["Document", "Status"], [
            ("Certificate of Status or equivalent authority-to-conduct-business documentation", PENDING),
            ("Articles / entity records", PENDING),
            ("Statement of Information", PENDING),
            ("EIN confirmation / W-9, if appropriate for submission/payment workflow", PENDING),
            ("Local business license / DBA, if applicable", PENDING),
            ("Corporate documentation owner", PENDING),
        ])}
        """,
        "06_AUTHORIZED_REPRESENTATIVE_AND_SIGNER_PACKET.md": f"""
        # Authorized Representative and Signer Packet

        {md_table(["Field", "Status"], [
            ("Authorized representative", PENDING),
            ("Signer authority", PENDING),
            ("Executive signer", PENDING),
            ("Board/entity authorization", PENDING),
            ("Signature date", PENDING),
        ])}

        Do not sign CDSS forms until authority and final clearance are confirmed.
        """,
        "07_VENDOR_FEES_AND_PAYMENT_PACKET.md": f"""
        # Vendor Fees and Payment Packet

        CDSS currently lists the RCFE CETP vendor initial application fee as $140.

        {md_table(["Payment Item", "Status"], [
            ("Vendor fee notice", "Official blank and filled draft copied to appendices."),
            ("Payment amount", "$140"),
            ("Payment method", PENDING),
            ("Check/money order number or VAP payment confirmation", PENDING),
            ("Finance/payment contact", PENDING),
        ])}
        """,
        "08_VENDOR_AUTOMATION_PLATFORM_STEPS.md": f"""
        # Vendor Automation Platform Steps

        CDSS encourages vendors to submit vendor applications, course requests, instructor changes, rosters, and course schedules through the Vendor Automation Platform.

        ## Internal Steps

        1. Register/sign in to the Vendor Automation Platform.
        2. Confirm this is not represented as approval; registration alone does not mean approved vendor status.
        3. Upload LIC 9141 and supporting documents when Compliance clears submission.
        4. Pay by credit card only if submitting online through the platform.
        5. Track notices and cure any incomplete notice within the 30-day window.

        Final account workflow: {PENDING_CDSS}.
        """,
        "09_VENDOR_APPLICATION_QA_CHECKLIST.md": f"""
        # Vendor Application QA Checklist

        {md_table(["QA Item", "Status"], [
            ("Vendor folder exists", "Prepared"),
            ("Vendor Markdown binder exists", "Prepared"),
            ("Vendor PDF binder exists", "Prepared"),
            ("Official LIC 9141 is present", "Prepared"),
            ("Filled LIC 9141 draft is present", "Prepared"),
            ("Vendor fee notice is present", "Prepared"),
            ("Corporate documentation checklist is present", "Prepared"),
            ("Missing fields are marked [PENDING CONFIRMATION]", "Prepared"),
            ("No course approval packets embedded as primary body content", "Prepared"),
            ("No fake form templates exist in the body", "Prepared"),
            ("Logo on cover uses ci-ion-logomark-white.svg or converted PNG", "Prepared"),
            ("No HTML generated", "Prepared"),
        ])}
        """,
        "10_VENDOR_APPLICATION_SUBMISSION_CHECKLIST.md": f"""
        # Vendor Application Submission Checklist

        {md_table(["Required Before Submission", "Status"], [
            ("Legal/entity details verified", PENDING),
            ("Corporate documents attached", PENDING),
            ("Authorized representative confirmed", PENDING),
            ("LIC 9141 checked", "Drafted; final check pending"),
            ("Vendor fee/payment packet checked", PENDING),
            ("Compliance clearance completed", PENDING),
            ("Dee approval completed", PENDING),
        ])}
        """,
        "11_VENDOR_APPLICATION_FINAL_APPROVAL_GATE.md": """
        # Vendor Application Final Approval Gate

        Execution may continue.

        Do NOT proceed with the following until final clearance:

        - Submit LIC 9141 to CDSS.
        - Sign CDSS forms.
        - Pay filing fee.
        - Publish vendor approval claims.
        - Use any CDSS vendor number before issued.
        - Represent CI Institute of Nursing as a CDSS-approved RCFE CETP vendor before approval.

        Required before submission:

        - Legal/entity details verified.
        - Corporate documents attached.
        - Authorized representative confirmed.
        - LIC 9141 checked.
        - Vendor fee/payment packet checked.
        - Compliance clearance completed.
        - Dee approval completed.
        """,
    }
    for name, text in section_texts.items():
        write(VENDOR / name, text)
    write(VENDOR / "appendices/corporate_documentation/CORPORATE_DOCUMENTATION_CHECKLIST.md", section_texts["05_CORPORATE_DOCUMENTATION_CHECKLIST.md"])
    for placeholder in [
        "CERTIFICATE_OF_STATUS_PLACEHOLDER.md",
        "ARTICLES_OR_ENTITY_RECORDS_PLACEHOLDER.md",
        "STATEMENT_OF_INFORMATION_PLACEHOLDER.md",
        "EIN_OR_W9_REFERENCE_PLACEHOLDER.md",
        "LOCAL_BUSINESS_LICENSE_OR_DBA_PLACEHOLDER.md",
    ]:
        write(
            VENDOR / "appendices/corporate_documentation" / placeholder,
            f"""
            # {placeholder.replace('_', ' ').replace('.md', '').title()}

            Status: {PENDING}

            Add the actual source-backed document here only after Compliance confirms it is appropriate for the LIC 9141 vendor application package.
            """,
        )
    write(VENDOR / "appendices/signer_authority/AUTHORIZED_SIGNER_PLACEHOLDER.md", section_texts["06_AUTHORIZED_REPRESENTATIVE_AND_SIGNER_PACKET.md"])
    write(VENDOR / "appendices/fee_packet/VENDOR_FEE_PACKET_README.md", section_texts["07_VENDOR_FEES_AND_PAYMENT_PACKET.md"])
    write(VENDOR / "appendices/vendor_automation_platform/VENDOR_AUTOMATION_PLATFORM_README.md", section_texts["08_VENDOR_AUTOMATION_PLATFORM_STEPS.md"])
    order = list(section_texts)
    cover = f"""# {INST['name']}

    ## RCFE Administrator Continuing Education Training Program (CETP)

    ## LIC 9141 Vendor Application Binder

    Prepared for Team Execution and Compliance Review

    Prepared by: {INST['prepared_by']}

    Project Oversight: {INST['oversight']}

    {INST['website']}

    {INST['physical']}

    Prepared: {TODAY}

    ---
    """
    binder = textwrap.dedent(cover).strip() + "\n\n" + "\n\n---\n\n".join((VENDOR / n).read_text(encoding="utf-8") for n in order)
    write(VENDOR / "RCFE_CETP_LIC9141_Vendor_Application_Binder.md", binder)
    make_doc_library(VENDOR)
    SimplePDF(VENDOR / "RCFE_CETP_LIC9141_Vendor_Application_Binder.pdf", "LIC 9141 Vendor Application Binder", "RCFE Administrator Continuing Education Training Program (CETP)\nLIC 9141 Vendor Application Binder", binder, VENDOR / "assets/ci-ion-logomark-white.svg").build()


def build_course_package():
    setup_common(COURSE, "LIC 9140 Course Approval Applications")
    for sub in ["course_packets", "appendices/course_fee_packet", "appendices/instructor_documentation", "appendices/lms_reviewer_access", "appendices/post_approval_operations"]:
        (COURSE / sub).mkdir(parents=True, exist_ok=True)
    old_official = COMBINED / "appendices/official_forms"
    old_course = COMBINED / "appendices/course_approval_packets"
    old_instr = COMBINED / "appendices/instructor_documentation"
    old_ops = COMBINED / "appendices/operations_forms"
    copy_if_exists(old_official / "LIC_9140_3-24_Request_for_Course_Approval_fillable.pdf", COURSE / "appendices/official_forms/LIC_9140_3-24_Request_for_Course_Approval_fillable.pdf")
    copy_if_exists(old_official / "LIC_9140A_Request_Add_Replace_Instructor_fillable.pdf", COURSE / "appendices/official_forms/LIC_9140A_Request_Add_Replace_Instructor_fillable.pdf")
    copy_if_exists(old_official / "LIC_9139_3-23_Renewal_CE_Course_Approval_fillable.pdf", COURSE / "appendices/official_forms/LIC_9139_3-23_Renewal_CE_Course_Approval_fillable.pdf")
    copy_if_exists(old_official / "Notice_of_Payment_Information_Course_Fees.pdf", COURSE / "appendices/official_forms/Notice_of_Payment_Information_Course_Fees.pdf")
    copy_if_exists(old_official / "RCFE_Core_of_Knowledge_09-19.pdf", COURSE / "appendices/official_forms/RCFE_Core_of_Knowledge_09-19.pdf")
    copy_if_exists(old_official / "ACB_Guide_for_Course_Review.pdf", COURSE / "appendices/official_forms/ACB_Guide_for_Course_Review.pdf")
    copy_if_exists(old_official / "ACS_Sample_Course_Outline_200505.pdf", COURSE / "appendices/official_forms/ACS_Sample_Course_Outline_200505.pdf")
    copy_if_exists(old_course / "Notice_of_Payment_Information_Course_Fees_FILLED_DRAFT.pdf", COURSE / "appendices/filled_drafts/Notice_of_Payment_Information_Course_Fees_FILLED_DRAFT.pdf")
    copy_if_exists(old_course / "Notice_of_Payment_Information_Course_Fees_FILLED_DRAFT.pdf", COURSE / "appendices/course_fee_packet/Notice_of_Payment_Information_Course_Fees_FILLED_DRAFT.pdf")
    copy_if_exists(old_instr / "APPENDIX_C_INSTRUCTOR_DOCUMENTATION_PACKET.md", COURSE / "appendices/instructor_documentation/INSTRUCTOR_DOCUMENTATION_PACKET.md")
    copy_if_exists(old_instr / "POST_APPROVAL_REFERENCE_ONLY_LIC_9140A_Request_Add_Replace_Instructor.pdf", COURSE / "appendices/instructor_documentation/POST_APPROVAL_REFERENCE_ONLY_LIC_9140A_Request_Add_Replace_Instructor.pdf")
    for p in old_ops.glob("*"):
        copy_if_exists(p, COURSE / "appendices/post_approval_operations" / p.name)

    for course_id, title, hours, _cat in COURSES:
        target = COURSE / "course_packets" / course_id
        target.mkdir(parents=True, exist_ok=True)
        packet_candidates = list(old_course.glob(f"{course_id}_*.md"))
        if packet_candidates:
            split_existing_packet(course_id, title, packet_candidates[0], target)
        copy_if_exists(old_course / f"LIC_9140_FILLED_DRAFT_{course_id}.pdf", target / f"LIC_9140_FILLED_DRAFT_{course_id}.pdf")
        copy_if_exists(old_course / f"LIC_9140_FILLED_DRAFT_{course_id}.pdf", COURSE / "appendices/filled_drafts" / f"LIC_9140_FILLED_DRAFT_{course_id}.pdf")

    course_rows = [(cid, title, hours, cat, f"${hours * 10}", "One LIC 9140 required for this course") for cid, title, hours, cat in COURSES]
    steps = "\n".join(f"{i}. {s}" for i, s in enumerate([
        "Confirm first-wave course list.",
        "Confirm each course fits one RCFE Core of Knowledge category.",
        "Complete one LIC 9140 per course.",
        "Attach course description, objectives, outline, teaching methods, sources, evaluation method, participant assessment, active participation method, records statement, and location/reviewer access.",
        "Attach instructor qualifications and disclosures.",
        "Complete course fee notice or Vendor Automation Platform payment process.",
        "Assemble each course packet.",
        "Run course-level QA.",
        "Obtain Compliance clearance.",
        "Obtain Dee approval.",
        "Submit through Vendor Automation Platform or mail.",
        "Track CDSS response.",
        "Cure incomplete notice within deadline if needed.",
        "Record course numbers only after issued.",
        "Activate post-approval roster, certificate, recordkeeping, and schedule controls.",
    ], 1))
    section_texts = {
        "00_COURSE_APPROVAL_EXECUTIVE_SUMMARY.md": f"""
        # Course Approval Executive Summary

        CI Institute of Nursing is preparing a separate LIC 9140 course approval package for the first-wave RCFE CETP courses. This package groups the first-wave courses for project management only. Each proposed course requires its own LIC 9140 and its own course approval packet.

        {md_table(["Area", "Status"], [
            ("First-wave courses", f"{len(COURSES)} proposed courses preserved from existing drafts."),
            ("Total proposed live-stream hours", TOTAL_HOURS),
            ("Course fee planning total", f"${TOTAL_COURSE_FEES} at $10 per unit/hour."),
            ("Vendor dependency", "Course approval depends on CDSS vendor approval; do not imply vendor approval has been issued."),
        ])}
        """,
        "01_COURSE_APPROVAL_PROJECT_CHARTER.md": f"""
        # Course Approval Project Charter

        {md_table(["Field", "Value"], [
            ("Institution", INST["name"]),
            ("Primary form", "LIC 9140 Request for Course Approval"),
            ("Purpose", "Obtain CDSS approval for each first-wave RCFE CETP course."),
            ("Vendor approval status", "[PENDING CDSS ISSUANCE]"),
            ("Physical Address", INST["physical"]),
            ("Records Retention Address", INST["records"]),
        ])}
        """,
        "02_COURSE_APPROVAL_TEAM_EXECUTION_SUMMARY.md": f"""
        # Course Approval Team Execution Summary

        {md_table(["Team / Owner", "Action", "Do Not"], [
            ("Course Development", "Review and complete each course packet.", "Do not combine courses under one LIC 9140."),
            ("Compliance", "Check LIC 9140 attachments, instructor docs, reviewer access, and claims.", "Do not approve submission with missing required attachments."),
            ("LMS / Reviewer Access", "Validate LMS/live-stream access without exposing passwords.", "Do not publish reviewer passwords in binder files."),
            ("Finance", f"Confirm ${TOTAL_COURSE_FEES} course fee planning total.", "Do not pay until final clearance."),
        ])}
        """,
        "03_COURSE_APPROVAL_CRITICAL_PATH.md": f"""
        # Course Approval Critical Path

        {md_table(["Task", "Dependency", "Status"], [
            ("Confirm course list and one-category mapping", "Course catalog", "Drafted"),
            ("Complete one LIC 9140 per course", "Official LIC 9140", "Drafts copied"),
            ("Attach complete course packet per course", "Course materials", "Drafted"),
            ("Attach instructor qualifications/disclosures", "Instructor slate", PENDING),
            ("Validate LMS/live-stream reviewer access", "Platform account", PENDING),
            ("Approve course fee packet", "Finance approval", PENDING),
            ("Compliance and Dee final approval", "All prior gates", PENDING),
        ])}
        """,
        "04_FIRST_WAVE_COURSE_CATALOG.md": f"# First-Wave Course Catalog\n\n{md_table(['Course ID', 'Title', 'Hours', 'Core Category', 'CDSS Fee', 'LIC 9140 Rule'], course_rows)}\n",
        "05_CORE_OF_KNOWLEDGE_CROSSWALK.md": f"# Core of Knowledge Crosswalk\n\n{md_table(['Course ID', 'Title', 'Mapped RCFE Core of Knowledge Category'], [(cid, title, cat) for cid, title, _hours, cat in COURSES])}\n",
        "06_LIC_9140_APPLICATION_TASKS.md": f"# LIC 9140 Application Tasks\n\n## Step-by-Step Course Approval Process\n\n{steps}\n\n## LIC 9140 Field Map\n\n{md_table(['Field', 'Status'], [('Vendor number', '[PENDING CDSS ISSUANCE]'), ('Course title', 'Course-specific values in each course packet'), ('Course hours', 'Course-specific whole-hour values'), ('Instructor', PENDING), ('Course dates', PENDING), ('Course pricing', PENDING), ('Reviewer access', PENDING)])}\n",
        "07_INSTRUCTOR_QUALIFICATION_PACKET.md": f"# Instructor Qualification Packet\n\n{md_table(['Item', 'Status'], [('Instructor names', PENDING), ('Instructor resumes', PENDING), ('Instructor credentials/licenses', PENDING), ('Instructor disclosure answers', PENDING), ('Instructor assignment by course segment', PENDING)])}\n",
        "08_LMS_AND_REVIEWER_ACCESS_PACKET.md": f"# LMS and Reviewer Access Packet\n\n{md_table(['Field', 'Status'], [('LMS URL', PENDING), ('Reviewer login', PENDING), ('Reviewer password handling instruction', 'Provide through controlled channel only; do not publish passwords.'), ('Live-stream platform', PENDING), ('Active participation verification method', 'Drafted in each course packet'), ('Self-paced requirements', 'Not used for current first-wave live-stream drafts; mark any self-paced conversion separately for CDSS review.')])}\n",
        "09_COURSE_FEES_AND_PAYMENT_PACKET.md": f"# Course Fees and Payment Packet\n\nCDSS currently lists CETP course fees as $10 per unit/hour for new or updated CETP courses.\n\n{md_table(['Item', 'Amount / Status'], [('Total first-wave hours', TOTAL_HOURS), ('Total CDSS course fee planning amount', f'${TOTAL_COURSE_FEES}'), ('Course fee notice', 'Official blank and filled draft copied'), ('Payment method', PENDING), ('Finance/payment contact', PENDING)])}\n",
        "10_RECORDKEEPING_AND_CERTIFICATE_CONTROLS.md": f"# Recordkeeping and Certificate Controls\n\n{md_table(['Control', 'Status'], [('LIC 9142A or equivalent roster', 'Copied to appendices/post_approval_operations'), ('Course Schedule Notification', 'Copied to appendices/post_approval_operations'), ('Certificate of Completion sample', 'Copied/prepared in appendices/post_approval_operations'), ('Recordkeeping templates', 'Copied to appendices/post_approval_operations'), ('Recordkeeping owner', PENDING), ('Records Retention Address', INST['records']), ('Vendor/course numbers on certificates', '[PENDING CDSS ISSUANCE]')])}\n",
        "11_COURSE_APPROVAL_QA_CHECKLIST.md": f"""
        # Course Approval QA Checklist

        {md_table(["QA Item", "Status"], [
            ("Course approval folder exists", "Prepared"),
            ("Course approval Markdown binder exists", "Prepared"),
            ("Course approval PDF binder exists", "Prepared"),
            ("Official LIC 9140 is present", "Prepared"),
            ("One filled LIC 9140 draft exists per proposed course", "Prepared"),
            ("Each course has a full course packet folder", "Prepared"),
            ("Instructor documentation packet exists", "Prepared"),
            ("Course fee packet exists", "Prepared"),
            ("Roster/schedule/certificate/recordkeeping appendices exist", "Prepared"),
            ("Missing fields are marked [PENDING CONFIRMATION]", "Prepared"),
            ("No fake form templates exist in the body", "Prepared"),
            ("Logo on cover uses ci-ion-logomark-white.svg or converted PNG", "Prepared"),
            ("No HTML generated", "Prepared"),
        ])}
        """,
        "12_COURSE_APPROVAL_SUBMISSION_CHECKLIST.md": f"# Course Approval Submission Checklist\n\n{md_table(['Required Before Submission', 'Status'], [('Course approval packets complete', 'Drafted; final review pending'), ('One LIC 9140 prepared per course', 'Prepared'), ('Instructor documentation complete', PENDING), ('LMS/reviewer access validated', PENDING), ('Participant evaluation and assessment method complete', 'Drafted'), ('Active participation method complete', 'Drafted'), ('Course fees checked', PENDING), ('Compliance clearance completed', PENDING), ('Dee approval completed', PENDING)])}\n",
        "13_COURSE_APPROVAL_FINAL_APPROVAL_GATE.md": """
        # Course Approval Final Approval Gate

        Execution may continue.

        Do NOT proceed with the following until final clearance:

        - Submit LIC 9140 applications to CDSS.
        - Sign CDSS course forms.
        - Pay course fees.
        - Publish course approval claims.
        - Issue administrator CE certificates.
        - Advertise courses for RCFE administrator renewal credit.
        - Use any CDSS course number before issued.

        Required before submission:

        - Course approval packets complete.
        - One LIC 9140 prepared per course.
        - Instructor documentation complete.
        - LMS/reviewer access validated.
        - Participant evaluation and assessment method complete.
        - Active participation method complete.
        - Course fees checked.
        - Compliance clearance completed.
        - Dee approval completed.
        """,
    }
    for name, text in section_texts.items():
        write(COURSE / name, text)
    write(COURSE / "appendices/instructor_documentation/INSTRUCTOR_DOCUMENTATION_CHECKLIST.md", section_texts["07_INSTRUCTOR_QUALIFICATION_PACKET.md"])
    write(COURSE / "appendices/lms_reviewer_access/LMS_REVIEWER_ACCESS_INSTRUCTIONS.md", section_texts["08_LMS_AND_REVIEWER_ACCESS_PACKET.md"])
    write(COURSE / "appendices/course_fee_packet/COURSE_FEE_PACKET_README.md", section_texts["09_COURSE_FEES_AND_PAYMENT_PACKET.md"])
    cover = f"""# {INST['name']}

    ## RCFE Administrator Continuing Education Training Program (CETP)

    ## LIC 9140 Course Approval Applications Binder

    Prepared for Team Execution and Compliance Review

    Prepared by: {INST['prepared_by']}

    Project Oversight: {INST['oversight']}

    {INST['website']}

    {INST['physical']}

    Prepared: {TODAY}

    ---
    """
    binder = textwrap.dedent(cover).strip() + "\n\n" + "\n\n---\n\n".join((COURSE / n).read_text(encoding="utf-8") for n in section_texts)
    binder += "\n\n---\n\n# Course Packet Index\n\n" + md_table(["Course ID", "Folder"], [(cid, f"course_packets/{cid}") for cid, *_ in COURSES])
    write(COURSE / "RCFE_CETP_LIC9140_Course_Approval_Binder.md", binder)
    make_doc_library(COURSE)
    SimplePDF(COURSE / "RCFE_CETP_LIC9140_Course_Approval_Binder.pdf", "LIC 9140 Course Approval Applications Binder", "RCFE Administrator Continuing Education Training Program (CETP)\nLIC 9140 Course Approval Applications Binder", binder, COURSE / "assets/ci-ion-logomark-white.svg").build()


def main():
    if not COMBINED.exists():
        raise SystemExit(f"Missing combined folder: {COMBINED}")
    if not ARCHIVE.exists():
        shutil.copytree(COMBINED, ARCHIVE)
    VENDOR.mkdir(parents=True, exist_ok=True)
    COURSE.mkdir(parents=True, exist_ok=True)
    build_vendor_package()
    build_course_package()
    print(f"Archive: {ARCHIVE}")
    print(f"Vendor package: {VENDOR}")
    print(f"Course package: {COURSE}")
    print(f"Course packets: {len(COURSES)}")
    print("No HTML generated")


if __name__ == "__main__":
    main()
