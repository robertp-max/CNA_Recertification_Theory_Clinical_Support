from pathlib import Path
import re
import shutil

from pypdf import PdfReader, PdfWriter
from pypdf.generic import BooleanObject, NameObject
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.colors import HexColor, white
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


ROOT = Path(r"C:\AI\Git\CNA_Recertification_Theory_Clinical_Support\RCFE CETP Vendor Packet for CDSS")
APP = ROOT / "appendices"
OFFICIAL = APP / "official_forms"
COURSE_DIR = APP / "course_approval_packets"
INSTR_DIR = APP / "instructor_documentation"
OPS_DIR = APP / "operations_forms"
RESEARCH_DIR = ROOT / "source_research"
NOTES_DIR = ROOT / "working_notes"
LOGO = Path(r"C:\AI\CIION\Builder\ci_ion_logo_cropped.png")
if not LOGO.exists():
    LOGO = Path(r"C:\AI\CIION\ci_ion_logo_cropped.png")

TODAY = "June 1, 2026"
PENDING = "[PENDING CONFIRMATION]"
PENDING_CDSS = "[PENDING CDSS CONFIRMATION]"

for folder in [ROOT, APP, OFFICIAL, COURSE_DIR, INSTR_DIR, OPS_DIR, RESEARCH_DIR, NOTES_DIR]:
    folder.mkdir(parents=True, exist_ok=True)

INST = {
    "name": "CI Institute of Nursing",
    "legal": "CI INSTITUTE OF NURSING",
    "program": "RCFE Administrator Continuing Education Training Program (CETP)",
    "binder": "CDSS Vendor Approval Project Binder",
    "purpose": "Prepared for Team Execution and Compliance Review",
    "prepared_by": "TJ Padilla",
    "oversight": "Dee Bustos",
    "website": "ciinstituteofnursing.com",
    "address": "419 E Hamilton Ave, Campbell, CA 95008",
    "mailing": "890 Santa Cruz Ave, Menlo Park, CA 94025-4641",
    "phone": "(650) 799-5744",
    "email": PENDING,
    "authorized_rep": PENDING,
    "vendor_number": PENDING,
}

SOURCE_URLS = [
    ("CDSS Administrator Certification Bureau Forms", "https://cdss.ca.gov/inforesources/community-care/administrator-certification/certification-forms", "Official source for vendor forms and payment notices."),
    ("CDSS FAQ - Vendors", "https://cdss.ca.gov/inforesources/community-care/administrator-certification/administrator-information/faq-vendors", "Fees, course format, outline requirements, certificates, records, cure process, and renewals."),
    ("CDSS Administrator Information", "https://www.cdss.ca.gov/inforesources/community-care/administrator-certification/administrator-information", "Renewal hours, live/self-paced limits, CEU rule, and contacts."),
    ("LIC 9141 Vendor Application/Renewal (6/23)", "https://www.cdss.ca.gov/Portals/9/CCL/ACP/2021/LIC9141%28fillable%29.PDF", "Vendor application/renewal form."),
    ("LIC 9140 Request for Course Approval (3/24)", "https://cdss.ca.gov/Portals/9/CCL/ACP/2023/LIC9140%28fillable%29.pdf?ver=2024-04-16-083715-430", "Course approval form."),
    ("LIC 9140A Request to Add or Replace Instructor (3/23)", "https://www.cdss.ca.gov/Portals/9/180427%20LIC%209140A%20%28Fillable%29.pdf?ver=2018-05-01-102125-973", "Post-approval instructor add/replace form."),
    ("LIC 9139 Renewal of Continuing Education Course Approval (3/23)", "https://www.cdss.ca.gov/Portals/9/CCL/ACP/2021/LIC9139%28fillable%29.PDF", "Unchanged CETP course renewal form."),
    ("LIC 9142A Roster of Participants (3/23)", "https://www.cdss.ca.gov/Portals/9/180427%20LIC%209142A%20%28Fillable%29.pdf?ver=2018-05-01-102950-963", "Roster or equivalent-information baseline."),
    ("Vendor Course Schedule Notification (2/2026)", "https://www.cdss.ca.gov/Portals/9/CCL/ACP/2023/CourseScheduleNotification.pdf?ver=2024-07-09-130539-477", "Quarterly course schedule notification."),
    ("Notice of Payment Information: Vendor Fees", "https://www.cdss.ca.gov/Portals/9/CCL/ACP/2023/NoticeofPaymentInformationVendorFees.pdf?ver=2024-07-09-130540-180", "Vendor fee payment notice."),
    ("Notice of Payment Information: Course Fees", "https://www.cdss.ca.gov/Portals/9/CCL/ACP/2023/NoticeofPaymentInformationCourseFees.pdf?ver=2024-07-09-130539-853", "CETP course fee payment notice."),
    ("RCFE Core of Knowledge (09/19)", "https://www.cdss.ca.gov/PORTALS/9/CCL/ACP/2019/CoreofKnowledgeRCFE.pdf", "Official RCFE Core of Knowledge categories."),
    ("PIN 23-14-CCLD", "https://cdss.ca.gov/Portals/9/CCLD/PINs/2023/CCLD/PIN-23-14-CCLD.pdf", "Course format terminology and live-stream authority."),
    ("ACB Guide for Course Review", "https://www.cdss.ca.gov/Portals/9/CCL/ACP/2023/ACBGuideforCourseReview.pdf?ver=2024-07-07-120508-043", "Course review attachment expectations."),
    ("ACS Sample Course Outline", "https://www.cdss.ca.gov/Portals/9/CCL/ACP/2019/200505%20Sample%20Course%20Outline_PDF.pdf", "Course outline model."),
    ("ACS Sample Certificate of Completion", "https://www.cdss.ca.gov/Portals/9/ACS%20Sample%20Certificate%20of%20Completion.pdf?ver=2018-10-05-144756-173", "Certificate content model."),
    ("Vendor Automation Platform", "https://ca-dss.prod.simpligov.com", "Online filing and payment platform."),
    ("Vendor Automation Platform Manual", "https://www.cdss.ca.gov/Portals/9/VendorAutomationManual.pdf", "Platform workflow reference; dated June 2022."),
]

COURSES = [
    {
        "id": "RCFE-CETP-001",
        "title": "RCFE Laws, Regulations, Policies, and Procedural Standards",
        "hours": 4,
        "cat": "Laws, Regulations, Policies, and Procedural Standards Impacting RCFE",
        "desc": "This course prepares RCFE administrators to identify and apply current California RCFE statutory, regulatory, policy, and procedural standards in daily operations, incident response, documentation, and compliance decision-making.",
        "obj": [
            "Identify the role of Health and Safety Code section 1569 and Title 22, Division 6, Chapter 8 in RCFE operations.",
            "Differentiate licensing, certification, mandated reporting, complaint, personal rights, and recordkeeping obligations.",
            "Apply a regulatory cross-check process before implementing a policy or operational change.",
            "Recognize situations requiring escalation before public claims, certificates, or CDSS submissions are made.",
        ],
        "outline": [
            ("Hour 1", "RCFE legal framework, CDSS/CCLD structure, certification context, and official CDSS resource navigation."),
            ("Hour 2", "Title 22 operating requirements, mandated reporting, confidentiality, complaint procedures, and personal rights."),
            ("Hour 3", "Policies, procedural controls, documentation standards, staff communication, and escalation triggers."),
            ("Hour 4", "Scenarios applying laws and procedural standards to admissions, incidents, records, and public-facing language."),
        ],
        "qa": [
            ("Which source should be checked first when confirming RCFE operating requirements?", "California Code of Regulations, Title 22, Division 6, Chapter 8."),
            ("What must occur before CDSS-approved RCFE CE status is advertised?", "CDSS approval must be issued and approval claims must be cleared."),
            ("Why avoid approval claims before CDSS action?", "Premature claims can mislead administrators and create compliance risk."),
            ("What is the control when rule interpretation is uncertain?", "Escalate to Compliance and mark pending CDSS confirmation."),
            ("Which Core category applies?", "Laws, Regulations, Policies, and Procedural Standards Impacting RCFE."),
        ],
    },
    {
        "id": "RCFE-CETP-002",
        "title": "Alzheimer's Disease and Related Dementias: Person-Centered Care Foundations",
        "hours": 4,
        "cat": "Managing Alzheimer's Disease and Related Dementias",
        "desc": "This course focuses on dementia-care foundations for RCFE administrators, including person-centered approaches, communication, behavior interpretation, individualized service planning, staff direction, and documentation expectations.",
        "obj": [
            "Describe common dementia-related changes that affect communication, behavior, safety, and care planning.",
            "Identify non-pharmacologic, person-centered strategies for responding to distress and unmet needs.",
            "Connect dementia care practices to RCFE assessment, service planning, supervision, and documentation responsibilities.",
            "Evaluate whether staff instructions align with resident dignity, safety, and regulatory expectations.",
        ],
        "outline": [
            ("Hour 1", "Dementia overview, common symptoms, progression, and administrator-level implications."),
            ("Hour 2", "Communication methods, behavior interpretation, unmet-needs framework, and de-escalation principles."),
            ("Hour 3", "Individualized service planning, family communication, activities, staff coaching, and documentation."),
            ("Hour 4", "Case scenarios: person-centered interventions, risk documentation, and escalation decisions."),
        ],
        "qa": [
            ("What is the first operational question when a resident with dementia becomes distressed?", "What unmet need, trigger, pain, fear, or environmental factor may be contributing?"),
            ("Why should dementia interventions be documented?", "Documentation supports continuity, supervision, accountability, and regulatory review."),
            ("Which approach best aligns with person-centered care?", "Adapting communication and environment to the resident's history, abilities, and needs."),
            ("What should staff do when dementia-related behavior creates immediate safety risk?", "Follow the approved safety protocol and escalate according to facility policy."),
            ("Which Core category applies?", "Managing Alzheimer's Disease and Related Dementias."),
        ],
    },
    {
        "id": "RCFE-CETP-003",
        "title": "Alzheimer's Disease and Related Dementias: Safety, Environment, and Documentation",
        "hours": 4,
        "cat": "Managing Alzheimer's Disease and Related Dementias",
        "desc": "This course expands dementia CE coverage by focusing on environmental safety, restricted access considerations, dangerous-item controls, staff monitoring, activity adaptation, care-plan updates, and incident documentation.",
        "obj": [
            "Identify environmental and supervision controls relevant to residents living with dementia.",
            "Evaluate dementia-care safety risks involving egress, dangerous items, nutrition, hydration, and activities.",
            "Document safety observations, interventions, family communication, and service-plan updates accurately.",
            "Recognize when a change in condition requires reassessment, care coordination, or leadership escalation.",
        ],
        "outline": [
            ("Hour 1", "Dementia-related safety risks, environmental cues, egress awareness, and dangerous-item controls."),
            ("Hour 2", "Activities, nutrition, hydration, sleep, wandering risk, and staffing communication."),
            ("Hour 3", "Observation, charting, care-plan updates, family communication, and reportable-event awareness."),
            ("Hour 4", "Scenario review: environmental controls, staff direction, documentation, and resident dignity."),
        ],
        "qa": [
            ("What type of control reduces dementia-related safety risk without undermining dignity?", "A person-centered environmental or supervision control tailored to the resident's assessed need."),
            ("When should a dementia care plan be updated?", "When observation, condition change, incident, or service need indicates the current plan is no longer sufficient."),
            ("Why are hydration and nutrition part of dementia safety planning?", "Changes in cognition can affect intake, cueing needs, and risk of decline."),
            ("What must staff avoid when describing dementia-related behavior?", "Judgmental labels that do not document objective observations and interventions."),
            ("How many dementia hours do RCFE administrators generally need within a 40-hour renewal cycle?", "At least 8 hours."),
        ],
    },
    {
        "id": "RCFE-CETP-004",
        "title": "Resident Rights, Dignity, Councils, and Abuse Prevention",
        "hours": 2,
        "cat": "Residents' Rights",
        "desc": "This course prepares administrators to operationalize RCFE resident rights through staff training, care planning, family communication, resident and family councils, abuse-prevention controls, and complaint response.",
        "obj": [
            "Identify key RCFE resident rights and operational practices that protect dignity and choice.",
            "Recognize staff behaviors and facility practices that may create resident-rights risk.",
            "Apply a complaint and abuse-prevention escalation process consistent with claim-safe training standards.",
        ],
        "outline": [
            ("Hour 1", "Resident rights, dignity, privacy, choice, councils, family engagement, and complaint pathways."),
            ("Hour 2", "Abuse-prevention controls, staff training scenarios, documentation, escalation, and corrective action."),
        ],
        "qa": [
            ("What is a core operational purpose of resident rights training?", "To ensure staff decisions preserve dignity, choice, privacy, and safety."),
            ("What should occur when a resident-rights complaint is received?", "Document, escalate, investigate, and respond according to approved policy."),
            ("Why are resident and family councils relevant?", "They support resident voice, issue identification, communication, and quality improvement."),
            ("What documentation style is most appropriate for rights concerns?", "Objective facts, actions taken, notifications, and follow-up."),
            ("Which Core category applies?", "Residents' Rights."),
        ],
    },
    {
        "id": "RCFE-CETP-005",
        "title": "Medication Management in RCFE Operations",
        "hours": 4,
        "cat": "Medication Management",
        "desc": "This course addresses administrator-level medication management responsibilities, including assistance with self-administered medications, storage, documentation, disposal, communication, error prevention, and escalation.",
        "obj": [
            "Differentiate administrator oversight duties from clinical medication decision-making.",
            "Identify common medication documentation, storage, disposal, and communication controls in RCFE operations.",
            "Recognize medication-error risk patterns and escalation triggers.",
            "Apply a medication-management audit checklist to sample scenarios.",
        ],
        "outline": [
            ("Hour 1", "RCFE medication management framework, resident profiles, and physician/pharmacy communication."),
            ("Hour 2", "Storage, labeling, documentation, disposal, PRN awareness, and controlled-access controls."),
            ("Hour 3", "Common errors, staff training, observation, escalation, and incident review."),
            ("Hour 4", "Scenario audit: records, missed doses, changed orders, refusals, and communication chain."),
        ],
        "qa": [
            ("What is an administrator-level medication control?", "Ensuring staff follow approved documentation, storage, communication, and escalation procedures."),
            ("What should occur when medication instructions are unclear?", "Escalate for clarification through the approved physician/pharmacy communication process."),
            ("Why are medication refusals documented?", "They affect resident safety, communication, follow-up, and regulatory review."),
            ("What should a medication audit check include?", "Orders, labels, logs, storage, disposal, staff training, and error follow-up."),
            ("Which Core category applies?", "Medication Management."),
        ],
    },
    {
        "id": "RCFE-CETP-006",
        "title": "Admission, Retention, Reappraisal, and Needs and Services Plans",
        "hours": 3,
        "cat": "Resident Admission, Retention, and Assessment Procedures",
        "desc": "This course prepares administrators to manage admission and retention decisions, pre-admission appraisal, physician documentation, reappraisal, needs and services plans, and escalation when resident needs change.",
        "obj": [
            "Identify records and decision points used in RCFE admission and retention review.",
            "Explain how appraisal, physician reports, and service planning guide resident care decisions.",
            "Recognize when a change in condition requires reappraisal, external coordination, or leadership review.",
            "Apply documentation controls to sample admission and retention scenarios.",
        ],
        "outline": [
            ("Hour 1", "Admission screening, pre-admission appraisal, physician report, resident characteristics, and fit review."),
            ("Hour 2", "Needs and services plans, reassessment triggers, documentation standards, and family/provider communication."),
            ("Hour 3", "Retention scenarios, relocation/eviction awareness, exception/escalation controls, and record audit."),
        ],
        "qa": [
            ("What is the function of a needs and services plan?", "It translates assessed resident needs into services, monitoring, documentation, and staff direction."),
            ("When should reappraisal be considered?", "When resident condition, behavior, care needs, or safety risks materially change."),
            ("Why should admission decisions be documented carefully?", "They support resident safety, regulatory compliance, and defensible operations."),
            ("What should happen when retention is uncertain?", "Escalate to leadership and Compliance before a final decision."),
            ("Which Core category applies?", "Resident Admission, Retention, and Assessment Procedures."),
        ],
    },
    {
        "id": "RCFE-CETP-007",
        "title": "Emergency Procedures and Physical Environment Controls",
        "hours": 2,
        "cat": "Managing the Physical Environment",
        "desc": "This course addresses administrator controls for physical environment safety, emergency procedures, maintenance, housekeeping, passageways, hazardous materials, water and temperature considerations, and disaster-plan readiness.",
        "obj": [
            "Identify physical-environment risks that affect resident safety in RCFE settings.",
            "Apply an emergency and environmental readiness checklist to sample facility scenarios.",
            "Recognize documentation and escalation controls for environmental hazards and disaster readiness.",
        ],
        "outline": [
            ("Hour 1", "Physical environment requirements, maintenance, housekeeping, passageways, hazardous storage, and safety checks."),
            ("Hour 2", "Emergency procedures, disaster-plan readiness, communication, drills, and environmental risk documentation."),
        ],
        "qa": [
            ("Why are clear passageways an administrator concern?", "They affect resident safety, emergency movement, and regulatory compliance."),
            ("What should happen when an environmental hazard is identified?", "Document, correct or mitigate, communicate, and escalate when risk remains."),
            ("What is a disaster-plan readiness control?", "Maintaining current procedures, staff awareness, drills, supplies, and communication plans."),
            ("Why are hazardous materials storage controls important?", "They reduce resident access risk and support safe operations."),
            ("Which Core category applies?", "Managing the Physical Environment."),
        ],
    },
    {
        "id": "RCFE-CETP-008",
        "title": "Staff Supervision, Training Records, and Administrator Accountability",
        "hours": 2,
        "cat": "Management/Supervision of Staff",
        "desc": "This course focuses on administrator responsibilities for staff supervision, role clarity, onboarding, training records, day/night coverage communication, performance follow-up, and documentation of corrective coaching.",
        "obj": [
            "Identify staff supervision controls that support resident care and regulatory readiness.",
            "Explain how training records, duty assignments, and communication tools support accountability.",
            "Apply a coaching and documentation approach to staff performance scenarios.",
        ],
        "outline": [
            ("Hour 1", "Staff supervision framework, role clarity, staffing patterns, day/night communication, and training requirements."),
            ("Hour 2", "Performance coaching, corrective documentation, personnel file controls, and supervision scenarios."),
        ],
        "qa": [
            ("What should staff supervision documentation show?", "Expectations, observations, coaching, follow-up, and escalation when needed."),
            ("Why are training records part of supervision?", "They show staff were prepared for assigned duties and support compliance review."),
            ("What is a risk of unclear role assignments?", "Gaps in care, inconsistent documentation, and reduced accountability."),
            ("When should a staffing issue be escalated?", "When resident safety, regulatory compliance, or coverage expectations may be affected."),
            ("Which Core category applies?", "Management/Supervision of Staff."),
        ],
    },
    {
        "id": "RCFE-CETP-009",
        "title": "RCFE Business Operations, Records, and Claim-Safe Communications",
        "hours": 2,
        "cat": "Business Operations",
        "desc": "This course addresses business-operation controls relevant to RCFE administrators, including records, fiscal and operational documentation, audit readiness, public-facing communication, and claim-safe management practices.",
        "obj": [
            "Identify RCFE business-operation records and controls that support compliant management.",
            "Distinguish operational communication from public claims that require approval confirmation.",
            "Apply an audit-readiness checklist to business records, training records, and public-facing statements.",
        ],
        "outline": [
            ("Hour 1", "Business-operation controls, records, fiscal documentation, plan-of-operation awareness, and audit readiness."),
            ("Hour 2", "Claim-safe communications, catalog and website language, approval-number controls, and scenario review."),
        ],
        "qa": [
            ("What makes a public RCFE CE claim unsafe before CDSS approval?", "It may imply approval, course credit, vendor number, or course number that has not been issued."),
            ("Why should records be organized before CDSS submission?", "Organized records reduce incomplete-response risk and support monitoring readiness."),
            ("What business-operation control supports audit readiness?", "A document library with owner, status, source, and retention rule."),
            ("What should happen before using a CDSS vendor number publicly?", "CDSS must issue the number and Compliance must clear the language."),
            ("Which Core category applies?", "Business Operations."),
        ],
    },
]

for course in COURSES:
    course["delivery"] = "Live-Stream initial delivery with LMS-supported materials and post-test"
    course["instructor"] = PENDING
    course["reviewer_access"] = PENDING
    course["participant_fee"] = PENDING
    course["records_address"] = INST["address"]
    course["cdss_fee"] = course["hours"] * 10

TOTAL_HOURS = sum(c["hours"] for c in COURSES)
TOTAL_COURSE_FEES = sum(c["cdss_fee"] for c in COURSES)


def md_table(headers, rows):
    out = ["| " + " | ".join(headers) + " |", "| " + " | ".join(["---"] * len(headers)) + " |"]
    for row in rows:
        out.append("| " + " | ".join(str(x).replace("\n", " ").replace("|", "/") for x in row) + " |")
    return "\n".join(out)


def team_block(name, provided, actions, output, dont):
    return (
        f"### {name}\n\nPROVIDED:\n\n"
        + "\n".join(f"- {x}" for x in provided)
        + "\n\nACTION STEPS:\n\n"
        + "\n".join(f"- {x}" for x in actions)
        + "\n\nOUTPUT:\n\n"
        + "\n".join(f"- {x}" for x in output)
        + "\n\nDO NOT:\n\n"
        + "\n".join(f"- {x}" for x in dont)
        + "\n"
    )


def course_packet(c):
    meta = md_table(
        ["Field", "Draft content"],
        [
            ("Course title", c["title"]),
            ("Requested hours", c["hours"]),
            ("Delivery method", c["delivery"]),
            ("RCFE Core of Knowledge category", c["cat"]),
            ("Instructor assignment by segment", c["instructor"]),
            ("Records maintained and address", f"Course outline, schedule, roster, completion documentation, evaluations, post-test results, instructor qualification documentation, and reviewer materials at {c['records_address']}"),
            ("LMS reviewer access instructions", c["reviewer_access"]),
            ("Certificate completion rule", "No certificate until CDSS approval, full attendance, participation verification, passing post-test, and issued vendor/course numbers."),
        ],
    )
    outline = md_table(["Segment", "Content", "Instructor assignment"], [(a, b, c["instructor"]) for a, b in c["outline"]])
    questions = md_table(["No.", "Question", "Answer key"], [(i + 1, q, a) for i, (q, a) in enumerate(c["qa"])])
    sources = [
        "CDSS RCFE Core of Knowledge Training Standard, 09/19.",
        "California Code of Regulations, Title 22, Division 6, Chapter 8, Residential Care Facilities for the Elderly.",
        "CDSS FAQ - Vendors, updated 7/29/2025.",
        "LIC 9140 Request for Course Approval, 3/24.",
        "PIN 23-14-CCLD, course format terminology and live-stream authority.",
        "ACB Guide for Course Review and ACS Sample Course Outline.",
    ]
    checklist = [
        "LIC 9140 filled draft prepared for RCFE CETP.",
        "Course description included.",
        "Measurable objectives included.",
        "Hour-by-hour course outline included.",
        "Teaching methods included.",
        "Instructor assignment by segment marked pending until instructor slate is confirmed.",
        "Participant evaluation method included.",
        "Participant assessment/post-test included.",
        "Works cited/source list included.",
        "Live-stream active participation verification method included.",
        "Records maintained and records address included.",
        "Reviewer access marked pending until LMS/live-stream account is validated.",
        "Certificate rule included.",
    ]
    return f"""## {c['id']} - {c['title']}

{meta}

### Course Description

{c['desc']}

### Measurable Learning Objectives

{chr(10).join(f"- {x}" for x in c["obj"])}

### Hour-by-Hour Outline

{outline}

### Teaching Methods

- Live lecture with visual slides and participant-facing handouts.
- Case scenarios and guided discussion using RCFE operations examples.
- Polls, chat prompts, verbal check-ins, and knowledge checks during the live-stream.
- Post-test administered through the LMS or controlled form after attendance is verified.

### Participant Evaluation Method

Participants complete a course and instructor evaluation covering objective clarity, relevance to RCFE administrator duties, instructor effectiveness, materials, technology, timing, and suggested improvements. Evaluation records are retained with the course file.

### Participant Assessment and Post-Test Method

Participants must attend the full approved duration, complete the active participation checks, and pass the post-test with a minimum score of 80 percent. Remediation may be provided once; the certificate is withheld until attendance, participation, and assessment requirements are satisfied.

### Post-Test Questions and Answer Key

{questions}

### Active Participation Verification Method

Live-stream attendance is verified through platform log-in records, attendance check at opening and closing, periodic polls or chat/verbal responses at least every 20 to 30 minutes, facilitator monitoring, and post-test completion. A participant who is absent, inactive, or disconnected for a material portion of the approved hours is not issued completion credit until the deficiency is resolved under an approved make-up approach.

### Works Cited and Source List

{chr(10).join(f"- {x}" for x in sources)}

### Course Approval Checklist

{chr(10).join(f"- {x}" for x in checklist)}
"""


def fill_pdf(template, output, values):
    reader = PdfReader(str(template))
    writer = PdfWriter()
    for page in reader.pages:
        writer.add_page(page)
    if "/AcroForm" in reader.trailer["/Root"]:
        writer._root_object.update({NameObject("/AcroForm"): reader.trailer["/Root"]["/AcroForm"]})
        try:
            writer.set_need_appearances_writer(True)
        except Exception:
            writer._root_object[NameObject("/AcroForm")].update({NameObject("/NeedAppearances"): BooleanObject(True)})
    for page in writer.pages:
        writer.update_page_form_field_values(page, values)
    with output.open("wb") as f:
        writer.write(f)


def safe_name(text):
    return re.sub(r"[^A-Za-z0-9]+", "_", text).strip("_")


# Source research index.
research_md = "# CDSS Source Research Index\n\nUpdated source baseline as of June 1, 2026. Official CDSS and California sources control over internal drafts.\n\n"
research_md += md_table(["Source", "URL", "Use"], SOURCE_URLS)
research_md += "\n\n## Current Requirements Confirmed\n\n"
research_md += "\n".join(
    [
        "- LIC 9141 Vendor Application/Renewal is listed by CDSS as revision 6/23.",
        "- LIC 9140 Request for Course Approval is listed by CDSS as revision 3/24.",
        "- LIC 9140A Request to Add or Replace Instructor is listed by CDSS as revision 3/23 and is a post-approval instructor change form unless needed for an approved-course change.",
        "- LIC 9139 Renewal of Continuing Education Course Approval is listed by CDSS as revision 3/23 and is for unchanged continuing education course renewals only.",
        "- LIC 9142A Roster of Participants is listed by CDSS as revision 3/23; CDSS allows an equivalent roster containing the same information.",
        "- Vendor Course Schedule Notification is revision 2/2026 and is due on or before the first day of each quarter when required.",
        "- RCFE administrator renewal requires 40 CETP hours; at least half must be in-person or live-stream unless a specific exception applies; RCFE-NHA renewal is 20 hours in any format.",
        "- ACB will not accept more than 10 administrator training hours in one day.",
        "- CETP courses must be in whole-hour increments and fit one Core of Knowledge subject.",
        "- RCFE renewal includes at least 8 hours related to Alzheimer's disease or other dementias; laws/regulations content should be covered by the first-wave catalog.",
        "- New or updated CETP course fee is $10 per hour; RCFE CETP vendor application/renewal fee is $140 under current CDSS fee guidance.",
        "- Incomplete application/payment cure window is 30 days from notice under CDSS FAQ guidance.",
        "- Records must be retained for 3 years from vendorship approval, course approval, or course offering, whichever is most recent.",
    ]
)
research_md += f"\n\n## Items Marked {PENDING_CDSS}\n\n"
research_md += "\n".join(
    [
        "- Payment notice forms do not show a printed revision date; the CDSS portal URL version timestamp is 2024-07-09.",
        "- LIC 9140 and LIC 9141 PDFs contain visible DRAFT text while being linked from the official CDSS Certification Forms page.",
        "- LIC 9139 timing differs between the form instruction and some source text; the binder uses the conservative 60-day renewal planning rule unless ACB confirms otherwise.",
        "- Final Vendor Automation Platform account workflow and reviewer access steps require current account access and assigned analyst confirmation.",
    ]
)
(RESEARCH_DIR / "CDSS_SOURCE_RESEARCH_INDEX.md").write_text(research_md + "\n", encoding="utf-8")

# Appendices.
(OFFICIAL / "APPENDIX_A_CDSS_VENDOR_APPLICATION_FORMS.md").write_text(
    f"""# Appendix A - CDSS Vendor Application Forms

Official CDSS forms are stored in this appendix folder. Filled drafts use only confirmed CI Institute of Nursing information and mark missing fields as {PENDING}.

## Included Official Forms

{md_table(['Form', 'Revision / status', 'Use'], [
    ('LIC 9141 Vendor Application/Renewal', '6/23', 'New RCFE CETP vendorship request'),
    ('Notice of Payment Information: Vendor Fees', 'CDSS portal timestamp 2024-07-09; printed revision not visible', 'Vendor application fee payment transmittal'),
    ('Supporting corporate document checklist', 'Internal checklist', 'Authority to conduct business and source-backed entity documentation'),
])}

## CI Institute of Nursing Fields Used

{md_table(['Field', 'Value'], [
    ('Organization / business name', INST['name']),
    ('Legal/business name from CIION source extract', INST['legal']),
    ('Address', INST['address']),
    ('Website', INST['website']),
    ('Phone', INST['phone']),
    ('Application type', 'New'),
    ('Program type', 'RCFE'),
    ('Program/vendor type', 'CETP'),
    ('Authorized representative/contact', PENDING),
    ('Email/fax', PENDING),
    ('Legal entity type / tax data / signer authority', PENDING),
])}

## Supporting Corporate Document Checklist

- Certificate of Status or equivalent authority-to-conduct-business documentation from the California Secretary of State: {PENDING}.
- Formation document or entity record matching the legal name used on LIC 9141: {PENDING}.
- Statement of Information or equivalent current entity filing: {PENDING}.
- Authorized representative/signature authority documentation: {PENDING}.
- IRS EIN confirmation, W-9, DBA, or local business license: include only if source-backed or requested by Finance/Compliance.
- No tax ID, signer, license, or corporate number is inserted without source confirmation.
""",
    encoding="utf-8",
)

appendix_c = f"""# Appendix C - Instructor Documentation

Instructor documentation is not complete until Compliance confirms the proposed instructor slate and source-backed qualifications. LIC 9140A is included as a post-approval reference only and is not treated as an active initial submission form unless ACB requires it for a course change.

{md_table(['Item', 'Status', 'Notes'], [
    ('Instructor roster', PENDING, 'One instructor of record and segment instructor must be assigned for each course segment.'),
    ('Current resume for each instructor', PENDING, 'Required support for LIC 9140 and any LIC 9140A change.'),
    ('Qualification pathway evidence', PENDING, 'Degree plus experience, four years experience, related California license plus experience, or qualifying administrator experience.'),
    ('Disclosure questions for each proposed instructor', PENDING, 'Professional license/certificate, facility license, facility employment, and legal/administrative action disclosures.'),
    ('Segment assignment matrix', PENDING, 'Each hour-by-hour course outline includes instructor assignment pending confirmation.'),
    ('LIC 9140A', 'Post-approval reference only', 'Use only to add or replace an instructor for an already approved course.'),
])}

## Instructor Qualification Standard

The proposed instructor must have verifiable knowledge and/or experience in the subject area and must meet at least one CDSS-recognized pathway: bachelor's or higher degree plus two years relevant experience; four years relevant experience; related California license/certification plus two years related experience; or at least four years as a California facility administrator within the last six years with substantial compliance and verifiable subject training.

## Do Not

- Do not invent instructor names, license numbers, resumes, or facility histories.
- Do not submit a course packet until instructor documentation is complete or ACB confirms an alternate acceptable process.
- Do not use LIC 9140A for the initial instructor slate unless ACB directs that use.
"""
(INSTR_DIR / "APPENDIX_C_INSTRUCTOR_DOCUMENTATION_PACKET.md").write_text(appendix_c, encoding="utf-8")

(OPS_DIR / "APPENDIX_D_POST_APPROVAL_OPERATING_FORMS.md").write_text(
    f"""# Appendix D - Post-Approval Operating Forms

These forms and templates support CI Institute of Nursing RCFE CE Operations after CDSS approval. They are not approval claims and must not be used to issue CE credit before CDSS issues the applicable vendor and course approvals.

{md_table(['Document', 'Use', 'Status'], [
    ('LIC 9142A Roster of Participants or equivalent roster', 'Attendance and completion documentation', 'Official form plus internal equivalent workbook'),
    ('Vendor Course Schedule Notification', 'Quarterly schedule notification to ACB when required', 'Official form plus filled draft'),
    ('Certificate of Completion sample', 'Post-approval certificate content control', 'Sample only; approval numbers pending CDSS issuance'),
    ('Recordkeeping log templates', 'Three-year retention and monitoring readiness', 'Internal operations workbook'),
])}

## Certificate Completion Sample Controls

The certificate sample must include vendor or authorized representative signature, vendor name, vendor number, course name, course number, approved course hours, date, time, and location. Vendor number and course number remain {PENDING} until issued by CDSS.

## Do Not

- Do not issue RCFE administrator CE certificates before CDSS approval.
- Do not use a vendor number, course number, or approved-status claim before CDSS issues it.
- Do not store roster, evaluation, or assessment records outside the approved records-control process.
""",
    encoding="utf-8",
)

(OPS_DIR / "Certificate_of_Completion_Sample_CIIN_RCFE_CETP.md").write_text(
    f"""# Certificate of Completion Sample - RCFE CETP

This is an internal post-approval certificate content sample based on CDSS expectations. It is not active and must not be issued before CDSS approval.

{md_table(['Certificate field', 'Draft value'], [
    ('Training provider', INST['name']),
    ('CDSS vendor number', '[PENDING CDSS ISSUANCE]'),
    ('Course name', '[PENDING COURSE-SPECIFIC VALUE]'),
    ('CDSS course number', '[PENDING CDSS ISSUANCE]'),
    ('Approved course hours', '[PENDING COURSE-SPECIFIC VALUE]'),
    ('Course date(s)', PENDING),
    ('Course time(s)', PENDING),
    ('Course location / live-stream platform', PENDING),
    ('Participant name', PENDING),
    ('Authorized representative signature', PENDING),
    ('Issue rule', 'Issue only after approval, full attendance, participation verification, and passing assessment.'),
])}
""",
    encoding="utf-8",
)

for c in COURSES:
    (COURSE_DIR / f"{c['id']}_{safe_name(c['title'])}.md").write_text(course_packet(c), encoding="utf-8")

# Workbooks.
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Roster Equivalent"
headers = ["Course Title", "CDSS Course Number", "Vendor Number", "Date", "Location/Platform", "Instructor", "Participant Last Name", "Participant First Name", "Phone", "Facility Name or License #", "Email", "Time In", "Time Out", "Completion Status", "Notes"]
ws.append(headers)
ws.append([PENDING, "[PENDING CDSS ISSUANCE]", "[PENDING CDSS ISSUANCE]", PENDING, PENDING, PENDING, PENDING, PENDING, PENDING, PENDING, PENDING, PENDING, PENDING, PENDING, "Use only after approval."])
for col in range(1, len(headers) + 1):
    ws.cell(1, col).font = Font(bold=True, color="FFFFFF")
    ws.cell(1, col).fill = PatternFill("solid", fgColor="8B1515")
    ws.column_dimensions[get_column_letter(col)].width = 20
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
wb.save(OPS_DIR / "RCFE_CETP_Roster_Equivalent.xlsx")

wb = openpyxl.Workbook()
for i, sheet in enumerate(["Course Files", "Instructor Files", "Evaluations", "Certificates", "Schedule Notices"]):
    ws = wb.active if i == 0 else wb.create_sheet(sheet)
    ws.title = sheet
    ws.append(["Record ID", "Course", "Date", "Record Type", "Required Retention", "Storage Location", "Owner", "Status", "Notes"])
    ws.append([PENDING, PENDING, PENDING, sheet, "3 years from vendorship approval, course approval, or course offering, whichever is most recent", INST["address"], PENDING, PENDING, "Do not use for approved operations until CDSS approval is issued."])
    for col in range(1, 10):
        ws.cell(1, col).font = Font(bold=True, color="FFFFFF")
        ws.cell(1, col).fill = PatternFill("solid", fgColor="8B1515")
        ws.column_dimensions[get_column_letter(col)].width = 22
        for cell in ws[get_column_letter(col)]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
wb.save(OPS_DIR / "RCFE_CETP_Recordkeeping_Log_Templates.xlsx")

for src, dst in [
    (OFFICIAL / "LIC_9140A_Request_Add_Replace_Instructor_fillable.pdf", INSTR_DIR / "POST_APPROVAL_REFERENCE_ONLY_LIC_9140A_Request_Add_Replace_Instructor.pdf"),
    (OFFICIAL / "LIC_9142A_3-23_Roster_of_Participants_fillable.pdf", OPS_DIR / "LIC_9142A_3-23_Roster_of_Participants_fillable.pdf"),
    (OFFICIAL / "Vendor_Course_Schedule_Notification_2-2026.pdf", OPS_DIR / "Vendor_Course_Schedule_Notification_2-2026.pdf"),
    (OFFICIAL / "ACS_Sample_Certificate_of_Completion.pdf", OPS_DIR / "ACS_Sample_Certificate_of_Completion.pdf"),
]:
    if src.exists():
        shutil.copyfile(src, dst)

# Filled drafts.
lic9141 = OFFICIAL / "LIC_9141_6-23_Vendor_Application_Renewal_fillable.pdf"
if lic9141.exists():
    fill_pdf(
        lic9141,
        OFFICIAL / "LIC_9141_FILLED_DRAFT_CIIN_RCFE_CETP.pdf",
        {
            "Check Box 123": "/Yes",
            "Check Box 126": "/Yes",
            "Check Box 101": "/Yes",
            "LIC9141_11": INST["name"],
            "LIC9141_12": INST["address"],
            "LIC9141_13": INST["authorized_rep"],
            "LIC9141_14": INST["phone"],
            "LIC9141_15": PENDING,
            "LIC9141_16": INST["email"],
            "LIC9141_17": INST["website"],
            "LIC9141_78": PENDING,
            "LIC9141_27": PENDING,
            "LIC9141_28": PENDING,
            "LIC9141_29": PENDING,
            "LIC9141_53": PENDING,
            "LIC9141_54": PENDING,
            "LIC9141_55": PENDING,
            "LIC9141_56": PENDING,
            "LIC9141_57": PENDING,
            "LIC9141_58": PENDING,
            "LIC9141_59": PENDING,
            "LIC9141_60": PENDING,
            "LIC9141_63": PENDING,
        },
    )

vendor_pay = OFFICIAL / "Notice_of_Payment_Information_Vendor_Fees.pdf"
if vendor_pay.exists():
    fill_pdf(
        vendor_pay,
        OFFICIAL / "Notice_of_Payment_Information_Vendor_Fees_FILLED_DRAFT.pdf",
        {
            "Approved Authorized Representative": PENDING,
            "Business Address Street City State Zip Code": INST["address"],
            "Telephone Number and Email": f"{INST['phone']} / {INST['email']}",
            "Facility and Program Type": "RCFE CETP",
            "RCFE CETP": "Yes",
            "Check or Money Order Number": PENDING,
            "Fee Amount": "$140",
            "Assigned Analyst If known": PENDING,
        },
    )

lic9140 = OFFICIAL / "LIC_9140_3-24_Request_for_Course_Approval_fillable.pdf"
if lic9140.exists():
    for c in COURSES:
        fill_pdf(
            lic9140,
            COURSE_DIR / f"LIC_9140_FILLED_DRAFT_{c['id']}.pdf",
            {
                "LIC9140_8": "/Yes",
                "LIC9140_9": PENDING,
                "LIC9140_10": INST["name"],
                "LIC9140_11": INST["address"],
                "LIC9140_12": PENDING,
                "LIC9140_13": INST["phone"],
                "LIC9140_14": PENDING,
                "LIC9140_15": INST["email"],
                "LIC9140_16": PENDING,
                "Text Field 20": c["title"],
                "LIC9140_17": str(c["hours"]),
                "LIC9140_18": PENDING,
                "LIC9140_19": c["participant_fee"],
                "Check Box 27": "/Yes",
                "Text Field 24": c["cat"],
                "Text Field 25": c["reviewer_access"],
                "Check Box 29": "/Yes",
                "Text Field 30": PENDING,
                "Text Field 31": PENDING,
                "Text Field 32": PENDING,
                "Text Field 37": PENDING,
            },
        )

course_pay = OFFICIAL / "Notice_of_Payment_Information_Course_Fees.pdf"
if course_pay.exists():
    values = {
        "Vendor Name and Numbers": f"{INST['name']} / {PENDING}",
        "Approved Authorized Representative": PENDING,
        "Business Address Street City State Zip Code": INST["address"],
        "Telephone Number and Email": f"{INST['phone']} / {INST['email']}",
        "Residential Care Facility for the Elderly RCFE": "Yes",
        "Fee Amount": f"${TOTAL_COURSE_FEES}",
        "Assigned Analyst If known": PENDING,
        "CheckMoney Order Total": f"Total course fees: ${TOTAL_COURSE_FEES}",
        "fill_49": f"${TOTAL_COURSE_FEES}",
    }
    row_fields = [
        ("Hiring Good Staff EXAMPLERow1", "2Row1", "Residential Care Facility for the Elderly RCFERow2", "4 hours 4x10Row1", "40Row1"),
        ("Hiring Good Staff EXAMPLERow2", "2Row2", "Residential Care Facility for the Elderly RCFERow3", "4 hours 4x10Row2", "40Row2"),
        ("Hiring Good Staff EXAMPLERow3", "2Row3", "Residential Care Facility for the Elderly RCFERow4", "4 hours 4x10Row3", "40Row3"),
        ("Hiring Good Staff EXAMPLERow4", "2Row4", "Residential Care Facility for the Elderly RCFERow5", "4 hours 4x10Row4", "40Row4"),
        ("Hiring Good Staff EXAMPLERow5", "2Row5", "Residential Care Facility for the Elderly RCFERow6", "4 hours 4x10Row5", "40Row5"),
        ("Hiring Good Staff EXAMPLERow1_2", "2Row1_2", "Residential Care Facility for the Elderly RCFERow2_2", "4 hours 4x10Row1_2", "40Row1_2"),
        ("Hiring Good Staff EXAMPLERow2_2", "2Row2_2", "Residential Care Facility for the Elderly RCFERow3_2", "4 hours 4x10Row2_2", "40Row2_2"),
        ("Hiring Good Staff EXAMPLERow3_2", "2Row3_2", "Residential Care Facility for the Elderly RCFERow4_2", "4 hours 4x10Row3_2", "40Row3_2"),
        ("Hiring Good Staff EXAMPLERow4_2", "2Row4_2", "Residential Care Facility for the Elderly RCFERow5_2", "4 hours 4x10Row4_2", "40Row4_2"),
    ]
    for c, fields in zip(COURSES, row_fields):
        values[fields[0]] = c["title"]
        values[fields[1]] = str(c["hours"])
        values[fields[2]] = "X"
        values[fields[3]] = str(c["hours"])
        values[fields[4]] = f"${c['cdss_fee']}"
    fill_pdf(course_pay, COURSE_DIR / "Notice_of_Payment_Information_Course_Fees_FILLED_DRAFT.pdf", values)

roster = OFFICIAL / "LIC_9142A_3-23_Roster_of_Participants_fillable.pdf"
if roster.exists():
    fill_pdf(
        roster,
        OPS_DIR / "LIC_9142A_FILLED_DRAFT_CIIN_RCFE_CETP.pdf",
        {
            "Check Box 20": "/Yes",
            "Text Field 12": INST["name"],
            "Text Field 13": PENDING,
            "Text Field 14": PENDING,
            "Text Field 15": PENDING,
            "Text Field 16": "Live-Stream / LMS",
            "Text Field 17": PENDING,
            "Text Field 18": PENDING,
        },
    )

schedule = OFFICIAL / "Vendor_Course_Schedule_Notification_2-2026.pdf"
if schedule.exists():
    fill_pdf(
        schedule,
        OPS_DIR / "Vendor_Course_Schedule_Notification_FILLED_DRAFT.pdf",
        {
            "Text9": INST["name"],
            "Text10": PENDING,
            "Text11": INST["address"],
            "Text12": INST["email"],
            "Text13": INST["phone"],
            "Vendor Authorized Representative SignatureRow1": PENDING,
            "TitleRow1": PENDING,
            "DateRow1": PENDING,
        },
    )

# Binder sections.
critical_rows = [
    ("Compliance", "Confirm current CDSS source baseline", "Compliance", "Official forms downloaded", "1 day", "Gate 1", "Drafted", "Outdated forms create return risk."),
    ("Forms / Corporate Documentation", "Complete LIC 9141 and corporate support", "Compliance + Finance", "Legal/entity/source docs", "2-4 days", "Gate 2", PENDING, "Application cannot be signed or submitted."),
    ("Course Development", "Finalize first-wave course packets", "Course Development", "Course list and CDSS source review", "4-7 days", "Gate 2", "Drafted", "LIC 9140 return or denial risk."),
    ("Instructor Documentation", "Attach instructor resumes and disclosures", "Compliance + Course Development", "Instructor slate", "2-5 days", "Gate 2", PENDING, "Course approvals remain incomplete."),
    ("LMS / Reviewer Access", "Validate live-stream/LMS reviewer access", "LMS Team", "Course packets and platform account", "2-4 days", "Gate 3", PENDING, "Reviewer cannot inspect online/live-stream delivery."),
    ("Website / Catalog", "Clear pending-approval catalog language", "Website / Catalog", "Course list and claim rules", "1-2 days", "Gate 3", PENDING, "Public claim risk."),
    ("Finance / Fees", "Approve vendor and course fee worksheet", "Finance", "Course hours and filing path", "1 day", "Gate 3", "Drafted", "Incomplete payment can trigger 30-day cure window."),
    ("Leadership Approval", "Dee final submission approval", "Leadership", "All prior gates", "1 day", "Final Gate", PENDING, "Submission not authorized."),
    ("CDSS Submission", "Submit LIC 9141 and LIC 9140 packets", "Compliance", "Final Gate", "1 day", "Submission", "Not started", "Late submission affects planned offering date."),
]

sections = {}
sections["00_EXECUTIVE_SUMMARY.md"] = f"""# Executive Summary

CI Institute of Nursing is preparing a dedicated RCFE Administrator Continuing Education Training Program approval track for CDSS review. This binder converts the existing RCFE/CDSS research into an internal execution packet for leadership review, compliance review, and team completion.

This RCFE CETP Project Binder intentionally separates the CDSS RCFE CETP approval track from previously developed CI Institute of Nursing CNA Recertification, HHA Initial, HHA Renewal, Moodle Build, Marketing, TTS, and CDPH application materials. Those materials remain useful institutional context, but they do not substitute for the RCFE CETP approval packet.

## What Is Being Turned Over

- Official CDSS form inventory and current source research.
- LIC 9141 filled draft for New / RCFE / CETP.
- LIC 9140 filled drafts and full attachment packets for {len(COURSES)} proposed first-wave courses.
- Instructor documentation packet with source-backed qualification requirements and pending placeholders.
- LMS/reviewer-access requirements for live-stream and LMS-supported delivery.
- Website/catalog claim-safety rules and final approval gate.
- Post-approval roster, schedule notification, certificate sample, and recordkeeping templates.

## Track Status

{md_table(['Area', 'Status', 'Executive note'], [
    ('Vendor application track', 'Drafted', 'LIC 9141 filled draft created with pending fields clearly marked.'),
    ('Course approval track', 'Drafted', f'{len(COURSES)} first-wave RCFE CETP course packets drafted with {TOTAL_HOURS} proposed live-stream hours.'),
    ('Official forms', 'Appendix only', 'Official CDSS forms and filled drafts are stored in appendices, not recreated as internal form templates in the binder body.'),
    ('Instructor documentation', 'Pending Compliance', 'Instructor names, resumes, qualification evidence, and disclosures remain pending source confirmation.'),
    ('LMS / reviewer access', 'Needs Update', 'Reviewer URL, credentials, demo live-stream access, and account validation remain pending.'),
    ('Finance / fees', 'Drafted', f'Planning fee: $140 vendor application fee plus ${TOTAL_COURSE_FEES} in proposed first-wave course fees.'),
    ('Leadership gate', 'Required Before Submission', 'Dee approval is required before submission, signatures, approval claims, certificates, or use of issued numbers.'),
])}

## Gate Status

- Execution may continue internally.
- No CDSS submission, signatures, approval claims, certificates, or issued-number use until final clearance.
- Missing fields remain marked {PENDING}; unresolved CDSS interpretations remain marked {PENDING_CDSS}.
"""

sections["01_PROJECT_CHARTER.md"] = f"""# Project Charter

{md_table(['Field', 'Value'], [
    ('Institution', INST['name']),
    ('Program track', INST['program']),
    ('Binder', INST['binder']),
    ('Purpose', INST['purpose']),
    ('Project Oversight', INST['oversight']),
    ('Website', INST['website']),
    ('Physical address', INST['address']),
    ('Mailing address from CIION source extract', INST['mailing']),
])}

## Objective

Prepare the CI Institute of Nursing RCFE CETP Approval Packet for CDSS review by organizing the official vendor application, first-wave course approvals, instructor documentation, LMS/reviewer access tasks, website/catalog claim controls, finance fee worksheet, recordkeeping controls, and final leadership gate into a single executive binder.

## In Scope

- RCFE CETP vendorship application planning.
- Official CDSS forms and filled drafts using source-backed information only.
- First-wave RCFE CETP course approval documentation.
- Instructor documentation requirements and pending placeholders.
- LMS/reviewer access requirements for live-stream and LMS-supported delivery.
- Website/catalog language controls for approval-safe communications.
- Fee planning, payment worksheets, submission checklist, and final gate.

## Out of Scope

- HTML executive summary.
- New web app, UI prototype, or certificate issuance workflow.
- CNA Recertification application edits.
- ContentV2 or standalone-course-mvp modifications.
- CDSS approval claims, vendor numbers, course numbers, or signatures before issuance/clearance.
"""

sections["02_TEAM_EXECUTION_SUMMARY.md"] = "# Team Execution Summary\n\n" + "\n".join(
    [
        team_block("Compliance", ["Current CDSS source research index.", "Official CDSS forms and filled draft inventory.", "Submission checklist and risk register."], ["Review all pending fields.", "Validate official form versions immediately before submission.", "Confirm instructor documentation is complete.", "Clear website/catalog language before publication."], ["Compliance-cleared RCFE CETP submission package."], ["Do not submit, sign, or publish approval claims before final clearance.", "Do not replace pending fields with inferred values."]),
        team_block("Course Development", [f"{len(COURSES)} first-wave course packets with objectives, outlines, post-tests, sources, and checklists.", "RCFE Core of Knowledge mapping."], ["Review course titles and hours for strategic fit.", "Assign instructors by segment.", "Attach resumes and source-backed qualification evidence.", "Build learner-facing materials after Compliance approves course scope."], ["Submission-ready LIC 9140 attachments for each first-wave course."], ["Do not split one course across multiple Core categories.", "Do not use generic instructions in place of a full course outline."]),
        team_block("Forms / Corporate Documentation", ["LIC 9141 filled draft.", "Vendor fee notice filled draft.", "Corporate document checklist."], ["Confirm legal/entity name and type.", "Obtain California authority-to-conduct-business documentation.", "Confirm signer authority and payment instrument details."], ["Final corporate support file for the RCFE CETP Approval Packet."], ["Do not invent tax IDs, license numbers, corporate numbers, signer names, or entity status."]),
        team_block("Instructor Documentation", ["Instructor documentation packet and LIC 9140A post-approval reference."], ["Confirm instructor roster.", "Collect resumes and licenses/certifications where applicable.", "Complete disclosure answers for each proposed instructor."], ["Course-specific instructor qualification packet."], ["Do not submit instructor declarations without source-backed resumes and disclosures."]),
        team_block("LMS / Reviewer Access", ["Live-stream/LMS participation-control requirements.", "Reviewer access fields marked pending in each LIC 9140 filled draft."], ["Create reviewer-access environment.", "Validate credentials.", "Prepare participation/audit report sample.", "Confirm post-test and completion-gate behavior."], ["Reviewer-ready LMS/live-stream access instructions."], ["Do not expose live reviewer passwords in open artifacts.", "Do not activate certificate issuance before approval."]),
        team_block("Website / Catalog", ["Claim-safe language controls.", "Approval gate rules."], ["Prepare pending-approval catalog copy.", "Remove any wording implying CDSS approval before issuance.", "Add final approved copy only after Compliance clearance."], ["Claim-safe RCFE CE catalog and website language."], ["Do not advertise approved status, vendor number, or course number before CDSS issuance."]),
        team_block("Finance / Fees", ["Vendor fee worksheet.", "Course fee worksheet.", f"Planning total: $140 plus ${TOTAL_COURSE_FEES} for first-wave course submissions."], ["Confirm payment method.", "Confirm check/money order or VAP payment path.", "Approve final fee worksheet."], ["Payment-ready application and course fee package."], ["Do not submit without reconciled fee worksheet and payment authority."]),
        team_block("Leadership Approval", ["Executive binder, critical path, risk register, final gate."], ["Review final package after Compliance clearance.", "Approve final submission package."], ["Final authorization to submit."], ["Do not approve submission until official forms, course packets, instructor docs, LMS access, website language, and fees are cleared."]),
        team_block("Post-Approval Operations", ["Roster, schedule notification, certificate sample, recordkeeping logs."], ["Prepare quarterly schedule process.", "Activate roster/evaluation/certificate controls after approval.", "Maintain records for required retention period."], ["Controlled RCFE CE operating workflow after approval."], ["Do not issue CE certificates or operate as approved before CDSS approval."]),
    ]
)

sections["03_CRITICAL_PATH_TO_SUBMISSION.md"] = f"""# Critical Path to Submission

Critical-path tasks block CDSS submission. Parallel work may continue, but the final submission package cannot move forward until all gates are cleared.

{md_table(['Workstream', 'Task', 'Owner', 'Dependency', 'Duration', 'Gate', 'Status', 'Risk if delayed'], critical_rows)}

## Parallel Work

- Course slide and handout preparation may continue after course packet scope is approved.
- LMS shell setup may continue while instructor documentation is collected.
- Website/catalog draft copy may continue while legal/entity and fee details are pending.
- Operations forms may be prepared but must remain inactive before CDSS approval.

## Final Approval Gates

- Gate 1: CDSS source and form version check complete.
- Gate 2: Official forms, corporate support, course packets, and instructor documents complete.
- Gate 3: Reviewer access, website/catalog language, and fee worksheet approved.
- Final Gate: Dee approves the final submission package.
"""

sections["04_COMPLIANCE_TASKS.md"] = "# Compliance Tasks\n\n" + team_block("Compliance", ["Current CDSS source research index.", "Official form inventory and filled draft set.", "Risk register and QA checklist.", "Final approval gate language."], ["Check every official form against the CDSS Certification Forms page immediately before submission.", "Confirm all pending fields with source-backed documentation.", "Review LIC 9140 attachments against the ACB course-review guide.", "Confirm no approval claims appear on website/catalog materials.", "Prepare incomplete-notice response workflow before submission."], ["Compliance-cleared CI Institute of Nursing RCFE CETP Approval Packet."], ["Do not submit incomplete forms when pending fields remain unresolved.", "Do not use unofficial form versions if CDSS updates the form page.", "Do not approve certificate issuance before CDSS approval."])

sections["05_COURSE_DEVELOPMENT_TASKS.md"] = (
    f"# Course Development Tasks\n\nThe first-wave catalog is designed to support RCFE administrator renewal needs while respecting the CDSS rule that a CETP course should fit one Core of Knowledge subject. The course package includes {TOTAL_HOURS} proposed live-stream hours and covers the law/regulation and dementia priorities.\n\n"
    + md_table(["Course ID", "Title", "Hours", "Core category", "Delivery", "CDSS fee"], [(c["id"], c["title"], c["hours"], c["cat"], c["delivery"], f"${c['cdss_fee']}") for c in COURSES])
    + "\n\n"
    + team_block("Course Development", ["Detailed course packets in Appendix B.", "LIC 9140 filled drafts for each proposed course.", "Post-test and answer key for each proposed course."], ["Review and approve first-wave course list.", "Confirm course materials are consistent with each approved outline.", "Assign instructors by segment and attach qualification evidence.", "Confirm LMS/live-stream participation controls before submission."], ["Submission-ready course approval packets."], ["Do not leave any LIC 9140 attachment as generic instructions.", "Do not map one CETP course to multiple Core categories.", "Do not issue certificates before approval."])
)

sections["06_LMS_TASKS.md"] = "# LMS Tasks\n\n" + team_block("LMS / Reviewer Access", ["Course delivery method set to live-stream with LMS-supported materials and post-test.", "Reviewer access fields marked pending in each LIC 9140 filled draft.", "Active participation verification method included in each course packet."], ["Create a reviewer-ready LMS area for each course.", "Provide reviewer account credentials through a controlled channel.", "Validate post-test scoring, attendance log exports, participation record exports, and completion gates.", "Prepare a live-stream demonstration link or scheduled access process if requested by ACB."], ["Validated reviewer access instructions for each LIC 9140 packet."], ["Do not publish reviewer passwords in binder files.", "Do not enable certificate generation until CDSS approval and Compliance clearance.", "Do not represent self-paced delivery unless the self-paced course is separately reviewed and approved."])

sections["07_WEBSITE_AND_CATALOG_TASKS.md"] = "# Website and Catalog Tasks\n\n" + team_block("Website / Catalog", ["Claim-safe language controls.", "Course list and pending-approval status."], ["Draft catalog copy that says the RCFE CETP approval packet is in preparation or pending, if public mention is necessary.", "Remove any implication that CDSS approval has been issued.", "Publish approved course numbers, vendor number, and certificate language only after CDSS issuance and Compliance clearance."], ["Claim-safe RCFE CE website/catalog content."], ["Do not advertise approved status.", "Do not list CDSS course numbers or vendor numbers before issued.", "Do not imply administrators may use the courses for renewal credit before approval."])

sections["08_CDSS_FORMS_AND_APPENDICES.md"] = f"""# CDSS Forms and Appendices

Official forms are stored in appendices and are not reproduced as generated text-box templates in the binder body. Filled drafts are marked clearly when source-backed information is missing.

{md_table(['Appendix', 'Folder', 'Contents'], [
    ('Appendix A', 'appendices/official_forms', 'LIC 9141, vendor fee notice, official CDSS form set, corporate document checklist.'),
    ('Appendix B', 'appendices/course_approval_packets', 'LIC 9140 filled drafts, one course packet per proposed first-wave course, course fee notice draft.'),
    ('Appendix C', 'appendices/instructor_documentation', 'Instructor documentation packet and LIC 9140A post-approval reference.'),
    ('Appendix D', 'appendices/operations_forms', 'LIC 9142A, schedule notification, certificate sample, roster equivalent, recordkeeping logs.'),
])}

## Current Official Form Inventory

{md_table(['Form / artifact', 'Current version confirmed', 'Packet use'], [
    ('LIC 9141 Vendor Application/Renewal', '6/23', 'Appendix A - new RCFE CETP vendorship.'),
    ('LIC 9140 Request for Course Approval', '3/24', 'Appendix B - one per first-wave course.'),
    ('LIC 9140A Request to Add or Replace Instructor', '3/23', 'Appendix C - post-approval reference only unless ACB requires otherwise.'),
    ('LIC 9139 Renewal of Continuing Education Course Approval', '3/23', 'Future renewal only for unchanged courses.'),
    ('LIC 9142A Roster of Participants', '3/23', 'Appendix D - post-approval operations.'),
    ('Vendor Course Schedule Notification', '2/2026', 'Appendix D - quarterly schedule process.'),
    ('Vendor fee payment notice', 'Printed revision not visible; CDSS URL timestamp 2024-07-09', 'Appendix A - fee transmittal.'),
    ('Course fee payment notice', 'Printed revision not visible; CDSS URL timestamp 2024-07-09', 'Appendix B - course fee transmittal.'),
])}

## Pending Form Fields

- Legal entity type and authority documentation: {PENDING}.
- Authorized representative/contact details: {PENDING}.
- Signature authority and signature date: {PENDING}.
- Vendor number, assigned analyst, and course numbers: pending CDSS issuance.
- Instructor names, qualification pathway evidence, and disclosure answers: {PENDING}.
- Reviewer access details: {PENDING}.
"""

sections["09_INSTRUCTOR_QUALIFICATION_PACKET.md"] = appendix_c.replace("# Appendix C - Instructor Documentation", "# Instructor Qualification Packet")
sections["10_COURSE_APPROVAL_DOCUMENTATION.md"] = "# Course Approval Documentation\n\nThis section contains the course approval documentation needed to support LIC 9140 attachments. Each proposed course has a corresponding filled LIC 9140 draft and a course-specific appendix packet.\n\n" + "\n".join(course_packet(c) for c in COURSES)

sections["11_POLICIES_AND_PROCEDURES.md"] = "# Policies and Procedures\n\n" + team_block("Policies and Procedures", ["Recordkeeping policy requirements from CDSS sources.", "Certificate content controls from CDSS FAQ and sample certificate.", "Live-stream participation and post-test rules embedded in course packets."], ["Finalize RCFE CETP recordkeeping SOP.", "Finalize attendance and active participation SOP.", "Finalize post-test, remediation, and certificate release SOP.", "Finalize approval-claims and website/catalog language SOP."], ["Operations-ready policy set after CDSS approval."], ["Do not issue CE certificates until all approved-course and participant-completion rules are met.", "Do not alter approved course content without determining whether a new LIC 9140 is required."])

sections["12_RECORDKEEPING_AND_OPERATIONS.md"] = "# Recordkeeping and Operations\n\nCI Institute of Nursing RCFE CE Operations will retain required records for three years from vendorship approval, course approval, or course offering, whichever is most recent.\n\n" + md_table(["Record type", "Minimum content", "Storage / owner"], [
    ("Course file", "Approved course title, number, outline, source list, schedule, and material set.", f"{INST['address']} / Compliance {PENDING}"),
    ("Instructor file", "Resume, qualification evidence, disclosure answers, segment assignment, approval records.", f"{INST['address']} / Compliance {PENDING}"),
    ("Roster/completion file", "LIC 9142A or equivalent roster, attendance logs, participation checks, completion status.", f"{INST['address']} / Operations {PENDING}"),
    ("Evaluation file", "Participant evaluation of course and instructor plus QA review notes.", f"{INST['address']} / QA {PENDING}"),
    ("Certificate file", "Certificate copy or issuance log with vendor/course number, hours, dates, times, and location.", f"{INST['address']} / Operations {PENDING}"),
]) + "\n\n" + team_block("Post-Approval Operations", ["LIC 9142A and equivalent roster workbook.", "Quarterly schedule notification form.", "Certificate sample and recordkeeping logs."], ["Activate roster and certificate controls only after approval.", "Submit quarterly schedule notifications when required.", "Prepare monitoring-ready records for each approved course offering."], ["Audit-ready RCFE CE operating file."], ["Do not retroactively issue certificates for unapproved offerings.", "Do not modify course delivery from the approved format without Compliance review."])

risk_rows = [
    ("Incomplete LIC 9141", "High", "Missing legal/entity, contact, fee, or signer data can trigger incomplete notice.", "Use pending-field matrix; collect source-backed records before signature."),
    ("Incomplete LIC 9140 attachments", "High", "Course request can be returned or delayed if required attachments are missing.", "QA every course against LIC 9140 and ACB guide."),
    ("Instructor documentation not source-backed", "High", "Course approval may not proceed if qualification pathway is unsupported.", "Collect resumes, licenses, disclosures, and segment assignment matrix."),
    ("Reviewer access not validated", "High", "ACB cannot review live-stream/LMS-supported delivery.", "Validate reviewer credentials, live-stream path, post-test, and activity logs."),
    ("Premature approval claims", "High", "Public materials could imply approval or renewal credit before CDSS action.", "Use claim-safe catalog language and final Compliance clearance."),
    ("Course category span", "Medium", "CDSS expects each CETP course to fit one Core of Knowledge subject.", "Keep each course mapped to one category; split topics when needed."),
    ("Payment mismatch", "Medium", "Incorrect fee can trigger incomplete notice or delay.", "Finance approves $140 vendor fee and $10/hour course fee worksheet."),
    ("Outdated official form", "Medium", "Using superseded forms can delay processing.", "Check CDSS forms page immediately before submission."),
    ("Recordkeeping gaps", "Medium", "Monitoring readiness depends on rosters, evaluations, instructor records, schedules, and completion records.", "Use Appendix D logs and three-year retention controls."),
]
sections["13_RISK_REGISTER.md"] = "# Risk Register\n\n" + md_table(["Risk", "Severity", "Why it matters", "Control"], risk_rows)

sections["14_QA_AND_INTERNAL_AUDIT.md"] = f"""# QA and Internal Audit

{md_table(['QA item', 'Status'], [
    ('Official form versions checked against CDSS Certification Forms page', PENDING),
    ('LIC 9141 filled and reviewed', 'Drafted; pending fields remain'),
    ('Vendor fee worksheet approved', 'Drafted; Finance approval pending'),
    ('One LIC 9140 draft prepared per course', 'Drafted'),
    ('Course packets include description, objectives, outline, methods, evaluation, post-test, sources, active participation, records, and checklist', 'Drafted'),
    ('Instructor resumes and disclosures attached', PENDING),
    ('Reviewer access validated', PENDING),
    ('Website/catalog language claim-safe', PENDING),
    ('All approval claims removed until CDSS approval', 'Required'),
    ('Dee final approval completed', PENDING),
])}

Any field that cannot be traced to CI Institute of Nursing source material, official CDSS guidance, or a leadership-confirmed value remains marked {PENDING}. Any unresolved regulatory interpretation remains marked {PENDING_CDSS}.
"""

sections["15_SUBMISSION_PACKAGE_CHECKLIST.md"] = f"""# Submission Package Checklist

## Required Before Submission

- Compliance completes final review.
- Official forms are filled and checked.
- Course approval packets are complete.
- Instructor documentation is complete.
- Catalog and website language are claim-safe.
- Reviewer access is validated, if online or self-paced/live-stream.
- Fee worksheet is approved by Finance.
- All required signatures are collected.
- Dee approves the final submission package.

## Package Components

{md_table(['Component', 'Location', 'Status'], [
    ('LIC 9141 filled draft', 'appendices/official_forms', 'Drafted; pending fields remain'),
    ('Vendor payment notice', 'appendices/official_forms', 'Drafted; payment details pending'),
    ('Corporate documentation checklist', 'appendices/official_forms', PENDING),
    ('LIC 9140 filled drafts', 'appendices/course_approval_packets', 'Drafted for each first-wave course'),
    ('Course approval attachments', 'appendices/course_approval_packets', 'Drafted'),
    ('Course fee notice', 'appendices/course_approval_packets', 'Drafted; payment details pending'),
    ('Instructor documentation packet', 'appendices/instructor_documentation', PENDING),
    ('LMS/reviewer access instructions', 'Course packets and LMS task section', PENDING),
    ('Operations forms', 'appendices/operations_forms', 'Prepared for post-approval use'),
])}
"""

sections["16_POST_APPROVAL_OPERATIONS.md"] = "# Post-Approval Operations\n\n" + team_block("Post-Approval Operations", ["Roster equivalent workbook.", "Recordkeeping log workbook.", "Certificate sample.", "Course schedule notification form."], ["Enter approved vendor number and course numbers only after CDSS issuance.", "Schedule courses consistent with approved format and approved hours.", "Maintain attendance, participation, assessment, evaluation, instructor, and certificate records.", "Submit quarterly schedules when required or requested."], ["Controlled CI Institute of Nursing RCFE CE Operations."], ["Do not use any approval number before CDSS issues it.", "Do not offer courses for administrator renewal credit before approval.", "Do not alter approved course format, hours, or content without Compliance review."])

sections["18_FINAL_APPROVAL_GATE.md"] = """# Final Approval Gate

Execution may continue.

Do NOT proceed with the following until final clearance:

- Submit any application to CDSS.
- Sign any CDSS forms.
- Publish approval claims.
- Issue administrator CE certificates.
- Advertise approved status.
- Use any CDSS vendor number or course approval number before issued.

Required before submission:

- Compliance completes final review.
- Official forms are filled and checked.
- Course approval packets are complete.
- Instructor documentation is complete.
- Catalog and website language are claim-safe.
- Reviewer access is validated, if online or self-paced/live-stream.
- Fee worksheet is approved by Finance.
- All required signatures are collected.
- Dee approves the final submission package.
"""

for name, content in sections.items():
    (ROOT / name).write_text(content, encoding="utf-8")

cover = f"""# {INST['name']}

## {INST['program']}

## {INST['binder']}

{INST['purpose']}

Prepared by:
{INST['prepared_by']}

Project Oversight:
{INST['oversight']}

{INST['website']}
{INST['address']}

Prepared: {TODAY}

---

"""
order = [
    "00_EXECUTIVE_SUMMARY.md",
    "01_PROJECT_CHARTER.md",
    "02_TEAM_EXECUTION_SUMMARY.md",
    "03_CRITICAL_PATH_TO_SUBMISSION.md",
    "04_COMPLIANCE_TASKS.md",
    "05_COURSE_DEVELOPMENT_TASKS.md",
    "06_LMS_TASKS.md",
    "07_WEBSITE_AND_CATALOG_TASKS.md",
    "08_CDSS_FORMS_AND_APPENDICES.md",
    "09_INSTRUCTOR_QUALIFICATION_PACKET.md",
    "10_COURSE_APPROVAL_DOCUMENTATION.md",
    "11_POLICIES_AND_PROCEDURES.md",
    "12_RECORDKEEPING_AND_OPERATIONS.md",
    "13_RISK_REGISTER.md",
    "14_QA_AND_INTERNAL_AUDIT.md",
    "15_SUBMISSION_PACKAGE_CHECKLIST.md",
    "16_POST_APPROVAL_OPERATIONS.md",
    "18_FINAL_APPROVAL_GATE.md",
]
(ROOT / "RCFE_CETP_Project_Binder.md").write_text(cover + "\n\n---\n\n".join((ROOT / x).read_text(encoding="utf-8") for x in order), encoding="utf-8")

# Linked library.
rows = []
def add_doc(section, path, typ, status, owner="", notes=""):
    rows.append([section, path.name, str(path), typ, status, owner, notes])
add_doc("Binder", ROOT / "RCFE_CETP_Project_Binder.md", "Markdown", "Generated", "Compliance", "Aggregated binder")
add_doc("Binder", ROOT / "RCFE_CETP_Project_Binder.pdf", "PDF", "Generated", "Compliance", "Polished binder")
for name in order:
    add_doc("Binder Sections", ROOT / name, "Markdown", "Generated", "Compliance", "Required section file")
for folder, section in [(OFFICIAL, "Appendix A / Official Forms"), (COURSE_DIR, "Appendix B / Course Approval"), (INSTR_DIR, "Appendix C / Instructor"), (OPS_DIR, "Appendix D / Operations"), (RESEARCH_DIR, "Source Research"), (NOTES_DIR, "Working Notes")]:
    for path in sorted(folder.glob("*")):
        if path.is_file():
            add_doc(section, path, path.suffix.replace(".", "").upper() or "File", "Generated/downloaded")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Linked Document Library"
ws.append(["Section", "Document Name", "Path", "Type", "Status", "Owner", "Notes"])
for row in rows:
    ws.append(row)
widths = [26, 48, 92, 14, 24, 18, 44]
for col in range(1, 8):
    ws.cell(1, col).font = Font(bold=True, color="FFFFFF")
    ws.cell(1, col).fill = PatternFill("solid", fgColor="8B1515")
    ws.column_dimensions[get_column_letter(col)].width = widths[col - 1]
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
wb.save(ROOT / "LINKED_DOCUMENT_LIBRARY.xlsx")
shutil.copyfile(ROOT / "LINKED_DOCUMENT_LIBRARY.xlsx", ROOT / "17_LINKED_DOCUMENT_LIBRARY.xlsx")


# PDF generation.
M = HexColor("#8B1515")
G = HexColor("#FFC107")
DK = HexColor("#212529")
GY = HexColor("#666666")
LG = HexColor("#DEE2E6")
SURF = HexColor("#F8F9FA")
RED = HexColor("#B71C1C")


class BinderPDF:
    def __init__(self, path):
        self.c = canvas.Canvas(str(path), pagesize=letter)
        self.c.setTitle("CI Institute of Nursing RCFE CETP Project Binder")
        self.c.setAuthor(INST["prepared_by"])
        self.w, self.h = letter
        self.page = 0
        self.land = False
        self.y = 0

    def cover(self):
        self.w, self.h = letter
        self.c.setPageSize(letter)
        self.c.setFillColor(M)
        self.c.rect(0, 0, self.w, self.h, fill=1, stroke=0)
        self.c.setFillColor(G)
        self.c.rect(0, self.h - 8, self.w, 8, fill=1, stroke=0)
        self.c.rect(0, 78, self.w, 3, fill=1, stroke=0)
        if LOGO.exists():
            try:
                self.c.drawImage(str(LOGO), self.w / 2 - 110, self.h - 180, width=220, height=80, preserveAspectRatio=True, mask="auto")
            except Exception:
                pass
        self.c.setFillColor(white)
        self.c.setFont("Helvetica-Bold", 30)
        self.c.drawCentredString(self.w / 2, self.h - 245, "CI Institute of Nursing")
        self.c.setStrokeColor(G)
        self.c.setLineWidth(2)
        self.c.line(self.w / 2 - 150, self.h - 265, self.w / 2 + 150, self.h - 265)
        self.c.setFont("Helvetica-Bold", 14)
        self.c.drawCentredString(self.w / 2, self.h - 300, "RCFE Administrator Continuing Education Training Program (CETP)")
        self.c.drawCentredString(self.w / 2, self.h - 320, "CDSS Vendor Approval Project Binder")
        self.c.setFillColor(HexColor("#F3DADA"))
        self.c.setFont("Helvetica", 11)
        self.c.drawCentredString(self.w / 2, self.h - 350, "Prepared for Team Execution and Compliance Review")
        self.c.setFillColor(white)
        self.c.setFont("Helvetica", 10)
        for i, line in enumerate(["Prepared by: TJ Padilla", "Project Oversight: Dee Bustos", TODAY, INST["website"], INST["address"]]):
            self.c.drawCentredString(self.w / 2, 205 - i * 17, line)
        self.c.showPage()
        self.page = 0

    def new_page(self, land=False):
        if self.page > 0:
            self.footer()
            self.c.showPage()
        self.land = land
        self.w, self.h = landscape(letter) if land else letter
        self.c.setPageSize((self.w, self.h))
        self.page += 1
        self.header()
        self.y = self.h - 66

    def header(self):
        self.c.setStrokeColor(M)
        self.c.setLineWidth(1.4)
        self.c.line(54, self.h - 34, self.w - 54, self.h - 34)
        self.c.setStrokeColor(G)
        self.c.setLineWidth(0.7)
        self.c.line(54, self.h - 37, self.w - 54, self.h - 37)
        self.c.setFillColor(GY)
        self.c.setFont("Helvetica", 7.2)
        self.c.drawString(54, self.h - 28, f"{INST['name']} | {INST['address']} | {INST['website']}")

    def footer(self):
        self.c.setStrokeColor(G)
        self.c.setLineWidth(0.5)
        self.c.line(54, 34, self.w - 54, 34)
        self.c.setFillColor(GY)
        self.c.setFont("Helvetica", 6.8)
        self.c.drawString(54, 24, "CI Institute of Nursing - RCFE CETP Project Binder / TJ Padilla")
        self.c.drawRightString(self.w - 54, 24, f"Page {self.page}")

    def need(self, n):
        if self.y - n < 50:
            self.new_page(self.land)

    def wrap(self, text, font, size, max_w):
        words = str(text).split()
        lines = []
        cur = ""
        for word in words:
            test = f"{cur} {word}".strip()
            if self.c.stringWidth(test, font, size) <= max_w:
                cur = test
            else:
                if cur:
                    lines.append(cur)
                cur = word
        if cur:
            lines.append(cur)
        return lines or [""]

    def h1(self, text):
        self.need(34)
        self.c.setFillColor(M)
        self.c.setFont("Helvetica-Bold", 15)
        self.c.drawString(60, self.y, text)
        self.y -= 8
        self.c.setStrokeColor(G)
        self.c.setLineWidth(0.8)
        self.c.line(60, self.y, self.w - 60, self.y)
        self.y -= 14

    def h2(self, text):
        self.need(22)
        self.c.setFillColor(M)
        self.c.setFont("Helvetica-Bold", 10.2)
        self.c.drawString(60, self.y, text)
        self.y -= 14

    def para(self, text, size=8.4, color=DK, space=6):
        lines = []
        for part in str(text).split("\n"):
            lines.extend(self.wrap(part, "Helvetica", size, self.w - 120))
        self.need(len(lines) * size * 1.35 + space)
        self.c.setFont("Helvetica", size)
        self.c.setFillColor(color)
        for line in lines:
            self.c.drawString(60, self.y, line)
            self.y -= size * 1.35
        self.y -= space

    def bullets(self, items, size=7.8):
        for item in items:
            lines = self.wrap(item, "Helvetica", size, self.w - 140)
            self.need(len(lines) * size * 1.3 + 3)
            self.c.setFont("Helvetica", size)
            self.c.setFillColor(DK)
            self.c.drawString(68, self.y, "-")
            for line in lines:
                self.c.drawString(80, self.y, line)
                self.y -= size * 1.3
            self.y -= 2
        self.y -= 4

    def box(self, title, lines, fill=SURF, stroke=G):
        h = 20 + 12 * len(lines)
        self.need(h + 8)
        self.c.setFillColor(fill)
        self.c.roundRect(60, self.y - h + 8, self.w - 120, h, 4, fill=1, stroke=0)
        self.c.setStrokeColor(stroke)
        self.c.setLineWidth(0.6)
        self.c.roundRect(60, self.y - h + 8, self.w - 120, h, 4, fill=0, stroke=1)
        self.c.setFillColor(M)
        self.c.setFont("Helvetica-Bold", 8.5)
        self.c.drawString(70, self.y - 6, title)
        self.c.setFillColor(DK)
        self.c.setFont("Helvetica", 7.8)
        y = self.y - 20
        for line in lines:
            self.c.drawString(76, y, line)
            y -= 11
        self.y -= h + 8

    def table(self, headers, data, widths=None, size=6.7, hsize=7.0):
        x0 = 60
        max_w = self.w - 120
        if widths is None:
            widths = [max_w / len(headers)] * len(headers)
        scale = max_w / sum(widths)
        widths = [w * scale for w in widths]

        def row_height(row, header=False):
            font = "Helvetica-Bold" if header else "Helvetica"
            sz = hsize if header else size
            return max(14, max(len(self.wrap(cell, font, sz, w - 6)) * sz * 1.22 + 8 for cell, w in zip(row, widths)))

        def draw_row(row, header=False):
            rh = row_height(row, header)
            if self.y - rh < 50:
                return False
            self.c.setFillColor(M if header else white)
            self.c.rect(x0, self.y - rh, sum(widths), rh, fill=1, stroke=0)
            x = x0
            for cell, w in zip(row, widths):
                self.c.setStrokeColor(LG)
                self.c.setLineWidth(0.35)
                self.c.rect(x, self.y - rh, w, rh, fill=0, stroke=1)
                font = "Helvetica-Bold" if header else "Helvetica"
                sz = hsize if header else size
                self.c.setFont(font, sz)
                self.c.setFillColor(white if header else DK)
                ty = self.y - sz - 4
                for line in self.wrap(cell, font, sz, w - 6):
                    self.c.drawString(x + 3, ty, line)
                    ty -= sz * 1.22
                x += w
            self.y -= rh
            return True

        self.need(28)
        if not draw_row(headers, True):
            self.new_page(self.land)
            draw_row(headers, True)
        for row in data:
            if not draw_row([str(c) for c in row], False):
                self.new_page(self.land)
                draw_row(headers, True)
                draw_row([str(c) for c in row], False)
        self.y -= 10

    def close(self):
        self.footer()
        self.c.save()


pdf = BinderPDF(ROOT / "RCFE_CETP_Project_Binder.pdf")
pdf.cover()
pdf.new_page()
pdf.h1("Executive Summary")
pdf.para("CI Institute of Nursing is preparing a dedicated RCFE Administrator Continuing Education Training Program approval track for CDSS review. This binder converts current RCFE/CDSS research into an internal execution packet for leadership review, compliance review, and team completion.")
pdf.box("Gate Status", ["Execution may continue internally.", "No CDSS submission, signatures, approval claims, certificates, or issued-number use until final clearance.", f"First-wave draft: {len(COURSES)} courses, {TOTAL_HOURS} hours, ${TOTAL_COURSE_FEES} course-fee planning total."])
pdf.h2("Status Matrix")
pdf.table(["Area", "Status", "Executive note"], [
    ("Vendor application", "Drafted", "LIC 9141 filled draft prepared with pending fields."),
    ("Course packets", "Drafted", f"{len(COURSES)} LIC 9140 course packets with post-tests and source lists."),
    ("Instructor docs", PENDING, "Names, resumes, qualifications, and disclosures required."),
    ("Reviewer access", PENDING, "LMS/live-stream account and access instructions required."),
    ("Finance", "Drafted", f"$140 vendor fee plus ${TOTAL_COURSE_FEES} course fees."),
    ("Leadership", PENDING, "Dee final approval required before submission."),
], [110, 85, 330])
pdf.h1("Project Charter")
pdf.table(["Field", "Value"], [("Institution", INST["name"]), ("Program", INST["program"]), ("Binder", INST["binder"]), ("Purpose", INST["purpose"]), ("Project Oversight", INST["oversight"]), ("Physical address", INST["address"]), ("Mailing address", INST["mailing"])], [160, 370])
pdf.para("The CI Institute of Nursing document library and repository materials were used as the institutional baseline for identifying existing CNA, HHA, Moodle, Marketing, and CDPH documentation. This RCFE CETP Project Binder separates the CDSS RCFE CETP approval track from those existing materials.")
pdf.h1("Team Execution Summary")
team_rows = [
    ("Compliance", "CDSS source index, forms, risk register", "Clear pending fields, form versions, website claims", "Submission-cleared package"),
    ("Course Development", "Course packets and post-tests", "Finalize materials, assign instructors", "LIC 9140 attachments"),
    ("Forms / Corporate Docs", "LIC 9141 draft and checklist", "Confirm legal/entity and signer data", "Corporate support file"),
    ("Instructor Documentation", "Qualification packet", "Collect resumes/disclosures", "Instructor support file"),
    ("LMS / Reviewer Access", "Participation controls", "Validate reviewer account", "Reviewer-ready access"),
    ("Website / Catalog", "Claim-safe rules", "Remove approval claims", "Claim-safe catalog"),
    ("Finance / Fees", "Fee worksheet", "Approve payment method", "Payment-ready package"),
    ("Leadership", "Binder and gate", "Approve after clearance", "Submission authorization"),
]
pdf.table(["Team", "Provided", "Action Steps", "Output"], team_rows, [95, 150, 170, 120], size=6.5)
pdf.new_page(land=True)
pdf.h1("Critical Path to Submission")
pdf.para("Full-page dependency view. Critical dependencies block CDSS submission; parallel work can continue while pending fields are resolved.", size=8.0)
pdf.table(["Workstream", "Task", "Owner", "Dependency", "Duration", "Gate", "Status", "Risk if delayed"], critical_rows, [95, 130, 85, 120, 60, 70, 80, 160], size=6.3, hsize=6.6)
pdf.box("Parallel Work", ["Course materials, LMS shells, website draft copy, and operations logs may continue before final submission.", "Final gate blocks CDSS submission, signatures, approval claims, certificates, and use of issued numbers."], fill=HexColor("#FFF8E1"))
pdf.new_page()
pdf.h1("Compliance Tasks")
pdf.bullets(["Check official CDSS form versions immediately before submission.", "Resolve all pending form fields with source-backed documentation.", "QA LIC 9140 attachments against the ACB course-review guide.", "Clear website/catalog language before publication.", "Prepare the 30-day incomplete-notice response workflow."])
pdf.h1("Course Development Tasks")
pdf.para("The first-wave catalog is designed around whole-hour RCFE CETP courses mapped to one RCFE Core of Knowledge subject each.")
pdf.table(["Course", "Hours", "Core category", "Delivery", "CDSS fee"], [(c["title"], str(c["hours"]), c["cat"], "Live-Stream + LMS", f"${c['cdss_fee']}") for c in COURSES], [165, 45, 170, 90, 55], size=6.2)
pdf.h1("LMS / Reviewer Access")
pdf.bullets(["Create reviewer-ready LMS area for each course.", "Validate credentials through a controlled channel.", "Prepare attendance, activity, and post-test reporting samples.", "Disable certificate issuance before approval."])
pdf.h1("Website and Catalog Controls")
pdf.bullets(["Use pending-approval language only if public mention is necessary.", "Remove language implying CDSS approval or administrator-renewal credit before issuance.", "Publish approved course numbers and vendor number only after CDSS issuance and Compliance clearance."])
pdf.h1("CDSS Forms and Appendices")
pdf.table(["Appendix", "Contents", "Status"], [("A", "LIC 9141, vendor fee notice, corporate checklist", "Drafted with pending fields"), ("B", "LIC 9140 drafts, course packets, course fee notice", "Drafted"), ("C", "Instructor docs and LIC 9140A post-approval reference", PENDING), ("D", "Roster, schedule notification, certificate sample, logs", "Prepared for post-approval")], [70, 310, 150], size=7)
pdf.h1("Instructor Qualification Packet")
pdf.para("Instructor documentation remains pending until the instructor slate, resumes, qualification pathway evidence, and disclosure answers are confirmed. LIC 9140A is included as a post-approval reference only unless ACB directs otherwise.")
pdf.table(["Requirement", "Status"], [("Instructor roster", PENDING), ("Current resume", PENDING), ("Qualification pathway evidence", PENDING), ("Disclosure questions", PENDING), ("Segment assignment", PENDING)], [250, 280], size=7)
pdf.h1("Course Approval Documentation")
for c in COURSES:
    pdf.h2(f"{c['id']} - {c['title']}")
    pdf.table(["Field", "Draft"], [("Hours", str(c["hours"])), ("Delivery", c["delivery"]), ("Core category", c["cat"]), ("Instructor", c["instructor"]), ("Reviewer access", c["reviewer_access"])], [105, 425], size=6.6)
    pdf.para(c["desc"], size=7.6)
    pdf.h2("Objectives")
    pdf.bullets(c["obj"], size=7.2)
    pdf.h2("Hour-by-Hour Outline")
    pdf.table(["Segment", "Content", "Instructor"], [(a, b, c["instructor"]) for a, b in c["outline"]], [70, 360, 100], size=6.5)
    pdf.h2("Assessment and Certificate Rule")
    pdf.para("Participants must attend the full approved duration, complete active participation checks, and pass the post-test with a minimum score of 80 percent. No certificate is issued until CDSS approval is received and vendor/course numbers are available.", size=7.4)
    pdf.h2("Post-Test Answer Key")
    pdf.table(["No.", "Question", "Answer"], [(str(i + 1), q, a) for i, (q, a) in enumerate(c["qa"])], [30, 250, 250], size=6.1)
pdf.h1("Policies and Procedures")
pdf.bullets(["Finalize recordkeeping SOP.", "Finalize attendance and active participation SOP.", "Finalize post-test, remediation, and certificate release SOP.", "Finalize approval-claims and catalog language SOP."])
pdf.h1("Recordkeeping and Operations")
pdf.table(["Record", "Minimum content", "Retention"], [("Course file", "Outline, materials, schedule, source list", "3 years"), ("Instructor file", "Resume, qualifications, disclosures", "3 years"), ("Roster/completion", "LIC 9142A/equivalent, participation, assessment", "3 years"), ("Evaluation", "Course/instructor evaluation", "3 years"), ("Certificate log", "Vendor/course number, hours, dates, times, location", "3 years")], [115, 300, 80], size=6.7)
pdf.h1("Risk Register")
pdf.table(["Risk", "Severity", "Control"], [(r[0], r[1], r[3]) for r in risk_rows], [150, 55, 325], size=6.5)
pdf.h1("QA and Internal Audit")
pdf.bullets(["Official form versions checked against CDSS Certification Forms page.", "All pending fields resolved or kept pending with no invented values.", "Course packets checked against LIC 9140 and ACB course-review guide.", "Instructor documentation complete.", "Reviewer access validated.", "Website/catalog language claim-safe.", "Dee final approval recorded before submission."])
pdf.h1("Submission Package Checklist")
pdf.bullets(["Compliance completes final review.", "Official forms are filled and checked.", "Course approval packets are complete.", "Instructor documentation is complete.", "Catalog and website language are claim-safe.", "Reviewer access is validated, if online or self-paced/live-stream.", "Fee worksheet is approved by Finance.", "All required signatures are collected.", "Dee approves the final submission package."])
pdf.h1("Post-Approval Operations")
pdf.bullets(["Enter approved vendor number and course numbers only after CDSS issuance.", "Use LIC 9142A or equivalent roster for attendance/completion records.", "Submit quarterly schedules when required.", "Maintain course, instructor, roster, evaluation, and certificate records for required retention period."])
pdf.h1("Final Approval Gate")
pdf.box("Execution may continue.", ["Do NOT submit any application to CDSS.", "Do NOT sign any CDSS forms.", "Do NOT publish approval claims or advertise approved status.", "Do NOT issue administrator CE certificates.", "Do NOT use any CDSS vendor number or course approval number before issued."], fill=HexColor("#FDECEC"), stroke=RED)
pdf.h2("Required Before Submission")
pdf.bullets(["Compliance completes final review.", "Official forms are filled and checked.", "Course approval packets are complete.", "Instructor documentation is complete.", "Catalog and website language are claim-safe.", "Reviewer access is validated, if online or self-paced/live-stream.", "Fee worksheet is approved by Finance.", "All required signatures are collected.", "Dee approves the final submission package."])
pdf.h1("Source Research")
pdf.table(["Source", "Use"], [(s[0], s[2]) for s in SOURCE_URLS], [220, 310], size=6.3)
pdf.close()

(NOTES_DIR / "GENERATION_SUMMARY.md").write_text(
    f"""# Generation Summary

Generated on {TODAY}.

- Main Markdown binder: RCFE_CETP_Project_Binder.md
- Main PDF binder: RCFE_CETP_Project_Binder.pdf
- Required section files: 00 through 18, with 17 provided as an XLSX document library.
- Official CDSS forms downloaded into appendices/official_forms and source_research.
- Filled official-form drafts were created only where fields were safe to populate; unknown values remain {PENDING}.
- No HTML output was generated.
""",
    encoding="utf-8",
)

print("Generated RCFE CETP binder artifacts")
print(f"Courses: {len(COURSES)}")
print(f"Total hours: {TOTAL_HOURS}")
print(f"Course fees: ${TOTAL_COURSE_FEES}")
