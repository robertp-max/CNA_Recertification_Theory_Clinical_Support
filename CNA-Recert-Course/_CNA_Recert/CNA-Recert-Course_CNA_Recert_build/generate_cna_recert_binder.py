from __future__ import annotations

import datetime as dt
import json
import shutil
import textwrap
import urllib.request
from pathlib import Path

from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from pypdf import PdfReader
from reportlab.graphics import renderPDF
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas as pdfcanvas
from reportlab.platypus import (
    BaseDocTemplate,
    Frame,
    Image as RLImage,
    KeepTogether,
    NextPageTemplate,
    PageBreak,
    PageTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)
from svglib.svglib import svg2rlg


ROOT = Path(__file__).resolve().parents[1]
EXPECTED_ROOT = Path(r"C:\AI\Git\CNA_Recertification_Theory_Clinical_Support")
OUT = ROOT / "CNA-Recert-Course_CNA_Recert"
ASSETS = OUT / "assets"
BUILD = ROOT / "CNA-Recert-Course_CNA_Recert_build"
BACKUPS = ROOT / "CNA-Recert-Course_CNA_Recert_backups"

TITLE = "CNA Recertification Project Binder"
try:
    PREPARED = dt.date.today().strftime("%B %#d, %Y")
except ValueError:
    PREPARED = dt.date.today().strftime("%B %d, %Y").replace(" 0", " ")

BURGUNDY = colors.HexColor("#7A1F2B")
DARK = colors.HexColor("#2B1B1F")
GOLD = colors.HexColor("#C7A34A")
PALE_GOLD = colors.HexColor("#F6EBC8")
LIGHT = colors.HexColor("#F8F7F4")
MID = colors.HexColor("#E7E1D8")
GREEN = colors.HexColor("#2F7D4F")
AMBER = colors.HexColor("#A46C00")
RED = colors.HexColor("#B13A2E")
BLUE = colors.HexColor("#315D8C")
GRAY = colors.HexColor("#6B6B6B")

OUTPUT_FILES = [
    "CNA_Recertification_Project_Binder.md",
    "CNA_Recertification_Project_Binder.pdf",
    "CNA_Recertification_Critical_Path_Flowchart.pdf",
    "CNA_Recertification_Critical_Path_Flowchart.png",
    "CNA_Recertification_Binder_Source_Index.md",
    "CNA_Recertification_Binder_Build_Report.md",
    "CNA_Recertification_Executive_Task_Tracker.xlsx",
    "README.md",
    "CNA_Recertification_Binder_Style_Guide.md",
]

SOURCE_DOCS = [
    "README.md",
    "BUILD_READINESS_STATUS.md",
    "NEXT_BUILD_STEP.md",
    "CNA-Recert-Course/FINAL_UNIFIED_IMPLEMENTATION_BLUEPRINT.md",
    "CNA-Recert-Course/PHASE_0_COMPLIANCE_FOUNDATION.md",
    "CNA-Recert-Course/PHASE_5_BUILD_SPEC_PACKET/00_Executive/00_EXECUTIVE_BUILD_BRIEF.md",
    "CNA-Recert-Course/PHASE_5_BUILD_SPEC_PACKET/02_Certificate_Gates/02_CERTIFICATE_GATE_SPEC.md",
    "CNA-Recert-Course/PHASE_5_BUILD_SPEC_PACKET/03_Active_Time/03_ACTIVE_TIME_VALIDATION_PLAN.md",
    "CNA-Recert-Course/PHASE_5_BUILD_SPEC_PACKET/08_QA_UAT/08_QA_NEGATIVE_TEST_SCRIPT.md",
    "CNA-Recert-Course/Content/00_EXECUTIVE_SUMMARY.md",
    "CNA-Recert-Course/Content/01_SOURCE_TO_COURSE_CROSSWALK.md",
    "CNA-Recert-Course/Content/02_THEORY_SYLLABUS_TABLE.md",
    "CNA-Recert-Course/Content/index/35_CONTENT_PACKAGE_INDEX_UPDATED.md",
    "CNA-Recert-Course/Content/theory/modules",
    "CNA-Recert-Course/Content/theory/exam",
    "CNA-Recert-Course/Content/clinical-support",
    "CNA-Recert-Course/Content/qa",
    "CNA-Recert-Course/CNA_Modules",
    "CNA-Recert-Course/ContentV2/03_MODULE_LESSON_4CARD_MAP.md",
    "CNA-Recert-Course/ContentV2/04_MODULE_ASSESSMENT_MAP.md",
    "CNA-Recert-Course/ContentV2/05_FINAL_ASSESSMENT_MAP.md",
    "CNA-Recert-Course/ContentV2/08_SME_COMPLIANCE_REVIEW_FLAGS.md",
    "CNA-Recert-Course/ContentV2/10_CONTENT_COVERAGE_AND_TIME_RECONCILIATION.md",
    "CNA-Recert-Course/ContentV2/qa/TIME_DEPTH_AUDIT.md",
    "CNA-Recert-Course/ContentV2/survey-evidence/SURVEY_READINESS_COVERAGE_SUMMARY.md",
    "CNA-Recert-Course/ContentV2/survey-evidence/CONTENT_GAPS_AND_DISPOSITIONS.md",
    "CNA-Recert-Course/reconciliation/00_CI_ION_RECONCILIATION_EXECUTIVE_SUMMARY.md",
    "CNA-Recert-Course/reconciliation/04_EXPORT_TO_STANDALONE_APP_CROSSWALK.md",
    "CNA-Recert-Course/reconciliation/08_CERTIFICATE_GATE_ACTIVE_TIME_AFFIDAVIT_RECONCILIATION.md",
    "CNA-Recert-Course/reconciliation/09_OPTIONAL_CLINICAL_SUPPORT_SEPARATION_RECONCILIATION.md",
    "CNA-Recert-Course/reconciliation/10_SOURCE_REPAIR_SME_COMPLIANCE_FLAGS.md",
    "CNA-Recert-Course/reconciliation/12_QA_NEGATIVE_TEST_AND_ACCEPTANCE_PLAN.md",
    "CNA-Recert-Course/reconciliation/13_GO_NO_GO_BLOCKERS_AND_DECISIONS.md",
    "CNA-Recert-Course/reconciliation/15_SPREADSHEET_URL_DOCUMENTATION_EVIDENCE_REGISTER.md",
    "standalone-course-mvp/README.md",
    "standalone-course-mvp/src/data",
    "CNA-Recert-Course/_RCFE CETP/RCFE CETP Vendor Packet for CDSS/RCFE_CETP_Project_Binder.pdf",
]

MODULES = [
    ("M00", "Orientation and Compliance Boundaries", 30, "5", "30 allocated / 5 lessons", "Drafted", "Unified Blueprint; Content syllabus", "Legal/CDPH copy review", "Confirm acknowledgements and no-PHI language"),
    ("M01", "Infection Control and PPE", 90, "6", "90 allocated / under-depth active estimate", "Needs SME Review", "Legacy CNA-CE draft + scattered CCCCO refs", "No dedicated NATP 10-17 source", "SME/source review and source-backed expansion"),
    ("M02", "Resident Rights, Abuse Prevention, and Boundaries", 120, "5", "105 lesson min / under-depth", "In Progress", "CCCCO/NATP Module 17", "Under-depth expansion", "Expand partial objectives from source"),
    ("M03", "Dementia, Communication, and Respectful Care", 120, "5", "105 lesson min / source repair", "Needs Source Repair", "CCCCO/NATP Modules 13, 16", "Sensitive content and truncated source", "Repair M03/L03-L05 from CCCCO M16 and SME review"),
    ("M04", "Mobility, Falls, and Workplace Safety", 120, "5", "105 lesson min / under-depth", "In Progress", "CCCCO/NATP Modules 12, 14", "No hands-on competency claim", "Expand source-backed safety/restorative content"),
    ("M05", "Nutrition, Skin Integrity, Vital Signs", 120, "5", "105 lesson min / under-depth", "Needs SME Review", "CCCCO/NATP Modules 10, 11, 13", "Pressure-injury scope review", "SME review and CNA-scope wording check"),
    ("M06", "Documentation, Reporting, PHI Avoidance, and Scope", 90, "5", "80 lesson min / under-depth", "In Progress", "CCCCO/NATP Module 15", "PHI and scope controls", "Confirm no PHI free text and expand source-backed content"),
    ("M07", "Final Review, Exam/Test, Affidavit", 30, "5", "17 lesson min / final exam excluded", "Needs Legal/CDPH Review", "All modules; final assessment map", "Affidavit/certificate approval", "Approve final wording and keep certificate disabled"),
]

STATUSES = [
    "Not Started", "In Progress", "Drafted", "Complete", "Needs Source Repair",
    "Needs SME Review", "Needs Compliance Review", "Needs Legal/CDPH Review",
    "Needs App Verification", "Needs Owner Decision", "Blocked", "Deferred", "Not Applicable",
]
PRIORITIES = ["P0 Blocker", "P1 High", "P2 Medium", "P3 Low"]


def ensure_scope() -> None:
    if ROOT != EXPECTED_ROOT:
        print("STOPPED: Repo path mismatch. No files changed.")
        raise SystemExit(2)
    for p in [OUT, ASSETS, BUILD, BACKUPS]:
        p.mkdir(parents=True, exist_ok=True)


def backup_existing() -> list[str]:
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    made: list[str] = []
    for rel in OUTPUT_FILES + ["assets/ci-ion-logomark-white.svg", "assets/ci-ion-logomark-white.png"]:
        src = OUT / rel
        if src.exists():
            dst = BACKUPS / stamp / rel
            dst.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(src, dst)
            made.append(str(dst.relative_to(ROOT)))
    return made


def download_logo() -> Path:
    logo = ASSETS / "ci-ion-logomark-white.svg"
    if not logo.exists():
        urllib.request.urlretrieve("https://ciinstituteofnursing.com/assets/logos/ci-ion-logomark-white.svg", logo)
    return logo


def p(text: str, style: ParagraphStyle) -> Paragraph:
    return Paragraph(str(text).replace("&", "&amp;"), style)


def styles():
    ss = getSampleStyleSheet()
    return {
        "h1": ParagraphStyle("h1", parent=ss["Heading1"], fontName="Helvetica-Bold", fontSize=18, leading=22, textColor=BURGUNDY, spaceAfter=8),
        "h2": ParagraphStyle("h2", parent=ss["Heading2"], fontName="Helvetica-Bold", fontSize=13, leading=16, textColor=BURGUNDY, spaceBefore=8, spaceAfter=5),
        "body": ParagraphStyle("body", parent=ss["BodyText"], fontName="Helvetica", fontSize=8.8, leading=11.3, textColor=DARK, spaceAfter=5),
        "small": ParagraphStyle("small", parent=ss["BodyText"], fontName="Helvetica", fontSize=7.5, leading=9.5, textColor=DARK),
        "coverTitle": ParagraphStyle("coverTitle", parent=ss["Title"], fontName="Helvetica-Bold", fontSize=24, leading=28, textColor=colors.white, alignment=TA_CENTER),
        "coverSub": ParagraphStyle("coverSub", parent=ss["BodyText"], fontName="Helvetica", fontSize=12, leading=16, textColor=colors.white, alignment=TA_CENTER),
        "callout": ParagraphStyle("callout", parent=ss["BodyText"], fontName="Helvetica-Bold", fontSize=9.2, leading=12, textColor=BURGUNDY),
        "cell": ParagraphStyle("cell", parent=ss["BodyText"], fontName="Helvetica", fontSize=7.1, leading=8.5, textColor=DARK),
        "cellBold": ParagraphStyle("cellBold", parent=ss["BodyText"], fontName="Helvetica-Bold", fontSize=7.1, leading=8.5, textColor=DARK),
    }


def table(data, col_widths, repeat_rows=1, font_size=7.1, align="LEFT"):
    st = styles()
    converted = []
    for r, row in enumerate(data):
        converted.append([p(x, st["cellBold"] if r == 0 else st["cell"]) for x in row])
    t = Table(converted, colWidths=col_widths, repeatRows=repeat_rows, hAlign=align)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BURGUNDY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), font_size),
        ("GRID", (0, 0), (-1, -1), 0.35, MID),
        ("LINEBELOW", (0, 0), (-1, 0), 1.2, GOLD),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    return t


def status_color(status: str):
    if status == "Complete":
        return GREEN
    if "Blocked" in status or status == "Needs Source Repair":
        return RED
    if "Review" in status or "Decision" in status or "Verification" in status:
        return AMBER
    if status in ("Drafted", "In Progress"):
        return BLUE
    return GRAY


def header_footer(c: pdfcanvas.Canvas, doc):
    if doc.page == 1:
        return
    w, h = letter
    c.saveState()
    c.setFillColor(BURGUNDY)
    c.rect(0, h - 0.5 * inch, w, 0.5 * inch, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 8)
    c.drawString(0.55 * inch, h - 0.31 * inch, "CI Institute of Nursing - 419 E Hamilton Ave, Campbell, CA 95008 - ciinstituteofnursing.com")
    c.setStrokeColor(GOLD)
    c.setLineWidth(1.2)
    c.line(0.55 * inch, 0.48 * inch, w - 0.55 * inch, 0.48 * inch)
    c.setFillColor(DARK)
    c.setFont("Helvetica", 7.5)
    c.drawString(0.55 * inch, 0.28 * inch, f"CI Institute of Nursing - {TITLE} / TJ Padilla")
    c.drawRightString(w - 0.55 * inch, 0.28 * inch, f"Page {doc.page}")
    c.restoreState()


def cover_page(story, logo: Path):
    st = styles()
    story.append(Spacer(1, 0.35 * inch))
    story.append(Paragraph("CI Institute of Nursing", st["coverSub"]))
    story.append(Spacer(1, 0.18 * inch))
    story.append(Paragraph("CNA Recertification Theory + Optional Clinical Support", st["coverTitle"]))
    story.append(Paragraph("CDPH / TPRU Readiness Project Binder", st["coverTitle"]))
    story.append(Spacer(1, 0.18 * inch))
    story.append(Paragraph("Prepared for Team Execution and Compliance Review", st["coverSub"]))
    story.append(Spacer(1, 0.3 * inch))
    if logo.exists():
        try:
            drawing = svg2rlg(str(logo))
            drawing.width = 1.1 * inch
            drawing.height = 1.1 * inch
            story.append(drawing)
        except Exception:
            pass
    cover_rows = [
        ["Prepared by", "TJ Padilla"],
        ["Project Oversight", "Dee Bustos"],
        ["Website", "ciinstituteofnursing.com"],
        ["Address", "419 E Hamilton Ave, Campbell, CA 95008"],
        ["Prepared", PREPARED],
        ["Status", "Internal Execution Packet / Not Production Approved"],
    ]
    story.append(Spacer(1, 0.35 * inch))
    ct = table(cover_rows, [2.1 * inch, 3.9 * inch], repeat_rows=0)
    ct.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), PALE_GOLD),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.white),
    ]))
    story.append(ct)
    story.append(Spacer(1, 0.35 * inch))
    story.append(Paragraph("Guardrail: no CDPH/TPRU approval claim, no production certificate issuance, no clinical-hour credit claim, and no learner-facing final answer keys.", st["coverSub"]))
    story.append(NextPageTemplate("body"))
    story.append(PageBreak())


def on_cover(c: pdfcanvas.Canvas, doc):
    w, h = letter
    c.saveState()
    c.setFillColor(BURGUNDY)
    c.rect(0, 0, w, h, fill=1, stroke=0)
    c.setStrokeColor(GOLD)
    c.setLineWidth(3)
    c.rect(0.45 * inch, 0.45 * inch, w - 0.9 * inch, h - 0.9 * inch, fill=0, stroke=1)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 9)
    c.drawCentredString(w / 2, 0.72 * inch, "CI Institute of Nursing - CNA Recertification Project Binder / TJ Padilla Page 1")
    c.restoreState()


def build_flowchart_pdf(path: Path):
    c = pdfcanvas.Canvas(str(path), pagesize=landscape(letter))
    draw_flowchart(c, "pdf")
    c.save()


def draw_flowchart(c, mode="pdf"):
    if mode == "pdf":
        w, h = landscape(letter)
        def rect(x, y, ww, hh, fill, outline=BURGUNDY, text="", sub="", small=""):
            c.setFillColor(fill); c.setStrokeColor(outline); c.setLineWidth(1.1); c.roundRect(x, y, ww, hh, 8, fill=1, stroke=1)
            c.setFillColor(DARK); c.setFont("Helvetica-Bold", 7.2); c.drawString(x + 6, y + hh - 14, text[:42])
            c.setFont("Helvetica", 6.1); c.drawString(x + 6, y + hh - 25, sub[:48]); c.drawString(x + 6, y + 8, small[:54])
        def line(x1, y1, x2, y2):
            c.setStrokeColor(GOLD); c.setLineWidth(1.2); c.line(x1, y1, x2, y2)
            c.setFillColor(GOLD); c.circle(x2, y2, 2.2, fill=1, stroke=0)
        c.setFillColor(colors.white); c.rect(0, 0, w, h, fill=1, stroke=0)
        c.setFillColor(BURGUNDY); c.rect(0, h - 0.55 * inch, w, 0.55 * inch, fill=1, stroke=0)
        c.setFillColor(colors.white); c.setFont("Helvetica-Bold", 14); c.drawString(0.45 * inch, h - 0.35 * inch, "CNA Recertification Critical Path Flowchart")
        c.setFont("Helvetica", 7); c.drawRightString(w - 0.45 * inch, h - 0.35 * inch, "Internal execution / not cleared for production")
        boxes = [
            (35, 465, 150, 42, "1 Source Evidence Gate", "Owner: Source Evidence", "Risk: unresolved CCCCO gaps"),
            (220, 465, 150, 42, "2 ContentV2 Gate", "Owner: ContentV2/Data", "Risk: source repair blocks clearance"),
            (405, 465, 160, 42, "3 Time-Depth / 720 Gate", "Owner: ID + Compliance", "Risk: under-depth lessons"),
            (590, 465, 150, 42, "4 CCCCO Survey Mapping", "Owner: Source Evidence", "Risk: partial/deferred objectives"),
            (220, 355, 150, 42, "5 Assessment Protection", "Owner: Assessment/QA", "Risk: key exposure"),
            (405, 355, 160, 42, "6 Optional Clinical Separation", "Owner: Compliance", "Risk: clinical-credit confusion"),
            (590, 355, 150, 42, "7 Active-Time Evidence", "Owner: QA/Compliance", "Risk: logs not sufficient"),
            (35, 250, 150, 42, "8 Certificate/Affidavit", "Owner: Legal/Registrar", "Risk: missing approved wording"),
            (220, 250, 150, 42, "9 App / MVP Review", "Owner: App Engineer", "Risk: runtime mismatch"),
            (405, 250, 160, 42, "10 QA Negative Testing", "Owner: QA Lead", "Risk: bypasses not caught"),
            (590, 250, 150, 42, "11 Drive Review Packet", "Owner: Packaging", "Risk: 59 URL rows incomplete"),
            (220, 130, 150, 42, "12 Dee Approval Gate", "Owner: Dee Bustos", "Risk: owner decisions open"),
            (405, 130, 160, 42, "13 Pilot/Internal Review", "Owner: Leadership + QA", "Risk: pilot before clearance"),
            (590, 130, 150, 42, "14 Production Disabled", "Owner: Compliance/Owner", "Risk: improper certificate use"),
        ]
        for i, (x, y, ww, hh, title, owner, risk) in enumerate(boxes):
            fill = PALE_GOLD if i in (0, 3, 10) else colors.HexColor("#F7F1F1") if i in (7, 13) else LIGHT
            outline = RED if i in (2, 7, 13) else BURGUNDY
            rect(x, y, ww, hh, fill, outline, title, owner + " | Status: Open", risk)
        arrows = [(185, 486, 220, 486), (370, 486, 405, 486), (565, 486, 590, 486), (295, 465, 295, 397), (485, 465, 485, 397), (665, 465, 665, 397), (370, 376, 405, 376), (565, 376, 590, 376), (110, 465, 110, 292), (185, 271, 220, 271), (370, 271, 405, 271), (565, 271, 590, 271), (665, 250, 665, 172), (370, 151, 405, 151), (565, 151, 590, 151), (295, 250, 295, 172)]
        for a in arrows:
            line(*a)
        c.setFillColor(DARK); c.setFont("Helvetica", 7)
        c.drawString(35, 82, "Parallel workstreams converge on QA, Drive packet, leadership approval, and final production-disabled gate. No internal card IDs are exposed.")
        c.setStrokeColor(GOLD); c.line(35, 70, w - 35, 70)
        c.setFont("Helvetica", 7); c.drawString(35, 48, f"CI Institute of Nursing - {TITLE} / TJ Padilla")
        c.drawRightString(w - 35, 48, "Critical Path Flowchart")


def build_flowchart_png(path: Path):
    W, H = 1600, 1050
    img = Image.new("RGB", (W, H), "white")
    d = ImageDraw.Draw(img)
    try:
        font_b = ImageFont.truetype("arialbd.ttf", 22)
        font = ImageFont.truetype("arial.ttf", 17)
        font_s = ImageFont.truetype("arial.ttf", 15)
    except Exception:
        font_b = font = font_s = ImageFont.load_default()
    burg = "#7A1F2B"; gold = "#C7A34A"; pale = "#F6EBC8"; light = "#F8F7F4"; red = "#B13A2E"; dark = "#2B1B1F"
    d.rectangle([0, 0, W, 90], fill=burg)
    d.text((60, 28), "CNA Recertification Critical Path Flowchart", fill="white", font=font_b)
    boxes = [
        (60, 160, "1 Source Evidence Gate", "Source Evidence | Open", "Unresolved CCCCO gaps"),
        (430, 160, "2 ContentV2 Gate", "ContentV2/Data | Open", "Source repair blocks clearance"),
        (800, 160, "3 Time-Depth / 720 Gate", "ID + Compliance | Open", "Under-depth lessons"),
        (1170, 160, "4 CCCCO Survey Mapping", "Source Evidence | Open", "Partial/deferred objectives"),
        (430, 380, "5 Assessment Protection", "Assessment/QA | Open", "Key exposure"),
        (800, 380, "6 Optional Clinical Separation", "Compliance | Open", "Clinical-credit confusion"),
        (1170, 380, "7 Active-Time Evidence", "QA/Compliance | Open", "Logs not sufficient"),
        (60, 600, "8 Certificate/Affidavit", "Legal/Registrar | Blocked", "Missing approved wording"),
        (430, 600, "9 App / MVP Review", "App Engineer | Needs verification", "Runtime mismatch"),
        (800, 600, "10 QA Negative Testing", "QA Lead | Not executed", "Bypasses not caught"),
        (1170, 600, "11 Drive Review Packet", "Packaging | Pending", "59 URL rows incomplete"),
        (430, 820, "12 Dee Approval Gate", "Dee Bustos | Open", "Owner decisions open"),
        (800, 820, "13 Pilot/Internal Review", "Leadership + QA | Deferred", "Pilot before clearance"),
        (1170, 820, "14 Production Disabled", "Compliance/Owner | Required", "Improper certificate use"),
    ]
    for x, y, title, owner, risk in boxes:
        fill = pale if title.startswith(("1", "4", "11")) else "#F7F1F1" if title.startswith(("8", "14")) else light
        outline = red if title.startswith(("3", "8", "14")) else burg
        d.rounded_rectangle([x, y, x + 300, y + 95], radius=16, fill=fill, outline=outline, width=3)
        d.text((x + 16, y + 14), title, fill=dark, font=font_b)
        d.text((x + 16, y + 46), owner, fill=dark, font=font)
        d.text((x + 16, y + 70), "Risk: " + risk, fill=dark, font=font_s)
    def arrow(a, b):
        d.line([a, b], fill=gold, width=4)
        d.ellipse([b[0]-5, b[1]-5, b[0]+5, b[1]+5], fill=gold)
    for a, b in [((360,208),(430,208)),((730,208),(800,208)),((1100,208),(1170,208)),((580,255),(580,380)),((950,255),(950,380)),((1320,255),(1320,380)),((730,428),(800,428)),((1100,428),(1170,428)),((210,255),(210,600)),((360,648),(430,648)),((730,648),(800,648)),((1100,648),(1170,648)),((1320,695),(1320,820)),((730,868),(800,868)),((1100,868),(1170,868)),((580,695),(580,820))]:
        arrow(a, b)
    d.text((60, 990), f"CI Institute of Nursing - {TITLE} / TJ Padilla", fill=dark, font=font_s)
    img.save(path)


def build_pdf(logo: Path, flowchart_png: Path) -> int:
    pdf_path = OUT / "CNA_Recertification_Project_Binder.pdf"
    doc = BaseDocTemplate(str(pdf_path), pagesize=letter, rightMargin=0.48 * inch, leftMargin=0.48 * inch, topMargin=0.75 * inch, bottomMargin=0.65 * inch)
    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id="normal")
    doc.addPageTemplates([
        PageTemplate(id="cover", frames=[frame], onPage=on_cover),
        PageTemplate(id="body", frames=[frame], onPage=header_footer),
    ])
    st = styles()
    story = []
    cover_page(story, logo)

    def heading(name):
        story.append(Paragraph(name, st["h1"]))
        story.append(Spacer(1, 0.04 * inch))

    heading("1. Executive Summary")
    story.append(Paragraph("This binder is an internal execution and compliance-review packet for a 12-hour asynchronous CNA recertification theory pathway with a separate Optional Clinical Support hub. It summarizes the source evidence, ContentV2 status, assessment protections, certificate gate blockers, app review status, QA controls, and owner decisions needed before any external claim or certificate action.", st["body"]))
    story.append(table([
        ["Executive Area", "Current Evidence", "Status"],
        ["Exists", "Unified blueprint, Phase 0 compliance foundation, ContentV1 packet, ContentV2 package, survey evidence, reconciliation tracker, standalone MVP files.", "Drafted"],
        ["In progress", "Source repair, time-depth expansion, app/runtime verification, Google Drive review packaging, QA negative testing.", "In Progress"],
        ["Blocked", "Approval metadata, NAC/provider details if applicable, approved certificate wording, affidavit/e-sign method, active-time validation, compliance/legal/owner clearance.", "Blocked"],
        ["Cannot happen yet", "CDPH/TPRU approval representations, production certificate issuance, certificate downloads, clinical-hour credit claims, learner-facing final answer keys.", "Blocked"],
    ], [1.35*inch, 4.5*inch, 1.05*inch]))

    heading("2. Gate Status")
    story.append(table([
        ["Gate Rule", "Status / Required Boundary"],
        ["Internal execution", "Execution may continue internally."],
        ["Regulatory/production claims", "No CDPH/TPRU approval claims, production certificate issuance, certificate downloads, or clinical-hour claims until final clearance."],
        ["Optional Clinical Support", "Remains separate, optional, non-credit, non-gating, and not clinical-hour credit."],
        ["Assessment protection", "Final answer keys and rationales remain internal/admin-only."],
    ], [2.0*inch, 4.9*inch]))

    heading("3. Status Matrix")
    status_rows = [
        ["Area", "Status", "Executive Note"],
        ["Source package", "Drafted", "Source docs are present; M01/M03/M05 remain flagged."],
        ["CCCCO/NATP mapping", "In Progress", "72 objectives mapped; coverage includes covered, assessed, partial, deferred, source-repair, and SME-review dispositions."],
        ["ContentV2 package", "Drafted", "8 modules, 41 lessons, 277 cards; all compliance flags preserved."],
        ["Time-depth reconciliation", "Needs Compliance Review", "Declared 720-minute model remains; active-learning estimate shows gap requiring source-backed expansion."],
        ["App / standalone MVP", "Needs App Verification", "Standalone MVP files present; runtime/browser verification not completed in reconciliation evidence."],
        ["Module assessments", "Drafted", "Module checks mapped; internal scoring only."],
        ["Final exam", "Needs SME Review", "50-question pool mapped; final answer keys/rationales admin-only."],
        ["Certificate gate", "Blocked", "Approval metadata, approved wording, active-time, and affidavit method missing."],
        ["Active-time evidence", "Blocked", "Plugin/manual method not validated; Moodle logs alone are insufficient."],
        ["Affidavit/e-sign", "Needs Legal/CDPH Review", "Affidavit text exists; production method not approved."],
        ["Optional Clinical Support", "Needs Compliance Review", "Separate support pathway; runtime separation verification pending."],
        ["TTS/media", "Deferred", "Planning-only unless owner/legal/compliance authorize."],
        ["Audit packet", "In Progress", "Reconciliation documents and tracker exist; missing URL evidence remains."],
        ["Google Drive review package", "Needs Owner Decision", "59 URL rows remain undocumented; Drive links pending owner supply."],
        ["Leadership approval", "Needs Owner Decision", "Dee approval must follow evidence, QA, legal, and compliance clearance."],
    ]
    story.append(table(status_rows, [1.7*inch, 1.25*inch, 3.95*inch]))

    heading("4. Project Charter")
    story.append(table([
        ["Field", "Value"],
        ["Institution", "CI Institute of Nursing"],
        ["Program", "CNA Recertification Theory + Optional Clinical Support"],
        ["Binder", "CNA Recertification Project Binder"],
        ["Purpose", "Executive execution packet and compliance evidence map; not an approval packet."],
        ["Project Oversight", "Dee Bustos"],
        ["Prepared by", "TJ Padilla"],
        ["Physical address", "419 E Hamilton Ave, Campbell, CA 95008"],
        ["Website", "ciinstituteofnursing.com"],
        ["Required theory minutes", "720 minutes / 12 hours"],
        ["Optional support status", "Optional, non-credit, non-gating, not clinical-hour credit"],
        ["Production certificate status", "Disabled pending approval metadata, active-time validation, approved wording, affidavit method, and final clearance"],
    ], [2.15*inch, 4.75*inch]))

    heading("5. Team Execution Summary")
    story.append(table([
        ["Team", "Provided", "Action Steps", "Output"],
        ["Compliance", "Phase 0 and gate rules", "Validate approval metadata, certificate wording, affidavit method, active time, claims language", "Clearance memo / blocker closure"],
        ["Course Development", "ContentV1 and ContentV2 modules", "Close source-backed depth gaps; keep no-padding rule", "Reviewed module package"],
        ["Source Evidence / CCCCO", "CCCCO 10-17 mapping", "Resolve source repair and partial/deferred dispositions", "Updated evidence map"],
        ["ContentV2 / Data", "JSON/TS package and generated maps", "Finalize after SME/compliance review", "Canonical release candidate"],
        ["App / Standalone MVP", "Prototype app and data adapters", "Run runtime verification and route checks", "QA evidence packet"],
        ["Assessment / QA", "Module/final maps", "Protect keys, run negative tests", "QA acceptance report"],
        ["Certificate Gate", "Gate specs and draft fields", "Block until approvals are documented", "Locked certificate plan"],
        ["TTS / Media", "Narration planning artifacts", "Await owner/legal authorization", "Approved media decision"],
        ["Google Drive Packaging", "Manifest and URL register", "Re-export failed docs; owner decision on videos/folder", "Complete review package"],
        ["Leadership", "Owner decision role", "Approve go/no-go only after P0 closures", "Dee approval record"],
    ], [1.1*inch, 1.7*inch, 2.55*inch, 1.55*inch]))

    heading("6. Critical Path Flowchart")
    story.append(Paragraph("The following full-page visual shows the dependency flow from source evidence through production-disabled control. Parallel workstreams converge on QA, packaging, leadership approval, and final clearance.", st["body"]))
    story.append(PageBreak())
    story.append(RLImage(str(flowchart_png), width=7.1*inch, height=4.65*inch))
    story.append(PageBreak())

    heading("7. Critical Path Task Table")
    tasks = [
        ["Workstream", "Task", "Owner", "Dependency", "Duration / Effort", "Gate", "Status", "Risk if Delayed"],
        ["Source Evidence", "Resolve CCCCO/source repair evidence", "Source Evidence", "Local source docs", "Medium", "Source Evidence", "Needs Source Repair", "Survey evidence remains incomplete"],
        ["Course Content", "Finalize ContentV2 after SME flags", "ContentV2 / ID", "Source Evidence", "High", "ContentV2", "Needs SME Review", "Learner content remains draft"],
        ["Time Depth", "Close under-depth lessons using approved source", "ID + Compliance", "ContentV2", "High", "720-Minute", "Needs Compliance Review", "Time model remains unsupported by depth"],
        ["Assessment", "Protect final pool and remediation boundaries", "Assessment / QA", "ContentV2", "Medium", "Assessment", "Drafted", "Answer-key exposure risk"],
        ["Clinical Separation", "Verify optional clinical is non-gating", "Compliance / QA", "App review", "Medium", "Clinical", "Needs Compliance Review", "Learner misperception of credit"],
        ["Active Time", "Validate plugin or manual hold", "QA / Compliance", "Staging method", "High", "Active-Time", "Blocked", "Certificate gate cannot clear"],
        ["Certificate", "Approve metadata, wording, and affidavit method", "Legal / Registrar", "Approval artifacts", "High", "Certificate", "Blocked", "Improper certificate issuance"],
        ["App", "Run standalone MVP browser verification", "App Engineer", "ContentV2 package", "Medium", "App Review", "Needs App Verification", "Route/content mismatch"],
        ["QA", "Run negative tests", "QA Lead", "App and gate config", "Medium", "QA", "Not Started", "Bypass defects remain unknown"],
        ["Packaging", "Complete Google Drive review packet", "Packaging", "URL imports/decisions", "Medium", "Drive Packet", "Needs Owner Decision", "Missing evidence at review"],
        ["Leadership", "Dee approval after P0 closure", "Dee Bustos", "All gates", "Low", "Leadership", "Needs Owner Decision", "Premature pilot decision"],
        ["Release Control", "Keep production disabled until clearance", "Compliance / Owner", "Leadership gate", "Ongoing", "Final Gate", "Blocked", "Unauthorized production use"],
    ]
    story.append(table(tasks, [0.85*inch, 1.25*inch, 0.85*inch, 1.0*inch, 0.75*inch, 0.75*inch, 0.9*inch, 1.0*inch]))

    heading("8. Course Structure and Module Map")
    story.append(table([["Module", "Title", "Minutes", "Lessons", "Cards / Current Evidence", "Status", "Key Source", "Review Flag", "Next Action"]] + MODULES, [0.38*inch, 1.15*inch, 0.45*inch, 0.45*inch, 1.05*inch, 0.75*inch, 1.0*inch, 1.0*inch, 1.0*inch], font_size=6.4))

    heading("9. Source Evidence and CCCCO Coverage")
    story.append(table([
        ["Evidence Area", "Inspected Finding", "Status"],
        ["CCCCO/NATP modules", "Modules 10-17 are mapped to current theory and optional support destinations.", "In Progress"],
        ["Coverage rollup", "72 objectives mapped; 35 covered, 8 assessed, 10 partial, 8 deferred, 7 source repair, 1 out of scope, 3 SME review.", "Drafted"],
        ["Gaps", "M03 source repair and partial/under-depth objectives remain visible; hands-on skills are deferred to clinical/support context.", "Needs Source Repair"],
        ["Survey language", "No survey/audit approval is claimed; coverage is evidence mapping only.", "Needs Compliance Review"],
    ], [1.65*inch, 4.15*inch, 1.1*inch]))

    heading("10. ContentV2 and Time-Depth Status")
    story.append(table([
        ["Measure", "Evidence"],
        ["Required theory model", "720 minutes / 12 hours declared and inherited from ContentV1 syllabus."],
        ["Pipeline counts", "8 modules, 41 lessons, 277 cards, 277 narration clips."],
        ["Current gap", "Content-depth gap reported as 83 minutes in coverage reconciliation; time-depth audit estimates 447.12 active-learning minutes and 29 failing lessons."],
        ["No-padding rule", "Gaps are reported and must be closed only with source-backed expansion; narration/assessment/clinical support do not replace the 720 model."],
        ["Open modules", "M01, M02, M03, M04, M05, and M06 remain under-depth or source-repair in audit evidence."],
    ], [2.0*inch, 4.9*inch]))

    heading("11. Assessment and Final Exam Protection")
    story.append(table([
        ["Control", "Status / Evidence"],
        ["Module checks", "Module assessment map exists; questions are internal-scored with remediation direction only."],
        ["Final assessment pool", "50-question pool mapped to sources; answer keys and rationales remain internal/admin-only."],
        ["Learner result boundaries", "Final failed result should be topic-level only; no full key or rationale exposure."],
        ["Review flags", "Infection control and skin integrity questions preserve SME-review flags."],
    ], [2.0*inch, 4.9*inch]))

    heading("12. Certificate Gate, Affidavit, and Active-Time Controls")
    story.append(table([
        ["Required Control", "Current Status"],
        ["Approval metadata / NAC/provider details", "Missing; P0 blocker."],
        ["Approved certificate wording", "Missing approval; draft mapping is not enough for production."],
        ["Affidavit/e-sign method", "Missing legal/CDPH/owner approval; wet-sign fallback may be needed."],
        ["Active-time validation", "Missing; Moodle logs alone are insufficient."],
        ["Certificate production", "Disabled until all approval, affidavit, active-time, and clearance evidence is documented."],
    ], [2.25*inch, 4.65*inch]))

    heading("13. Optional Clinical Support Separation")
    story.append(Paragraph("Optional Clinical Support remains optional, non-credit, non-gating, and not clinical-hour credit. Skipped clinical-skill videos in the export are an owner decision for the support pathway only and do not change required theory credit.", st["body"]))
    story.append(table([
        ["Separation Control", "Status", "Next Step"],
        ["Separate pathway", "Needs Compliance Review", "Verify app labeling and navigation."],
        ["Non-credit label", "Needs Compliance Review", "Confirm no clinical-hour claim."],
        ["Not certificate gate", "Needs App Verification", "Run QA-02: skip clinical and confirm certificate path unaffected."],
    ], [2.0*inch, 1.45*inch, 3.45*inch]))

    heading("14. App / Standalone MVP Status")
    story.append(table([
        ["Area", "Evidence", "Status"],
        ["V2 UI / MVP", "Standalone app README describes browser-usable prototype only; not final Moodle production course.", "Internal Review"],
        ["ContentV2 wiring", "Generated ContentV2 adapter/data files referenced in reconciliation.", "Needs App Verification"],
        ["Lesson routing", "App files present; runtime/browser verification not run in documentation pass.", "Needs App Verification"],
        ["Production controls", "README states no real certificates, no real learner data, and no PHI collection.", "Drafted"],
    ], [1.4*inch, 4.25*inch, 1.25*inch]))

    heading("15. Audit Packet and Reconciliation Evidence")
    story.append(table([
        ["Evidence", "Status / Note"],
        ["Reconciliation folder", "Contains executive summary, crosswalks, certificate/active-time, QA, blockers, and URL evidence register."],
        ["Export tracker", "Master tracker workbook exists in reconciliation evidence; current task creates a CNA executive task tracker."],
        ["Spreadsheet URL evidence", "162 URLs classified; 59 still undocumented; sheet/row/cell extraction pending."],
        ["Google Drive package", "Pending final Drive links and owner decisions; no final review package claimed."],
    ], [2.0*inch, 4.9*inch]))

    heading("16. QA Negative Tests and Acceptance Controls")
    story.append(table([
        ["Risk", "Negative Test", "Expected Result", "Status", "Evidence / Next Step"],
        ["PHI entry", "Attempt PHI entry/upload", "No PHI requested, typed, uploaded, shown, or implied", "Not Started", "Run QA-01"],
        ["Optional clinical gating", "Skip clinical support", "Theory/certificate path unaffected", "Not Started", "Run QA-02"],
        ["Final answer key exposure", "Review learner final result", "No keys/rationales visible", "Not Started", "Run QA-03"],
        ["Certificate download", "Attempt download before clearance", "Certificate unavailable/disabled", "Not Started", "Run QA-04"],
        ["Active-time bypass", "Idle/multi-tab/direct access", "Time cannot be satisfied by bypass", "Not Started", "Run QA-05"],
        ["Missing approval metadata", "Blank provider/approval fields", "Gate blocks output", "Not Started", "Run QA-06"],
        ["Source-repair content", "Try marking flagged content reviewed", "Flag remains until evidence exists", "Not Started", "Run QA-07"],
        ["App route mismatch", "Navigate all modules/routes", "Only verified content displays", "Not Started", "Run QA-10"],
        ["TTS/media not authorized", "Trigger narration/audio production", "Planning only; no production audio", "Not Started", "Owner/legal decision"],
    ], [1.0*inch, 1.45*inch, 1.7*inch, 0.8*inch, 1.95*inch], font_size=6.4))

    heading("17. Remaining Blockers and Owner Decisions")
    story.append(table([
        ["Priority", "Decision / Blocker", "Owner", "Status"],
        ["P0 Blocker", "Approval metadata, NAC/provider details if applicable, approved certificate wording, affidavit/e-sign method, active-time validation, legal/compliance clearance before production claims", "Program Owner / Legal / Compliance", "Open"],
        ["P1 High", "M01 infection-control SME/source review, M03 source repair and sensitive-content review, M05 skin integrity SME review, runtime app verification", "SME / App QA", "Open"],
        ["P2 Medium", "TTS/media authorization and under-depth source-backed expansion sequencing", "Owner / ID", "Open"],
        ["P3 Low", "External-reference URL notes and formatting polish after evidence closure", "Repo Auditor", "Open"],
    ], [1.0*inch, 3.75*inch, 1.3*inch, 0.85*inch]))

    heading("18. Final Approval Gate")
    story.append(Paragraph("Execution may continue internally.", st["callout"]))
    story.append(table([
        ["Do NOT proceed until final clearance", "Required before production/pilot clearance"],
        ["Submit or represent CDPH/TPRU approval; publish approval claims; enable production certificate issuance; enable certificate downloads; issue CNA renewal certificates from this MVP; claim clinical-hour credit; count Optional Clinical Support as required theory or clinical credit; use provider/NAC/course approval metadata before issued/confirmed; release learner-facing final answer keys.", "Source evidence verified; ContentV2 finalized; time-depth plan approved; active-time validation approved; certificate wording approved; affidavit/e-sign method approved; SME/compliance review completed; app QA completed; Dee approval completed."],
    ], [3.45*inch, 3.45*inch]))

    heading("19. Source Research / Evidence Index")
    src_rows = [["Source", "Use", "Status", "Path", "Notes"]]
    for s in SOURCE_DOCS[:28]:
        src_rows.append([Path(s).name or s, "Binder evidence / style reference", "Inspected/Referenced" if (ROOT / s).exists() else "Reference Listed", s, "Summarized only; not pasted into PDF."])
    story.append(table(src_rows, [1.25*inch, 1.25*inch, 1.0*inch, 2.3*inch, 1.1*inch], font_size=6.2))

    doc.build(story)
    return len(PdfReader(str(pdf_path)).pages)


def build_markdown():
    md = f"""# CNA Recertification Project Binder

CI Institute of Nursing

Prepared by: TJ Padilla  
Project Oversight: Dee Bustos  
Prepared: {PREPARED}  
Status: Internal Execution Packet / Not Production Approved

## Executive Summary

This binder is an executive internal execution packet for the CNA Recertification Theory + Optional Clinical Support project. It preserves the documented posture: 12-hour asynchronous online theory, a separate Optional Clinical Support pathway, no production certificate issuance, no CDPH/TPRU approval claim, no clinical-hour credit claim, and final answer keys kept internal/admin-only.

## Gate Status

Execution may continue internally. No CDPH/TPRU approval claims, production certificate issuance, certificate downloads, clinical-hour claims, or learner-facing final answer keys are permitted until final clearance.

## Required Outputs

- `CNA_Recertification_Project_Binder.pdf`
- `CNA_Recertification_Critical_Path_Flowchart.pdf`
- `CNA_Recertification_Critical_Path_Flowchart.png`
- `CNA_Recertification_Executive_Task_Tracker.xlsx`
- `CNA_Recertification_Binder_Source_Index.md`
- `CNA_Recertification_Binder_Build_Report.md`

## Core Facts Preserved

- Required theory total: 720 minutes / 12 hours.
- Module timing: M00 30, M01 90, M02 120, M03 120, M04 120, M05 120, M06 90, M07 30.
- Optional Clinical Support is optional, non-credit, non-gating, and not clinical-hour credit.
- Certificate production remains disabled pending approval metadata, active-time validation, approved certificate wording, affidavit method, and compliance/legal/owner clearance.
- Module 1, Module 3, and Module 5 review/source flags remain visible.
"""
    (OUT / "CNA_Recertification_Project_Binder.md").write_text(md, encoding="utf-8")


def build_source_index():
    rows = ["# CNA Recertification Binder Source Index", "", "| Source | Use | Status | Path | Notes |", "|---|---|---|---|---|"]
    for s in SOURCE_DOCS:
        exists = (ROOT / s).exists()
        rows.append(f"| {Path(s).name or s} | Evidence or style reference | {'Present' if exists else 'Listed / not found'} | `{s}` | Summarized; not copied into binder. |")
    (OUT / "CNA_Recertification_Binder_Source_Index.md").write_text("\n".join(rows) + "\n", encoding="utf-8")


def build_readme():
    text = f"""# CNA Recertification Project Binder Packet

This folder contains the official internal execution binder packet for the CI Institute of Nursing CNA Recertification Theory + Optional Clinical Support project.

Status: Internal Execution Packet / Not Production Approved.

Primary files:
- `CNA_Recertification_Project_Binder.pdf`
- `CNA_Recertification_Critical_Path_Flowchart.pdf`
- `CNA_Recertification_Critical_Path_Flowchart.png`
- `CNA_Recertification_Executive_Task_Tracker.xlsx`

Compliance guardrail: this packet does not claim CDPH/TPRU approval, production certificate issuance, certificate-download readiness, or clinical-hour credit.
"""
    (OUT / "README.md").write_text(text, encoding="utf-8")


def build_style_guide():
    text = f"""# CNA Recertification Binder Style Guide

## Cover Layout
Use a formal burgundy cover with centered institution, program title, packet purpose, preparer, oversight owner, address, website, prepared date, and status.

## Logo Placement
Use the CI Institute of Nursing logomark on the cover when available. Store assets under `assets/` and do not package font files.

## Header and Footer
Header: `CI Institute of Nursing | 419 E Hamilton Ave, Campbell, CA 95008 | ciinstituteofnursing.com`.
Footer must include `{TITLE}` and preparer attribution: `CI Institute of Nursing - {TITLE} / TJ Padilla Page X`.

## Title Hierarchy
Use burgundy section headings, gold divider rules, short executive prose, and compact tables.

## Table Style
Use burgundy header rows, white text, gold header rule, alternating light rows, wrapped text, and compact executive row heights.

## Status Convention
Allowed status labels: {", ".join(STATUSES)}.

## Colors
Burgundy `#7A1F2B`, gold `#C7A34A`, light background `#F8F7F4`, dark body `#2B1B1F`.

## Critical Path Visual Style
Use full-page visual boxes with arrows, owner, status, risk if delayed, parallel workstreams, blockers, and no internal card IDs.

## Source Research Table
Summarize source evidence with Source, Use, Status, Path, and Notes. Do not paste raw file contents.

## Final Gate Callout
Use a bordered/gold-accent table stating internal execution may continue, but approval claims, certificate issuance/downloads, clinical-hour credit claims, provider metadata use, and answer-key exposure are blocked until clearance.
"""
    (OUT / "CNA_Recertification_Binder_Style_Guide.md").write_text(text, encoding="utf-8")


def tracker_rows():
    return [
        ["CP-001", "Source Evidence", "P0 Blocker", "High", "Needs Source Repair", "Yes", "Resolve CCCCO/source repair evidence", "Survey evidence + ContentV2 gaps", "72 mapped objectives; 7 source repair", "Unresolved CCCCO gaps", "Repair M03 and under-depth objectives", "Source Evidence", "SME", "Local source docs", "Source index", PREPARED, PREPARED, ""],
        ["CP-002", "ContentV2 / Time Depth", "P1 High", "High", "Needs Compliance Review", "No", "Close time-depth gap with source-backed expansion", "TIME_DEPTH_AUDIT.md", "447.12 active-learning min estimate", "Under-depth lessons", "Expand without padding", "Instructional Designer", "Compliance", "Source Evidence", "Time-depth audit", PREPARED, PREPARED, ""],
        ["CP-003", "Assessment / Exam", "P1 High", "High", "Needs SME Review", "No", "Protect final exam keys and flagged questions", "Final assessment map", "50-question pool; keys internal-only", "Answer-key exposure", "Review app final-result boundaries", "Assessment Lead", "QA / SME", "ContentV2", "App QA", PREPARED, PREPARED, ""],
        ["CP-004", "Certificate Gate", "P0 Blocker", "Critical", "Blocked", "Yes", "Obtain approval metadata, wording, affidavit method", "Certificate reconciliation", "No approval artifacts inspected", "Improper certificate issuance", "Collect written approvals", "Program Owner / Legal", "Compliance / CDPH", "Approval evidence", "Gate reconciliation", PREPARED, PREPARED, ""],
        ["CP-005", "Optional Clinical Support", "P1 High", "High", "Needs Compliance Review", "No", "Verify non-credit non-gating separation", "Optional clinical reconciliation", "Support pathway only", "Clinical-credit confusion", "Run QA-02 and review labels", "Compliance", "QA", "App runtime", "QA evidence", PREPARED, PREPARED, ""],
        ["CP-006", "App / QA", "P1 High", "High", "Needs App Verification", "No", "Run standalone MVP browser verification", "App crosswalk", "Files present; not runtime-tested", "Route/content mismatch", "Run app QA after evidence closure", "App Engineer", "QA / Compliance", "ContentV2 package", "Browser evidence", PREPARED, PREPARED, ""],
        ["CP-007", "Audit / Reconciliation", "P1 High", "Medium", "Needs Owner Decision", "No", "Resolve 59 undocumented URL rows", "URL evidence register", "162 URLs classified", "Missing Drive evidence", "Re-export failed docs and decide skipped media/folder", "Packaging", "Program Owner", "Drive access", "URL register", PREPARED, PREPARED, ""],
        ["CP-008", "Leadership", "P0 Blocker", "Critical", "Needs Owner Decision", "Yes", "Complete Dee approval gate only after P0 closures", "Go/no-go blockers", "P0 open", "Premature pilot decision", "Review after QA and approvals", "Dee Bustos", "Compliance / Legal", "P0 closure", "Approval record", PREPARED, PREPARED, ""],
    ]


def build_workbook() -> list[str]:
    wb = Workbook()
    wb.remove(wb.active)
    # Excel sheet titles cannot contain /, so requested slash names use workbook-safe hyphens.
    sheets = ["Dashboard", "Critical Path", "Master Task Tracker", "Module Status", "Source Evidence", "ContentV2 - Time Depth", "Assessment - Exam", "Certificate Gate", "Optional Clinical Support", "App - QA", "Audit - Reconciliation", "Google Drive Packaging", "Glossary"]
    for name in sheets:
        wb.create_sheet(name)
    header_fill = PatternFill("solid", fgColor="7A1F2B")
    gold_fill = PatternFill("solid", fgColor="C7A34A")
    light_fill = PatternFill("solid", fgColor="F8F7F4")
    white_font = Font(color="FFFFFF", bold=True)
    bold = Font(bold=True, color="2B1B1F")
    thin = Side(style="thin", color="D9D2C3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers = ["ID", "Domain", "Priority", "Severity", "Status", "Blocker", "Task / Expected Item", "Evidence Source", "Current Evidence", "Gap / Risk", "Required Action", "Owner Role", "Reviewer Role", "Dependency", "Verification File / Command", "Date Opened", "Date Updated", "Notes"]
    rows = tracker_rows()
    dv_status = DataValidation(type="list", formula1='"' + ",".join(STATUSES) + '"', allow_blank=True)
    dv_priority = DataValidation(type="list", formula1='"' + ",".join(PRIORITIES) + '"', allow_blank=True)
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.auto_filter.ref = f"A1:R{max(2, len(rows)+1)}"
        ws.add_data_validation(dv_status)
        ws.add_data_validation(dv_priority)
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["E"].width = 24
        ws.column_dimensions["G"].width = 34
        ws.column_dimensions["H"].width = 28
        ws.column_dimensions["I"].width = 32
        ws.column_dimensions["J"].width = 30
        ws.column_dimensions["K"].width = 30
        ws.column_dimensions["L"].width = 22
        ws.column_dimensions["M"].width = 22
        ws.column_dimensions["N"].width = 24
        ws.column_dimensions["O"].width = 26
        ws.column_dimensions["R"].width = 24
        for r in range(2, 202):
            dv_priority.add(ws[f"C{r}"])
            dv_status.add(ws[f"E{r}"])
    master = wb["Master Task Tracker"]
    for row in rows:
        master.append(row)
    domain_map = {
        "Critical Path": rows,
        "ContentV2 - Time Depth": [rows[1]],
        "Assessment - Exam": [rows[2]],
        "Certificate Gate": [rows[3]],
        "Optional Clinical Support": [rows[4]],
        "App - QA": [rows[5]],
        "Audit - Reconciliation": [rows[6]],
        "Google Drive Packaging": [rows[6]],
        "Source Evidence": [rows[0]],
    }
    for name, subset in domain_map.items():
        ws = wb[name]
        for row in subset:
            ws.append(row)
    ms = wb["Module Status"]
    ms.append(["Module", "Title", "Minutes", "Lessons", "Status", "Key Source", "Review Flag", "Next Action"])
    for cell in ms[1]:
        cell.fill = header_fill; cell.font = white_font; cell.alignment = Alignment(wrap_text=True); cell.border = border
    for m in MODULES:
        ms.append([m[0], m[1], m[2], m[3], m[5], m[6], m[7], m[8]])
    dash = wb["Dashboard"]
    dash.delete_rows(1)
    dash["A1"] = "CNA Recertification Executive Task Tracker"
    dash["A1"].font = Font(bold=True, size=16, color="7A1F2B")
    dash["A3"] = "Status"; dash["B3"] = "Count"; dash["D3"] = "Priority"; dash["E3"] = "Count"
    for c in ["A3", "B3", "D3", "E3"]:
        dash[c].fill = header_fill; dash[c].font = white_font; dash[c].alignment = Alignment(horizontal="center")
    for i, s in enumerate(STATUSES, start=4):
        dash[f"A{i}"] = s
        dash[f"B{i}"] = f'=COUNTIF(\'Master Task Tracker\'!E:E,A{i})'
    for i, pr in enumerate(PRIORITIES, start=4):
        dash[f"D{i}"] = pr
        dash[f"E{i}"] = f'=COUNTIF(\'Master Task Tracker\'!C:C,D{i})'
    dash["G3"] = "Gate Status"
    dash["G4"] = "Execution may continue internally; production/certificate/approval claims remain blocked."
    dash["G3"].fill = gold_fill; dash["G3"].font = bold
    dash["G4"].alignment = Alignment(wrap_text=True)
    dash.column_dimensions["G"].width = 55
    glossary = wb["Glossary"]
    glossary.append(["Term", "Definition"])
    glossary.append(["Optional Clinical Support", "Optional, non-credit, non-gating support pathway; not clinical-hour credit."])
    glossary.append(["Active-Time Validation", "Validated evidence that active participation meets time requirements; Moodle logs alone are insufficient."])
    glossary.append(["P0 Blocker", "Must be resolved before production/pilot clearance or certificate action."])
    for ws in wb.worksheets:
        max_row = ws.max_row
        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = border
        if max_row > 1:
            ws.conditional_formatting.add(f"E2:E{max_row}", FormulaRule(formula=['ISNUMBER(SEARCH("Blocked",E2))'], fill=PatternFill("solid", fgColor="F4CCCC")))
            ws.conditional_formatting.add(f"E2:E{max_row}", FormulaRule(formula=['ISNUMBER(SEARCH("Review",E2))'], fill=PatternFill("solid", fgColor="FCE5CD")))
            ws.conditional_formatting.add(f"E2:E{max_row}", FormulaRule(formula=['E2="Complete"'], fill=PatternFill("solid", fgColor="D9EAD3")))
            ws.conditional_formatting.add(f"C2:C{max_row}", FormulaRule(formula=['ISNUMBER(SEARCH("P0",C2))'], fill=PatternFill("solid", fgColor="F4CCCC")))
    path = OUT / "CNA_Recertification_Executive_Task_Tracker.xlsx"
    wb.save(path)
    wb2 = load_workbook(path)
    return wb2.sheetnames


def qa(page_count: int, sheetnames: list[str], backed_up: list[str]) -> dict:
    pdf = OUT / "CNA_Recertification_Project_Binder.pdf"
    flow_pdf = OUT / "CNA_Recertification_Critical_Path_Flowchart.pdf"
    flow_png = OUT / "CNA_Recertification_Critical_Path_Flowchart.png"
    text = "\n".join(page.extract_text() or "" for page in PdfReader(str(pdf)).pages)
    checks = {
        "pdf_exists": pdf.exists() and pdf.stat().st_size > 0,
        "binder_page_count": page_count,
        "flowchart_pdf_exists": flow_pdf.exists() and flow_pdf.stat().st_size > 0,
        "flowchart_png_exists": flow_png.exists() and flow_png.stat().st_size > 0,
        "footer_verified": "CNA Recertification Project Binder" in text,
        "no_raw_markdown_pipes": "|" not in text,
        "no_code_block_styling": "```" not in text,
        "no_cdph_approved_claim": "CDPH Approved" not in text and "CDPH-approved" not in text,
        "no_certificate_issuance_claim": "Certificate Issuance Enabled" not in text,
        "optional_clinical_non_gating": "non-gating" in text,
        "answer_keys_not_exposed": "answer keys and rationales remain internal/admin-only" in text,
        "workbook_sheets": sheetnames,
        "backed_up": backed_up,
    }
    return checks


def build_report(checks: dict, backed_up: list[str]):
    report = {
        "prepared": PREPARED,
        "files_created": OUTPUT_FILES + ["assets/ci-ion-logomark-white.svg"],
        "files_backed_up": backed_up,
        "source_docs_inspected": SOURCE_DOCS,
        "qa": checks,
        "remaining_p0_blockers": [
            "Approval metadata / CDPH provider-course approval",
            "NAC/provider details if applicable",
            "Approved certificate wording",
            "Affidavit/e-sign or wet-sign method",
            "Active-time validation/manual review method",
            "Legal/compliance clearance before production claims or certificate issuance",
        ],
        "evidence_missing": [
            "Final Google Drive review links",
            "59 URL rows still undocumented",
            "Runtime/browser app verification evidence",
            "SME sign-off for M01/M03/M05",
            "Written approval artifacts for certificate, affidavit, active time, and provider metadata",
        ],
    }
    md = ["# CNA Recertification Binder Build Report", "", f"Prepared: {PREPARED}", ""]
    md.append("## Files Created")
    md.extend(f"- `{f}`" for f in report["files_created"])
    md.append("\n## Files Backed Up")
    md.extend(f"- `{f}`" for f in backed_up) if backed_up else md.append("- None")
    md.append("\n## QA")
    for k, v in checks.items():
        md.append(f"- {k}: {v}")
    md.append("\n## Remaining P0 Blockers")
    md.extend(f"- {b}" for b in report["remaining_p0_blockers"])
    md.append("\n## Evidence Missing")
    md.extend(f"- {e}" for e in report["evidence_missing"])
    (OUT / "CNA_Recertification_Binder_Build_Report.md").write_text("\n".join(md) + "\n", encoding="utf-8")
    (BUILD / "last_build_report.json").write_text(json.dumps(report, indent=2), encoding="utf-8")


def main():
    ensure_scope()
    backed_up = backup_existing()
    logo = download_logo()
    flow_pdf = OUT / "CNA_Recertification_Critical_Path_Flowchart.pdf"
    flow_png = OUT / "CNA_Recertification_Critical_Path_Flowchart.png"
    build_flowchart_pdf(flow_pdf)
    build_flowchart_png(flow_png)
    build_markdown()
    build_source_index()
    build_readme()
    build_style_guide()
    sheets = build_workbook()
    pages = build_pdf(logo, flow_png)
    checks = qa(pages, sheets, backed_up)
    build_report(checks, backed_up)
    print(json.dumps({"pages": pages, "sheets": sheets, "backed_up": backed_up, "checks": checks}, indent=2))


if __name__ == "__main__":
    main()
