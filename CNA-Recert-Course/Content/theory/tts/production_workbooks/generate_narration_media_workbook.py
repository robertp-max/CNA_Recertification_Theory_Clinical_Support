from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(r"C:\AI\Git\CNA_Recertification_Theory_Clinical_Support")
CONTENT_ROOT = ROOT / "CNA-Recert-Course" / "Content"
OUTPUT_DIR = CONTENT_ROOT / "theory" / "tts" / "production_workbooks"
OUTPUT_PATH = OUTPUT_DIR / "CNA_RECERT_ALL_MODULES_NARRATION_MEDIA_PRODUCTION.xlsx"

SOURCE_FILES = {
    "index_primary": CONTENT_ROOT / "23_CONTENT_PACKAGE_INDEX.md",
    "index_updated": CONTENT_ROOT / "index" / "35_CONTENT_PACKAGE_INDEX_UPDATED.md",
    "tts_package": CONTENT_ROOT / "theory" / "tts" / "34_TTS_NARRATION_PACKAGE_COMPLETE.md",
    "module_00": CONTENT_ROOT / "theory" / "modules" / "04_THEORY_MODULE_00_ORIENTATION_FULL.md",
    "module_01": CONTENT_ROOT / "theory" / "modules" / "05_THEORY_MODULE_01_INFECTION_CONTROL_FULL.md",
    "module_02": CONTENT_ROOT / "theory" / "modules" / "24_THEORY_MODULE_02_RESIDENT_RIGHTS_ABUSE_PREVENTION_FULL.md",
    "module_03": CONTENT_ROOT / "theory" / "modules" / "25_THEORY_MODULE_03_DEMENTIA_COMMUNICATION_CULTURAL_RESPECT_FULL.md",
    "module_04": CONTENT_ROOT / "theory" / "modules" / "26_THEORY_MODULE_04_MOBILITY_FALLS_WORKPLACE_SAFETY_FULL.md",
    "module_05": CONTENT_ROOT / "theory" / "modules" / "27_THEORY_MODULE_05_NUTRITION_SKIN_INTEGRITY_VITAL_SIGNS_FULL.md",
    "module_06": CONTENT_ROOT / "theory" / "modules" / "28_THEORY_MODULE_06_DOCUMENTATION_CHANGE_OF_CONDITION_SCOPE_FULL.md",
    "module_07": CONTENT_ROOT / "theory" / "modules" / "29_THEORY_MODULE_07_REVIEW_FINAL_EXAM_AFFIDAVIT_FULL.md",
    "exam_pool": CONTENT_ROOT / "theory" / "exam" / "30_FINAL_EXAM_POOL_50_COMPLETE.md",
    "clinical_support": CONTENT_ROOT / "clinical-support" / "32_CLINICAL_SUPPORT_FULL_CONTENT.md",
    "source_verification": CONTENT_ROOT / "source-verification" / "SOURCE_TITLE_VERIFICATION_PASS.md",
    "source_time_notice": CONTENT_ROOT / "source-verification" / "TIME_ALLOCATION_CORRECTION_NOTICE.md",
    "mvp_module_data": ROOT / "standalone-course-mvp" / "src" / "data" / "moduleContent.ts",
    "mvp_course_data": ROOT / "standalone-course-mvp" / "src" / "data" / "courseData.ts",
    "mvp_clinical_data": ROOT / "standalone-course-mvp" / "src" / "data" / "clinicalSupportContent.ts",
    "mvp_main": ROOT / "standalone-course-mvp" / "src" / "main.tsx",
}

MODULE_META = {
    "M00": "Module 0: Orientation and Compliance Boundaries",
    "M01": "Module 1: Infection Control and PPE",
    "M02": "Module 2: Resident Rights, Abuse Prevention, and Boundaries",
    "M03": "Module 3: Dementia, Communication, and Respectful Care",
    "M04": "Module 4: Mobility, Falls, and Workplace Safety",
    "M05": "Module 5: Nutrition, Skin Integrity, Vital Signs, and Observation",
    "M06": "Module 6: Documentation, Reporting, PHI Avoidance, and Scope",
    "MFINAL": "Final Review, Exam/Test, Affidavit, and Certificate Status",
    "CLIN": "Optional Clinical Support Hub",
}

MODULE_STATUS = {
    "M00": "NEEDS COMPLIANCE REVIEW",
    "M01": "NEEDS SME REVIEW",
    "M02": "NEEDS COMPLIANCE REVIEW",
    "M03": "NEEDS SOURCE REVIEW",
    "M04": "NEEDS COMPLIANCE REVIEW",
    "M05": "NEEDS SME REVIEW",
    "M06": "NEEDS COMPLIANCE REVIEW",
    "MFINAL": "NEEDS COMPLIANCE REVIEW",
    "CLIN": "OPTIONAL SUPPORT ONLY",
}

MODULE_ORDER = {"M00": 0, "M01": 1, "M02": 2, "M03": 3, "M04": 4, "M05": 5, "M06": 6, "MFINAL": 7, "CLIN": 8}

STATUS_FILL = {
    "COMPLETE": PatternFill("solid", fgColor="C6EFCE"),
    "NEEDS SME REVIEW": PatternFill("solid", fgColor="FFF2CC"),
    "NEEDS SOURCE REVIEW": PatternFill("solid", fgColor="FFE699"),
    "NEEDS COMPLIANCE REVIEW": PatternFill("solid", fgColor="F8CBAD"),
    "PLACEHOLDER SOURCE MISSING": PatternFill("solid", fgColor="F4CCCC"),
    "OPTIONAL SUPPORT ONLY": PatternFill("solid", fgColor="D9E1F2"),
}

RISKY_TERMS = [
    "clinical hours",
    "clinical credit",
    "competency validation",
    "cdph approved",
    "complete renewal",
    "certified",
    "official certificate",
    "production certificate",
    "real certificate",
    "guaranteed",
]


@dataclass
class Scene:
    module_code: str
    scene_id: str
    slide_type: str
    scene_title: str
    narration: str
    tts_seconds: int
    card_seconds: int
    status: str
    location: str
    flux_prompt: str
    detailed_image_description: str
    ltx_prompt: str
    ltx_count: int

    @property
    def module_name(self) -> str:
        return MODULE_META[self.module_code]

    @property
    def estimated_video_length(self) -> str:
        return f"{self.card_seconds} seconds"


def read_text(path: Path) -> str:
    if not path.exists():
        return ""
    return path.read_text(encoding="utf-8", errors="replace")


def clean_text(value: str) -> str:
    replacements = {
        "â€”": " - ",
        "â€“": "-",
        "â€™": "'",
        "â€œ": '"',
        "â€\x9d": '"',
        "â€¢": "-",
        "âœ…": "",
        "âš‘": "",
        "Â": "",
        "\u00a0": " ",
        "\r": " ",
        "\n": " ",
        "\t": " ",
    }
    text = value
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def sanitize_risky(value: str) -> str:
    text = clean_text(value)
    phrase_replacements = {
        "clinical-hour credit": "renewal-credit context",
        "clinical hour credit": "renewal-credit context",
        "clinical hours": "in-person practice requirements",
        "clinical credit": "renewal-credit context",
        "competency validation": "skills review",
        "cdph approved": "state review pending",
        "complete renewal": "finish renewal steps",
        "official certificate": "course completion record",
        "production certificate": "course completion record",
        "real certificate": "course completion record",
        "guaranteed": "expected",
    }
    for old, new in phrase_replacements.items():
        text = re.sub(re.escape(old), new, text, flags=re.IGNORECASE)

    # Replace person-name style tokens with de-identified language.
    text = re.sub(r"\b(Mr|Mrs|Ms)\.\s+[A-Za-z][A-Za-z'\-]*\b", "a fictional resident", text, flags=re.IGNORECASE)

    text = re.sub(r"\b[Cc]ertified\b", "CNA", text)
    text = re.sub(r"\bCDPH\b", "state authority", text)
    return clean_text(text)


def words_to_seconds(text: str) -> int:
    words = len([w for w in re.split(r"\s+", text) if w])
    seconds = int(round((words / 145.0) * 60.0))
    return max(10, seconds)


def estimate_card_seconds(tts_seconds: int, slide_type: str) -> int:
    base = int(round(tts_seconds * 0.55))
    if slide_type in {"orientation", "phi-warning", "certificate-gate", "affidavit", "final-exam-preview"}:
        base = int(round(tts_seconds * 0.45))
    return max(12, min(35, base))


def estimate_ltx_count(scene_title: str, slide_type: str) -> int:
    title = scene_title.lower()
    if slide_type in {"orientation", "phi-warning", "certificate-gate", "affidavit", "final-exam-preview", "stakeholder-note", "qa-note"}:
        return 1
    if any(term in title for term in ["transfer", "fall", "abuse", "emergency", "pressure", "aspiration", "choking", "dementia"]):
        return 4
    if slide_type in {"challenge", "debrief", "remediation", "clinical-support"}:
        return 2
    return 2


def classify_slide_type(title: str, module_code: str, default: str = "content") -> str:
    t = title.lower()
    if module_code == "CLIN":
        if "phi" in t or "warning" in t:
            return "phi-warning"
        return "clinical-support"
    if "orientation" in t or "welcome" in t or "navigate" in t:
        return "orientation"
    if "phi" in t or "compliance" in t:
        return "phi-warning"
    if "challenge" in t or "knowledge check" in t or "quiz" in t or "check" in t:
        return "challenge"
    if "debrief" in t:
        return "debrief"
    if "remediation" in t:
        return "remediation"
    if "final exam" in t or "exam" in t:
        return "final-exam-preview"
    if "affidavit" in t:
        return "affidavit"
    if "certificate" in t or "next steps" in t:
        return "certificate-gate"
    return default


def status_for_scene(module_code: str, scene_title: str) -> str:
    if module_code == "CLIN":
        return "OPTIONAL SUPPORT ONLY"
    if module_code == "M05" and re.search(r"skin|pressure", scene_title, flags=re.IGNORECASE):
        return "NEEDS SOURCE REVIEW"
    return MODULE_STATUS.get(module_code, "NEEDS SOURCE REVIEW")


def module_code_from_number(module_number: int) -> str:
    if module_number == 7:
        return "MFINAL"
    return f"M{module_number:02d}"


def make_location(module_code: str, scene_id: str, slide_type: str, scene_title: str) -> str:
    module_path = module_code
    return sanitize_risky(
        f"CNA-Recert/{module_path}/{scene_id} - {slide_type} | {MODULE_META[module_code]} / {scene_title}"
    )


def make_flux_prompt(scene: Scene) -> str:
    base = (
        "Photorealistic documentary healthcare training still, mobile-friendly composition, "
        "clear subject framing, natural light, respectful CNA workplace context, no text overlay, "
        "no logos, no patient records, no identifying details."
    )
    context = (
        f"Scene focus: {scene.scene_title}. "
        f"Topic: {scene.module_name}. "
        "Use professional uniforms, realistic posture, and safety-aligned actions. "
        "If the scene is compliance-heavy, represent it with a checklist or dashboard visual."
    )
    if scene.slide_type in {"phi-warning", "certificate-gate", "affidavit", "final-exam-preview"}:
        context += " Include a clean training screen, clipboard, or compliance board metaphor with teal/orange UI accents only in screens."
    return sanitize_risky(f"{base} {context}")


def make_detailed_image_description(scene: Scene) -> str:
    details = (
        f"Setting: {scene.module_name} training context in a long-term care or classroom-style environment. "
        "People: diverse adult CNAs and staff in standard scrubs with name badges hidden or non-readable. "
        "Posture: calm, attentive, and safety-focused body language. "
        "Lighting: soft natural or neutral indoor lighting with high clarity. "
        "Mood: respectful, practical, and professional. "
        "Safety details: no unsafe procedures, no exposed records, no resident identifiers, no chart screenshots, and no medication labels shown."
    )
    if scene.slide_type == "clinical-support":
        details += " Emphasize optional clinical support, skills refresh, practice support, and documentation support that is separate from online CE gate logic."
    if scene.slide_type == "phi-warning":
        details += " Include a visible no-PHI checklist icon set (names blocked, faces blurred in examples, records blocked) without any real data."
    return sanitize_risky(details)


def make_ltx_prompt(scene: Scene) -> str:
    movement = (
        f"Create a {scene.card_seconds}-second realistic training clip for scene '{scene.scene_title}'. "
        "Begin with a medium-wide establishing shot, then a slow push-in on the key action. "
        "Use smooth handheld-stabilized or tripod movement, no abrupt cuts, no blur effects, no dramatic chaos. "
        "Keep actions clinically safe and realistic, with clear educational intent and mobile-safe framing."
    )
    if scene.slide_type in {"phi-warning", "certificate-gate", "affidavit", "final-exam-preview"}:
        movement += " Use a visual metaphor: training dashboard, checklist completion, certificate gate panel, or attestation review screen."
    if scene.slide_type == "clinical-support":
        movement += " Reinforce that this is optional clinical support, skills refresh, and practice support not tied to online CE certificate gate logic."
    return sanitize_risky(movement)


def parse_tts_segments(tts_text: str) -> list[dict[str, str | int]]:
    pattern = re.compile(
        r"Segment ID:\s*TTS-(?P<module>\d)-(?P<segment>\d{3})\s+Module:\s*(?P<module_check>\d)\s+Lesson:\s*(?P<lesson>.*?)\s+Estimated Duration:\s*(?P<duration>\d+)\s*seconds\s+Transcript:\s*\"(?P<transcript>(?:\\.|[^\"\\])*)\"\s+Pronunciation Notes:\s*(?P<notes>.*?)(?=\n\s*Segment ID:|\n\s*MODULE\s+\d|\Z)",
        flags=re.IGNORECASE | re.DOTALL,
    )
    segments: list[dict[str, str | int]] = []
    for match in pattern.finditer(tts_text):
        module_number = int(match.group("module"))
        if module_number != int(match.group("module_check")):
            continue
        transcript = match.group("transcript").replace('\\"', '"')
        segments.append(
            {
                "module": module_number,
                "segment": int(match.group("segment")),
                "lesson": clean_text(match.group("lesson")),
                "duration": int(match.group("duration")),
                "transcript": sanitize_risky(transcript),
            }
        )
    return sorted(segments, key=lambda item: (int(item["module"]), int(item["segment"])))


def parse_module_checks(ts_text: str) -> dict[int, list[dict[str, str]]]:
    module_checks: dict[int, list[dict[str, str]]] = defaultdict(list)
    module_blocks = re.finditer(
        r"\n  \{\n    id: \"m(?P<mid>\d)\"(?P<body>.*?)(?=\n  \{\n    id: \"m\d\"|\n\];)",
        ts_text,
        flags=re.DOTALL,
    )
    for block in module_blocks:
        module_id = int(block.group("mid"))
        body = block.group("body")
        checks_match = re.search(r"scenarioChecks:\s*\[(?P<checks>.*?)\],\s*remediation:", body, flags=re.DOTALL)
        module_remediation = ""
        remediation_match = re.search(r"remediation:\s*\"((?:\\.|[^\"\\])*)\"", body)
        if remediation_match:
            module_remediation = sanitize_risky(remediation_match.group(1).replace('\\"', '"'))

        if not checks_match:
            continue
        checks_text = checks_match.group("checks")
        check_blocks = re.finditer(
            r"\{\s*title:\s*\"(?P<title>(?:\\.|[^\"\\])*)\"\s*,\s*prompt:\s*\"(?P<prompt>(?:\\.|[^\"\\])*)\"(?P<tail>.*?)\}",
            checks_text,
            flags=re.DOTALL,
        )
        for check in check_blocks:
            title = sanitize_risky(check.group("title").replace('\\"', '"'))
            prompt = sanitize_risky(check.group("prompt").replace('\\"', '"'))
            tail = check.group("tail")
            feedback_match = re.search(r"feedback:\s*\"((?:\\.|[^\"\\])*)\"", tail)
            remediation_match_tail = re.search(r"remediation:\s*\"((?:\\.|[^\"\\])*)\"", tail)
            feedback = sanitize_risky(feedback_match.group(1).replace('\\"', '"')) if feedback_match else ""
            remediation = (
                sanitize_risky(remediation_match_tail.group(1).replace('\\"', '"'))
                if remediation_match_tail
                else module_remediation
            )
            module_checks[module_id].append(
                {
                    "title": title,
                    "prompt": prompt,
                    "feedback": feedback,
                    "remediation": remediation,
                }
            )
    return module_checks


def parse_final_screens(module7_text: str) -> list[dict[str, str]]:
    screens: list[dict[str, str]] = []
    screen_re = re.compile(
        r"^\s*Screen\s+(?P<id>7\.\d+\.\d+)\s+(?:\-|–|—|â€”)?\s*(?P<title>.+)$",
        flags=re.MULTILINE | re.IGNORECASE,
    )
    matches = list(screen_re.finditer(module7_text))
    for idx, match in enumerate(matches):
        block_start = match.end()
        block_end = matches[idx + 1].start() if idx + 1 < len(matches) else len(module7_text)
        block = module7_text[block_start:block_end]
        lines = []
        for raw_line in block.splitlines():
            candidate = clean_text(raw_line)
            if not candidate:
                continue
            if any(
                candidate.startswith(prefix)
                for prefix in [
                    "On-Screen Text",
                    "Moodle Activity Type",
                    "Learner-facing questionnaire fields",
                    "Admin/facilitator notes",
                    "Resources",
                    "Field",
                    "Submit button",
                ]
            ):
                continue
            if candidate.startswith("Source") or candidate.startswith("Evidence"):
                continue
            lines.append(candidate)
        narration = sanitize_risky(" ".join(lines[:10]))
        narration_words = narration.split()
        if len(narration_words) > 190:
            narration = " ".join(narration_words[:190]).strip() + "."
        screens.append({"screen_ref": match.group("id"), "title": clean_text(match.group("title")), "narration": narration})
    return screens


def parse_clinical_units(clin_text: str) -> list[dict[str, str]]:
    units: list[dict[str, str]] = []
    unit_re = re.compile(r"^CLINICAL SUPPORT UNIT\s+(?P<num>\d+):\s*(?P<title>.+)$", flags=re.MULTILINE)
    matches = list(unit_re.finditer(clin_text))
    for idx, match in enumerate(matches):
        block_start = match.end()
        block_end = matches[idx + 1].start() if idx + 1 < len(matches) else len(clin_text)
        block = clin_text[block_start:block_end]
        lines = []
        for raw_line in block.splitlines():
            candidate = clean_text(raw_line)
            if not candidate:
                continue
            if candidate.startswith("Moodle Activity Type") or candidate.startswith("Learner-Facing Intro"):
                continue
            if candidate.startswith("Facilitator/Admin Notes"):
                continue
            if re.match(r"^Check\s+\d+", candidate):
                continue
            lines.append(candidate)
        narration = " ".join(lines[:8])
        narration = sanitize_risky(narration)
        if len(narration.split()) > 170:
            narration = " ".join(narration.split()[:170]).strip() + "."
        units.append({"unit": match.group("num"), "title": clean_text(match.group("title")), "narration": narration})
    return units


def next_scene_id(counter_map: dict[str, dict[str, int]], module_code: str, kind: str) -> str:
    counter_map[module_code][kind] += 1
    return f"{module_code}-{kind}{counter_map[module_code][kind]:02d}"


def scene_sort_key(scene: Scene) -> tuple[int, int, str]:
    kind_order = {"S": 0, "Q": 1, "D": 2, "A": 3}
    scene_parts = scene.scene_id.split("-")
    suffix = scene_parts[-1]
    kind = suffix[0] if suffix else "S"
    num = int(re.sub(r"\D", "", suffix) or "0")
    return MODULE_ORDER.get(scene.module_code, 99), kind_order.get(kind, 9), f"{num:04d}"


def build_scenes(
    tts_segments: list[dict[str, str | int]],
    module_checks: dict[int, list[dict[str, str]]],
    final_screens: list[dict[str, str]],
    clinical_units: list[dict[str, str]],
) -> list[Scene]:
    scenes: list[Scene] = []
    counters: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))

    for segment in tts_segments:
        module_number = int(segment["module"])
        module_code = module_code_from_number(module_number)
        lesson = str(segment["lesson"])
        transcript = sanitize_risky(str(segment["transcript"]))

        kind = "S"
        if module_code == "MFINAL":
            if "exam" in lesson.lower():
                kind = "Q"
            elif "affidavit" in lesson.lower() or "certificate" in lesson.lower():
                kind = "A"

        slide_type = classify_slide_type(lesson, module_code)
        scene_id = next_scene_id(counters, module_code, kind)
        tts_seconds = int(segment["duration"])
        card_seconds = estimate_card_seconds(tts_seconds, slide_type)
        status = status_for_scene(module_code, lesson)

        placeholder_scene = Scene(
            module_code=module_code,
            scene_id=scene_id,
            slide_type=slide_type,
            scene_title=lesson,
            narration=transcript,
            tts_seconds=tts_seconds,
            card_seconds=card_seconds,
            status=status,
            location="",
            flux_prompt="",
            detailed_image_description="",
            ltx_prompt="",
            ltx_count=0,
        )
        placeholder_scene.location = make_location(module_code, scene_id, slide_type, lesson)
        placeholder_scene.ltx_count = estimate_ltx_count(lesson, slide_type)
        placeholder_scene.flux_prompt = make_flux_prompt(placeholder_scene)
        placeholder_scene.detailed_image_description = make_detailed_image_description(placeholder_scene)
        placeholder_scene.ltx_prompt = make_ltx_prompt(placeholder_scene)
        scenes.append(placeholder_scene)

    for module_number in range(0, 7):
        module_code = module_code_from_number(module_number)
        checks = module_checks.get(module_number, [])
        if not checks:
            if module_code == "M00":
                checks = [
                    {
                        "title": "Orientation acknowledgment check",
                        "prompt": "Confirm understanding of online CE boundaries, no-PHI policy, and certificate-gate requirements before starting Module 1.",
                        "feedback": "Review online CE cap, no-PHI boundaries, and required acknowledgements.",
                        "remediation": "Re-read orientation disclosures and complete acknowledgements before continuing.",
                    }
                ]
            else:
                checks = [
                    {
                        "title": f"{MODULE_META[module_code]} challenge check",
                        "prompt": "Apply the module's safety guidance to a realistic CNA shift scenario and choose the safest immediate action.",
                        "feedback": "Reinforce timely reporting, safety boundaries, and role-appropriate decisions.",
                        "remediation": "Review key lesson points before retaking this challenge.",
                    }
                ]

        for check in checks:
            q_title = sanitize_risky(check.get("title", "Challenge check"))
            q_narration = sanitize_risky(
                f"Challenge check for {MODULE_META[module_code]}. {check.get('prompt', '')} Pause briefly, select the safest response, and verify your reasoning against scope and safety standards."
            )
            q_seconds = words_to_seconds(q_narration)
            q_slide = "challenge"
            q_scene_id = next_scene_id(counters, module_code, "Q")
            q_status = status_for_scene(module_code, q_title)

            q_scene = Scene(
                module_code=module_code,
                scene_id=q_scene_id,
                slide_type=q_slide,
                scene_title=q_title,
                narration=q_narration,
                tts_seconds=q_seconds,
                card_seconds=estimate_card_seconds(q_seconds, q_slide),
                status=q_status,
                location="",
                flux_prompt="",
                detailed_image_description="",
                ltx_prompt="",
                ltx_count=0,
            )
            q_scene.location = make_location(module_code, q_scene_id, q_slide, q_title)
            q_scene.ltx_count = estimate_ltx_count(q_title, q_slide)
            q_scene.flux_prompt = make_flux_prompt(q_scene)
            q_scene.detailed_image_description = make_detailed_image_description(q_scene)
            q_scene.ltx_prompt = make_ltx_prompt(q_scene)
            scenes.append(q_scene)

            d_title = sanitize_risky(f"{q_title} debrief")
            d_narration = sanitize_risky(
                f"Debrief for {MODULE_META[module_code]}. {check.get('feedback', '')} Remediation guidance: {check.get('remediation', 'Review the related lesson and retry.')}"
            )
            d_seconds = words_to_seconds(d_narration)
            d_slide = "debrief"
            d_scene_id = next_scene_id(counters, module_code, "D")
            d_status = status_for_scene(module_code, d_title)

            d_scene = Scene(
                module_code=module_code,
                scene_id=d_scene_id,
                slide_type=d_slide,
                scene_title=d_title,
                narration=d_narration,
                tts_seconds=d_seconds,
                card_seconds=estimate_card_seconds(d_seconds, d_slide),
                status=d_status,
                location="",
                flux_prompt="",
                detailed_image_description="",
                ltx_prompt="",
                ltx_count=0,
            )
            d_scene.location = make_location(module_code, d_scene_id, d_slide, d_title)
            d_scene.ltx_count = estimate_ltx_count(d_title, d_slide)
            d_scene.flux_prompt = make_flux_prompt(d_scene)
            d_scene.detailed_image_description = make_detailed_image_description(d_scene)
            d_scene.ltx_prompt = make_ltx_prompt(d_scene)
            scenes.append(d_scene)

    for screen in final_screens:
        title = sanitize_risky(screen["title"])
        narration = sanitize_risky(screen["narration"])
        screen_ref = screen["screen_ref"]

        if screen_ref.startswith("7.3"):
            kind = "Q"
            slide_type = "final-exam-preview"
        elif screen_ref.startswith("7.4") or "affidavit" in title.lower():
            kind = "A"
            slide_type = "affidavit"
        elif screen_ref.startswith("7.5") or "certificate" in title.lower() or "next steps" in title.lower():
            kind = "A"
            slide_type = "certificate-gate"
        else:
            kind = "S"
            slide_type = "content"

        scene_id = next_scene_id(counters, "MFINAL", kind)
        seconds = words_to_seconds(narration)
        final_scene = Scene(
            module_code="MFINAL",
            scene_id=scene_id,
            slide_type=slide_type,
            scene_title=title,
            narration=narration,
            tts_seconds=seconds,
            card_seconds=estimate_card_seconds(seconds, slide_type),
            status="NEEDS COMPLIANCE REVIEW",
            location="",
            flux_prompt="",
            detailed_image_description="",
            ltx_prompt="",
            ltx_count=0,
        )
        final_scene.location = make_location("MFINAL", scene_id, slide_type, title)
        final_scene.ltx_count = estimate_ltx_count(title, slide_type)
        final_scene.flux_prompt = make_flux_prompt(final_scene)
        final_scene.detailed_image_description = make_detailed_image_description(final_scene)
        final_scene.ltx_prompt = make_ltx_prompt(final_scene)
        scenes.append(final_scene)

    clin_intro = sanitize_risky(
        "This optional clinical support hub provides skills refresh, practice support, and documentation support. "
        "It is not part of the online CE certificate gate, and it is not renewal credit unless written approval exists."
    )
    clin_scene_id = next_scene_id(counters, "CLIN", "S")
    clin_scene = Scene(
        module_code="CLIN",
        scene_id=clin_scene_id,
        slide_type="clinical-support",
        scene_title="Optional clinical support overview",
        narration=clin_intro,
        tts_seconds=words_to_seconds(clin_intro),
        card_seconds=estimate_card_seconds(words_to_seconds(clin_intro), "clinical-support"),
        status="OPTIONAL SUPPORT ONLY",
        location="",
        flux_prompt="",
        detailed_image_description="",
        ltx_prompt="",
        ltx_count=0,
    )
    clin_scene.location = make_location("CLIN", clin_scene_id, clin_scene.slide_type, clin_scene.scene_title)
    clin_scene.ltx_count = estimate_ltx_count(clin_scene.scene_title, clin_scene.slide_type)
    clin_scene.flux_prompt = make_flux_prompt(clin_scene)
    clin_scene.detailed_image_description = make_detailed_image_description(clin_scene)
    clin_scene.ltx_prompt = make_ltx_prompt(clin_scene)
    scenes.append(clin_scene)

    phi_narration = sanitize_risky(
        "PHI warning. Do not upload names, faces, dates of birth, addresses, chart captures, medication records, or record numbers. "
        "Use de-identified learning documentation only inside optional clinical support practice areas."
    )
    phi_scene_id = next_scene_id(counters, "CLIN", "S")
    phi_scene = Scene(
        module_code="CLIN",
        scene_id=phi_scene_id,
        slide_type="phi-warning",
        scene_title="No-PHI upload and documentation warning",
        narration=phi_narration,
        tts_seconds=words_to_seconds(phi_narration),
        card_seconds=estimate_card_seconds(words_to_seconds(phi_narration), "phi-warning"),
        status="OPTIONAL SUPPORT ONLY",
        location="",
        flux_prompt="",
        detailed_image_description="",
        ltx_prompt="",
        ltx_count=0,
    )
    phi_scene.location = make_location("CLIN", phi_scene_id, phi_scene.slide_type, phi_scene.scene_title)
    phi_scene.ltx_count = estimate_ltx_count(phi_scene.scene_title, phi_scene.slide_type)
    phi_scene.flux_prompt = make_flux_prompt(phi_scene)
    phi_scene.detailed_image_description = make_detailed_image_description(phi_scene)
    phi_scene.ltx_prompt = make_ltx_prompt(phi_scene)
    scenes.append(phi_scene)

    for unit in clinical_units:
        unit_title = sanitize_risky(unit["title"])
        unit_narration = sanitize_risky(unit["narration"])
        scene_id = next_scene_id(counters, "CLIN", "S")
        unit_scene = Scene(
            module_code="CLIN",
            scene_id=scene_id,
            slide_type="clinical-support",
            scene_title=unit_title,
            narration=unit_narration,
            tts_seconds=words_to_seconds(unit_narration),
            card_seconds=estimate_card_seconds(words_to_seconds(unit_narration), "clinical-support"),
            status="OPTIONAL SUPPORT ONLY",
            location="",
            flux_prompt="",
            detailed_image_description="",
            ltx_prompt="",
            ltx_count=0,
        )
        unit_scene.location = make_location("CLIN", scene_id, unit_scene.slide_type, unit_scene.scene_title)
        unit_scene.ltx_count = estimate_ltx_count(unit_scene.scene_title, unit_scene.slide_type)
        unit_scene.flux_prompt = make_flux_prompt(unit_scene)
        unit_scene.detailed_image_description = make_detailed_image_description(unit_scene)
        unit_scene.ltx_prompt = make_ltx_prompt(unit_scene)
        scenes.append(unit_scene)

    if not any(s.module_code == "M00" and s.slide_type == "phi-warning" for s in scenes):
        phi_scene_id = next_scene_id(counters, "M00", "S")
        narration = sanitize_risky(
            "No-PHI boundary reminder. Never enter names, faces, addresses, record identifiers, or chart captures in course activities. "
            "Use only de-identified examples for learning responses."
        )
        m00_phi_scene = Scene(
            module_code="M00",
            scene_id=phi_scene_id,
            slide_type="phi-warning",
            scene_title="No-PHI boundary reminder",
            narration=narration,
            tts_seconds=words_to_seconds(narration),
            card_seconds=estimate_card_seconds(words_to_seconds(narration), "phi-warning"),
            status="NEEDS COMPLIANCE REVIEW",
            location="",
            flux_prompt="",
            detailed_image_description="",
            ltx_prompt="",
            ltx_count=0,
        )
        m00_phi_scene.location = make_location("M00", phi_scene_id, m00_phi_scene.slide_type, m00_phi_scene.scene_title)
        m00_phi_scene.ltx_count = estimate_ltx_count(m00_phi_scene.scene_title, m00_phi_scene.slide_type)
        m00_phi_scene.flux_prompt = make_flux_prompt(m00_phi_scene)
        m00_phi_scene.detailed_image_description = make_detailed_image_description(m00_phi_scene)
        m00_phi_scene.ltx_prompt = make_ltx_prompt(m00_phi_scene)
        scenes.append(m00_phi_scene)

    scenes = sorted(scenes, key=scene_sort_key)
    return scenes


def freeze_and_filter(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


def style_headers(ws) -> None:
    header_fill = PatternFill("solid", fgColor="DCE6F1")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_wrap(ws, columns: Iterable[int]) -> None:
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col in columns:
            row[col - 1].alignment = Alignment(vertical="top", wrap_text=True)


def set_column_widths(ws, widths: dict[int, float]) -> None:
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width


def mmss(total_seconds: int) -> str:
    minutes, seconds = divmod(total_seconds, 60)
    return f"{minutes:02d}:{seconds:02d}"


def build_runtime_summary(scenes: list[Scene]) -> list[tuple[str, str]]:
    status_counts = Counter(scene.status for scene in scenes)
    type_counts = Counter(scene.slide_type for scene in scenes)
    total_tts = sum(scene.tts_seconds for scene in scenes)
    total_card = sum(scene.card_seconds for scene in scenes)
    total_ltx = sum(scene.ltx_count for scene in scenes)

    missing_narration = sum(1 for scene in scenes if not scene.narration.strip() and scene.status != "PLACEHOLDER SOURCE MISSING")
    missing_flux = sum(1 for scene in scenes if not scene.flux_prompt.strip())
    missing_ltx = sum(1 for scene in scenes if not scene.ltx_prompt.strip())

    modules_included = ", ".join(sorted({scene.module_code for scene in scenes}, key=lambda item: MODULE_ORDER[item]))
    optional_count = sum(1 for scene in scenes if scene.module_code == "CLIN")
    review_required = sum(1 for scene in scenes if scene.status in {"NEEDS SME REVIEW", "NEEDS SOURCE REVIEW", "NEEDS COMPLIANCE REVIEW", "PLACEHOLDER SOURCE MISSING"})

    rows: list[tuple[str, str]] = [
        ("Total cards", str(len(scenes))),
        ("TTS total seconds", str(total_tts)),
        ("TTS total mm:ss", mmss(total_tts)),
        ("Card duration total seconds", str(total_card)),
        ("Card duration total mm:ss", mmss(total_card)),
        ("Total LTX generations", str(total_ltx)),
        ("Missing narration count", str(missing_narration)),
        ("Missing FLUX prompt count", str(missing_flux)),
        ("Missing LTX prompt count", str(missing_ltx)),
        ("Modules included", modules_included),
        ("Optional clinical support scenes count", str(optional_count)),
        ("Review-required scenes count", str(review_required)),
    ]

    for status in [
        "COMPLETE",
        "NEEDS SME REVIEW",
        "NEEDS SOURCE REVIEW",
        "NEEDS COMPLIANCE REVIEW",
        "PLACEHOLDER SOURCE MISSING",
        "OPTIONAL SUPPORT ONLY",
    ]:
        rows.append((f"Status count - {status}", str(status_counts.get(status, 0))))

    for slide_type in sorted(type_counts.keys()):
        rows.append((f"Slide type count - {slide_type}", str(type_counts[slide_type])))

    rows.extend(
        [
            ("Workbook generated date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Source content root", str(CONTENT_ROOT)),
            ("Output path", str(OUTPUT_PATH)),
            (
                "Notes",
                "Prototype-only planning workbook. Preserve review flags. Optional clinical support is separate from online CE certificate gate and not renewal credit unless written approval exists.",
            ),
        ]
    )

    return rows


def write_workbook(scenes: list[Scene]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "TTS + Prompts Production"
    ws2 = wb.create_sheet("Narration Only")
    ws3 = wb.create_sheet("Image + Video Prompts")
    ws4 = wb.create_sheet("Runtime Summary")

    ws1.append(
        [
            "Module",
            "Location",
            "TTS Narration Script",
            "FLUX1Dev Image Gen Prompt",
            "Detailed Image Description",
            "LTX Video Gen Prompt",
            "Estimated Video Length",
            "Status",
        ]
    )

    ws2.append(["Scene ID", "Slide Type", "Scene Title", "TTS Narration Script", "TTS Pacing (s)", "Card Dur (s)"])
    ws3.append(
        [
            "Scene ID",
            "Slide Type",
            "Scene Title",
            "FLUX1Dev Image Prompt",
            "Detailed Image Description",
            "LTX Video Gen Prompt",
            "LTX Count",
        ]
    )
    ws4.append(["Metric", "Value"])

    for scene in scenes:
        ws1.append(
            [
                scene.module_name,
                scene.location,
                scene.narration,
                scene.flux_prompt,
                scene.detailed_image_description,
                scene.ltx_prompt,
                scene.estimated_video_length,
                scene.status,
            ]
        )

        ws2.append(
            [
                scene.scene_id,
                scene.slide_type,
                scene.scene_title,
                scene.narration,
                scene.tts_seconds,
                scene.card_seconds,
            ]
        )

        ws3.append(
            [
                scene.scene_id,
                scene.slide_type,
                scene.scene_title,
                scene.flux_prompt,
                scene.detailed_image_description,
                scene.ltx_prompt,
                scene.ltx_count,
            ]
        )

    for metric, value in build_runtime_summary(scenes):
        ws4.append([metric, value])

    for ws in [ws1, ws2, ws3, ws4]:
        style_headers(ws)
        freeze_and_filter(ws)
        ws.row_dimensions[1].height = 26

    apply_wrap(ws1, [2, 3, 4, 5, 6])
    apply_wrap(ws2, [4])
    apply_wrap(ws3, [4, 5, 6])
    apply_wrap(ws4, [1, 2])

    set_column_widths(ws1, {1: 38, 2: 72, 3: 78, 4: 78, 5: 78, 6: 78, 7: 20, 8: 28})
    set_column_widths(ws2, {1: 16, 2: 20, 3: 44, 4: 90, 5: 14, 6: 14})
    set_column_widths(ws3, {1: 16, 2: 20, 3: 44, 4: 86, 5: 86, 6: 86, 7: 12})
    set_column_widths(ws4, {1: 44, 2: 120})

    for row_idx in range(2, ws1.max_row + 1):
        status = ws1.cell(row=row_idx, column=8).value
        fill = STATUS_FILL.get(str(status), None)
        if fill:
            ws1.cell(row=row_idx, column=8).fill = fill

    wb.save(OUTPUT_PATH)


def find_risky_hits(scenes: list[Scene]) -> dict[str, int]:
    flattened_parts: list[str] = []
    for scene in scenes:
        flattened_parts.extend(
            [
                scene.narration,
                scene.flux_prompt,
                scene.detailed_image_description,
                scene.ltx_prompt,
            ]
        )
    flattened = "\n".join(flattened_parts)
    hits: dict[str, int] = {}
    for term in RISKY_TERMS:
        count = len(re.findall(re.escape(term), flattened, flags=re.IGNORECASE))
        if count:
            hits[term] = count
    return hits


def find_phi_like_hits(scenes: list[Scene]) -> dict[str, int]:
    text = "\n".join(scene.narration for scene in scenes)
    patterns = {
        "possible_name_prefix": r"\b(Mr|Mrs|Ms)\.\s+[A-Za-z]+",
        "ssn_like": r"\b\d{3}-\d{2}-\d{4}\b",
        "dob_like": r"\b\d{1,2}/\d{1,2}/\d{2,4}\b",
        "mrn_like": r"\bMRN[-\s]?[A-Za-z0-9]{3,}\b",
    }
    hits: dict[str, int] = {}
    for label, pattern in patterns.items():
        count = len(re.findall(pattern, text, flags=re.IGNORECASE))
        if count:
            hits[label] = count
    return hits


def validate_output(scenes: list[Scene]) -> dict[str, object]:
    wb = load_workbook(OUTPUT_PATH)
    required_sheets = ["TTS + Prompts Production", "Narration Only", "Image + Video Prompts", "Runtime Summary"]
    missing_sheets = [name for name in required_sheets if name not in wb.sheetnames]

    scene_counts_by_module = Counter(scene.module_code for scene in scenes)
    status_counts = Counter(scene.status for scene in scenes)

    missing_narration = [scene.scene_id for scene in scenes if not scene.narration.strip() and scene.status != "PLACEHOLDER SOURCE MISSING"]
    missing_flux = [scene.scene_id for scene in scenes if not scene.flux_prompt.strip()]
    missing_ltx = [scene.scene_id for scene in scenes if not scene.ltx_prompt.strip()]

    csv_in_output_dir = sorted(str(path.name) for path in OUTPUT_DIR.glob("*.csv"))

    risky_hits = find_risky_hits(scenes)
    phi_hits = find_phi_like_hits(scenes)

    return {
        "workbook_exists": OUTPUT_PATH.exists(),
        "workbook_loadable": True,
        "missing_sheets": missing_sheets,
        "csv_in_output_dir": csv_in_output_dir,
        "scene_counts_by_module": dict(scene_counts_by_module),
        "status_counts": dict(status_counts),
        "missing_narration_count": len(missing_narration),
        "missing_flux_count": len(missing_flux),
        "missing_ltx_count": len(missing_ltx),
        "missing_narration_scene_ids": missing_narration[:20],
        "risky_wording_hits": risky_hits,
        "phi_like_hits": phi_hits,
    }


def collect_source_inspection_list() -> list[str]:
    inspected = []
    for key, path in SOURCE_FILES.items():
        if path.exists():
            inspected.append(str(path.relative_to(ROOT)).replace("\\", "/"))
        else:
            inspected.append(f"{path.relative_to(ROOT).as_posix()} [missing]")
    return inspected


def main() -> None:
    tts_text = read_text(SOURCE_FILES["tts_package"])
    module_ts = read_text(SOURCE_FILES["mvp_module_data"])
    module7_text = read_text(SOURCE_FILES["module_07"])
    clinical_text = read_text(SOURCE_FILES["clinical_support"])

    tts_segments = parse_tts_segments(tts_text)
    module_checks = parse_module_checks(module_ts)
    final_screens = parse_final_screens(module7_text)
    clinical_units = parse_clinical_units(clinical_text)

    scenes = build_scenes(tts_segments, module_checks, final_screens, clinical_units)
    write_workbook(scenes)
    validation = validate_output(scenes)

    summary = {
        "output_workbook": str(OUTPUT_PATH),
        "total_scenes": len(scenes),
        "sheets": ["TTS + Prompts Production", "Narration Only", "Image + Video Prompts", "Runtime Summary"],
        "source_files_inspected": collect_source_inspection_list(),
        "validation": validation,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
    }

    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
