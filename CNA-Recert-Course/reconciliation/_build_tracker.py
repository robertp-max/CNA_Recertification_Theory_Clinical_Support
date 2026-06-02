#!/usr/bin/env python3
"""
CI-ION Course Reconciliation Master Tracker generator.

Reads the local CI-ION export manifest.json (the required evidence source) and
builds a polished multi-sheet Excel workbook plus a markdown URL evidence
register fragment.

WRITE SCOPE: only writes inside CNA-Recert-Course\\reconciliation.
This script does NOT modify any source content, app code, ContentV2, or the export.
It only INSPECTS the export (read-only) and emits reconciliation artifacts.
"""
from __future__ import annotations

import json
import re
import datetime as dt
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import BarChart, Reference

# ----------------------------------------------------------------------------
# Paths
# ----------------------------------------------------------------------------
HERE = Path(__file__).resolve().parent                      # ...\reconciliation
EXPORT_DIR = (HERE.parent /
              "CI-ION-course-export-CI-ION - Course Structures - Contents-20260601-114428")
MANIFEST = EXPORT_DIR / "manifest.json"
OUT_XLSX = HERE / "CI_ION_Course_Reconciliation_Master_Tracker.xlsx"
OUT_URL_MD = HERE / "_url_evidence_register_rows.md"

TODAY = dt.date.today().isoformat()

# ----------------------------------------------------------------------------
# Style palette (modern compliance dashboard)
# ----------------------------------------------------------------------------
C_HEADER_BG = "1F2937"      # slate-800 dark header
C_HEADER_FG = "FFFFFF"
C_TITLE_BG = "111827"       # near-black
C_BAND = "F3F4F6"           # light band
C_CARD_BG = "EEF2FF"        # indigo-50 card
C_CARD_ACCENT = "4338CA"    # indigo-700
C_BLOCKER = "DC2626"        # red
C_BLOCKER_BG = "FEE2E2"
C_WARN_BG = "FEF3C7"        # amber
C_OK_BG = "DCFCE7"          # green
C_INFO_BG = "DBEAFE"        # blue
C_SECTION = "0F766E"        # teal section header

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=C_HEADER_FG)
TITLE_FONT = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
SUB_FONT = Font(name="Calibri", size=10, italic=True, color="E5E7EB")
CELL_FONT = Font(name="Calibri", size=10, color="111827")
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def header_fill():
    return PatternFill("solid", fgColor=C_HEADER_BG)


# ----------------------------------------------------------------------------
# Load + classify the manifest
# ----------------------------------------------------------------------------
manifest = json.loads(MANIFEST.read_text(encoding="utf-8"))

cfg = manifest["config"]
links = manifest["links_found"]
files_added = manifest["files_added"]
folders_not_exp = manifest.get("folders_found_but_not_exported", [])
skipped = manifest.get("skipped", [])
errors = manifest.get("errors", [])

added_ids = {f.get("id"): f for f in files_added if f.get("id")}
skipped_ids = {s.get("id"): s for s in skipped if s.get("id")}
error_ids = {e.get("id"): e for e in errors if e.get("id")}
folder_ids = {f.get("id"): f for f in folders_not_exp if f.get("id")}


def extract_id(url: str):
    """Return (kind, id, anchor) for a Google URL."""
    m = re.search(r"/document/d/([^/]+)", url)
    if m:
        anchor = None
        a = re.search(r"tab=(t\.[a-z0-9]+)", url)
        if a:
            anchor = a.group(1)
        return ("document", m.group(1), anchor)
    m = re.search(r"/spreadsheets/d/([^/]+)", url)
    if m:
        return ("spreadsheet", m.group(1), None)
    m = re.search(r"/file/d/([^/]+)", url)
    if m:
        return ("drive_file", m.group(1), None)
    m = re.search(r"/folders/([^/?]+)", url)
    if m:
        return ("folder", m.group(1), None)
    return ("external", None, None)


HOST_PURPOSE = {
    "calendar.app.google": "Scheduling / booking reference",
    "chat.whatsapp.com": "Student/cohort communication channel",
    "discover.trinitydc.edu": "External Moodle how-to blog reference",
    "docs.moodle.org": "Moodle documentation reference (grade hiding)",
    "learn.ciinstituteofnursing.com": "Live Moodle course shell (id=13)",
    "maps.app.goo.gl": "Facility/location map reference",
    "www.figma.com": "CI-ION user-flow design board",
}

URL_TYPE = {
    "document": "Google Doc",
    "spreadsheet": "Google Sheet",
    "drive_file": "Google Drive file",
    "folder": "Google Drive folder",
    "external": "External link",
}


def classify(url):
    kind, gid, anchor = extract_id(url)
    host = re.sub(r"^https?://", "", url).split("/")[0]
    row = {
        "url": url,
        "kind": kind,
        "id": gid,
        "anchor": anchor,
        "host": host,
        "copied_file": "",
        "copied_url": "",
        "dest": "",
        "skipped_reason": "",
        "copy_error": "",
        "equiv_local": "",
        "needs_new_doc": "Yes",
        "status": "Needs Document Import",
        "purpose": "Needs owner decision",
        "priority": "P2 High",
    }

    if gid and gid in added_ids:
        f = added_ids[gid]
        row["copied_file"] = f.get("zip_name", "")
        row["copied_url"] = f.get("source_url", url)
        row["dest"] = "linked-files/"
        row["status"] = "Copied/imported document exists and is mapped"
        row["needs_new_doc"] = "No"
        row["priority"] = "P3 Low"
        nm = f.get("source_name", "")
        if "cccco-na-model-curriculum" in (f.get("zip_name") or ""):
            row["purpose"] = "CCCCO NA model curriculum source module"
        elif nm:
            row["purpose"] = f"Source document: {nm}"
    elif gid and gid in skipped_ids:
        s = skipped_ids[gid]
        if s.get("type") == "video_file":
            row["status"] = "Video was skipped because video copying was disabled"
        else:
            row["status"] = "Skipped by Copy Config"
        row["skipped_reason"] = s.get("reason", "")
        row["purpose"] = f"Skipped media/large file: {s.get('name','')}"
        row["needs_new_doc"] = "Yes"
        row["priority"] = "P2 High"
    elif gid and gid in folder_ids:
        fo = folder_ids[gid]
        row["status"] = "Folder was found but not copied because folder copying was disabled"
        row["skipped_reason"] = fo.get("reason", "")
        row["purpose"] = f"Folder not exported: {fo.get('name','')}"
        row["needs_new_doc"] = "Yes"
        row["priority"] = "P1 High"
    elif gid and gid in error_ids:
        e = error_ids[gid]
        err = e.get("error", "")
        row["copy_error"] = err
        if "No item with the given ID" in err:
            row["status"] = "Copy failed due to access or missing item"
            row["purpose"] = "Source not accessible / not found at export time"
            row["priority"] = "P1 High"
        else:
            row["status"] = "Copied/imported file exists but needs extraction or conversion"
            row["purpose"] = "Native Google Doc/Sheet could not be converted on export"
            row["priority"] = "P1 High"
        row["needs_new_doc"] = "Yes"
    elif kind == "external":
        row["status"] = "External reference only"
        row["purpose"] = HOST_PURPOSE.get(host, "External reference")
        row["needs_new_doc"] = "Maybe"
        row["priority"] = "P3 Low"
    return row


# Build URL rows with duplicate detection (canonical by id else url)
seen_keys = {}
url_rows = []
for i, url in enumerate(sorted(links), start=1):
    r = classify(url)
    key = r["id"] or url
    if key in seen_keys:
        canonical = seen_keys[key]
        r["status"] = "Duplicate URL or same source document with different tab/anchor"
        r["needs_new_doc"] = "No"
        r["priority"] = "P3 Low"
        r["purpose"] = f"Duplicate / alternate anchor of URL-{canonical:03d}"
    else:
        seen_keys[key] = i
    r["uid"] = f"URL-{i:03d}"
    url_rows.append(r)

# URL metrics
n_urls = len(url_rows)
n_copied = sum(1 for r in url_rows if r["status"].startswith("Copied/imported document exists"))
n_skipped = sum(1 for r in url_rows if "skipped" in r["status"].lower())
n_errored = sum(1 for r in url_rows if r["status"] in (
    "Copy failed due to access or missing item",
    "Copied/imported file exists but needs extraction or conversion"))
n_dupe = sum(1 for r in url_rows if r["status"].startswith("Duplicate"))
n_external = sum(1 for r in url_rows if r["status"] == "External reference only")
n_folder = sum(1 for r in url_rows if r["status"].startswith("Folder was found"))
n_undocumented = sum(1 for r in url_rows if r["needs_new_doc"] == "Yes")

URL_METRICS = dict(
    total=n_urls, copied=n_copied, skipped=n_skipped, errored=n_errored,
    dupe=n_dupe, external=n_external, folder=n_folder, undocumented=n_undocumented,
)

print("URL METRICS:", URL_METRICS)
print("files_added:", len(files_added), "skipped:", len(skipped),
      "errors:", len(errors), "folders_not_exported:", len(folders_not_exp))

# ----------------------------------------------------------------------------
# Allowed enumerations
# ----------------------------------------------------------------------------
STATUS_VALUES = [
    "Not Started", "In Progress", "Needs Export Verification", "Needs Source Repair",
    "Needs SME Review", "Needs Compliance Review", "Needs Legal/CDPH Review",
    "Needs App Verification", "Needs Document Import",
    "Blocked - Source Access / Copy Error", "Skipped by Copy Config", "Blocked",
    "Complete", "Deferred", "Not Applicable",
]
PRIORITY_VALUES = ["P0 Blocker", "P1 High", "P2 Medium", "P3 Low"]
PROGRAM_VALUES = ["CNA Recert", "RCFE Admin", "Case Manager", "Cross-Program"]
SEVERITY_VALUES = ["Critical", "High", "Medium", "Low", "Info"]
BLOCKER_VALUES = ["Yes", "No"]

EXPORT = "CI-ION-course-export ... -20260601-114428"

# ----------------------------------------------------------------------------
# Master Tracker rows
# Columns order (see MT_COLS)
# ----------------------------------------------------------------------------
MT_COLS = [
    "ID", "Program", "Domain", "Priority", "Severity", "Status", "Blocker",
    "Expected Item", "Expected Source / Control Doc", "Actual Export Evidence",
    "Actual App Evidence", "Spreadsheet URL Evidence", "Gap / Mismatch",
    "Required Action", "Deliverable To Create", "Owner Role", "Reviewer Role",
    "Dependencies", "Verification Command or File Path", "Date Opened",
    "Date Updated", "Notes",
]


def MR(idn, program, domain, prio, sev, status, blocker, expected, exp_src,
       exp_ev, app_ev, url_ev, gap, action, deliver, owner, reviewer, deps,
       verify, notes):
    return [idn, program, domain, prio, sev, status, blocker, expected, exp_src,
            exp_ev, app_ev, url_ev, gap, action, deliver, owner, reviewer, deps,
            verify, TODAY, TODAY, notes]


EXP_PATH = ("CNA-Recert-Course\\CI-ION-course-export-CI-ION - Course "
            "Structures - Contents-20260601-114428")

MASTER_ROWS = [
    MR("REC-001", "CNA Recert", "Export Verification", "P1 High", "High",
       "Complete", "No",
       "Confirm exact CI-ION export folder exists locally",
       "Owner instruction; Phase 1",
       "Folder present: manifest.json + CI-ION - Course Structures - Contents.xlsx + linked-files/ (41 files total)",
       "n/a",
       "n/a",
       "None - folder inspected",
       "Record inventory in 01_CI_ION_EXPORT_INVENTORY.md",
       "01_CI_ION_EXPORT_INVENTORY.md", "Repo Auditor", "Program Owner",
       "None",
       f"Get-ChildItem -LiteralPath '{EXP_PATH}' -Recurse",
       "Export is a Google Drive linked-doc + active-spreadsheet export (Apps Script), not a Moodle .mbz backup."),
    MR("REC-002", "CNA Recert", "Export Verification", "P1 High", "High",
       "Complete", "No",
       "Confirm manifest.json present and inspected",
       "Phase 3 evidence rule",
       "manifest.json (53,719 bytes) inspected: config, links_found(162), files_added(40), skipped(36), errors(22), folders_found_but_not_exported(1)",
       "n/a", "All 162 links registered (see Spreadsheet URL Evidence sheet)",
       "None", "Map every URL to an evidence status",
       "15_SPREADSHEET_URL_DOCUMENTATION_EVIDENCE_REGISTER.md",
       "Repo Auditor", "Program Owner", "REC-001",
       f"{EXP_PATH}\\manifest.json",
       "Source spreadsheet id 1RCylz7nAc8JOuc... exported as .xlsx."),
    MR("REC-003", "CNA Recert", "Source Spreadsheet", "P1 High", "High",
       "Needs Export Verification", "No",
       "Extract structure/tabs from CI-ION - Course Structures - Contents.xlsx",
       "Active spreadsheet export",
       "CI-ION - Course Structures - Contents.xlsx (455,709 bytes) present, not yet parsed into reconciliation tables",
       "n/a", "Spreadsheet is the master source for URL list",
       "Spreadsheet sheet/row/cell positions for each URL not yet extracted",
       "Parse .xlsx tabs; map URLs to sheet/row/cell; backfill URL Evidence register positional columns",
       "01_CI_ION_EXPORT_INVENTORY.md (appendix)", "Repo Auditor", "Program Owner",
       "REC-002",
       f"{EXP_PATH}\\CI-ION - Course Structures - Contents.xlsx",
       "Positional columns currently 'Needs verification' in URL register."),
    MR("REC-004", "CNA Recert", "Spreadsheet URL Evidence", "P1 High", "High",
       "Needs Document Import", "Yes",
       "Every spreadsheet URL has a documentation/evidence status",
       "Phase 3 hard rule",
       f"162 URLs classified: {URL_METRICS['copied']} copied, {URL_METRICS['skipped']} skipped, "
       f"{URL_METRICS['errored']} copy/convert errors, {URL_METRICS['dupe']} duplicate/anchor, "
       f"{URL_METRICS['external']} external, {URL_METRICS['folder']} folder-not-copied",
       "n/a",
       f"{URL_METRICS['undocumented']} URLs still need document import/extraction",
       "Native Google Docs failed docx conversion; videos + folder skipped by config",
       "Import/convert native Google Docs; owner decision on folder + videos",
       "15_SPREADSHEET_URL_DOCUMENTATION_EVIDENCE_REGISTER.md",
       "Repo Auditor", "Program Owner / SME", "REC-002",
       "See 'Spreadsheet URL Evidence' sheet",
       "Do not treat spreadsheet rows as documented just because a URL exists."),
    MR("REC-005", "CNA Recert", "Source Documents", "P2 Medium", "Medium",
       "Complete", "No",
       "Inventory local CNA source documentation packet",
       "Phase 2",
       "CCCCO NA model curriculum modules 1-17 PDFs present in export linked-files/",
       "n/a", "Maps to course source list (NATP 10-17 primary)",
       "Export carries CCCCO modules 1-17; course uses 10-17 as primary",
       "Crosswalk export PDFs to repo source docs",
       "02_EXPORT_TO_SOURCE_DOC_CROSSWALK.md", "Repo Auditor", "SME",
       "REC-001",
       "CNA-Recert-Course\\Content\\01_SOURCE_TO_COURSE_CROSSWALK.md",
       "Local Content packet docs 00-23 + theory modules present."),
    MR("REC-006", "CNA Recert", "ContentV2", "P2 Medium", "Medium",
       "Complete", "No",
       "Confirm ContentV2 present locally and reconcile to source+export",
       "Phase 2",
       "n/a (export predates ContentV2)",
       "ContentV2 present: courseContentV2.json, schema, generated.ts, narration CSVs, docs 00-11",
       "n/a",
       "ContentV2 not cross-checked against export source PDFs",
       "Crosswalk ContentV2 lessons to source modules + export PDFs",
       "03_EXPORT_TO_CONTENTV2_CROSSWALK.md", "Repo Auditor", "SME",
       "REC-005",
       "CNA-Recert-Course\\ContentV2\\data\\courseContentV2.json",
       "ContentV2 is present locally (not missing)."),
    MR("REC-007", "CNA Recert", "App Wiring", "P1 High", "High",
       "Needs App Verification", "No",
       "Confirm standalone-course-mvp consumes ContentV2 + reflects gates",
       "Phase 2",
       "n/a",
       "standalone-course-mvp/src/data: contentV2.generated.ts, examPool.ts, certificate/gate pages present",
       "n/a",
       "App-to-export/source binding not verified in running build",
       "Run app build; verify module/exam/gate wiring against ContentV2",
       "04_EXPORT_TO_STANDALONE_APP_CROSSWALK.md", "App Engineer", "QA",
       "REC-006",
       "standalone-course-mvp\\src\\data\\contentV2.generated.ts",
       "Working tree dirty - many app files modified (see _backups)."),
]

# CNA modules M00-M07
MODULES = [
    ("REC-010", "M00 Orientation", 30,
     "Module 0 orientation: identity fields, online-cap ack, no-PHI ack, certificate prerequisites",
     "04_THEORY_MODULE_00_ORIENTATION_FULL.md",
     "CI-ION_ Pre-Orientation Student Catalog Test.pdf; CI Institute of Nursing Student Catalog 3-17-2026.docx",
     "Identity/cap/PHI ack are compliance-gated; verify wording approved"),
    ("REC-011", "M01 Infection Control", 90,
     "Module 1 infection control theory",
     "05_THEORY_MODULE_01_INFECTION_CONTROL_FULL.md",
     "No dedicated CCCCO module 1-17 = infection control; scattered + legacy CNA-CE-001",
     "SME/source review REQUIRED - infection control not a discrete NATP 10-17 module (GAP_LOG)"),
    ("REC-012", "M02 Resident Rights / Abuse Prevention", 120,
     "Module 2 resident rights & abuse prevention",
     "24_THEORY_MODULE_02_RESIDENT_RIGHTS_ABUSE_PREVENTION_FULL.md",
     "cccco-na-model-curriculum-module-17.pdf; PM Day 1/2 Module 2 Patient Resident Rights.pdf",
     "Confirm abuse-reporting wording against current CA statute"),
    ("REC-013", "M03 Dementia / Communication / Cultural Respect", 120,
     "Module 3 dementia, communication, cultural respect",
     "25_THEORY_MODULE_03_DEMENTIA_COMMUNICATION_CULTURAL_RESPECT_FULL.md",
     "cccco-na-model-curriculum-module-13.pdf (Long Term Care Resident)",
     "SOURCE REPAIR RISK - confirm local source M03 is complete/uninterrupted"),
    ("REC-014", "M04 Mobility / Falls / Workplace Safety", 120,
     "Module 4 mobility, falls, workplace safety",
     "26_THEORY_MODULE_04_MOBILITY_FALLS_WORKPLACE_SAFETY_FULL.md",
     "cccco-na-model-curriculum-module-14.pdf (Rehabilitative); module-12.pdf (Emergency)",
     "Confirm restorative/mobility mapping to rehab module"),
    ("REC-015", "M05 Nutrition / Skin Integrity / Vital Signs", 120,
     "Module 5 nutrition, skin integrity, vital signs",
     "27_THEORY_MODULE_05_NUTRITION_SKIN_INTEGRITY_VITAL_SIGNS_FULL.md",
     "cccco-na-model-curriculum-module-10.pdf (Vital Signs); module-11.pdf (Nutrition)",
     "SME review REQUIRED - skin integrity / pressure injury content"),
    ("REC-016", "M06 Documentation / Change of Condition / Scope", 90,
     "Module 6 documentation, change of condition, scope of practice",
     "28_THEORY_MODULE_06_DOCUMENTATION_CHANGE_OF_CONDITION_SCOPE_FULL.md",
     "cccco-na-model-curriculum-module-15.pdf; Day 3/6 Module 15.1/15.2 Observation and Charting.pdf",
     "Ensure no PHI in charting examples"),
    ("REC-017", "M07 Review / Final Exam / Affidavit", 30,
     "Module 7 review, final exam, affidavit",
     "29_THEORY_MODULE_07_REVIEW_FINAL_EXAM_AFFIDAVIT_FULL.md",
     "CI-ION NATP Final Exam.pdf",
     "Keep answer keys internal; affidavit wording = legal review"),
]
for idn, name, mins, exp, src, exp_ev, note in MODULES:
    sme = "Needs SME Review" if ("SME" in note or "REPAIR" in note) else "In Progress"
    blocker = "No"
    sev = "High" if ("SME" in note or "REPAIR" in note) else "Medium"
    MASTER_ROWS.append(MR(
        idn, "CNA Recert", "Theory Module", "P2 Medium", sev, sme, blocker,
        f"{name} ({mins} min target)",
        f"CNA-Recert-Course\\Content\\theory\\modules\\{src}",
        exp_ev,
        "ContentV2 lesson + standalone module page",
        "Mapped via crosswalk",
        note,
        "Verify module content vs source PDF + ContentV2",
        "06_MODULE_TIME_AND_STRUCTURE_RECONCILIATION.md; 10_SOURCE_REPAIR_SME_COMPLIANCE_FLAGS.md",
        "SME (RN/Clinical Educator)", "Compliance",
        "REC-005, REC-006",
        f"CNA-Recert-Course\\Content\\theory\\modules\\{src}",
        f"Timing target {mins} min; total required model = 720 min / 12 h."))

MASTER_ROWS += [
    MR("REC-020", "CNA Recert", "Module Timing", "P2 Medium", "Medium",
       "In Progress", "No",
       "Reconcile module time model to 720 min / 12 h",
       "FINAL_UNIFIED_IMPLEMENTATION_BLUEPRINT.md",
       "Source PDFs do not assert per-module minutes",
       "ContentV2 time recon doc present (10_CONTENT_COVERAGE_AND_TIME_RECONCILIATION)",
       "n/a",
       "Need to confirm sum M00..M07 = 30+90+120+120+120+120+90+30 = 720",
       "Validate time model totals and active-time mapping",
       "06_MODULE_TIME_AND_STRUCTURE_RECONCILIATION.md", "Instructional Designer", "Compliance",
       "REC-010..REC-017",
       "CNA-Recert-Course\\ContentV2\\10_CONTENT_COVERAGE_AND_TIME_RECONCILIATION.md",
       "Do not claim active-time validation."),
    MR("REC-021", "CNA Recert", "Assessment", "P1 High", "High",
       "Needs SME Review", "No",
       "Final exam / 50-question pool reconciliation",
       "09_FINAL_EXAM_BLUEPRINT.md; 30_FINAL_EXAM_POOL_50_COMPLETE.md",
       "CI-ION NATP Final Exam.pdf present in export",
       "standalone-course-mvp examPool.ts present",
       "Mapped",
       "Export final exam vs repo 50-pool not item-matched; answer keys must stay internal",
       "Item-map export exam to repo pool; verify 50 items + pass mark",
       "07_ASSESSMENT_AND_QUIZ_BANK_RECONCILIATION.md", "SME", "Compliance",
       "REC-005",
       "CNA-Recert-Course\\Content\\theory\\exam\\30_FINAL_EXAM_POOL_50_COMPLETE.md",
       "Do not expose answer keys in learner-facing docs."),
    MR("REC-022", "CNA Recert", "Module Checks", "P2 Medium", "Medium",
       "In Progress", "No",
       "Module knowledge checks / interaction matrix",
       "08_MODULE_KNOWLEDGE_CHECK_BLUEPRINT.md; 12_INTERACTION_CHECK_MATRIX.csv",
       "Not directly represented in export",
       "ContentV2 04_MODULE_ASSESSMENT_MAP + v2ModuleQuiz.ts",
       "n/a",
       "Confirm each module has required check + interaction",
       "Reconcile checks vs modules",
       "07_ASSESSMENT_AND_QUIZ_BANK_RECONCILIATION.md", "Instructional Designer", "SME",
       "REC-021",
       "CNA-Recert-Course\\Content\\08_MODULE_KNOWLEDGE_CHECK_BLUEPRINT.md",
       ""),
    MR("REC-023", "CNA Recert", "Affidavit", "P0 Blocker", "Critical",
       "Needs Legal/CDPH Review", "Yes",
       "Affidavit / final statement + e-sign acceptance",
       "13_AFFIDAVIT_TEXT.md",
       "Not in export",
       "standalone CertificateGatePage / affidavit step",
       "n/a",
       "Affidavit wording + e-signature acceptance not legally approved",
       "Legal review of affidavit wording; decide e-sign vs wet-sign",
       "08_CERTIFICATE_GATE_ACTIVE_TIME_AFFIDAVIT_RECONCILIATION.md", "Legal/Compliance", "Program Owner",
       "REC-024",
       "CNA-Recert-Course\\Content\\13_AFFIDAVIT_TEXT.md",
       "Do not invent e-signature acceptance."),
    MR("REC-024", "CNA Recert", "Certificate Gate", "P0 Blocker", "Critical",
       "Needs Legal/CDPH Review", "Yes",
       "Certificate gate behavior + approved certificate wording",
       "14_CERTIFICATE_FIELD_MAPPING.csv",
       "Not in export",
       "standalone CertificateGatePage.tsx present (prototype)",
       "n/a",
       "No CDPH approval, no approved cert wording, no NAC/provider metadata",
       "Obtain provider/course approval, cert wording, provider IDs in writing",
       "08_CERTIFICATE_GATE_ACTIVE_TIME_AFFIDAVIT_RECONCILIATION.md; 13_GO_NO_GO_BLOCKERS_AND_DECISIONS.md",
       "Program Owner / Legal", "CDPH",
       "None",
       "CNA-Recert-Course\\Content\\14_CERTIFICATE_FIELD_MAPPING.csv",
       "Certificate production BLOCKED. Do not enable live issuance."),
    MR("REC-025", "CNA Recert", "Active Time", "P0 Blocker", "Critical",
       "Needs Compliance Review", "Yes",
       "Active-time validation / manual admin review control",
       "ACTIVE_TIME_POC_TEST_PLAN.md (referenced)",
       "Not in export",
       "App active-time control not validated",
       "n/a",
       "Active-time plugin not validated; Moodle logs alone insufficient",
       "Validate active-time in staging or use manual review hold",
       "08_CERTIFICATE_GATE_ACTIVE_TIME_AFFIDAVIT_RECONCILIATION.md", "QA/Technical", "Compliance",
       "REC-024",
       "CNA-Recert-Course\\Content\\00_EXECUTIVE_SUMMARY.md",
       "Do not rely on Moodle logs alone as active participation evidence."),
    MR("REC-026", "CNA Recert", "Optional Clinical", "P1 High", "High",
       "Needs Compliance Review", "No",
       "Optional Clinical Support is separate, optional, non-credit, non-gating",
       "03_CLINICAL_SUPPORT_SYLLABUS_TABLE.md; 32_CLINICAL_SUPPORT_FULL_CONTENT.md",
       "CNA skill videos skipped in export (35 video_file)",
       "standalone ClinicalHubPage.tsx present",
       "Mapped (videos skipped by config)",
       "Must verify clinical hub is non-gating and labeled non-credit",
       "Confirm separation + labeling; do not count as renewal clinical-hour credit",
       "09_OPTIONAL_CLINICAL_SUPPORT_SEPARATION_RECONCILIATION.md", "Compliance", "Program Owner",
       "REC-024",
       "CNA-Recert-Course\\Content\\clinical-support\\32_CLINICAL_SUPPORT_FULL_CONTENT.md",
       "Optional Clinical Support is NOT CA clinical-hour credit and NOT certificate-gating."),
    MR("REC-027", "CNA Recert", "Optional Clinical", "P2 Medium", "Medium",
       "In Progress", "No",
       "Optional confidence checks (non-graded, non-gating)",
       "33_OPTIONAL_CLINICAL_CONFIDENCE_CHECKS_COMPLETE.md",
       "Not in export",
       "standalone confidence-check flows",
       "n/a",
       "Confirm confidence checks are non-credit and clearly optional",
       "Reconcile confidence checks separation",
       "09_OPTIONAL_CLINICAL_SUPPORT_SEPARATION_RECONCILIATION.md", "Instructional Designer", "Compliance",
       "REC-026",
       "CNA-Recert-Course\\Content\\clinical-support\\confidence-checks\\33_OPTIONAL_CLINICAL_CONFIDENCE_CHECKS_COMPLETE.md",
       ""),
    MR("REC-028", "CNA Recert", "PHI Avoidance", "P1 High", "High",
       "Needs Compliance Review", "No",
       "PHI warnings before uploads / free-text / documentation support",
       "19_COMPLIANCE_REVIEW_CHECKLIST.md",
       "Not in export",
       "standalone PhiWarningBlock.tsx present",
       "n/a",
       "Verify PHI warning shown before every upload/free-text surface",
       "Audit all free-text/upload surfaces for PHI warning + no-PHI ack",
       "10_SOURCE_REPAIR_SME_COMPLIANCE_FLAGS.md; 12_QA_NEGATIVE_TEST_AND_ACCEPTANCE_PLAN.md",
       "Compliance/Privacy", "Program Owner",
       "REC-007",
       "standalone-course-mvp\\src\\components\\ui\\PhiWarningBlock.tsx",
       "Do not collect, request, display, or generate PHI."),
    MR("REC-029", "CNA Recert", "Audit Packet", "P2 Medium", "Medium",
       "In Progress", "No",
       "Audit packet evidence map (layered evidence model)",
       "00_EXECUTIVE_SUMMARY.md (layered evidence)",
       "manifest.json is itself export-provenance evidence",
       "App evidence exports TBD",
       "Manifest = copy provenance",
       "Assemble audit evidence chain (identity, completion, interaction, active-time, exam, affidavit, admin)",
       "Build audit evidence map",
       "11_AUDIT_PACKET_EVIDENCE_MAP.md", "Compliance", "Program Owner",
       "REC-024, REC-025",
       "CNA-Recert-Course\\reconciliation\\11_AUDIT_PACKET_EVIDENCE_MAP.md",
       ""),
    MR("REC-030", "CNA Recert", "QA", "P2 Medium", "Medium",
       "Not Started", "No",
       "QA negative tests + acceptance plan",
       "PROTOTYPE_1_ACCEPTANCE_CRITERIA.md (referenced)",
       "n/a", "App QA pending",
       "n/a",
       "Negative tests for gate bypass, PHI entry, optional-as-required not authored here",
       "Author QA negative-test + acceptance plan",
       "12_QA_NEGATIVE_TEST_AND_ACCEPTANCE_PLAN.md", "QA", "Program Owner",
       "REC-007",
       "CNA-Recert-Course\\reconciliation\\12_QA_NEGATIVE_TEST_AND_ACCEPTANCE_PLAN.md",
       ""),
    MR("REC-031", "CNA Recert", "SME Flags", "P1 High", "High",
       "Needs SME Review", "No",
       "Consolidated SME review flags (M01 infection, M05 skin, dementia language)",
       "18_SME_REVIEW_CHECKLIST.md",
       "CCCCO module PDFs available for SME cross-check",
       "n/a", "n/a",
       "SME sign-off not recorded",
       "RN/Clinical Educator review of flagged modules",
       "10_SOURCE_REPAIR_SME_COMPLIANCE_FLAGS.md", "SME (RN)", "Compliance",
       "REC-011, REC-015",
       "CNA-Recert-Course\\Content\\18_SME_REVIEW_CHECKLIST.md",
       "Mark unknowns Needs verification."),
    MR("REC-032", "CNA Recert", "Source Repair", "P1 High", "High",
       "Needs Source Repair", "No",
       "Module 3 source completeness (interrupted-content risk)",
       "25_THEORY_MODULE_03_... + 21_SOURCE_CONFLICT_LOG.md",
       "cccco-na-model-curriculum-module-13.pdf available",
       "n/a", "n/a",
       "Local source may contain incomplete/interrupted M03 content",
       "Verify M03 source completeness; repair from CCCCO module 13 if needed",
       "10_SOURCE_REPAIR_SME_COMPLIANCE_FLAGS.md", "Instructional Designer", "SME",
       "REC-013",
       "CNA-Recert-Course\\Content\\theory\\modules\\25_THEORY_MODULE_03_DEMENTIA_COMMUNICATION_CULTURAL_RESPECT_FULL.md",
       ""),
    MR("REC-033", "CNA Recert", "Compliance/Legal", "P0 Blocker", "Critical",
       "Needs Legal/CDPH Review", "Yes",
       "CDPH provider/course approval + NAC/provider metadata",
       "PHASE_0_COMPLIANCE_FOUNDATION.md",
       "Not in export", "Not in app",
       "n/a",
       "No CDPH approval; no provider/NAC IDs; no approved hours/title",
       "Obtain written CDPH approval + provider metadata before any certificate",
       "13_GO_NO_GO_BLOCKERS_AND_DECISIONS.md", "Program Owner / Legal", "CDPH",
       "None",
       "CNA-Recert-Course\\PHASE_0_COMPLIANCE_FOUNDATION.md",
       "Do not invent CDPH approval, provider IDs, hours, titles."),
    MR("REC-034", "CNA Recert", "Narration/TTS", "P2 Medium", "Medium",
       "Needs Legal/CDPH Review", "No",
       "TTS / cloned-voice narration authorization",
       "16_TTS_NARRATION_PACKAGE.md; ContentV2 narration CSVs",
       "Not in export",
       "narrationManifest.ts present (planning)",
       "n/a",
       "TTS/narration authorization not confirmed",
       "Obtain owner/legal authorization before TTS production",
       "14_DOCUMENTATION_STANDARD_FOR_ALL_PROGRAMS.md", "Program Owner / Legal", "Compliance",
       "None",
       "CNA-Recert-Course\\ContentV2\\narration\\narration_master.csv",
       "TTS is planning-only unless authorization documented."),
    MR("REC-035", "CNA Recert", "Missing Documentation", "P2 Medium", "Medium",
       "In Progress", "No",
       "Missing reconciliation documentation register",
       "This reconciliation task",
       "n/a", "n/a", "n/a",
       "Several reconciliation docs newly created; some marked incomplete pending evidence",
       "Track missing docs + author missing-doc prompts",
       "05_MISSING_DOCUMENTATION_REGISTER.md; CLAUDE_MISSING_DOCUMENTATION_PROMPTS.md",
       "Repo Auditor", "Program Owner",
       "None",
       "CNA-Recert-Course\\reconciliation\\05_MISSING_DOCUMENTATION_REGISTER.md",
       ""),
    MR("REC-040", "CNA Recert", "Spreadsheet URL Evidence", "P1 High", "High",
       "Blocked - Source Access / Copy Error", "No",
       "Resolve native Google Doc export-conversion failures",
       "manifest errors[]",
       "20 native Google Docs + 1 Google Sheet failed docx/xlsx conversion; 1 doc 'No item found'",
       "n/a",
       "These 22 source items are NOT documented locally",
       "Native Google Docs/Sheet could not be converted to Office formats on export",
       "Re-export Google Docs as PDF (FALLBACK already true) or manual download",
       "15_SPREADSHEET_URL_DOCUMENTATION_EVIDENCE_REGISTER.md", "Program Owner", "Repo Auditor",
       "REC-004",
       f"{EXP_PATH}\\manifest.json (errors array)",
       "Conversion error: google-apps.document -> docx not supported by Apps Script export."),
    MR("REC-041", "CNA Recert", "Spreadsheet URL Evidence", "P2 Medium", "Medium",
       "Skipped by Copy Config", "No",
       "35 CNA skill / tutorial videos skipped by config",
       "manifest skipped[]",
       "35 video_file entries skipped (SKIP_VIDEO_FILES=true); 1 large Google Doc (63MB) over MAX_FILE_MB=45",
       "n/a",
       "Clinical skill videos + instructor tutorial doc not in local export",
       "Skill videos + 63MB instructor doc excluded by export config thresholds",
       "Owner decision: separate video transfer vs intentional exclusion",
       "15_SPREADSHEET_URL_DOCUMENTATION_EVIDENCE_REGISTER.md", "Program Owner", "Repo Auditor",
       "REC-004",
       f"{EXP_PATH}\\manifest.json (skipped array)",
       "Videos relate to Optional Clinical Support, not required theory credit."),
    MR("REC-042", "CNA Recert", "Spreadsheet URL Evidence", "P1 High", "Medium",
       "Needs Document Import", "Yes",
       "Admission Exam folder found but not exported",
       "manifest folders_found_but_not_exported[]",
       "Folder 'Admission Exam' (id 1cyDJYkaXaQx...) not copied (INCLUDE_LINKED_FOLDERS=false)",
       "n/a",
       "Admission Exam contents undocumented locally",
       "Folder copy disabled (INCLUDE_LINKED_FOLDERS=false) so contents never enumerated",
       "Owner decision on whether Admission Exam is in scope for recert theory",
       "15_SPREADSHEET_URL_DOCUMENTATION_EVIDENCE_REGISTER.md", "Program Owner", "Repo Auditor",
       "REC-004",
       f"{EXP_PATH}\\manifest.json (folders_found_but_not_exported)",
       "Folder copying disabled to avoid huge ZIP/video exports."),
    # RCFE Admin + Case Manager placeholders
    MR("REC-050", "RCFE Admin", "Program Setup", "P2 Medium", "Medium",
       "Not Started", "No",
       "RCFE Administrator documentation standard (placeholder)",
       "14_DOCUMENTATION_STANDARD_FOR_ALL_PROGRAMS.md (CNA format as template)",
       "No RCFE source folder found in repo",
       "n/a", "n/a",
       "No RCFE Admin source content present locally",
       "Create placeholder tracker rows + missing-doc prompts only; do not invent content",
       "14_DOCUMENTATION_STANDARD_FOR_ALL_PROGRAMS.md", "Program Owner", "RCFE SME",
       "None",
       "n/a - source not present",
       "Do not copy CNA compliance assumptions into RCFE Admin."),
    MR("REC-051", "RCFE Admin", "Source Authority", "P1 High", "High",
       "Not Started", "Yes",
       "RCFE regulatory/source authority crosswalk (placeholder)",
       "Title 22 / CDSS CCLD (to be supplied by owner)",
       "n/a", "n/a", "n/a",
       "No RCFE regulatory source provided; must not be invented",
       "Owner supplies RCFE source authority before content",
       "14_DOCUMENTATION_STANDARD_FOR_ALL_PROGRAMS.md", "Program Owner", "RCFE SME / Legal",
       "REC-050",
       "n/a",
       "Do not invent RCFE regulatory requirements or CE approval."),
    MR("REC-060", "Case Manager", "Program Setup", "P2 Medium", "Medium",
       "Not Started", "No",
       "Case Manager Training documentation standard (placeholder)",
       "14_DOCUMENTATION_STANDARD_FOR_ALL_PROGRAMS.md (CNA format as template)",
       "No Case Manager source folder found in repo",
       "n/a", "n/a",
       "No Case Manager source content present locally",
       "Create placeholder tracker rows + missing-doc prompts only; do not invent content",
       "14_DOCUMENTATION_STANDARD_FOR_ALL_PROGRAMS.md", "Program Owner", "Case Mgmt SME",
       "None",
       "n/a - source not present",
       "Do not copy CNA compliance assumptions into Case Manager training."),
    MR("REC-061", "Case Manager", "Source Authority", "P1 High", "High",
       "Not Started", "Yes",
       "Case Manager regulatory/source authority crosswalk (placeholder)",
       "To be supplied by owner",
       "n/a", "n/a", "n/a",
       "No Case Manager source/authority provided; must not be invented",
       "Owner supplies Case Manager source authority before content",
       "14_DOCUMENTATION_STANDARD_FOR_ALL_PROGRAMS.md", "Program Owner", "Case Mgmt SME / Legal",
       "REC-060",
       "n/a",
       "Do not invent Case Manager regulatory requirements or CE approval."),
    MR("REC-070", "Cross-Program", "Documentation Standard", "P2 Medium", "Medium",
       "In Progress", "No",
       "Reusable 26-artifact documentation standard for all 3 programs",
       "CNA Content packet as format template",
       "n/a", "n/a", "n/a",
       "Standard defined; RCFE/Case Manager columns mostly 'Missing / Needs Source'",
       "Maintain Program Documentation Matrix",
       "14_DOCUMENTATION_STANDARD_FOR_ALL_PROGRAMS.md", "Repo Auditor", "Program Owner",
       "REC-050, REC-060",
       "CNA-Recert-Course\\reconciliation\\14_DOCUMENTATION_STANDARD_FOR_ALL_PROGRAMS.md",
       "CNA format only; no CNA compliance assumptions ported."),
]
print("MASTER_ROWS:", len(MASTER_ROWS))

# ----------------------------------------------------------------------------
# Static data for other sheets
# ----------------------------------------------------------------------------
EXPORT_INVENTORY = [
    ["EXP-01", "Spreadsheet (active export)", "CI-ION - Course Structures - Contents.xlsx",
     ".xlsx", "455,709", "Master source spreadsheet exported to xlsx", "Present", "Parse tabs/URLs"],
    ["EXP-02", "Manifest", "manifest.json", ".json", "53,719",
     "Export provenance: config, links, files, skipped, errors", "Present", "Primary evidence"],
    ["EXP-03", "Source curriculum", "linked-files/cccco-na-model-curriculum-module-1..17.pdf",
     ".pdf", "17 files", "CCCCO NA Model Curriculum modules 1-17", "Present", "SME crosswalk"],
    ["EXP-04", "Syllabus", "linked-files/CI Institute of Nursing - Course Syllabus.pdf",
     ".pdf", "799,354", "Provider syllabus", "Present", "Compliance review"],
    ["EXP-05", "Catalog", "linked-files/CI Institute of Nursing Student Catalog 3-17-2026.docx",
     ".docx", "2,025,744", "Student catalog", "Present", "Compliance review"],
    ["EXP-06", "Final exam", "linked-files/CI-ION NATP Final Exam.pdf",
     ".pdf", "165,675", "Final exam source", "Present", "Keep answer key internal"],
    ["EXP-07", "Pre-orientation test", "linked-files/CI-ION_ Pre-Orientation Student Catalog Test.pdf",
     ".pdf", "98,812", "Pre-orientation/catalog test", "Present", "Map to M00"],
    ["EXP-08", "Skills day materials", "linked-files/PM Day */Day * Module *.pdf",
     ".pdf", "~13 files", "Patient care skills/procedures/rights/charting day decks", "Present", "Optional clinical mapping"],
    ["EXP-09", "Cover art", "linked-files/Cover_*_Text.png",
     ".png", "4 files", "Section cover images", "Present", "Branding only"],
    ["EXP-10", "Live Scan example", "linked-files/Request for Live Scan Service (Example) .pdf",
     ".pdf", "919,190", "Background-check example form", "Present", "No PHI - example only"],
    ["EXP-11", "Linked folder", "Admission Exam (Drive folder)",
     "folder", "n/a", "Folder found but NOT exported (folder copy disabled)", "NOT copied", "Owner decision"],
    ["EXP-12", "Skipped videos", "35 CNA skill + Moodle tutorial videos",
     ".mp4", "n/a", "Skipped (SKIP_VIDEO_FILES=true)", "NOT copied", "Owner decision"],
    ["EXP-13", "Skipped large doc", "Tutorials - Instructor & Coordinator (Google Doc)",
     "gdoc", "63 MB", "Exceeds MAX_FILE_MB=45", "NOT copied", "Owner decision"],
    ["EXP-14", "Failed conversions", "20 native Google Docs + 1 Google Sheet",
     "gdoc/gsheet", "n/a", "docx/xlsx conversion not supported on export", "Copy error", "Re-export as PDF"],
    ["EXP-15", "Missing item", "1 Google Doc id 1mNA503...",
     "gdoc", "n/a", "No item found / no permission at export time", "Copy error", "Access/recover"],
]
EXPORT_INV_COLS = ["ID", "Category", "Item / Pattern", "Type", "Size (bytes)",
                   "Description", "Copy Status", "Action"]

URL_COLS = [
    "URL Evidence ID", "Spreadsheet Name", "Sheet Name", "Row Number", "Cell Address",
    "Spreadsheet Item / Label", "Source URL", "URL Type", "Source Document/File ID",
    "Source Tab/Anchor ID", "Expected Documentation Purpose", "Copy Status",
    "Copied File Name", "Copied File URL", "Destination Folder", "Skipped Reason",
    "Copy Error", "Equivalent Local/Repo Documentation", "Needs New Documentation?",
    "Required Action", "Owner Role", "Priority", "Verification Notes",
]

MISSING_DOCS = [
    ["MD-01", "CNA Recert", "Export inventory", "01_CI_ION_EXPORT_INVENTORY.md", "Created", "P2 Medium",
     "Full export inventory", "Repo Auditor"],
    ["MD-02", "CNA Recert", "Export->Source crosswalk", "02_EXPORT_TO_SOURCE_DOC_CROSSWALK.md", "Created", "P2 Medium",
     "Map export PDFs to repo source docs", "Repo Auditor"],
    ["MD-03", "CNA Recert", "Export->ContentV2 crosswalk", "03_EXPORT_TO_CONTENTV2_CROSSWALK.md", "Created", "P2 Medium",
     "Map export to ContentV2 lessons", "Repo Auditor"],
    ["MD-04", "CNA Recert", "Export->App crosswalk", "04_EXPORT_TO_STANDALONE_APP_CROSSWALK.md", "Created", "P1 High",
     "Map ContentV2 to standalone app", "App Engineer"],
    ["MD-05", "CNA Recert", "URL evidence register", "15_SPREADSHEET_URL_DOCUMENTATION_EVIDENCE_REGISTER.md", "Created", "P1 High",
     "Every URL mapped to status", "Repo Auditor"],
    ["MD-06", "CNA Recert", "Source spreadsheet parse", "01 appendix (positional URL map)", "Needs evidence", "P1 High",
     "Extract sheet/row/cell of each URL", "Repo Auditor"],
    ["MD-07", "CNA Recert", "Re-export failed Google Docs", "linked-files PDFs (22 items)", "Blocked", "P1 High",
     "Convert native docs to PDF/text", "Program Owner"],
    ["MD-08", "CNA Recert", "Certificate approval packet", "PHASE_0 / CDPH submission", "Blocked", "P0 Blocker",
     "CDPH provider/course approval", "Program Owner / Legal"],
    ["MD-09", "RCFE Admin", "Full documentation set", "RCFE Admin Docs sheet", "Needs source", "P2 Medium",
     "26-artifact set (placeholder)", "Program Owner"],
    ["MD-10", "Case Manager", "Full documentation set", "Case Manager Docs sheet", "Needs source", "P2 Medium",
     "26-artifact set (placeholder)", "Program Owner"],
]
MISSING_COLS = ["ID", "Program", "Documentation Area", "Target File / Location",
                "Status", "Priority", "Description", "Owner Role"]

MODULE_TIME = [
    ["M00", "Orientation", 30, "04_THEORY_MODULE_00_ORIENTATION_FULL.md", "Pre-Orientation Catalog Test; Catalog/Syllabus", "In Progress"],
    ["M01", "Infection Control", 90, "05_THEORY_MODULE_01_INFECTION_CONTROL_FULL.md", "No discrete CCCCO module (GAP)", "Needs SME Review"],
    ["M02", "Resident Rights / Abuse Prevention", 120, "24_..._FULL.md", "CCCCO module-17; Day 1/2 Module 2", "In Progress"],
    ["M03", "Dementia / Communication / Cultural Respect", 120, "25_..._FULL.md", "CCCCO module-13", "Needs Source Repair"],
    ["M04", "Mobility / Falls / Workplace Safety", 120, "26_..._FULL.md", "CCCCO module-14; module-12", "In Progress"],
    ["M05", "Nutrition / Skin Integrity / Vital Signs", 120, "27_..._FULL.md", "CCCCO module-10; module-11", "Needs SME Review"],
    ["M06", "Documentation / Change of Condition / Scope", 90, "28_..._FULL.md", "CCCCO module-15; Observation & Charting", "In Progress"],
    ["M07", "Review / Final Exam / Affidavit", 30, "29_..._FULL.md", "CI-ION NATP Final Exam.pdf", "Needs Legal/CDPH Review"],
]
MODULE_TIME_COLS = ["Module", "Title", "Target Minutes", "Repo Source File",
                    "Export Source Evidence", "Status"]

ASSESSMENT = [
    ["AS-01", "Module knowledge checks", "08_MODULE_KNOWLEDGE_CHECK_BLUEPRINT.md", "ContentV2 04_MODULE_ASSESSMENT_MAP", "In Progress", "Each module needs check"],
    ["AS-02", "Interaction check matrix", "12_INTERACTION_CHECK_MATRIX.csv", "n/a", "In Progress", "Verify per-module interaction"],
    ["AS-03", "Final exam blueprint", "09_FINAL_EXAM_BLUEPRINT.md", "CI-ION NATP Final Exam.pdf", "Needs SME Review", "50-question pool"],
    ["AS-04", "Final exam pool (50)", "30_FINAL_EXAM_POOL_50_COMPLETE.md", "examPool.ts", "Needs SME Review", "Item-map to export exam"],
    ["AS-05", "Quiz bank master", "ContentV2 source_copy/csv/31_QUIZ_BANK_MASTER", "n/a", "In Progress", "Keys internal only"],
]
ASSESSMENT_COLS = ["ID", "Assessment Item", "Repo Source", "Export/App Evidence", "Status", "Notes"]

CERT_GATE = [
    ["CG-01", "Identity (legal name + CNA cert #)", "Gate prerequisite", "M00 / CertificateGatePage", "Needs Compliance Review", "DOB excluded unless approved"],
    ["CG-02", "Module completion", "Completion tracking", "ContentV2 / app", "In Progress", ""],
    ["CG-03", "Interaction checks", "Interaction matrix", "app", "In Progress", ""],
    ["CG-04", "Active-time validation", "Active-time plugin / manual review", "Not validated", "Needs Compliance Review", "Not Moodle logs alone"],
    ["CG-05", "Final exam pass", "Exam pool + pass mark", "examPool.ts", "Needs SME Review", "Pass mark = needs verification"],
    ["CG-06", "Affidavit / e-sign", "13_AFFIDAVIT_TEXT.md", "CertificateGatePage", "Needs Legal/CDPH Review", "E-sign acceptance unresolved"],
    ["CG-07", "Admin clearance", "Manual admin review", "app", "In Progress", ""],
    ["CG-08", "Certificate wording", "14_CERTIFICATE_FIELD_MAPPING.csv", "DRAFT only", "Needs Legal/CDPH Review", "BLOCKED - no CDPH approval"],
    ["CG-09", "Provider/NAC metadata", "PHASE_0_COMPLIANCE_FOUNDATION.md", "Not present", "Needs Legal/CDPH Review", "Do not invent"],
]
CERT_GATE_COLS = ["ID", "Gate Control", "Control Doc", "App/Export Evidence", "Status", "Notes"]

OPT_CLIN = [
    ["OC-01", "Clinical Support is separate course section", "OPTIONAL_CLINICAL_HUB_STUB_SPEC", "ClinicalHubPage.tsx", "Needs Compliance Review", "Non-gating"],
    ["OC-02", "Labeled non-credit (not CA clinical-hour)", "03_CLINICAL_SUPPORT_SYLLABUS_TABLE.md", "app labeling", "Needs Compliance Review", "Must be explicit"],
    ["OC-03", "Confidence checks non-graded", "33_OPTIONAL_CLINICAL_CONFIDENCE_CHECKS_COMPLETE.md", "app", "In Progress", ""],
    ["OC-04", "Skill videos = support only", "32_CLINICAL_SUPPORT_FULL_CONTENT.md", "35 videos skipped in export", "Skipped by Copy Config", "Owner decision on transfer"],
    ["OC-05", "Not a certificate prerequisite", "REQUIRED_OPTIONAL_SEPARATION_CHECK", "gate logic", "Needs Compliance Review", "Unless written approval cited"],
]
OPT_CLIN_COLS = ["ID", "Separation Control", "Control Doc", "App/Export Evidence", "Status", "Notes"]

APP_WIRING = [
    ["AW-01", "ContentV2 consumed by app", "contentV2.generated.ts (2.1MB)", "Needs App Verification", "Confirm build uses generated content"],
    ["AW-02", "Module pages", "ModulesPage.tsx / LessonPlayerPage.tsx", "Needs App Verification", "Modified in working tree"],
    ["AW-03", "Exam pool", "examPool.ts", "Needs App Verification", ""],
    ["AW-04", "Certificate gate", "CertificateGatePage.tsx", "Needs App Verification", "Prototype only - not live issuance"],
    ["AW-05", "Clinical hub", "ClinicalHubPage.tsx", "Needs App Verification", "Non-gating"],
    ["AW-06", "PHI warning component", "PhiWarningBlock.tsx", "Needs App Verification", "Must precede uploads/free-text"],
    ["AW-07", "Final assessment splash", "FinalAssessmentSplashPage.tsx", "Needs App Verification", ""],
    ["AW-08", "Narration player", "NarrationPlayer.tsx / narrationManifest.ts", "Needs App Verification", "TTS planning-only"],
]
APP_WIRING_COLS = ["ID", "App Capability", "File Evidence", "Status", "Notes"]

AUDIT_EV = [
    ["AE-01", "Identity capture", "M00 identity fields", "Needs App Verification", "Name + CNA cert #"],
    ["AE-02", "Completion evidence", "Completion tracking export", "In Progress", "Layered, not logs alone"],
    ["AE-03", "Interaction evidence", "Interaction/check logs", "In Progress", ""],
    ["AE-04", "Active-time evidence", "Active-time records", "Needs Compliance Review", "Plugin not validated"],
    ["AE-05", "Exam result evidence", "Exam attempt/pass record", "In Progress", "Keys internal"],
    ["AE-06", "Affidavit evidence", "Signed affidavit record", "Needs Legal/CDPH Review", "E-sign unresolved"],
    ["AE-07", "Admin clearance evidence", "Manual review log", "In Progress", ""],
    ["AE-08", "Export provenance", "manifest.json", "Complete", "Copy provenance for source docs"],
]
AUDIT_EV_COLS = ["ID", "Evidence Type", "Source / Record", "Status", "Notes"]

QA_TESTS = [
    ["QA-01", "Certificate cannot issue without all gates", "Attempt cert with missing affidavit", "Blocked at gate", "Not Started"],
    ["QA-02", "Optional clinical cannot gate certificate", "Skip clinical, attempt cert", "Cert path unaffected", "Not Started"],
    ["QA-03", "PHI warning appears before uploads/free-text", "Open every upload/free-text surface", "Warning + ack shown", "Not Started"],
    ["QA-04", "No PHI fields requested anywhere", "Inspect all forms", "No DOB/MRN/patient data", "Not Started"],
    ["QA-05", "Answer keys not learner-visible", "Inspect exam UI/source", "Keys server/admin only", "Not Started"],
    ["QA-06", "Active-time not satisfied by idle", "Idle through module", "Active-time not credited", "Not Started"],
    ["QA-07", "Partial-credit disclaimer shown", "Review M00 + cert copy", "Online-cap ack present", "Not Started"],
    ["QA-08", "Mobile/accessibility", "Mobile + screen reader pass", "Text alternatives present", "Not Started"],
]
QA_COLS = ["ID", "Negative / Acceptance Test", "Procedure", "Expected Result", "Status"]

# Program documentation matrix (26 artifacts x 3 programs)
DOC_ARTIFACTS = [
    "Executive Summary", "Regulatory / Source Authority Crosswalk", "Syllabus Table",
    "Source Content Audit", "Module Lesson Card Map", "Module Assessment Map",
    "Final Assessment Map", "ContentV2 Schema Alignment Notes",
    "ContentV2 JSON Generation Plan / Generated Package", "Narration Production Guide",
    "Narration Master CSV Specification", "TTS Import CSV Specification",
    "Media Prompt Placeholder Map", "SME Review Flags", "Compliance Review Flags",
    "App Integration Notes", "Certificate / Completion Gate Specification",
    "Active-Time Validation Plan", "Audit Packet / Evidence Retention Plan",
    "Accessibility and Mobile QA Checklist", "Export Reconciliation Tracker",
    "Operations and Support Guide", "Go/No-Go Checklist",
    "Spreadsheet URL Evidence Register", "Source Import / Copy Manifest",
    "Missing Source Access Log",
]
# CNA status per artifact (Exists / Partial / Missing)
CNA_DOC_STATUS = [
    "Exists", "Exists", "Exists", "Exists (ContentV2 01)", "Exists (ContentV2 03)",
    "Exists (ContentV2 04)", "Exists (ContentV2 05)", "Exists (ContentV2 02)",
    "Exists (courseContentV2.json)", "Exists (ContentV2 06)", "Exists (narration_master.csv)",
    "Exists (tts_narration_import.csv)", "Exists (ContentV2 07)", "Exists (18 + ContentV2 08)",
    "Exists (19 + ContentV2 08)", "Exists (ContentV2 09)", "Exists (14 + gate spec)",
    "Partial - not validated", "Partial - this packet", "Exists (17)",
    "Created (this tracker)", "Missing - needs ops guide", "Created (13_GO_NO_GO)",
    "Created (15 register)", "Exists (manifest.json)", "Created (missing-doc register)",
]
PROG_DOC_COLS = ["#", "Documentation Artifact", "CNA Recert", "RCFE Admin", "Case Manager", "Notes"]

GLOSSARY = [
    ["CCCCO", "California Community Colleges Chancellor's Office - source of NA Model Curriculum"],
    ["NATP", "Nurse Assistant Training Program"],
    ["CDPH", "California Department of Public Health - CNA certification authority"],
    ["TPRU", "Training Program Review Unit (CDPH)"],
    ["NAC", "Nurse Assistant Certification / provider identifiers"],
    ["CE", "Continuing Education"],
    ["PHI", "Protected Health Information - must never be collected/displayed"],
    ["TTS", "Text-to-Speech narration (planning-only unless authorized)"],
    ["ContentV2", "Second-generation structured course content (JSON + generated TS)"],
    ["Active-Time", "Validated learner active engagement time (not idle, not logs alone)"],
    ["Optional Clinical Support", "Non-credit, non-gating skills refresh - NOT CA clinical-hour credit"],
    ["Manifest", "manifest.json export provenance file (config, links, files, skipped, errors)"],
    ["Copy Config", "Apps Script export configuration controlling what was/ wasn't copied"],
    ["Source Repair", "Re-deriving incomplete/interrupted module content from authoritative source"],
    ["Go/No-Go", "Decision gate on production readiness (currently NO-GO for certificates)"],
]
GLOSSARY_COLS = ["Term", "Definition"]

RCFE_DOCS = [["RCFE-%02d" % (i + 1), a, "Missing - Needs Source", "P2 Medium",
              "Placeholder. Do not invent RCFE content/regulatory/CE approval."]
             for i, a in enumerate(DOC_ARTIFACTS)]
RCFE_COLS = ["ID", "Documentation Artifact", "Status", "Priority", "Notes"]
CASE_DOCS = [["CM-%02d" % (i + 1), a, "Missing - Needs Source", "P2 Medium",
              "Placeholder. Do not invent Case Manager content/regulatory/CE approval."]
             for i, a in enumerate(DOC_ARTIFACTS)]
CASE_COLS = RCFE_COLS

# ----------------------------------------------------------------------------
# Workbook rendering helpers
# ----------------------------------------------------------------------------
wb = openpyxl.Workbook()


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = header_fill()
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 34


def write_table(ws, cols, rows, widths=None, freeze="A2", start_row=1,
                zebra=True):
    for ci, name in enumerate(cols, start=1):
        ws.cell(row=start_row, column=ci, value=name)
    style_header(ws, len(cols), row=start_row)
    r = start_row + 1
    for row in rows:
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = CELL_FONT
            cell.alignment = WRAP_TOP
            cell.border = BORDER
            if zebra and (r % 2 == 0):
                cell.fill = PatternFill("solid", fgColor=C_BAND)
        ws.row_dimensions[r].height = 30
        r += 1
    if widths:
        for ci, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w
    if freeze:
        ws.freeze_panes = freeze
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(cols))}{r-1}"
    return r - 1  # last data row


def add_status_cf(ws, col_letter, first, last):
    rng = f"{col_letter}{first}:{col_letter}{last}"
    red = PatternFill("solid", fgColor=C_BLOCKER_BG)
    amber = PatternFill("solid", fgColor=C_WARN_BG)
    green = PatternFill("solid", fgColor=C_OK_BG)
    blue = PatternFill("solid", fgColor=C_INFO_BG)
    for kw, fill in [
        ("Blocked", red), ("Needs Legal/CDPH Review", red),
        ("Needs Compliance Review", amber), ("Needs SME Review", amber),
        ("Needs Source Repair", amber), ("Needs Document Import", amber),
        ("Needs Export Verification", blue), ("Needs App Verification", blue),
        ("Skipped by Copy Config", blue), ("In Progress", blue),
        ("Complete", green),
    ]:
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'ISNUMBER(SEARCH("{kw}",{col_letter}{first}))'],
            fill=fill, stopIfTrue=True))


def add_priority_cf(ws, col_letter, first, last):
    rng = f"{col_letter}{first}:{col_letter}{last}"
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="containsText", formula=['"P0"'], fill=PatternFill("solid", fgColor=C_BLOCKER_BG)))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("P0",{col_letter}{first}))'],
        fill=PatternFill("solid", fgColor=C_BLOCKER_BG), font=Font(bold=True, color=C_BLOCKER)))


def add_blocker_cf(ws, col_letter, first, last):
    rng = f"{col_letter}{first}:{col_letter}{last}"
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=['"Yes"'], fill=PatternFill("solid", fgColor=C_BLOCKER_BG),
        font=Font(bold=True, color=C_BLOCKER)))


def add_validation(ws, col_letter, values, first, last):
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(values), allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first}:{col_letter}{last}")


# ----------------------------------------------------------------------------
# Sheet: Master Tracker
# ----------------------------------------------------------------------------
ws_mt = wb.active
ws_mt.title = "Master Tracker"
widths_mt = [9, 13, 20, 11, 9, 24, 8, 30, 32, 38, 26, 22, 34, 38, 34, 20, 18, 16, 38, 12, 12, 38]
last = write_table(ws_mt, MT_COLS, MASTER_ROWS, widths=widths_mt, freeze="C2")
# Column letters
col = {name: get_column_letter(i + 1) for i, name in enumerate(MT_COLS)}
add_priority_cf(ws_mt, col["Priority"], 2, last)
add_status_cf(ws_mt, col["Status"], 2, last)
add_blocker_cf(ws_mt, col["Blocker"], 2, last)
add_validation(ws_mt, col["Status"], STATUS_VALUES, 2, last)
add_validation(ws_mt, col["Priority"], PRIORITY_VALUES, 2, last)
add_validation(ws_mt, col["Program"], PROGRAM_VALUES, 2, last)
add_validation(ws_mt, col["Severity"], SEVERITY_VALUES, 2, last)
add_validation(ws_mt, col["Blocker"], BLOCKER_VALUES, 2, last)
MT_LAST = last

# ----------------------------------------------------------------------------
# Sheet: Spreadsheet URL Evidence
# ----------------------------------------------------------------------------
ws_url = wb.create_sheet("Spreadsheet URL Evidence")
url_table = []
for r in url_rows:
    url_table.append([
        r["uid"], "CI-ION - Course Structures & Contents", "Needs verification",
        "Needs verification", "Needs verification", "Needs verification",
        r["url"], URL_TYPE.get(r["kind"], "External link"), r["id"] or "",
        r["anchor"] or "", r["purpose"], r["status"], r["copied_file"],
        r["copied_url"], r["dest"], r["skipped_reason"], r["copy_error"],
        r["equiv_local"], r["needs_new_doc"], "See Copy Status / Master Tracker",
        "Program Owner", r["priority"], f"From manifest.json links_found; host={r['host']}",
    ])
widths_url = [12, 26, 14, 12, 12, 22, 50, 16, 26, 18, 30, 30, 34, 40, 14, 30, 40, 28, 14, 28, 16, 11, 34]
url_last = write_table(ws_url, URL_COLS, url_table, widths=widths_url, freeze="A2")
ucol = {name: get_column_letter(i + 1) for i, name in enumerate(URL_COLS)}
add_status_cf(ws_url, ucol["Copy Status"], 2, url_last)
add_priority_cf(ws_url, ucol["Priority"], 2, url_last)
# hyperlinks for Source URL
for rr in range(2, url_last + 1):
    c = ws_url[f'{ucol["Source URL"]}{rr}']
    if isinstance(c.value, str) and c.value.startswith("http"):
        c.hyperlink = c.value
        c.font = Font(name="Calibri", size=9, color="1D4ED8", underline="single")

# ----------------------------------------------------------------------------
# Other content sheets
# ----------------------------------------------------------------------------
def simple_sheet(title, cols, rows, widths, freeze="A2"):
    ws = wb.create_sheet(title)
    lastr = write_table(ws, cols, rows, widths=widths, freeze=freeze)
    return ws, lastr

ws_ei, ei_last = simple_sheet("Export Inventory", EXPORT_INV_COLS, EXPORT_INVENTORY,
                              [9, 22, 46, 12, 14, 44, 14, 24])
add_status_cf(ws_ei, get_column_letter(EXPORT_INV_COLS.index("Copy Status") + 1), 2, ei_last)

ws_md, md_last = simple_sheet("Missing Documentation", MISSING_COLS, MISSING_DOCS,
                              [8, 13, 28, 44, 16, 12, 40, 22])
add_status_cf(ws_md, get_column_letter(MISSING_COLS.index("Status") + 1), 2, md_last)
add_priority_cf(ws_md, get_column_letter(MISSING_COLS.index("Priority") + 1), 2, md_last)

ws_tm, tm_last = simple_sheet("CNA Module-Time Map", MODULE_TIME_COLS, MODULE_TIME,
                              [9, 40, 14, 36, 40, 22])
add_status_cf(ws_tm, get_column_letter(MODULE_TIME_COLS.index("Status") + 1), 2, tm_last)
# total row
trow = tm_last + 2
ws_tm.cell(row=trow, column=2, value="TOTAL REQUIRED THEORY (target)").font = Font(bold=True)
ws_tm.cell(row=trow, column=3, value=sum(m[2] for m in MODULE_TIME)).font = Font(bold=True)
ws_tm.cell(row=trow, column=4, value="= 720 min / 12.0 h").font = Font(bold=True, color=C_CARD_ACCENT)

ws_as, as_last = simple_sheet("Assessment Reconciliation", ASSESSMENT_COLS, ASSESSMENT,
                              [9, 30, 40, 34, 22, 30])
add_status_cf(ws_as, get_column_letter(ASSESSMENT_COLS.index("Status") + 1), 2, as_last)

ws_cg, cg_last = simple_sheet("Certificate Gate", CERT_GATE_COLS, CERT_GATE,
                              [9, 36, 34, 30, 26, 36])
add_status_cf(ws_cg, get_column_letter(CERT_GATE_COLS.index("Status") + 1), 2, cg_last)

ws_oc, oc_last = simple_sheet("Optional Clinical Separation", OPT_CLIN_COLS, OPT_CLIN,
                              [9, 40, 34, 30, 26, 32])
add_status_cf(ws_oc, get_column_letter(OPT_CLIN_COLS.index("Status") + 1), 2, oc_last)

ws_aw, aw_last = simple_sheet("App Wiring", APP_WIRING_COLS, APP_WIRING,
                              [9, 30, 36, 24, 34])
add_status_cf(ws_aw, get_column_letter(APP_WIRING_COLS.index("Status") + 1), 2, aw_last)

ws_ae, ae_last = simple_sheet("Audit Evidence", AUDIT_EV_COLS, AUDIT_EV,
                              [9, 28, 34, 24, 30])
add_status_cf(ws_ae, get_column_letter(AUDIT_EV_COLS.index("Status") + 1), 2, ae_last)

ws_qa, qa_last = simple_sheet("QA Negative Tests", QA_COLS, QA_TESTS,
                              [9, 40, 38, 34, 14])
add_status_cf(ws_qa, get_column_letter(QA_COLS.index("Status") + 1), 2, qa_last)

# Program Documentation Matrix
prog_rows = []
for i, art in enumerate(DOC_ARTIFACTS):
    prog_rows.append([i + 1, art, CNA_DOC_STATUS[i], "Missing - Needs Source",
                      "Missing - Needs Source",
                      "RCFE/Case Manager: do not invent; CNA format template only"])
ws_pm, pm_last = simple_sheet("Program Documentation Matrix", PROG_DOC_COLS, prog_rows,
                              [5, 44, 30, 24, 24, 44])

ws_rc, rc_last = simple_sheet("RCFE Admin Docs", RCFE_COLS, RCFE_DOCS,
                              [10, 44, 24, 12, 50])
ws_cm, cm_last = simple_sheet("Case Manager Docs", CASE_COLS, CASE_DOCS,
                              [10, 44, 24, 12, 50])
ws_gl, gl_last = simple_sheet("Glossary", GLOSSARY_COLS, GLOSSARY, [26, 80])

# ----------------------------------------------------------------------------
# Dashboard
# ----------------------------------------------------------------------------
def count_status(substr):
    return sum(1 for r in MASTER_ROWS if substr.lower() in str(r[5]).lower())

def count_exact(substr):
    return sum(1 for r in MASTER_ROWS if str(r[5]) == substr)

def count_prog(p):
    return sum(1 for r in MASTER_ROWS if r[1] == p)

total_rows = len(MASTER_ROWS)
m = {
    "Total tracker rows": total_rows,
    "Completed rows": count_exact("Complete"),
    "Blocked rows": sum(1 for r in MASTER_ROWS if str(r[6]) == "Yes"),
    "P0 blockers": sum(1 for r in MASTER_ROWS if "P0" in str(r[3])),
    "Needs export verification": count_status("Needs Export Verification"),
    "Needs document import": count_status("Needs Document Import"),
    "Blocked - source access/copy errors": count_status("Source Access"),
    "Skipped by copy config": count_status("Skipped by Copy Config"),
    "Needs source repair": count_status("Needs Source Repair"),
    "Needs SME review": count_status("Needs SME Review"),
    "Needs compliance review": count_status("Needs Compliance Review"),
    "Needs legal/CDPH review": count_status("Needs Legal/CDPH Review"),
    "Needs app verification": count_status("Needs App Verification"),
    "Missing documentation count": len(MISSING_DOCS),
    "Total spreadsheet URLs found": URL_METRICS["total"],
    "Copied/imported URL count": URL_METRICS["copied"],
    "Skipped URL count": URL_METRICS["skipped"],
    "Errored URL count": URL_METRICS["errored"],
    "Duplicate/anchor URL count": URL_METRICS["dupe"],
    "Still undocumented URL count": URL_METRICS["undocumented"],
    "External-only URL count": URL_METRICS["external"],
    "Folder-not-copied count": URL_METRICS["folder"],
    "CNA Recert rows": count_prog("CNA Recert"),
    "RCFE Admin rows": count_prog("RCFE Admin"),
    "Case Manager rows": count_prog("Case Manager"),
    "Cross-Program rows": count_prog("Cross-Program"),
}

ws_db = wb.create_sheet("Dashboard")
# Title band
ws_db.merge_cells("A1:F2")
t = ws_db["A1"]
t.value = "CI-ION COURSE RECONCILIATION — MASTER DASHBOARD"
t.font = TITLE_FONT
t.alignment = Alignment(horizontal="left", vertical="center")
for c in range(1, 7):
    ws_db.cell(row=1, column=c).fill = PatternFill("solid", fgColor=C_TITLE_BG)
    ws_db.cell(row=2, column=c).fill = PatternFill("solid", fgColor=C_TITLE_BG)
ws_db.merge_cells("A3:F3")
s = ws_db["A3"]
s.value = (f"Generated {TODAY}  •  Export: CI-ION-course-export-...-20260601-114428  •  "
           f"Status: NOT production-ready — certificate issuance BLOCKED")
s.font = Font(italic=True, size=10, color="B91C1C")
s.alignment = Alignment(horizontal="left", vertical="center")

# Metric cards grid
card_items = list(m.items())
start = 5
cols_per_row = 3
cw = [28, 12, 28, 12, 28, 12]
for i, w in enumerate(cw, start=1):
    ws_db.column_dimensions[get_column_letter(i)].width = w

r = start
i = 0
while i < len(card_items):
    for cpos in range(cols_per_row):
        if i >= len(card_items):
            break
        label, val = card_items[i]
        lc = ws_db.cell(row=r, column=1 + cpos * 2, value=label)
        vc = ws_db.cell(row=r, column=2 + cpos * 2, value=val)
        # color emphasis
        emph = C_CARD_BG
        fg = C_CARD_ACCENT
        if any(k in label for k in ["P0", "Blocked", "legal/CDPH", "undocumented", "Errored"]):
            emph = C_BLOCKER_BG; fg = C_BLOCKER
        elif "Completed" in label or "Copied" in label:
            emph = C_OK_BG; fg = "166534"
        lc.fill = PatternFill("solid", fgColor=emph)
        vc.fill = PatternFill("solid", fgColor=emph)
        lc.font = Font(size=10, bold=True, color="111827")
        vc.font = Font(size=16, bold=True, color=fg)
        lc.alignment = Alignment(wrap_text=True, vertical="center")
        vc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border = BORDER; vc.border = BORDER
        i += 1
    ws_db.row_dimensions[r].height = 38
    r += 1

# URL evidence breakdown table for chart
chart_start = r + 1
ws_db.cell(row=chart_start, column=1, value="URL Evidence Breakdown").font = Font(bold=True, size=12, color=C_SECTION)
hdr = chart_start + 1
ws_db.cell(row=hdr, column=1, value="Category")
ws_db.cell(row=hdr, column=2, value="Count")
url_break = [
    ("Copied/imported", URL_METRICS["copied"]),
    ("Skipped (config)", URL_METRICS["skipped"]),
    ("Copy/convert errors", URL_METRICS["errored"]),
    ("Duplicate/anchor", URL_METRICS["dupe"]),
    ("External only", URL_METRICS["external"]),
    ("Folder not copied", URL_METRICS["folder"]),
]
rr = hdr + 1
for name, val in url_break:
    ws_db.cell(row=rr, column=1, value=name)
    ws_db.cell(row=rr, column=2, value=val)
    rr += 1
for cc in (1, 2):
    ws_db.cell(row=hdr, column=cc).fill = header_fill()
    ws_db.cell(row=hdr, column=cc).font = HEADER_FONT

chart = BarChart()
chart.type = "bar"
chart.title = "Spreadsheet URL Evidence Status"
chart.height = 7
chart.width = 16
data = Reference(ws_db, min_col=2, min_row=hdr, max_row=rr - 1)
cats = Reference(ws_db, min_col=1, min_row=hdr + 1, max_row=rr - 1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.legend = None
ws_db.add_chart(chart, f"D{chart_start}")

# Reorder: Dashboard first
wb.move_sheet("Dashboard", -(len(wb.sheetnames) - 1))
ws_db.sheet_view.showGridLines = False

wb.save(OUT_XLSX)
print("WROTE:", OUT_XLSX)
print("Sheets:", wb.sheetnames)

# ----------------------------------------------------------------------------
# Emit complete URL evidence register markdown (doc 15)
# ----------------------------------------------------------------------------
OUT_DOC15 = HERE / "15_SPREADSHEET_URL_DOCUMENTATION_EVIDENCE_REGISTER.md"
um = URL_METRICS
head = f"""# 15 — Spreadsheet URL Documentation Evidence Register

## Purpose

Map **every** URL discovered in the CI-ION source spreadsheet to a documentation/evidence status.
A URL appearing in the spreadsheet does NOT mean the linked item is documented locally.

## Inputs Reviewed

- `manifest.json` → `links_found` (162 URLs), `files_added`, `skipped`, `errors`,
  `folders_found_but_not_exported`.
- `linked-files\\` directory (copied evidence).

## Method

Each URL is classified by cross-referencing its Google Drive/Docs ID against the manifest's
copied/skipped/error/folder records. Duplicate IDs and alternate `tab=` anchors of the same
document are collapsed to a canonical row; the rest are marked as duplicates.

## Evidence Summary

| Status bucket | Count |
|---|---:|
| Total URLs found | {um['total']} |
| Copied/imported document exists and is mapped | {um['copied']} |
| Skipped (video / oversized — copy config) | {um['skipped']} |
| Copy/convert error (needs extraction or access) | {um['errored']} |
| Duplicate URL / alternate tab anchor | {um['dupe']} |
| External reference only | {um['external']} |
| Folder found but not copied | {um['folder']} |
| **Still requires document import / owner decision** | **{um['undocumented']}** |

## Gaps Found

- Positional columns (sheet name / row / cell) are **Needs verification** until the `.xlsx` is
  parsed (REC-003). The full positional register lives in the `Spreadsheet URL Evidence` sheet of
  `CI_ION_Course_Reconciliation_Master_Tracker.xlsx`.
- {um['errored']} native Google Docs/Sheet links failed Office conversion and are undocumented.
- {um['skipped']} media/oversized items and {um['folder']} folder were skipped by config.

## Owner / Action Needed

- Program Owner: re-export failed Google Docs as PDF; decide on videos/folder.
- Repo Auditor: parse `.xlsx` to backfill sheet/row/cell positions.

## Blocker Status

**Status: Incomplete — evidence missing** for positional columns and {um['undocumented']}
undocumented URLs. Do not treat the spreadsheet URL set as fully documented.

## Next Verification Step

Parse `CI-ION - Course Structures - Contents.xlsx` to attach each URL to its sheet/row/cell.

---

## Full URL Evidence Register ({um['total']} URLs)

| URL Evidence ID | Source URL | URL Type | Source ID | Tab/Anchor | Copy Status | Needs New Doc? | Priority |
|---|---|---|---|---|---|---|---|
"""
lines = [head.rstrip("\n")]
for r in url_rows:
    u = r["url"]
    short = u if len(u) <= 78 else u[:75] + "..."
    lines.append("| {} | {} | {} | {} | {} | {} | {} | {} |".format(
        r["uid"], short, URL_TYPE.get(r["kind"], "External"),
        (r["id"] or "")[:22], r["anchor"] or "", r["status"],
        r["needs_new_doc"], r["priority"]))
OUT_DOC15.write_text("\n".join(lines) + "\n", encoding="utf-8")
print("WROTE:", OUT_DOC15, "rows:", len(url_rows))
if OUT_URL_MD.exists():
    OUT_URL_MD.unlink()

# Save metrics for reuse in markdown authoring
(HERE / "_metrics.json").write_text(json.dumps({
    "url_metrics": URL_METRICS,
    "master_rows": total_rows,
    "dashboard": m,
    "files_added": len(files_added),
    "skipped": len(skipped),
    "errors": len(errors),
    "folders_not_exported": len(folders_not_exp),
}, indent=2), encoding="utf-8")
print("METRICS:", json.dumps(m))



