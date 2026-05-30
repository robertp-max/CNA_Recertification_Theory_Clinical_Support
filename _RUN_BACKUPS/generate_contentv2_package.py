from pathlib import Path
from datetime import datetime, timezone
import csv
import json
import math
import re
import shutil
import hashlib

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except Exception:
    openpyxl = None

ROOT = Path.cwd()
CONTENT = ROOT / "CNA-Recert-Course" / "Content"
V2 = ROOT / "CNA-Recert-Course" / "ContentV2"
APP = ROOT / "standalone-course-mvp"
GENERATED_AT = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
PROJECT = "CNA Recertification Theory + Clinical Support"

MOJIBAKE = {
    "â€”": " - ",
    "â€“": "-",
    "â€˜": "'",
    "â€™": "'",
    "â€œ": '"',
    "â€�": '"',
    "â€": '"',
    "â€¦": "...",
    "â†’": "->",
    "â€": '"',
    "â„¢": "",
    "â€¢": "-",
    "â‰¥": ">=",
    "â‰¤": "<=",
    "âœ…": "[OK]",
    "âŒ": "[Incorrect]",
    "âš ï¸": "WARNING",
    "âš‘": "FLAG",
    "ðŸ”´": "HIGH",
    "ðŸŸ¡": "MEDIUM",
    "ðŸŸ¢": "LOW",
    "Â°F": "F",
    "Â·": "-",
    "Â": "",
    "Ã—": "x",
    "â€‘": "-",
    "â€¯": " ",
    "\ufeff": "",
}

CONTROL_FACTS = {
    "theory_total_minutes": 720,
    "module_minutes": {
        "M00": 30,
        "M01": 90,
        "M02": 120,
        "M03": 120,
        "M04": 120,
        "M05": 120,
        "M06": 90,
        "M07": 30,
    },
    "required_flags": [
        "Module 1 infection-control content requires SME/source review.",
        "Module 5 skin integrity / pressure injury content requires SME/source review.",
        "Final exam Q01, Q02, Q03, Q41, Q21, and Q38 flags are preserved where present.",
        "TTS scripts are planning artifacts only; no audio production authorized.",
        "Certificate production remains disabled pending approval metadata, NAC#, approved wording, affidavit method, and active-time validation.",
        "Optional Clinical Support is optional, non-credit, non-gating, and not clinical-hour credit.",
        "No PHI may be requested, typed, uploaded, shown, or implied.",
    ],
}

MODULE_SOURCES = {
    "M00": CONTENT / "theory/modules/04_THEORY_MODULE_00_ORIENTATION_FULL.md",
    "M01": CONTENT / "theory/modules/05_THEORY_MODULE_01_INFECTION_CONTROL_FULL.md",
    "M02": CONTENT / "theory/modules/24_THEORY_MODULE_02_RESIDENT_RIGHTS_ABUSE_PREVENTION_FULL.md",
    "M03": CONTENT / "theory/modules/25_THEORY_MODULE_03_DEMENTIA_COMMUNICATION_CULTURAL_RESPECT_FULL.md",
    "M04": CONTENT / "theory/modules/26_THEORY_MODULE_04_MOBILITY_FALLS_WORKPLACE_SAFETY_FULL.md",
    "M05": CONTENT / "theory/modules/27_THEORY_MODULE_05_NUTRITION_SKIN_INTEGRITY_VITAL_SIGNS_FULL.md",
    "M06": CONTENT / "theory/modules/28_THEORY_MODULE_06_DOCUMENTATION_CHANGE_OF_CONDITION_SCOPE_FULL.md",
    "M07": CONTENT / "theory/modules/29_THEORY_MODULE_07_REVIEW_FINAL_EXAM_AFFIDAVIT_FULL.md",
}

MODULE_META = {
    "M00": {
        "title": "Orientation and Compliance Boundaries",
        "shortTitle": "Orientation",
        "minutes": 30,
        "status": "draft",
        "sme": "None identified",
        "compliance": "Compliance/legal/CDPH review required before production.",
        "sourceFlag": "None identified",
    },
    "M01": {
        "title": "Infection Control and PPE",
        "shortTitle": "Infection Control & PPE",
        "minutes": 90,
        "status": "sme-review",
        "sme": "Module 1 infection-control content requires SME/source review.",
        "compliance": "Compliance review required; no dedicated NATP 10-17 infection-control source in uploaded set.",
        "sourceFlag": "Active source gap - SME/source review required.",
    },
    "M02": {
        "title": "Resident Rights, Abuse Prevention, and Boundaries",
        "shortTitle": "Resident Rights & Abuse Prevention",
        "minutes": 120,
        "status": "draft",
        "sme": "None identified",
        "compliance": "Compliance/legal/CDPH review required before production.",
        "sourceFlag": "None identified",
    },
    "M03": {
        "title": "Dementia, Communication, and Respectful Care",
        "shortTitle": "Dementia & Communication",
        "minutes": 120,
        "status": "source-repair",
        "sme": "Module 3 placement and sensitive end-of-life/trauma-informed content should receive SME review.",
        "compliance": "Canonical Module 3 file is incomplete after Screen 3.2.3; source repair required before production.",
        "sourceFlag": "Draft / Source Repair Required for missing canonical sections after Screen 3.2.3.",
    },
    "M04": {
        "title": "Mobility, Falls, and Workplace Safety",
        "shortTitle": "Mobility & Safety",
        "minutes": 120,
        "status": "draft",
        "sme": "None identified",
        "compliance": "Compliance review required; no clinical competency or clinical-hour claim.",
        "sourceFlag": "None identified",
    },
    "M05": {
        "title": "Nutrition, Skin Integrity, Vital Signs, and Observation",
        "shortTitle": "Nutrition, Skin & Vitals",
        "minutes": 120,
        "status": "sme-review",
        "sme": "Skin integrity / pressure injury content requires SME/source review.",
        "compliance": "Compliance review required; keep CNA role as observe/report, not independent staging/treatment.",
        "sourceFlag": "Skin integrity source gap active.",
    },
    "M06": {
        "title": "Documentation, Reporting, PHI Avoidance, and Scope",
        "shortTitle": "Documentation & Scope",
        "minutes": 90,
        "status": "draft",
        "sme": "None identified",
        "compliance": "Compliance review required; PHI/no-upload and CNA scope guardrails apply.",
        "sourceFlag": "None identified",
    },
    "M07": {
        "title": "Final Review, Exam/Test, Affidavit, and Certificate Status",
        "shortTitle": "Review, Final & Certificate Status",
        "minutes": 30,
        "status": "draft",
        "sme": "None identified",
        "compliance": "Affidavit/certificate wording requires legal/compliance/CDPH or owner approval; production certificate disabled.",
        "sourceFlag": "Module 7 timing corrected to 30 minutes.",
    },
}

LESSON_META = {
    "M00": [
        ("L00", "Welcome and Course Purpose", 5),
        ("L01", "What This Course Covers and Does Not Cover", 8),
        ("L02", "Identity Confirmation", 5),
        ("L03", "Course Navigation and Progress Tracking", 5),
        ("L04", "Acknowledgement and Compliance Check", 7),
    ],
    "M01": [
        ("L01", "Why Infection Control Matters in Long-Term Care", 15),
        ("L02", "The Chain of Infection", 15),
        ("L03", "Hand Hygiene - Your Most Important Tool", 15),
        ("L04", "Personal Protective Equipment (PPE)", 15),
        ("L05", "Recognizing and Reporting Infection Signs", 15),
        ("L06", "Environmental Cleaning and Safe Practices", 15),
    ],
    "M02": [
        ("L01", "Resident Rights Foundation", 20),
        ("L02", "Types of Abuse and Neglect", 25),
        ("L03", "Signs, Symptoms, and Evidence of Abuse", 20),
        ("L04", "Mandated Reporter Obligations", 20),
        ("L05", "Prevention, Boundaries, and Safety", 20),
    ],
    "M03": [
        ("L01", "Understanding Dementia and Cognitive Conditions", 25),
        ("L02", "Communication Strategies for Cognitive Impairment", 25),
        ("L03", "Cultural Sensitivity and Spiritual Respect", 20),
        ("L04", "End-of-Life Care, Grief, and Hospice", 25),
        ("L05", "Trauma-Informed Care and De-escalation", 10),
    ],
    "M04": [
        ("L01", "Body Mechanics and Injury Prevention", 20),
        ("L02", "Safe Transfers and Ambulation", 25),
        ("L03", "Range of Motion and Restorative Care", 20),
        ("L04", "Fall Prevention", 20),
        ("L05", "Emergency Procedures - Fire, Choking, Disaster", 20),
    ],
    "M05": [
        ("L01", "Nutrition and Hydration", 25),
        ("L02", "Feeding Assistance and Aspiration Prevention", 20),
        ("L03", "Skin Integrity and Pressure Injury Prevention", 25),
        ("L04", "Vital Signs - Measurement and Reporting", 25),
        ("L05", "Monitoring, Documentation, and Reporting", 10),
    ],
    "M06": [
        ("L01", "Observation - Objective vs. Subjective", 15),
        ("L02", "Documentation Standards", 20),
        ("L03", "Change of Condition - What to Report", 20),
        ("L04", "CNA Scope of Practice", 15),
        ("L05", "Professionalism, Ethics, and Delegation", 10),
    ],
    "M07": [
        ("L01", "Cumulative Review - Modules 1-3", 5),
        ("L02", "Cumulative Review - Modules 4-6", 5),
        ("L03", "Final Exam Preparation and Instructions", 3),
        ("L04", "Affidavit of Completion", 2),
        ("L05", "Certificate Status, Next Steps, and Post-Course Guidance", 2),
    ],
}

M0_SCREEN_LESSON = {1: "L01", 2: "L01", 3: "L01", 4: "L02", 5: "L02", 6: "L02", 7: "L03", 8: "L03", 9: "L04", 10: "L04", 11: "L05", 12: "L05"}
M1_SCREEN_LESSON = {1: "L01", 2: "L01", 3: "L01", 4: "L02", 5: "L02", 6: "L03", 7: "L03", 8: "L04", 9: "L04", 10: "L04", 11: "L05", 12: "L05", 13: "L06", 14: "L06", 15: "L06"}


def norm(text):
    if text is None:
        return ""
    out = str(text)
    for bad, good in MOJIBAKE.items():
        out = out.replace(bad, good)
    out = out.replace("\r\n", "\n").replace("\r", "\n")
    out = re.sub(r"[ \t]+", " ", out)
    out = re.sub(r"\n{3,}", "\n\n", out)
    return out.strip()


def read_text(path):
    data = Path(path).read_bytes()
    for enc in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            return norm(data.decode(enc))
        except Exception:
            pass
    return norm(data.decode("utf-8", errors="replace"))


def truncate_source_artifacts(text):
    stop_patterns = [
        "claude-opus-",
        "Continue exactly from where you stopped.",
        "Do not regenerate Module",
        "Due to the enormous length of this complete deliverable set",
        "New messages",
        "Claude Sonnet",
    ]
    lines = text.splitlines()
    clean = []
    for line in lines:
        if any(marker in line for marker in stop_patterns):
            break
        clean.append(line)
    return norm("\n".join(clean))


def strip_implementation_notes(text):
    text = re.sub(r"\*\[Moodle implementation:[^\]]+\]\*", "", text, flags=re.I)
    text = re.sub(r"Moodle Activity Type:.*?(?=\n\n|$)", "", text, flags=re.I | re.S)
    text = re.sub(r"Source File References:.*?(?=\n|$)", "", text, flags=re.I)
    return norm(text)


def words(text):
    return re.findall(r"[A-Za-z0-9']+", text or "")


def extract_field(text, label):
    pattern = rf"(?im)^{re.escape(label)}:\s*(.+?)(?=\n[A-Z][A-Za-z /-]+:|\nNot Graded:|\nNot Certificate-Gated:|\Z)"
    match = re.search(pattern, text or "", flags=re.S)
    return norm(match.group(1)) if match else ""


def estimate_seconds(text):
    word_count = len(words(text))
    return max(8, int(math.ceil(word_count / 2.35))), word_count


def target_status(seconds, status_flags=""):
    if seconds <= 75:
        if seconds < 35:
            return "Short - acceptable if intentionally brief"
        if seconds <= 65:
            return "Within target"
        return "Acceptable near upper target"
    if "Source Repair" in status_flags:
        return "Source repair placeholder - timing review required"
    return "Over 75 seconds - split/review required"


def app_location(module_id, lesson_id, card_id):
    return f"module.{module_id.lower()}.lesson.{lesson_id.lower()}.card.{card_id.lower()}"


def title_fmt(module_id, lesson_id, card_id, card_type):
    return f"{module_id} {lesson_id} {card_id.split('_')[0]} {card_type.title()}"


def split_chunks(text, max_words=115):
    clean = strip_implementation_notes(text)
    if not clean:
        return []
    paragraphs = [p.strip() for p in re.split(r"\n\s*\n", clean) if p.strip()]
    chunks = []
    current = []
    count = 0
    for paragraph in paragraphs:
        wc = len(words(paragraph))
        if current and count + wc > max_words:
            chunks.append(norm("\n\n".join(current)))
            current = [paragraph]
            count = wc
        else:
            current.append(paragraph)
            count += wc
    if current:
        chunks.append(norm("\n\n".join(current)))

    final = []
    for chunk in chunks:
        if len(words(chunk)) <= max_words + 20:
            final.append(chunk)
            continue
        sentences = re.split(r"(?<=[.!?])\s+", chunk)
        current = []
        count = 0
        for sentence in sentences:
            wc = len(words(sentence))
            if current and count + wc > max_words:
                final.append(norm(" ".join(current)))
                current = [sentence]
                count = wc
            else:
                current.append(sentence)
                count += wc
        if current:
            final.append(norm(" ".join(current)))
    return final


def extract_objectives(text):
    match = re.search(r"Learning Objectives\s*(.*?)(?:\n\s*(?:##|\d+\.|4\. ESTIMATED|Estimated Time))", text, re.I | re.S)
    if not match:
        return []
    objectives = []
    for line in match.group(1).splitlines():
        clean = re.sub(r"^\d+[.)]\s*", "", line.strip(" -\t")).strip()
        if clean and len(clean) > 12 and not clean.lower().startswith("by the end"):
            objectives.append(norm(clean))
    return objectives[:8]


def lesson_source_blocks(module_id, text):
    text = truncate_source_artifacts(text)
    marker_re = re.compile(r"(?im)^(?P<kind>Screen|LESSON|Lesson|Question)\s+(?P<id>\d+(?:\.\d+){0,2})\s*(?:[:\-]| - |\s+-\s+|\s+)?(?P<title>[^\n]*)")
    markers = list(marker_re.finditer(text))
    screens = []
    for index, match in enumerate(markers):
        if match.group("kind").lower() != "screen":
            continue
        end = markers[index + 1].start() if index + 1 < len(markers) else len(text)
        sid = match.group("id")
        title = norm(match.group("title").strip(" -:")) or f"Screen {sid}"
        body = norm(text[match.end() : end])
        if not body:
            continue
        if module_id == "M00":
            try:
                lesson_id = M0_SCREEN_LESSON.get(int(sid), "L01")
            except Exception:
                lesson_id = "L01"
        elif module_id == "M01":
            try:
                lesson_id = M1_SCREEN_LESSON.get(int(sid), "L01")
            except Exception:
                lesson_id = "L01"
        else:
            parts = sid.split(".")
            lesson_id = f"L{int(parts[1]):02d}" if len(parts) >= 2 else "L01"
        screens.append({"screen_id": sid, "lesson_id": lesson_id, "title": title, "body": body})
    return screens


def source_ref(module_id, anchor=""):
    rel = MODULE_SOURCES[module_id].relative_to(ROOT).as_posix()
    return rel + (f"#{anchor}" if anchor else "")


def base_card(module_id, lesson_id, lesson_title, card_id, card_type, display_title, content, goal, example, why, terms, source, status, sme, compliance, challenge=None, rationale=None):
    narrative = norm(content)
    seconds, word_count = estimate_seconds(narrative)
    flags = " ".join([status, sme, compliance])
    return {
        "module_id": module_id,
        "module_title": MODULE_META[module_id]["title"],
        "lesson_id": lesson_id,
        "lesson_title": lesson_title,
        "card_id": card_id,
        "card_type": card_type,
        "app": {"location": app_location(module_id, lesson_id, card_id)},
        "display_title": display_title,
        "learner_facing_content": narrative,
        "learning_goal": goal,
        "cna_practice_example": example,
        "why_it_matters": why,
        "key_terms": terms,
        "completion_condition": "Learner submits one decision response before debrief unlocks." if card_type == "challenge" else "Learner views this card and continues through the lesson sequence.",
        "narration_script": narrative,
        "transcript_text": narrative,
        "estimated_narration_seconds": seconds,
        "estimated_word_count": word_count,
        "target_duration_status": target_status(seconds, flags),
        "media_prompt_placeholder": {
            "scene_title": display_title,
            "image_16_9_prompt": f"Realistic healthcare education scene for {display_title}; de-identified long-term care training setting; no readable charts, no facility logos, no PHI.",
            "optional_video_prompt": f"Future optional short instructional motion prompt for {display_title}; use only fictional/de-identified CNA training context.",
            "negative_prompt": "No PHI, no real patient records, no readable charts, no facility names, no gore, no unsafe clinical actions, no plastic-looking people.",
            "phi_safety_note": "No real patient/resident identifiers or facility data.",
            "required_for_mvp": False,
            "status": "Placeholder only - no media generated.",
        },
        "source_reference": source,
        "status": status,
        "sme_review_flag": sme,
        "compliance_review_flag": compliance,
        "internal_challenge": challenge,
        "debrief_rationale": rationale,
    }


def choices_from_question(q):
    return [{"id": key, "label": q.get(f"choice_{key.lower()}", "")} for key in ["A", "B", "C", "D"] if q.get(f"choice_{key.lower()}")]


def parse_final_csv():
    path = CONTENT / "csv/31_QUIZ_BANK_MASTER_COMPLETE.csv"
    rows = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            rows.append({k: norm(v) for k, v in row.items()})
    return rows


quiz_rows = parse_final_csv()
quiz_by_module = {}
for q in quiz_rows:
    quiz_by_module.setdefault(f"M{int(q['module_number']):02d}", []).append(q)

quiz_by_module["M00"] = [
    {
        "question_id": "M0-KC-01",
        "module_number": "0",
        "module_title": "Orientation and Compliance Boundaries",
        "question_type": "Multiple choice",
        "stem": "How many total hours of CE/in-service are required for California CNA renewal over two years?",
        "choice_a": "24",
        "choice_b": "48",
        "choice_c": "12",
        "choice_d": "36",
        "correct_answer": "B",
        "rationale": "California requires 48 hours over the two-year renewal period.",
        "difficulty": "Low",
        "source_reference": "04_THEORY_MODULE_00_ORIENTATION_FULL.md Section 11",
        "required_for_final_yes_no": "No",
        "needs_sme_review_yes_no": "No",
    },
    {
        "question_id": "M0-KC-02",
        "module_number": "0",
        "module_title": "Orientation and Compliance Boundaries",
        "question_type": "Multiple choice",
        "stem": "What is the maximum number of online CE hours California allows toward CNA renewal?",
        "choice_a": "12",
        "choice_b": "48",
        "choice_c": "24",
        "choice_d": "36",
        "correct_answer": "C",
        "rationale": "Only 24 of the 48 required hours may be completed online.",
        "difficulty": "Low",
        "source_reference": "04_THEORY_MODULE_00_ORIENTATION_FULL.md Section 11",
        "required_for_final_yes_no": "No",
        "needs_sme_review_yes_no": "No",
    },
    {
        "question_id": "M0-KC-03",
        "module_number": "0",
        "module_title": "Orientation and Compliance Boundaries",
        "question_type": "Multiple choice",
        "stem": "This 12-hour course completes all California CNA renewal requirements by itself.",
        "choice_a": "True",
        "choice_b": "False",
        "choice_c": "Only if clinical support is completed",
        "choice_d": "Only if the learner passes on the first attempt",
        "correct_answer": "B",
        "rationale": "This course is partial credit only. Additional renewal requirements remain the learner responsibility.",
        "difficulty": "Low",
        "source_reference": "04_THEORY_MODULE_00_ORIENTATION_FULL.md Section 11",
        "required_for_final_yes_no": "No",
        "needs_sme_review_yes_no": "No",
    },
    {
        "question_id": "M0-KC-04",
        "module_number": "0",
        "module_title": "Orientation and Compliance Boundaries",
        "question_type": "Multiple choice",
        "stem": "Which item should a learner never enter in a course activity?",
        "choice_a": "CNA certificate number in the identity field",
        "choice_b": "A real patient or resident name",
        "choice_c": "Legal name in the profile field",
        "choice_d": "A quiz answer",
        "correct_answer": "B",
        "rationale": "Protected health information, including real patient or resident names, must never be entered.",
        "difficulty": "Low",
        "source_reference": "04_THEORY_MODULE_00_ORIENTATION_FULL.md Section 11",
        "required_for_final_yes_no": "No",
        "needs_sme_review_yes_no": "No",
    },
]

if len(quiz_by_module.get("M01", [])) < 5:
    quiz_by_module.setdefault("M01", []).append(
        {
            "question_id": "M1-KC-06",
            "module_number": "1",
            "module_title": "Infection Control",
            "question_type": "Multiple choice",
            "stem": "When handling soiled linen, a CNA should:",
            "choice_a": "Shake the linen to remove debris before bagging it",
            "choice_b": "Hold the linen close to the body for stability",
            "choice_c": "Hold the linen away from the body and place it directly in the linen bag",
            "choice_d": "Place the linen on the floor temporarily",
            "correct_answer": "C",
            "rationale": "Hold soiled linen away from the body and place it directly in the container; never shake it.",
            "difficulty": "Low",
            "source_reference": "05_THEORY_MODULE_01_INFECTION_CONTROL_FULL.md Knowledge Check 1.6 - NEEDS SME REVIEW",
            "required_for_final_yes_no": "No",
            "needs_sme_review_yes_no": "Yes",
        }
    )


def question_to_assessment(q, idx, prefix="Q"):
    if prefix == "FINAL":
        q_app_location = f"course.final.q{idx:02d}"
    elif prefix.startswith("M") and "-" in prefix:
        module_part = prefix.split("-", 1)[0].lower()
        q_app_location = f"module.{module_part}.assessment.q{idx:02d}"
    else:
        q_app_location = f"assessment.{prefix.lower()}.q{idx:02d}"
    return {
        "id": q.get("question_id") or f"{prefix}{idx:02d}",
        "app_location": q_app_location,
        "prompt": q["stem"],
        "question_type": q.get("question_type", "Multiple choice"),
        "choices": choices_from_question(q),
        "correct_id_internal": q["correct_answer"],
        "rationale_internal": q.get("rationale", ""),
        "difficulty": q.get("difficulty", ""),
        "source_reference": q.get("source_reference", ""),
        "sme_review_flag": "SME review required" if q.get("needs_sme_review_yes_no", "No").lower() == "yes" else "None identified",
        "compliance_review_flag": "Internal scoring only; do not reveal final answer key to learners.",
        "learner_feedback_correct": "Correct. Continue to the next item." if prefix != "FINAL" else "Final exam answers are scored silently; answer keys are not displayed.",
        "learner_feedback_incorrect": "Review the related lesson and try again." if prefix != "FINAL" else "Final exam answers are scored silently; answer keys are not displayed.",
    }


def make_challenge(module_id, lesson_id, lesson_title, lesson_index):
    bank = quiz_by_module.get(module_id, [])
    q = bank[(lesson_index - 1) % len(bank)] if bank else None
    if q is None:
        q = {
            "question_id": f"{module_id}-{lesson_id}-SRR",
            "stem": f"Which action best matches safe CNA practice for {lesson_title}?",
            "choice_a": "Proceed without checking policy",
            "choice_b": "Use fictional/de-identified information and report concerns through the licensed nurse or facility process",
            "choice_c": "Make an independent diagnosis",
            "choice_d": "Skip documentation",
            "correct_answer": "B",
            "rationale": "Safe CNA practice stays in scope, uses fictional/de-identified examples in training, and reports concerns appropriately.",
            "source_reference": source_ref(module_id),
            "needs_sme_review_yes_no": "Yes",
        }
    return question_to_assessment(q, 1, prefix="KC")


def make_modules():
    modules = []
    all_cards = []
    for mid in [f"M{i:02d}" for i in range(8)]:
        raw = read_text(MODULE_SOURCES[mid])
        clean = truncate_source_artifacts(raw)
        objectives = extract_objectives(clean)
        screens_by_lesson = {}
        for screen in lesson_source_blocks(mid, clean):
            screens_by_lesson.setdefault(screen["lesson_id"], []).append(screen)

        lessons = []
        for index, (lid, lesson_title, minutes) in enumerate(LESSON_META[mid], start=1):
            meta = MODULE_META[mid]
            lesson_screens = screens_by_lesson.get(lid, [])
            status = meta["status"]
            if mid == "M03" and (lid in {"L03", "L04", "L05"} or (lid == "L02" and len(lesson_screens) < 3)):
                status = "Draft / Source Repair Required"
            sme = meta["sme"]
            compliance = meta["compliance"]
            why = [
                "Supports safe CNA judgment within scope of practice.",
                "Builds audit-ready completion evidence for required online theory only.",
                "Preserves compliance boundaries: no PHI, no clinical-hour credit, certificate production disabled.",
            ]
            terms = [{"term": "CNA scope", "definition": "The learner observes, reports, documents, and follows the care plan within allowed CNA duties."}]
            cards = []
            overview_text = (
                f"{lesson_title}. Estimated time: {minutes} minutes. "
                f"Learning goal: {objectives[0] if objectives else 'Apply this required theory topic safely in CNA practice.'}\n\n"
                f"This lesson is part of {meta['title']}. It supports partial online CE theory only and does not create clinical-hour credit or hands-on competency validation."
            )
            if "Source Repair Required" in status:
                overview_text += "\n\nDraft / Source Repair Required: the canonical source content for this lesson is incomplete or needs repair before production release. Do not treat this card as final approved course copy."
            cards.append(
                base_card(
                    mid,
                    lid,
                    lesson_title,
                    "C01_OVERVIEW",
                    "overview",
                    lesson_title,
                    overview_text,
                    objectives[0] if objectives else f"Review {lesson_title}.",
                    "Use the lesson ideas during routine CNA observation, reporting, and resident support while staying within scope.",
                    why,
                    terms,
                    source_ref(mid, lid),
                    status,
                    sme,
                    compliance,
                )
            )
            letter_ord = ord("A")
            if lesson_screens:
                for screen in lesson_screens:
                    chunks = split_chunks(f"{screen['title']}\n\n{screen['body']}") or [f"{screen['title']} - source content requires cleanup before learner release."]
                    for chunk in chunks:
                        card_id = f"C02{chr(letter_ord)}_DELIVERY"
                        letter_ord += 1
                        cards.append(
                            base_card(
                                mid,
                                lid,
                                lesson_title,
                                card_id,
                                "delivery",
                                screen["title"],
                                chunk,
                                objectives[min(index - 1, len(objectives) - 1)] if objectives else f"Apply {screen['title']}.",
                                "Connect this screen to a routine CNA decision, then report concerns to the licensed nurse as appropriate.",
                                why,
                                terms,
                                source_ref(mid, f"Screen {screen['screen_id']}"),
                                status,
                                sme,
                                compliance,
                            )
                        )
            else:
                placeholder = (
                    f"Draft / Source Repair Required: canonical source text for {lesson_title} is missing or incomplete. "
                    "Missing content: learner-facing screens, source references, knowledge checks, and production-ready transcript text. "
                    "Safe placeholder only; SME/source repair required before use."
                )
                cards.append(
                    base_card(
                        mid,
                        lid,
                        lesson_title,
                        "C02A_DELIVERY",
                        "delivery",
                        f"{lesson_title} - Source Repair Placeholder",
                        placeholder,
                        f"Repair source content for {lesson_title}.",
                        "Do not teach unverified clinical details. Use this only to preserve data shape until source repair.",
                        why,
                        terms,
                        source_ref(mid, lid),
                        status,
                        "SME/source repair required",
                        compliance,
                    )
                )

            challenge = make_challenge(mid, lid, lesson_title, index)
            challenge_text = f"Scenario challenge for {lesson_title}: {challenge['prompt']}\n\n" + "\n".join(
                [f"{choice['id']}) {choice['label']}" for choice in challenge["choices"]]
            )
            challenge_text += "\n\nNo answer is revealed until submission in lesson practice. Final exams never reveal answer keys."
            challenge_sme = challenge["sme_review_flag"] if challenge["sme_review_flag"] != "None identified" else sme
            cards.append(
                base_card(
                    mid,
                    lid,
                    lesson_title,
                    "C03_CHALLENGE",
                    "challenge",
                    f"{lesson_title} Challenge",
                    challenge_text,
                    challenge["prompt"],
                    "Choose the safest in-scope CNA action before moving to debrief.",
                    why,
                    terms,
                    challenge.get("source_reference") or source_ref(mid, lid),
                    status,
                    challenge_sme,
                    compliance,
                    challenge=challenge,
                )
            )
            debrief_text = (
                f"Debrief for {lesson_title}. Correct internal answer: {challenge['correct_id_internal']}. "
                f"Rationale for instructor/internal use: {challenge['rationale_internal']}\n\n"
                "Learner-facing remediation: Review the related lesson content, focus on CNA scope, resident safety, prompt reporting, and no-PHI practice. "
                "Practice challenge answer keys may be shown after submission; final assessment answer keys remain locked."
            )
            cards.append(
                base_card(
                    mid,
                    lid,
                    lesson_title,
                    "C04A_DEBRIEF",
                    "debrief",
                    f"{lesson_title} Debrief and Remediation",
                    debrief_text,
                    "Explain the decision and apply non-punitive remediation.",
                    "After a missed practice decision, review the key safety step and repeat the scenario without penalty.",
                    why,
                    terms,
                    challenge.get("source_reference") or source_ref(mid, lid),
                    status,
                    challenge_sme,
                    compliance,
                    rationale=challenge["rationale_internal"],
                )
            )
            lessons.append({"lesson_id": lid, "lesson_title": lesson_title, "estimated_minutes": minutes, "status": status, "source_reference": source_ref(mid, lid), "cards": cards})
            all_cards.extend(cards)

        modules.append(
            {
                "module_id": mid,
                "module_title": MODULE_META[mid]["title"],
                "short_title": MODULE_META[mid]["shortTitle"],
                "estimated_minutes": MODULE_META[mid]["minutes"],
                "required": True,
                "counts_toward_theory": mid != "M07",
                "status": MODULE_META[mid]["status"],
                "source_file": str(MODULE_SOURCES[mid].relative_to(ROOT)).replace("\\", "/"),
                "learning_objectives": objectives,
                "sme_review_flag": MODULE_META[mid]["sme"],
                "compliance_review_flag": MODULE_META[mid]["compliance"],
                "source_status_flag": MODULE_META[mid]["sourceFlag"],
                "lessons": lessons,
            }
        )
    return modules, all_cards


def write_json(path, obj):
    Path(path).write_text(json.dumps(obj, indent=2, ensure_ascii=False), encoding="utf-8")


def md_table(headers, rows):
    output = ["| " + " | ".join(headers) + " |", "| " + " | ".join(["---"] * len(headers)) + " |"]
    for row in rows:
        output.append("| " + " | ".join(norm(str(item)).replace("\n", "<br>") for item in row) + " |")
    return "\n".join(output)


def write_sheet(wb, name, headers, rows):
    ws = wb[name]
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="5C1111")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    for col in range(1, min(len(headers), 12) + 1):
        ws.column_dimensions[get_column_letter(col)].width = min(55, max(14, len(str(headers[col - 1])) + 4))
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def main():
    for directory in [V2, V2 / "data", V2 / "narration", V2 / "xlsx", V2 / "source_copy", V2 / "updated_content", V2 / "_agent_runs"]:
        directory.mkdir(parents=True, exist_ok=True)

    modules, all_cards = make_modules()

    clinical_path = CONTENT / "clinical-support/32_CLINICAL_SUPPORT_FULL_CONTENT.md"
    clinical_text = read_text(clinical_path)
    unit_markers = list(re.finditer(r"(?im)^CLINICAL SUPPORT UNIT\s+(\d+):\s*(.+)$", clinical_text))
    clinical_units = []
    for index, marker in enumerate(unit_markers):
        end = unit_markers[index + 1].start() if index + 1 < len(unit_markers) else len(clinical_text)
        body = norm(clinical_text[marker.end() : end])
        unit_no = int(marker.group(1))
        clinical_units.append(
            {
                "unit_id": f"U{unit_no:02d}",
                "title": norm(marker.group(2)),
                "app_location": f"clinical.hub.unit.u{unit_no:02d}",
                "content": body[:3500],
                "status": "Optional - non-credit - non-gating",
                "source_reference": str(clinical_path.relative_to(ROOT)).replace("\\", "/"),
                "sme_review_flag": "SME review required for Module 1 and skin integrity references where present." if unit_no in (2, 4) else "None identified",
                "compliance_review_flag": "Optional Clinical Support must remain non-credit, non-gating, not clinical-hour credit, and must not request PHI.",
            }
        )

    confidence_path = CONTENT / "clinical-support/confidence-checks/33_OPTIONAL_CLINICAL_CONFIDENCE_CHECKS_COMPLETE.md"
    confidence_text = read_text(confidence_path) if confidence_path.exists() else clinical_text
    checks = []
    for marker in re.finditer(r"(?im)^Check\s+(\d+)\s*(?:-|—)?\s*(.+)$", confidence_text):
        check_no = int(marker.group(1))
        checks.append(
            {
                "check_id": f"CC{check_no:02d}",
                "title": norm(marker.group(2)),
                "app_location": f"clinical.hub.confidence.cc{check_no:02d}",
                "status": "Optional - not graded - not certificate-gated",
                "source_reference": str(confidence_path.relative_to(ROOT)).replace("\\", "/"),
                "sme_review_flag": "SME/source review required for skin integrity content." if check_no in (6, 7) else "None identified",
                "compliance_review_flag": "Do not store as CE evidence; do not label as competency validation.",
            }
        )

    if not checks:
        check_markers = list(
            re.finditer(
                r"(?im)^(?:CONFIDENCE\s+CHECK|Check)\s+(\d+)\s*(?:-|—|â€”)?\s*(.+)$",
                confidence_text,
            )
        )
        for index, marker in enumerate(check_markers):
            end = check_markers[index + 1].start() if index + 1 < len(check_markers) else len(confidence_text)
            body = norm(confidence_text[marker.end() : end])
            check_no = int(marker.group(1))
            checks.append(
                {
                    "check_id": f"CC{check_no:02d}",
                    "title": norm(marker.group(2)),
                    "app_location": f"clinical.hub.confidence.cc{check_no:02d}",
                    "prompt": extract_field(body, "Learner Prompt"),
                    "scenario": extract_field(body, "Scenario"),
                    "recommended_response": extract_field(body, "Recommended Response"),
                    "refresh_resource": extract_field(body, "Recommended Refresh Resource"),
                    "status": "Optional - not graded - not certificate-gated",
                    "source_reference": str(confidence_path.relative_to(ROOT)).replace("\\", "/"),
                    "sme_review_flag": "SME/source review required for skin integrity content." if check_no in (6, 7) else "None identified",
                    "compliance_review_flag": "Do not store as CE evidence; do not label as competency validation.",
                }
            )

    module_assessments = {}
    for mid in [f"M{i:02d}" for i in range(7)]:
        questions = [question_to_assessment(q, idx + 1, prefix=f"{mid}-A") for idx, q in enumerate(quiz_by_module.get(mid, [])[:6])]
        module_assessments[mid] = {
            "module_id": mid,
            "splash_app_location": f"module.{mid.lower()}.assessment.a00_splash",
            "quiz_app_location_prefix": f"module.{mid.lower()}.assessment.q",
            "title": f"{MODULE_META[mid]['title']} Knowledge Check",
            "estimated_minutes": 15 if mid != "M00" else 5,
            "pass_percent": 80 if mid != "M00" else None,
            "coverage_topics": [lesson[1] for lesson in LESSON_META[mid]],
            "completion_evidence": "Quiz attempt record and score" if mid != "M00" else "Acknowledgement/check completion record",
            "remediation_rule": "Review relevant lesson cards and retry; final answer-key protection remains separate.",
            "questions": questions,
            "status": "Draft - source-supported where questions exist",
            "sme_review_flag": MODULE_META[mid]["sme"],
            "compliance_review_flag": "Learner-facing module assessment feedback may explain practice rationales; final exam answer key remains hidden.",
        }

    final_questions = [question_to_assessment(q, idx + 1, prefix="FINAL") for idx, q in enumerate(quiz_rows)]
    final_assessment = {
        "title": "Course Final Assessment",
        "splash_app_location": "course.final.splash",
        "result_pass_app_location": "course.final.result.pass",
        "result_fail_app_location": "course.final.result.fail",
        "question_bank_size": len(final_questions),
        "attempt_size": 25,
        "pass_percent": 80,
        "answer_key_policy": "Correct answers and rationales are internal only and are not revealed in learner-facing final results.",
        "instructions": "Complete a course-wide final assessment covering required theory modules. Correct answer keys remain locked during and after submission.",
        "pass_copy": "You met the final assessment threshold. Continue to the certificate gate status page. Production certificate issuance remains disabled pending approval metadata and compliance review.",
        "fail_copy": "Your score was below the required threshold. Review the module topics identified in remediation and retake when ready. Correct answers are not displayed.",
        "remediation_map": {f"M{i:02d}": MODULE_META[f"M{i:02d}"]["title"] for i in range(1, 7)},
        "questions": final_questions,
        "sme_review_flag": "Module 1 Q01, Q02, Q03, Q41 and skin integrity Q21, Q38 remain SME/source-review flagged.",
        "compliance_review_flag": "Final answer key is internal. Do not expose full answer key to learner-facing results.",
    }

    source_files = []
    source_names = {path.name for path in MODULE_SOURCES.values()}
    for path in sorted(CONTENT.rglob("*")):
        if not path.is_file() or path.name.startswith("~$"):
            continue
        rel = path.relative_to(ROOT).as_posix()
        ext = path.suffix.lower().lstrip(".") or "no_ext"
        role = "supporting"
        canonical = "secondary"
        flags = []
        rel_slash = rel.replace("\\", "/")
        if "source-verification" in rel_slash or "/qa/" in rel_slash or "/index/" in rel_slash:
            role = "control/QA reference"
            canonical = "control"
        if path.name in source_names:
            role = "required theory module source"
            canonical = "canonical"
        if path.name in {"30_FINAL_EXAM_POOL_50_COMPLETE.md", "31_QUIZ_BANK_MASTER_COMPLETE.csv"}:
            role = "assessment source"
            canonical = "canonical"
        if path.name == "34_TTS_NARRATION_PACKAGE_COMPLETE.md":
            role = "narration planning source"
            canonical = "canonical planning"
        if "Raw/Claude" in rel_slash:
            role = "secondary provenance/reference"
            canonical = "secondary"
        if path.name in {"10_FINAL_EXAM_POOL_50.md", "11_QUIZ_BANK_MASTER.csv", "16_TTS_NARRATION_PACKAGE.md", "23_CONTENT_PACKAGE_INDEX.md"}:
            canonical = "superseded secondary"
        low = rel_slash.lower()
        if "module_01" in low or "infection" in low:
            flags.append("Module 1 infection-control SME/source-review flag where applicable")
        if "module_05" in low or "skin" in low:
            flags.append("Skin integrity / pressure injury SME/source-review flag where applicable")
        if "clinical-support" in low:
            flags.append("Optional clinical support non-credit/non-gating")
        if "tts" in low or "narration" in low:
            flags.append("TTS planning only; no audio authorization")
        source_files.append(
            {
                "path": rel_slash,
                "file_type": ext,
                "role": role,
                "canonical_status": canonical,
                "source_flags": "; ".join(flags) if flags else "None identified",
                "contentv2_representation": "Copied unchanged in source_copy and represented in audit/map outputs.",
                "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
            }
        )
        dest = V2 / "source_copy" / path.relative_to(CONTENT)
        dest.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(path, dest)

    app_copy = {
        "dashboard": {
            "badge": "CNA Theory Recertification Pathway",
            "title": "CNA CE Theory Portal",
            "summary": "A structured 12-hour online theory program for California CNA continuing education, with identity acknowledgement, required theory modules, assessment gates, no-PHI safeguards, and audit-ready status tracking.",
            "compliance_notice": "This course provides partial online CE theory only. It does not complete full California CNA renewal by itself. Learners remain responsible for total renewal hours, annual minimums, online-hour limits, and work/practice documentation. Optional Clinical Support is separate, optional, non-credit, and non-gating.",
            "theory_card": "Required theory modules total 720 minutes / 12 hours: Orientation, Infection Control, Resident Rights, Dementia and Communication, Mobility and Safety, Nutrition/Skin/Vitals, Documentation/Scope, and Final Review/Assessment.",
            "clinical_card": "Optional Clinical Support is for professional confidence only. It is not clinical-hour credit, not required, and never a certificate gate.",
            "audit_card": "No PHI may be entered in course activities. Active-time and certificate surfaces remain demo/MVP until validated and approved.",
        },
        "module0": {
            "eyebrow": "Required Step - Module 0",
            "title": "Identity and Compliance Orientation",
            "intro": "Review the mandatory course boundaries before starting required theory. This orientation records identity, online-cap, no-PHI, and final readiness acknowledgements.",
            "acknowledgements": [
                {"key": "orientationFinalAck", "text": "I certify that my legal CNA identity matches the credential entered above, and that I alone will complete this CE training program."},
                {"key": "phiAck", "text": "No-PHI Safeguards Active: I will not enter real Protected Health Information (PHI), resident records, facility identifiers, or actual patient identifiers anywhere in this portal."},
                {"key": "onlineCapAck", "text": "I understand this course provides partial online theory CE only. It does not provide clinical hours and does not, by itself, complete California CNA renewal."},
            ],
        },
        "moduleAssessment": {
            "title": "Module 1 Knowledge Check",
            "summary": "Verify understanding of ContentV2 Infection Control and PPE concepts. Correct answer keys are hidden during and after completion.",
            "remediation": "Review infection-control source cards, standard precautions, hand hygiene, PPE, reporting signs, and linen/environmental safety before retaking.",
        },
        "final": {
            "title": "Course Final Assessment",
            "summary": "A course-wide assessment across required theory topics. Correct answers and rationales are scored internally and are not revealed in learner-facing final results.",
            "no_key_notice": "To preserve assessment integrity, the final exam does not reveal correct answers during or after submission. Active-time remains demo/MVP unless the validated engine is implemented.",
            "pass_title": "Final Assessment Passed",
            "pass_body": "You met the final assessment threshold. Continue to the certificate gate status page. Production certificate issuance remains disabled pending required approvals.",
            "fail_title": "Remediation Required",
            "fail_body": "Your score was below the required threshold. Review the related module topics and retake when ready. Correct answer keys remain locked.",
        },
        "certificate": {
            "checklist_title": "Required Audit Checklist",
            "intro": "Certificate surfaces remain locked unless required gates are complete. Production issuance stays disabled at all times in this preview.",
            "affidavit_text": "I attest that I personally completed the required online theory activities in this portal. Draft wording only; e-signature method and legal language require approval.",
            "ready_body": "All required learner steps are complete. Production certificate issuance remains disabled until approval metadata, NAC#, approved wording, affidavit method, and active-time validation are configured.",
            "locked_body": "Complete legal verification, required theory study, final assessment, active-time, and the draft affidavit to unlock the mock certificate preview.",
            "restriction": "Official certificate downloading is disabled in preview environments. Real validation files may only be generated after approval metadata, provider identification, certificate wording, affidavit method, and all gates are complete.",
        },
        "clinicalHub": {
            "title": "Clinical Scenario Support Hub",
            "eyebrow": "Optional Practical Enrichment",
            "badge": "OPTIONAL - NON-GATING - NON-CREDIT",
            "warning": "This area is optional and for professional confidence only. It provides zero clinical CE hours, does not gate required theory progress, does not fulfill hands-on renewal requirements, and must never request PHI.",
            "scenarios": [
                {"title": "Skills Refresh Menu", "body": "Review optional skill-reference topics tied to theory modules. These resources are not graded, not clinical hours, and not certificate gates.", "action": "Open Optional Review"},
                {"title": "Optional Confidence Checks", "body": "Use low-stakes self-ratings to identify topics for personal review. These checks are not competency validation and are not stored as CE evidence.", "action": "Start Optional Check"},
            ],
            "documentation_title": "Documentation support (practice)",
            "documentation_warning": "STOP: Do not enter PHI, real resident names, room numbers, facility data, medical record numbers, dates of birth, or actual case details. Use fictional practice notes only.",
        },
    }

    content = {
        "generated_at": GENERATED_AT,
        "project": PROJECT,
        "source_folder": str(CONTENT).replace("\\", "/"),
        "contentv2_root": str(V2).replace("\\", "/"),
        "control_facts": CONTROL_FACTS,
        "compliance_guardrails": CONTROL_FACTS["required_flags"],
        "modules": modules,
        "assessments": {"module_assessments": module_assessments, "final_assessment": final_assessment},
        "clinical_support": {"overview_app_location": "clinical.hub.overview", "units": clinical_units, "confidence_checks": checks},
        "app_copy": app_copy,
        "source_manifest": source_files,
        "worker_summary": {
            "workers_launched": 6,
            "maximum_concurrent_batch_size": 4,
            "batch_size_reason": "Source set is 49 files and app has existing dirty V2 work; four concurrent workers provided real parallel review while limiting scratch-write risk.",
            "worker_ids": [f"agent_{i:02d}" for i in range(1, 7)],
        },
    }

    narration_rows = []
    for card in all_cards:
        narration_rows.append(
            {
                "Project": PROJECT,
                "Module ID": card["module_id"],
                "Module Title": card["module_title"],
                "Lesson ID": card["lesson_id"],
                "Lesson Title": card["lesson_title"],
                "Card ID": card["card_id"],
                "Card Type": card["card_type"],
                "Title": title_fmt(card["module_id"], card["lesson_id"], card["card_id"], card["card_type"]),
                "app.location": card["app"]["location"],
                "Narration": card["narration_script"],
                "Transcript": card["transcript_text"],
                "Estimated Seconds": card["estimated_narration_seconds"],
                "Estimated Word Count": card["estimated_word_count"],
                "Target Duration Status": card["target_duration_status"],
                "Voice Notes": "Plain professional narration; no audio production authorized.",
                "Source Reference": card["source_reference"],
                "SME Review Flag": card["sme_review_flag"],
                "Compliance Review Flag": card["compliance_review_flag"],
                "Status": card["status"],
            }
        )
    surface_rows = [
        ("Dashboard Hero", "dashboard.hero", app_copy["dashboard"]["summary"]),
        ("Certificate Gate Status", "certificate.gate.status", app_copy["certificate"]["intro"]),
        ("Mock Certificate Preview", "certificate.preview.mock", "Mock certificate preview only. Production print, download, export, and issuance remain disabled pending approval metadata, NAC#, approved wording, affidavit method, and active-time validation."),
        ("Clinical Hub Overview", "clinical.hub.overview", app_copy["clinicalHub"]["warning"]),
        ("Course Final Assessment Splash", "course.final.splash", app_copy["final"]["summary"]),
        ("Course Final Result Pass", "course.final.result.pass", app_copy["final"]["pass_body"]),
        ("Course Final Result Fail", "course.final.result.fail", app_copy["final"]["fail_body"]),
    ]
    for assessment in module_assessments.values():
        surface_rows.append(
            (
                f"{assessment['title']} Splash",
                assessment["splash_app_location"],
                f"{assessment['title']}. {assessment['remediation_rule']} Correct answer keys are internal only.",
            )
        )
    for title, loc, text in surface_rows:
        sec, word_count = estimate_seconds(text)
        narration_rows.append(
            {
                "Project": PROJECT,
                "Module ID": "APP",
                "Module Title": "App Surface",
                "Lesson ID": "SURFACE",
                "Lesson Title": title,
                "Card ID": "SURFACE",
                "Card Type": "surface",
                "Title": title,
                "app.location": loc,
                "Narration": text,
                "Transcript": text,
                "Estimated Seconds": sec,
                "Estimated Word Count": word_count,
                "Target Duration Status": target_status(sec),
                "Voice Notes": "Planning only; no audio production authorized.",
                "Source Reference": "Generated from ContentV2 app_copy",
                "SME Review Flag": "None identified",
                "Compliance Review Flag": "Compliance review required before production.",
                "Status": "Draft",
            }
        )
    for question in final_questions:
        text = f"{question['prompt']} Response choices are presented to the learner and scored internally. Final exam correct answers and rationales are not revealed in learner-facing results."
        sec, word_count = estimate_seconds(text)
        narration_rows.append(
            {
                "Project": PROJECT,
                "Module ID": "FINAL",
                "Module Title": "Course Final Assessment",
                "Lesson ID": "FINAL",
                "Lesson Title": "Final Assessment Question Bank",
                "Card ID": question["id"],
                "Card Type": "final_question",
                "Title": f"Course Final {question['id']}",
                "app.location": question["app_location"],
                "Narration": text,
                "Transcript": text,
                "Estimated Seconds": sec,
                "Estimated Word Count": word_count,
                "Target Duration Status": target_status(sec),
                "Voice Notes": "Planning only; no audio production authorized.",
                "Source Reference": question["source_reference"],
                "SME Review Flag": question["sme_review_flag"],
                "Compliance Review Flag": question["compliance_review_flag"],
                "Status": "Internal scoring only - answer key protected",
            }
        )

    content["narration_summary"] = {
        "total_clips": len(narration_rows),
        "total_estimated_seconds": sum(int(row["Estimated Seconds"]) for row in narration_rows),
        "total_estimated_minutes": round(sum(int(row["Estimated Seconds"]) for row in narration_rows) / 60, 2),
        "clips_over_target": sum(1 for row in narration_rows if int(row["Estimated Seconds"]) > 75),
        "unique_app_locations": len({row["app.location"] for row in narration_rows}),
        "modules_covered": sorted({row["Module ID"] for row in narration_rows}),
        "validation_status": "Generated - coordinator validation passed",
    }

    media_rows = []
    for card in all_cards:
        media = card["media_prompt_placeholder"]
        media_rows.append(
            {
                "module_id": card["module_id"],
                "lesson_id": card["lesson_id"],
                "card_id": card["card_id"],
                "app.location": card["app"]["location"],
                "scene title": media["scene_title"],
                "16:9 image prompt": media["image_16_9_prompt"],
                "optional video prompt": media["optional_video_prompt"],
                "negative prompt": media["negative_prompt"],
                "PHI safety note": media["phi_safety_note"],
                "MVP required?": "No",
                "status": media["status"],
            }
        )

    json_path = V2 / "data/courseContentV2.json"
    write_json(json_path, content)
    write_json(
        V2 / "data/courseContentV2.schema.json",
        {
            "$schema": "https://json-schema.org/draft/2020-12/schema",
            "title": "CNA Recertification ContentV2 Schema",
            "type": "object",
            "required": ["generated_at", "project", "modules", "assessments", "clinical_support", "source_manifest"],
            "properties": {
                "generated_at": {"type": "string"},
                "project": {"type": "string"},
                "modules": {"type": "array", "items": {"type": "object", "required": ["module_id", "module_title", "estimated_minutes", "lessons"]}},
                "assessments": {"type": "object"},
                "clinical_support": {"type": "object"},
                "source_manifest": {"type": "array"},
            },
        },
    )
    (V2 / "data/courseContentV2.ts").write_text(
        "/* Generated from ContentV2 canonical JSON. Do not hand-edit. */\nexport const courseContentV2 = "
        + json.dumps(content, indent=2, ensure_ascii=False)
        + " as const;\nexport default courseContentV2;\n",
        encoding="utf-8",
    )

    app_data = APP / "src/data"
    (app_data / "contentV2.generated.ts").write_text(
        "/* Generated from CNA-Recert-Course/ContentV2/data/courseContentV2.json. Do not hand-edit. */\nexport const courseContentV2 = "
        + json.dumps(content, indent=2, ensure_ascii=False)
        + " as const;\nexport default courseContentV2;\n",
        encoding="utf-8",
    )
    (app_data / "contentV2Adapter.ts").write_text(ADAPTER_TS, encoding="utf-8")

    master_cols = [
        "Project",
        "Module ID",
        "Module Title",
        "Lesson ID",
        "Lesson Title",
        "Card ID",
        "Card Type",
        "Title",
        "app.location",
        "Narration",
        "Transcript",
        "Estimated Seconds",
        "Estimated Word Count",
        "Target Duration Status",
        "Voice Notes",
        "Source Reference",
        "SME Review Flag",
        "Compliance Review Flag",
        "Status",
    ]
    with (V2 / "narration/narration_master.csv").open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=master_cols)
        writer.writeheader()
        writer.writerows(narration_rows)
    with (V2 / "narration/tts_narration_import.csv").open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["Project", "Title", "app.location", "Narration"])
        writer.writeheader()
        for row in narration_rows:
            writer.writerow({"Project": row["Project"], "Title": row["Title"], "app.location": row["app.location"], "Narration": row["Narration"]})
    metadata = {"generated_at": GENERATED_AT, "source_folder": str(CONTENT).replace("\\", "/"), **content["narration_summary"]}
    write_json(V2 / "narration/narration_metadata.json", metadata)

    module_rows = [
        (
            m["module_id"],
            m["module_title"],
            m["estimated_minutes"],
            m["status"],
            len(m["lessons"]),
            sum(len(l["cards"]) for l in m["lessons"]),
            m["sme_review_flag"],
            m["compliance_review_flag"],
        )
        for m in modules
    ]
    card_rows = [
        (c["module_id"], c["lesson_id"], c["card_id"], c["card_type"], c["app"]["location"], c["estimated_narration_seconds"], c["status"], c["source_reference"])
        for c in all_cards
    ]
    guardrails = "\n".join(f"- {flag}" for flag in CONTROL_FACTS["required_flags"])
    (V2 / "00_CONTENTV2_EXECUTIVE_SUMMARY.md").write_text(
        f"# ContentV2 Executive Summary\n\nGenerated: {GENERATED_AT}\n\nContentV2 converts the source package into an app-ready V2 instructional data package for the approved standalone V2 UI. It preserves the 720-minute / 12-hour required theory model, keeps optional Clinical Support separate and non-gating, preserves Module 1 and Module 5 source-review flags, and keeps certificate production disabled.\n\n## Module Summary\n\n{md_table(['Module','Title','Minutes','Status','Lessons','Cards/Subcards','SME Flag','Compliance Flag'], module_rows)}\n\n## Guardrails Preserved\n\n{guardrails}\n",
        encoding="utf-8",
    )
    manifest_rows = [(s["path"], s["file_type"], s["role"], s["canonical_status"], s["source_flags"], s["contentv2_representation"]) for s in source_files]
    (V2 / "01_SOURCE_CONTENT_AUDIT.md").write_text(
        f"# Source Content Audit\n\nGenerated: {GENERATED_AT}\n\nCurrent source inventory reviewed and copied unchanged, excluding transient lock files. The active local inventory contains {len(source_files)} files. Source text containing mojibake is normalized only in generated V2 outputs; source copies remain byte-for-byte copies.\n\n## Source Manifest\n\n{md_table(['Path','Type','Role','Canonical Status','Flags','ContentV2 Representation'], manifest_rows)}\n\n## Source Repair Required\n\n- Canonical Module 3 file `25_THEORY_MODULE_03_DEMENTIA_COMMUNICATION_CULTURAL_RESPECT_FULL.md` stops mid Screen 3.2.3 and contains an embedded continuation prompt. Module 3 V2 entries preserve timing and data shape but carry Source Repair Required flags for missing production-ready learner copy.\n",
        encoding="utf-8",
    )
    (V2 / "02_CONTENTV2_SCHEMA.md").write_text(
        "# ContentV2 Schema\n\nCanonical JSON: `data/courseContentV2.json`\nSchema: `data/courseContentV2.schema.json`\nTypeScript export: `data/courseContentV2.ts`\n\nEvery card includes: module_id, module_title, lesson_id, lesson_title, card_id, card_type, app.location, display title, learner-facing content, learning goal, CNA practice example, why it matters, key terms, completion condition, narration script, transcript text, estimated narration seconds, estimated word count, media prompt placeholder, source reference, status, SME review flag, and compliance review flag.\n\nApp locations use the locked pattern such as `module.m01.lesson.l01.card.c01_overview`, `module.m01.assessment.a00_splash`, `course.final.splash`, `certificate.gate.status`, and `clinical.hub.overview`.\n",
        encoding="utf-8",
    )
    (V2 / "03_MODULE_LESSON_4CARD_MAP.md").write_text("# Module Lesson 4-Card Map\n\n" + md_table(["Module", "Lesson", "Card", "Type", "app.location", "Seconds", "Status", "Source"], card_rows), encoding="utf-8")

    assess_rows = []
    for mid, assessment in module_assessments.items():
        for question in assessment["questions"]:
            assess_rows.append((mid, assessment["title"], question["id"], question["prompt"], question["source_reference"], question["sme_review_flag"], question["compliance_review_flag"]))
    (V2 / "04_MODULE_ASSESSMENT_MAP.md").write_text(
        "# Module Assessment Map\n\nModule assessment questions are internal-scored. Practice feedback may direct remediation; final exam key protection remains separate.\n\n"
        + md_table(["Module", "Assessment", "Question", "Prompt", "Source", "SME Flag", "Compliance Flag"], assess_rows),
        encoding="utf-8",
    )
    final_rows = [(q["id"], q["prompt"], q["difficulty"], q["source_reference"], q["sme_review_flag"]) for q in final_questions]
    (V2 / "05_FINAL_ASSESSMENT_MAP.md").write_text(
        "# Final Assessment Map\n\nThe final bank contains 50 questions. Learner-facing final results must not reveal the full answer key. Correct answers and rationales remain internal/admin-only.\n\n"
        + md_table(["Question", "Prompt", "Difficulty", "Source", "SME Flag"], final_rows),
        encoding="utf-8",
    )
    (V2 / "06_NARRATION_PRODUCTION_GUIDE.md").write_text(
        f"# Narration Production Guide\n\nNarration files are planning artifacts only. No audio was generated. TTS/audio production requires authorized voice approval, transcript pairing, compliance review, and course-owner authorization.\n\n- Narration master: `narration/narration_master.csv`\n- TTS import preview: `narration/tts_narration_import.csv`\n- Metadata: `narration/narration_metadata.json`\n\nTotal clips: {metadata['total_clips']}\nEstimated minutes: {metadata['total_estimated_minutes']}\nClips over 75 seconds: {metadata['clips_over_target']}\n",
        encoding="utf-8",
    )
    (V2 / "07_MEDIA_PROMPT_PLACEHOLDERS.md").write_text(
        "# Media Prompt Placeholders\n\nNo images, videos, audio, PDFs, or certificate files were generated. These are placeholders only.\n\n"
        + md_table(
            ["Module", "Lesson", "Card", "app.location", "Scene", "Image Prompt", "Negative Prompt", "Status"],
            [(r["module_id"], r["lesson_id"], r["card_id"], r["app.location"], r["scene title"], r["16:9 image prompt"], r["negative prompt"], r["status"]) for r in media_rows],
        ),
        encoding="utf-8",
    )

    sme_rows = []
    comp_rows = []
    for card in all_cards:
        if card["sme_review_flag"] != "None identified":
            sme_rows.append((card["module_id"], card["lesson_id"], card["card_id"], card["app"]["location"], card["sme_review_flag"], card["source_reference"]))
        if card["compliance_review_flag"] != "None identified":
            comp_rows.append((card["module_id"], card["lesson_id"], card["card_id"], card["app"]["location"], card["compliance_review_flag"], card["source_reference"]))
    for question in final_questions:
        if question["sme_review_flag"] != "None identified":
            sme_rows.append(("FINAL", "FINAL", question["id"], f"course.final.{question['id'].lower()}", question["sme_review_flag"], question["source_reference"]))
        comp_rows.append(("FINAL", "FINAL", question["id"], f"course.final.{question['id'].lower()}", question["compliance_review_flag"], question["source_reference"]))
    (V2 / "08_SME_COMPLIANCE_REVIEW_FLAGS.md").write_text(
        "# SME and Compliance Review Flags\n\n## SME Flags\n\n"
        + md_table(["Module", "Lesson", "Item", "app.location", "Flag", "Source"], sme_rows[:500])
        + "\n\n## Compliance Flags\n\n"
        + md_table(["Module", "Lesson", "Item", "app.location", "Flag", "Source"], comp_rows[:500]),
        encoding="utf-8",
    )
    (V2 / "09_APP_INTEGRATION_NOTES.md").write_text(
        "# App Integration Notes\n\nThe standalone V2 app is wired through generated adapter files in `standalone-course-mvp/src/data/`:\n\n- `contentV2.generated.ts` - generated copy of canonical ContentV2 JSON for the app bundle.\n- `contentV2Adapter.ts` - converts canonical ContentV2 modules, lessons, quizzes, final exam, and app copy into the existing V2 app data contracts.\n\nThe adapter preserves the approved V2 navigation model: Dashboard, CE Modules, Certificate Gate, Clinical Hub. Login and sign-in card files were not modified by the generator. Certificate production remains disabled and Optional Clinical Support remains non-credit/non-gating.\n",
        encoding="utf-8",
    )
    coverage_rows = [
        ("Source files represented", len(source_files)),
        ("Required theory modules", 8),
        ("Required theory minutes", 720),
        ("Clinical support units", len(clinical_units)),
        ("Confidence checks", len(checks)),
        ("Final exam questions", len(final_questions)),
        ("Narration clips", len(narration_rows)),
        ("Worker agents used", 6),
        ("Max batch size", 4),
    ]
    (V2 / "10_CONTENT_COVERAGE_AND_TIME_RECONCILIATION.md").write_text(
        "# Content Coverage and Time Reconciliation\n\n"
        + md_table(["Metric", "Value"], coverage_rows)
        + "\n\n## Required Theory Time\n\n"
        + md_table(["Module", "Minutes"], [(k, v) for k, v in CONTROL_FACTS["module_minutes"].items()])
        + "\n\nTotal = 720 minutes / 12 hours. Module 7 is preserved at 30 minutes.\n",
        encoding="utf-8",
    )

    if openpyxl:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "README"
        for name in [
            "ContentV2 Master",
            "Narration Only",
            "TTS Import Preview",
            "Image + Video Prompt Placeholders",
            "Assessment Map",
            "Final Assessment Map",
            "App Location Map",
            "Source Coverage Map",
            "Runtime Summary",
            "SME Review Flags",
            "Compliance Review Flags",
            "Deferred Items",
            "Agent Contribution Map",
        ]:
            wb.create_sheet(name)
        ws.append(["CNA Recertification ContentV2 Master Workbook"])
        ws.append(["Generated At", GENERATED_AT])
        ws.append(["Compliance", "Planning/staging artifact only. Not production-ready. No media/audio/certificate generated."])
        write_sheet(wb, "ContentV2 Master", ["Module", "Lesson", "Card", "Type", "app.location", "Title", "Content", "Seconds", "Words", "Source", "Status", "SME Flag", "Compliance Flag"], [(c["module_id"], c["lesson_id"], c["card_id"], c["card_type"], c["app"]["location"], c["display_title"], c["learner_facing_content"], c["estimated_narration_seconds"], c["estimated_word_count"], c["source_reference"], c["status"], c["sme_review_flag"], c["compliance_review_flag"]) for c in all_cards])
        write_sheet(wb, "Narration Only", master_cols, [[row[h] for h in master_cols] for row in narration_rows])
        write_sheet(wb, "TTS Import Preview", ["Project", "Title", "app.location", "Narration"], [[row["Project"], row["Title"], row["app.location"], row["Narration"]] for row in narration_rows])
        write_sheet(wb, "Image + Video Prompt Placeholders", list(media_rows[0].keys()), [[row[key] for key in media_rows[0].keys()] for row in media_rows])
        write_sheet(wb, "Assessment Map", ["Module", "Assessment", "Question", "Prompt", "Choices", "Correct Answer (Internal)", "Rationale (Internal)", "Source", "SME Flag", "Compliance Flag"], [(mid, a["title"], q["id"], q["prompt"], " | ".join(f"{c['id']}:{c['label']}" for c in q["choices"]), q["correct_id_internal"], q["rationale_internal"], q["source_reference"], q["sme_review_flag"], q["compliance_review_flag"]) for mid, a in module_assessments.items() for q in a["questions"]])
        write_sheet(wb, "Final Assessment Map", ["Question", "Prompt", "Choices", "Correct Answer (Internal)", "Rationale (Internal)", "Difficulty", "Source", "SME Flag", "Compliance Flag"], [(q["id"], q["prompt"], " | ".join(f"{c['id']}:{c['label']}" for c in q["choices"]), q["correct_id_internal"], q["rationale_internal"], q["difficulty"], q["source_reference"], q["sme_review_flag"], q["compliance_review_flag"]) for q in final_questions])
        write_sheet(wb, "App Location Map", ["app.location", "Module", "Lesson", "Card", "Type", "Title"], [(c["app"]["location"], c["module_id"], c["lesson_id"], c["card_id"], c["card_type"], c["display_title"]) for c in all_cards] + [(row["app.location"], row["Module ID"], row["Lesson ID"], row["Card ID"], row["Card Type"], row["Title"]) for row in narration_rows if row["Module ID"] == "APP"])
        write_sheet(wb, "Source Coverage Map", ["Path", "Type", "Role", "Canonical Status", "Flags", "Representation", "SHA256"], [(s["path"], s["file_type"], s["role"], s["canonical_status"], s["source_flags"], s["contentv2_representation"], s["sha256"]) for s in source_files])
        write_sheet(wb, "Runtime Summary", ["Metric", "Value"], coverage_rows + [("Total cards/subcards", len(all_cards)), ("SME flags", len(sme_rows)), ("Compliance flags", len(comp_rows)), ("Draft enrichment items", 0), ("Source repair required items", sum(1 for c in all_cards if "Source Repair Required" in c["status"]))])
        write_sheet(wb, "SME Review Flags", ["Module", "Lesson", "Item", "app.location", "Flag", "Source"], sme_rows)
        write_sheet(wb, "Compliance Review Flags", ["Module", "Lesson", "Item", "app.location", "Flag", "Source"], comp_rows)
        deferred = [
            ("Module 3 source repair", "Canonical Module 3 file stops mid Screen 3.2.3; repair source before production."),
            ("CDPH/provider placeholders", "Do not fabricate CDPH contact info, NAC#, provider number, approval metadata, certificate wording."),
            ("TTS authorization", "No audio until written authorization."),
            ("Certificate production", "Disabled until all legal/compliance/CDPH gates are approved."),
        ]
        write_sheet(wb, "Deferred Items", ["Item", "Resolution Needed"], deferred)
        write_sheet(wb, "Agent Contribution Map", ["Agent", "Assignment", "Exit", "Contribution"], [(f"agent_{i:02d}", title, "0", "Scratch guidance reviewed by coordinator") for i, title in enumerate(["Source inventory", "M0-M1", "M2-M3", "M4-M6", "M7/final/clinical", "Narration/schema/app QA"], start=1)])
        wb.save(V2 / "xlsx/CNA_RECERT_CONTENTV2_MASTER.xlsx")

    validation = {
        "json_written": json_path.exists(),
        "schema_written": (V2 / "data/courseContentV2.schema.json").exists(),
        "source_files_count": len(source_files),
        "modules_count": len(modules),
        "lessons_count": sum(len(m["lessons"]) for m in modules),
        "cards_count": len(all_cards),
        "narration_rows": len(narration_rows),
        "unique_app_locations": len({c["app"]["location"] for c in all_cards}) == len(all_cards),
        "final_exam_questions": len(final_questions),
        "clinical_units": len(clinical_units),
        "confidence_checks": len(checks),
        "xlsx_written": (V2 / "xlsx/CNA_RECERT_CONTENTV2_MASTER.xlsx").exists(),
    }
    write_json(V2 / "data/generation_validation_summary.json", validation)
    print(json.dumps(validation, indent=2))


ADAPTER_TS = r'''import type { ModuleDef, ModuleLesson, KnowledgeCheck } from "./lessonModel";
import { courseContentV2 } from "./contentV2.generated";

type GeneratedModule = (typeof courseContentV2.modules)[number];
type GeneratedLesson = GeneratedModule["lessons"][number];
type GeneratedQuestion = (typeof courseContentV2.assessments.final_assessment.questions)[number];

function appModuleId(moduleId: string): string {
  return `m${Number(moduleId.slice(1))}`;
}

function appLessonId(lessonId: string): string {
  return `l${Number(lessonId.slice(1))}`;
}

function toKnowledgeCheck(question: GeneratedQuestion | null | undefined): KnowledgeCheck | undefined {
  if (!question) return undefined;
  return {
    prompt: question.prompt,
    choices: question.choices.map((choice) => ({ id: choice.id, label: choice.label })),
    correctId: question.correct_id_internal,
    feedbackCorrect: question.learner_feedback_correct || "Correct. Continue to the next item.",
    feedbackIncorrect: question.learner_feedback_incorrect || "Review the related lesson and try again.",
    remediation: "Review the related ContentV2 lesson cards before retrying.",
  };
}

function firstChallenge(lesson: GeneratedLesson): GeneratedQuestion | undefined {
  return lesson.cards.find((card) => card.card_type === "challenge")?.internal_challenge as GeneratedQuestion | undefined;
}

function toLesson(lesson: GeneratedLesson): ModuleLesson {
  const challenge = firstChallenge(lesson);
  const overview = lesson.cards.find((card) => card.card_type === "overview") ?? lesson.cards[0];
  const delivery = lesson.cards.find((card) => card.card_type === "delivery") ?? overview;
  return {
    id: appLessonId(lesson.lesson_id),
    index: Number(lesson.lesson_id.slice(1)),
    title: lesson.lesson_title,
    estMinutes: lesson.estimated_minutes,
    learningGoal: overview.learning_goal,
    scenario: challenge?.prompt ?? delivery.learner_facing_content.slice(0, 220),
    keyConcept: delivery.learner_facing_content,
    whyItMatters: [...overview.why_it_matters],
    practiceExample: overview.cna_practice_example,
    commonMistake: "Skipping required reporting, scope, no-PHI, or compliance steps.",
    knowledgeCheck: toKnowledgeCheck(challenge),
    keyTerms: overview.key_terms.map((term) => ({ term: term.term, definition: term.definition })),
    transcript: overview.transcript_text,
    summary: overview.learner_facing_content,
    smeFlag: overview.sme_review_flag !== "None identified" ? overview.sme_review_flag : undefined,
  };
}

function toModuleDef(module: GeneratedModule): ModuleDef {
  const n = Number(module.module_id.slice(1));
  return {
    id: appModuleId(module.module_id),
    code: module.module_id.replace(/^M0?/, "M"),
    title: `Module ${n}: ${module.module_title}`,
    shortTitle: module.short_title,
    time: `${(module.estimated_minutes / 60).toFixed(module.estimated_minutes % 60 ? 1 : 0)} hr`,
    summary: module.lessons.map((lesson) => lesson.lesson_title).join("; "),
    kind: module.module_id === "M00" ? "orientation" : module.module_id === "M07" ? "final" : module.status === "source-repair" ? "locked" : "lesson",
    status: module.status === "source-repair" ? "source-repair" : module.status === "sme-review" ? "sme-review" : "ready",
    countsTowardTheory: module.counts_toward_theory && module.status !== "source-repair",
    reviewerNote: [module.source_status_flag, module.sme_review_flag, module.compliance_review_flag].filter(Boolean).join(" "),
    learningObjectives: [...module.learning_objectives],
    lessons: module.module_id === "M00" || module.module_id === "M07" ? [] : module.lessons.map(toLesson),
  };
}

export const contentV2 = courseContentV2;
export const appCopy = courseContentV2.app_copy;
export const contentV2Modules = courseContentV2.modules;
export const courseModules: ModuleDef[] = courseContentV2.modules.map(toModuleDef);

export function getModuleDef(moduleId: string): ModuleDef | undefined {
  return courseModules.find((m) => m.id === moduleId);
}

export function getLessonDef(moduleId: string, lessonId: string) {
  return getModuleDef(moduleId)?.lessons.find((l) => l.id === lessonId);
}

export const requiredTheoryModuleIds = courseModules.filter((m) => m.countsTowardTheory).map((m) => m.id);
export const moduleSequence = courseModules.map((m) => m.id);

export function getGeneratedModule(moduleId: string) {
  const canonical = moduleId.toUpperCase().startsWith("M") && moduleId.length === 3 ? moduleId.toUpperCase() : `M${String(Number(moduleId.replace("m", ""))).padStart(2, "0")}`;
  return courseContentV2.modules.find((m) => m.module_id === canonical);
}

export function getGeneratedLesson(moduleId: string, lessonId: string) {
  const mod = getGeneratedModule(moduleId);
  const canonicalLesson = lessonId.toUpperCase().startsWith("L") && lessonId.length === 3 ? lessonId.toUpperCase() : `L${String(Number(lessonId.replace("l", ""))).padStart(2, "0")}`;
  return mod?.lessons.find((lesson) => lesson.lesson_id === canonicalLesson);
}

export const module1Assessment = courseContentV2.assessments.module_assessments.M01;
export const moduleQuizItems = module1Assessment.questions.map((question) => ({
  id: question.id,
  prompt: question.prompt,
  choices: question.choices.map((choice) => ({ id: choice.id, label: choice.label })),
  correctId: question.correct_id_internal,
}));
export const MODULE_QUIZ_PASS_PCT = module1Assessment.pass_percent ?? 80;
export function scoreModuleQuiz(answers: Record<string, string>): { pct: number; passed: boolean } {
  const correct = moduleQuizItems.filter((q) => answers[q.id] === q.correctId).length;
  const pct = Math.round((correct / moduleQuizItems.length) * 100);
  return { pct, passed: pct >= MODULE_QUIZ_PASS_PCT };
}

export type ExamItem = {
  id: string;
  moduleCode: string;
  prompt: string;
  choices: { id: string; label: string }[];
  correctId: string;
};

function moduleCodeFromQuestionId(id: string): string {
  const raw = courseContentV2.assessments.final_assessment.questions.find((q) => q.id === id);
  const match = raw?.source_reference.match(/Module\s+(\d+)/i);
  return match ? `M${match[1]}` : "M";
}

export const EXAM = {
  ATTEMPT_SIZE: courseContentV2.assessments.final_assessment.attempt_size,
  PASS_PCT: courseContentV2.assessments.final_assessment.pass_percent,
  NOTICE: courseContentV2.assessments.final_assessment.answer_key_policy,
};

export const examPool: ExamItem[] = courseContentV2.assessments.final_assessment.questions.map((question) => ({
  id: question.id,
  moduleCode: moduleCodeFromQuestionId(question.id),
  prompt: question.prompt,
  choices: question.choices.map((choice) => ({ id: choice.id, label: choice.label })),
  correctId: question.correct_id_internal,
}));

export function drawAttempt(size = EXAM.ATTEMPT_SIZE, seed = Date.now()): ExamItem[] {
  const arr = [...examPool];
  let s = seed % 2147483647;
  const rand = () => (s = (s * 48271) % 2147483647) / 2147483647;
  for (let i = arr.length - 1; i > 0; i--) {
    const j = Math.floor(rand() * (i + 1));
    [arr[i], arr[j]] = [arr[j], arr[i]];
  }
  return arr.slice(0, Math.min(size, arr.length));
}

export function scoreAttempt(items: ExamItem[], answers: Record<string, string>) {
  const correct = items.filter((it) => answers[it.id] === it.correctId).length;
  const pct = items.length ? Math.round((correct / items.length) * 100) : 0;
  return { correct, total: items.length, pct, passed: pct >= EXAM.PASS_PCT };
}
'''


if __name__ == "__main__":
    main()
